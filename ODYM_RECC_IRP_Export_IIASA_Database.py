# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for parsing and checking in the IIASA IRP database.

See https://github.com/iiasa/irp-internal-workflow for naming conventions.

"""
# Import required libraries:
import os
import xlrd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt 
import pylab
import pandas as pd
import uuid
import shutil

import RECC_Paths # Import path file

# Create UUID of script run
Current_UUID = str(uuid.uuid4())
RECC_Paths.results_path_save = os.path.join(RECC_Paths.results_path_eval,'RECC_Results_' + Current_UUID)
if not os.path.exists(RECC_Paths.results_path_save): # Create scrip run results directory.
    os.makedirs(RECC_Paths.results_path_save)

# Set sector
Sect = 'pav'
#Sect = 'reb'

# Definitions
SpecPath      = os.path.join(RECC_Paths.results_path,'IRP - IIASA database variable template proposal_13_06_21.xlsx') 
SpecFile      = xlrd.open_workbook(SpecPath)
if Sect == 'reb':
    SpecSheet     = SpecFile.sheet_by_name('RECC_Export_reb')
if Sect == 'pav':
    SpecSheet     = SpecFile.sheet_by_name('RECC_Export_pav')

IIASA_ScNames  = []
RECC_ScFolders = []
RECC_ScIndex   = []
r = 2
while True: # Scan column for content until either sheet ends or cell content is empty string ''
    try:
        if len(SpecSheet.cell_value(r, 1)) > 0:
            IIASA_ScNames.append(SpecSheet.cell_value(r, 1))
            RECC_ScFolders.append(SpecSheet.cell_value(r, 2))
            RECC_ScIndex.append(int(SpecSheet.cell_value(r, 3)))
    except:
             break
    r += 1
    
IIASA_IndNames = []
IIASA_IndUnits = []
RECC_SysDefLoc = []
RECC_Ind_Unit  = []
RECC_ConvFactor= []
RECC_Res_Match = []
r = 2
while True: # Scan column for content until either sheet ends or cell content is empty string ''
    try:
        if len(SpecSheet.cell_value(r, 4)) > 0:
            IIASA_IndNames.append( SpecSheet.cell_value(r, 4))
            IIASA_IndUnits.append( SpecSheet.cell_value(r, 5))
            RECC_SysDefLoc.append( SpecSheet.cell_value(r, 6))
            RECC_Ind_Unit.append(  SpecSheet.cell_value(r, 7))
            RECC_ConvFactor.append(SpecSheet.cell_value(r, 8))
            RECC_Res_Match.append([])
    except:
             break
    r += 1

for m in range(0,len(RECC_Res_Match)):
    r = 9
    while True: # Scan columns for content
        try:
            if len(SpecSheet.cell_value(m+2, r)) > 0:
                RECC_Res_Match[m].append(SpecSheet.cell_value(m+2, r))
        except:
                 break
        r += 1

# Define result arrays and write header
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = Sect # pav or reb
#ws1.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
ws1.cell(row=1, column=1).value = 'model'
ws1.cell(row=1, column=2).value = 'scenario'
ws1.cell(row=1, column=3).value = 'region'
ws1.cell(row=1, column=4).value = 'variable'
ws1.cell(row=1, column=5).value = 'unit'
ws1.cell(row=1, column=6).value = 'subannual'
for m in range(0,46):
    ws1.cell(row=1, column=7+m).value = m+2015

# Gather data in a triple loop: for all scenarios, for all IIASA indicators, and for all n:1 RECC->IIASA Indicators
ri = 2    
for S in range(0,len(IIASA_ScNames)): # IIASA scenario index
    # Open scenario result file:
    Path = os.path.join(RECC_Paths.results_path,RECC_ScFolders[S],'SysVar_TotalGHGFootprint.xls')
    Resultfile   = xlrd.open_workbook(Path)
    Resultsheet  = Resultfile.sheet_by_name('TotalGHGFootprint')
    Resultsheet1 = Resultfile.sheet_by_name('Cover')
    UUID         = Resultsheet1.cell_value(3,2)
    Resultfile2  = xlrd.open_workbook(os.path.join(RECC_Paths.results_path,RECC_ScFolders[S],'ODYM_RECC_ModelResults_' + UUID + '.xlsx'))
    Resultsheet2 = Resultfile2.sheet_by_name('Model_Results')   
    for I in range(0,len(IIASA_IndNames)): # IIASA indicator index
        # Write line:
        ws1.cell(row=ri, column=1).value = 'ODYM-RECC v2.4'
        ws1.cell(row=ri, column=2).value = IIASA_ScNames[S]
        ws1.cell(row=ri, column=3).value = 'Global'
        ws1.cell(row=ri, column=4).value = IIASA_IndNames[I]
        ws1.cell(row=ri, column=5).value = IIASA_IndUnits[I]
        ws1.cell(row=ri, column=6).value = 'Year'
        # Collect data
        NuLi = np.zeros((1,46)) # Numbers for this line
        for N in range(0,len(RECC_Res_Match[I])):
            # First, get the position indices for the different result variables:
            rowi = 1
            while True:
                if Resultsheet2.cell_value(rowi, 0) == RECC_Res_Match[I][N]:
                    break
                rowi += 1
            for t in range(0,46): # time until 2060
                NuLi[0,t] += Resultsheet2.cell_value(rowi + RECC_ScIndex[S],t+7) * RECC_ConvFactor[I]
    
        # Write data to transfer file          
        for Y in range(0,46):
            ws1.cell(row=ri, column=Y+7).value = NuLi[0,Y]
        ri +=1
    

# Export list data
book.save(os.path.join(RECC_Paths.results_path_save,'RECC_ModelResults_IIASA_Export_'+ Sect + '_' + Current_UUID + '.xlsx'))               
                 
                 
#
#
#