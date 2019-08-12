# -*- coding: utf-8 -*-
"""
Created on Fri Jun 14 05:18:48 2019

@author: spauliuk
"""

"""

File RECC_ScenarioEvaluate.py

Script that runs the sensitivity and scnenario comparison scripts for different settings.

"""

# Import required libraries:
import os
import xlrd
import openpyxl

import RECC_Paths # Import path file
import RECC_G7IC_Cascade_PassVehicles_V2_0
import RECC_G7IC_Cascade_ResBuildings_V2_0
import RECC_G7IC_Sensitivity_PassVehicles_V2_0
import RECC_G7IC_Sensitivity_ResBuildings_V2_0

#ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
#ScenarioSetting = 'Evaluate_RECC_Cascade'
#ScenarioSetting = 'Evaluate_GroupTestRun'
ScenarioSetting = 'Evaluate_RECC_Sensitivity'

# open scenario sheet
ModelConfigListFile  = xlrd.open_workbook(os.path.join(RECC_Paths.recc_path,'RECC_ModelConfig_List_V2_1.xlsx'))
ModelEvalListSheet = ModelConfigListFile.sheet_by_name(ScenarioSetting)

# open result summary file
mywb = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_Template.xlsx'))

#Read control lines and execute main model script
Row  = 1

NoofCascadeSteps_pav     = 7 # 7 for vehs. and 6 for buildings
NoofCascadeSteps_reb     = 6 # 7 for vehs. and 6 for buildings
NoofSensitivitySteps_pav = 11 # no of different sensitivity analysis cases for pav
NoofSensitivitySteps_reb = 10 # no of different sensitivity analysis cases for reb

# search for script config list entry
while ModelEvalListSheet.cell_value(Row, 1) != 'ENDOFLIST':
    if ModelEvalListSheet.cell_value(Row, 1) != '':
        PassVehList = []
        ResBldsList = []
        RegionalScope = ModelEvalListSheet.cell_value(Row, 1)
        Setting       = ModelEvalListSheet.cell_value(Row, 2) # cascade or sensitivity
        print(RegionalScope)
        
    if Setting == 'Cascade_pav':
        for m in range(0,int(NoofCascadeSteps_pav)):
            PassVehList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC scenario comparison
        ASummaryV, AvgDecadalEmsV = RECC_G7IC_Cascade_PassVehicles_V2_0.main(RegionalScope,PassVehList)
        # write results summary to Excel
        Vsheet = mywb[RegionalScope + '_Vehicles']
        for r in range(0,3):
            for c in range(0,7):
                Vsheet.cell(row = r+3,  column = c+5).value  = ASummaryV[r,c]       
                Vsheet.cell(row = r+9,  column = c+5).value  = ASummaryV[r+3,c]       
                Vsheet.cell(row = r+15, column = c+5).value = ASummaryV[r+6,c]
                for d in range(0,4):
                    Vsheet.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalEmsV[r,c,d]
        
    if Setting == 'Cascade_reb':
        for m in range(0,int(NoofCascadeSteps_reb)):
            ResBldsList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC scenario comparison
        ASummaryB, AvgDecadalEmsB = RECC_G7IC_Cascade_ResBuildings_V2_0.main(RegionalScope,ResBldsList)
        # write results summary to Excel
        Bsheet = mywb[RegionalScope + '_Buildings']
        for r in range(0,3):
            for c in range(0,6):
                Bsheet.cell(row = r+3, column = c+5).value  = ASummaryB[r,c]       
                Bsheet.cell(row = r+9, column = c+5).value  = ASummaryB[r+3,c]       
                Bsheet.cell(row = r+15, column = c+5).value = ASummaryB[r+6,c]     
                for d in range(0,4):
                    Bsheet.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalEmsB[r,c,d]
                    
    if Setting == 'Sensitivity_pav':
        for m in range(0,int(NoofSensitivitySteps_pav)):
            PassVehList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC sensitivity analysis for pav
        CumEmsV_Sens, AnnEmsV2030_Sens, AnnEmsV2050_Sens, AvgDecadalEms = RECC_G7IC_Sensitivity_PassVehicles_V2_0.main(RegionalScope,PassVehList)        
        # write results summary to Excel
        Ssheet = mywb['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,11):
                Ssheet.cell(row = r+4, column = c+6).value  = AnnEmsV2030_Sens[r,c]
                Ssheet.cell(row = r+9, column = c+6).value  = AnnEmsV2050_Sens[r,c]
                Ssheet.cell(row = r+14,column = c+6).value  = CumEmsV_Sens[r,c]
                for d in range(0,4):
                    Ssheet.cell(row = d*3 + r + 19,column = c+6).value  = AvgDecadalEms[r,c,d]
                    
    if Setting == 'Sensitivity_reb':
        for m in range(0,int(NoofSensitivitySteps_reb)):
            ResBldsList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC sensitivity analysis for reb
        CumEmsB_Sens, AnnEmsB2030_Sens, AnnEmsB2050_Sens, AvgDecadalEms = RECC_G7IC_Sensitivity_ResBuildings_V2_0.main(RegionalScope,ResBldsList)        
        # write results summary to Excel
        Ssheet = mywb['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,10):
                Ssheet.cell(row = r+35,column = c+6).value  = AnnEmsB2030_Sens[r,c]
                Ssheet.cell(row = r+40,column = c+6).value  = AnnEmsB2050_Sens[r,c]
                Ssheet.cell(row = r+45,column = c+6).value  = CumEmsB_Sens[r,c]        
                for d in range(0,4):
                    Ssheet.cell(row = d*3 + r + 50,column = c+6).value  = AvgDecadalEms[r,c,d]   
                    
    # forward counter   
    if Setting == 'Cascade_pav':
        Row += NoofCascadeSteps_pav
    if Setting == 'Cascade_reb':
        Row += NoofCascadeSteps_reb        
    if Setting == 'Sensitivity_pav':
        Row += NoofSensitivitySteps_pav
    if Setting == 'Sensitivity_reb':
        Row += NoofSensitivitySteps_reb
        
    
mywb.save(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_12August2019S.xlsx'))
    
    
    
    
#
#