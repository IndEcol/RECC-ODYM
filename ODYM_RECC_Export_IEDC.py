# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for parsing and checking in the IIASA IRP database.

See https://github.com/iiasa/irp-internal-workflow for naming conventions.

"""
# Import required libraries:
import os
import openpyxl
import numpy as np
import uuid
import datetime
import shutil
import RECC_Paths # Import path file

def TableSectionToDict(Sheet,r,c):
    """
    Parameters
    ----------
    Sheet : Pointer to Excel worksheet (openpyxl)
    r : row index of upper left corner
    c : col index of upper left corner

    Returns
    -------
    res : Dictionary of lists, where list names are the labels of the first row and the list entries are the respective entries below in each row.

    """
    res = {}
    while True:
        Head = SpecsS.cell(r,c).value
        if Head is not None:
            res[Head] = []
            i = 1
            while True:
                Item = SpecsS.cell(r+i,c).value
                if Item is not None:
                    res[Head].append(Item)                
                    i +=1
                else:
                    break
            c +=1
        else: 
            break
    return res

# Create UUID of script run
Current_UUID = str(uuid.uuid4())

# Common definitions
datestring = datetime.datetime.today().strftime('%Y-%m-%d')
ScenLabels = ['LED','SSP1','SSP2']

# Define list with all RECC->iedc datasets to be exported, (un)comment those to be (not) exported.
DSList = [] 
#DSList.append('1_F_MaterialProduction_RECCv2.4')
#DSList.append('1_F_GHG_BySector_RECCv2.4')
#DSList.append('2_IUS_ResidentlBldngs_RECCv2.4')
DSList.append('1_F_MaterialFlows_RECCv2.4')

# open file with export specifications
ExportSpecsFile  = openpyxl.load_workbook(os.path.join(RECC_Paths.iedc_export_path,'Export_Define_RECC_v2.4.xlsx'))

# iterate over all datasets selected above
for DS in DSList:
    print('   ')
    print('Exporting values for ' + DS)
    # copy and rename template, open and add data cover information
    shutil.copy(os.path.join(RECC_Paths.iedc_export_path,'IEDC_RECCv2.4_Template.xlsx'), os.path.join(RECC_Paths.iedc_export_path,DS + '.xlsx'))
    SpecsS     = ExportSpecsFile[DS]
    ExportFile = openpyxl.load_workbook(os.path.join(RECC_Paths.iedc_export_path, DS + '.xlsx'))    
    CoverS     = ExportFile['Cover']
    DataS      = ExportFile['Data']
    rowA       = [] # list of row aspects for data table
    colA       = [] # list of col aspects for data table
    for m in range(4,73):
        if SpecsS.cell(m,1).value == 'x': # entry is dataset-specific and must be transferred to template
            CoverS.cell(m,4).value = SpecsS.cell(m,4).value
        CoverS.cell(7,8).value   = SpecsS.cell(7,8).value            
        CoverS.cell(7,9).value   = SpecsS.cell(7,9).value                    
        CoverS.cell(10,9).value  = SpecsS.cell(10,9).value                    
        CoverS.cell(10,11).value = SpecsS.cell(10,11).value                    
    for m in range(0,10):
        rowA.append(SpecsS.cell(12+m,6).value)
        colA.append(SpecsS.cell(12+m,8).value)
        for n in range(0,4):
            CoverS.cell(12+m,6+n).value = SpecsS.cell(12+m,6+n).value
    CoverS.cell(61,4).value = datestring
    CoverS.cell(62,4).value = datestring
    rowA = [i for i in rowA if i is not None]
    colA = [i for i in colA if i is not None]
    resrowcount = len(colA) + 1
    rescolcount = len(rowA) + 1
    foldercount = 26 # offest
    
    # write column labels for time series:
    for m in range(0,45):
        DataS.cell(1,rescolcount+m).value = 2016+m
        
    # load the indictors to read and the related aspects
    IndDict = TableSectionToDict(SpecsS,25,13)
    
    # load scenario results and write to dataset file
    while True:
        if SpecsS.cell(foldercount,6).value is not None:
            # open model run result file
            print(SpecsS.cell(foldercount,6).value)
            ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.res_v2_4_archive,SpecsS.cell(foldercount,6).value)) if filename.startswith('ODYM_RECC_ModelResults_')]
            Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.res_v2_4_archive,SpecsS.cell(foldercount,6).value,ResFile[0]))
            Resultsheet2 = Resultfile2['Model_Results']
            
            # Get indicator locations in file
            IndIndices   = []
            for m in range(0,len(IndDict['Indicator list'])):
                count = 1
                while True:
                    if Resultsheet2.cell(count,1).value == IndDict['Indicator list'][m]:
                        break 
                    count += 1
                IndIndices.append(count)
                
            # loop over all indicators and scenarios to extract and write values
            for m in range(0,len(IndDict['Indicator list'])): # all indicators
                for S in range(0,3): # three socioeconomic scenarios
                
                    # first, the labels:
                    for A in range(0,len(rowA)):
                        if rowA[A] == 'region':
                            DataS.cell(resrowcount,A+1).value = SpecsS.cell(foldercount,7).value
                        elif rowA[A] == 'end-use sector':
                            DataS.cell(resrowcount,A+1).value = SpecsS.cell(foldercount,8).value
                        elif rowA[A] == 'strategy':
                            DataS.cell(resrowcount,A+1).value = SpecsS.cell(foldercount,9).value
                        elif rowA[A] == 'scenario':
                            DataS.cell(resrowcount,A+1).value = ScenLabels[S] 
                        else:
                            DataS.cell(resrowcount,A+1).value = IndDict[rowA[A]][m]
                            
                    # then, the data:
                    for t in range(0,45):
                        DataS.cell(resrowcount,rescolcount+t).value = Resultsheet2.cell(IndIndices[m]+1+2*S,9+t).value 
                    resrowcount +=1
        else:
            break
        foldercount +=1
        
    # Add number of rows to export file:
    CoverS.cell(10,9).value = resrowcount-len(colA)-1
    
    # close export file for current parameter
    ExportFile.save(os.path.join(RECC_Paths.iedc_export_path, DS + '.xlsx'))     
    
ExportSpecsFile.close()        
        
#
#
#