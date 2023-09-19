# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for parsing and checking in the IIASA IRP database.

See https://github.com/iiasa/irp-internal-workflow for naming conventions.

"""
# Import required libraries:
import os
import openpyxl
import numpy as np
import pandas as pd
import uuid
import nomenclature
import pyam
from pathlib import Path
import RECC_Paths # Import path file
P_irp = RECC_Paths.irp_path

# Create UUID of script run
Current_UUID = str(uuid.uuid4())
RECC_Paths.results_path_save = Path(RECC_Paths.results_path_eval) / ('RECC_Results_' + Current_UUID)
if not os.path.exists(RECC_Paths.results_path_save): # Create scrip run results directory.
    os.makedirs(RECC_Paths.results_path_save)

# Set sector: separate operation for passenger vehicles pav and residential buildings reb, as the two are organised in separate scenario folders with the same data structure.
Sect = 'pav'
#Sect = 'reb'

# Definitions
SpecPath      = Path(RECC_Paths.results_path) / 'IRP - IIASA database variable template proposal_08_07_21.xlsx'
SpecFile      = openpyxl.load_workbook(SpecPath)
if Sect == 'reb':
    SpecSheet     = SpecFile['RECC_Export_reb']
if Sect == 'pav':
    SpecSheet     = SpecFile['RECC_Export_pav']

IIASA_ScNames  = []
RECC_ScFolders = []
RECC_ScIndex   = []
r = 2
while True: # Scan column for content until either sheet ends or cell content is empty string ''
    try:
        if len(SpecSheet.cell(r+1, 2).value) > 0:
            IIASA_ScNames.append(SpecSheet.cell(r+1, 2).value)
            RECC_ScFolders.append(SpecSheet.cell(r+1, 3).value)
            RECC_ScIndex.append(int(SpecSheet.cell(r+1, 4).value))
    except:
             break
    r += 1
    
IIASA_IndNames = []
IIASA_IndUnits = []
RECC_SysDefLoc = []
RECC_Ind_Unit  = []
RECC_ConvFactor= []
RECC_Res_Match = []
r = 2
while True: # Scan column for content until either sheet ends or cell content is empty string ''
    try:
        if len(SpecSheet.cell(r+1, 5).value) > 0:
            IIASA_IndNames.append( SpecSheet.cell(r+1, 5).value)
            IIASA_IndUnits.append( SpecSheet.cell(r+1, 6).value)
            RECC_SysDefLoc.append( SpecSheet.cell(r+1, 7).value)
            RECC_Ind_Unit.append(  SpecSheet.cell(r+1, 8).value)
            RECC_ConvFactor.append(SpecSheet.cell(r+1, 9).value)
            RECC_Res_Match.append([])
    except:
             break
    r += 1

for m in range(0,len(RECC_Res_Match)):
    r = 9
    while True: # Scan columns for content
        try:
            if len(SpecSheet.cell(m+3, r+1).value) > 0:
                RECC_Res_Match[m].append(SpecSheet.cell(m+3, r+1).value)
        except:
                 break
        r += 1

# Create dataframe with result data:
RECC_DF = pd.DataFrame([ ],
    columns=['scenario', 'variable', 'unit'] + [i for i in range(2015, 2061)])

# Gather data in a triple loop: for all scenarios, for all IIASA indicators, and for all n:1 RECC->IIASA Indicators
ri = 2    
for S in range(0,len(IIASA_ScNames)): # IIASA scenario index
    # Open scenario result file:
    # Find result fild with common start but unique UUID in name:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,RECC_ScFolders[S])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,RECC_ScFolders[S],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    for I in range(0,len(IIASA_IndNames)): # IIASA indicator index
        # Define df row
        df_row = {}
        df_row['scenario'] = IIASA_ScNames[S]
        df_row['variable'] = IIASA_IndNames[I]
        df_row['unit']     = IIASA_IndUnits[I]
        # Collect data
        NuLi = np.zeros((1,46)) # Numbers for this line
        for N in range(0,len(RECC_Res_Match[I])):
            # First, get the position indices for the different result variables:
            rowi = 1
            while True:
                if Resultsheet2.cell(rowi+1, 1).value == RECC_Res_Match[I][N]:
                    break
                rowi += 1
            for t in range(0,46): # time until 2060
                NuLi[0,t] += Resultsheet2.cell(rowi + RECC_ScIndex[S]+1,t+8).value * RECC_ConvFactor[I]
    
        # Write data to transfer file          
        for Y in range(0,46):
            df_row[2015+Y] = NuLi[0,Y]
        ri +=1
        RECC_DF = RECC_DF.append(df_row, ignore_index=True)

"""Load local copy of IRP_GRO_IIASA_DB nomenclature"""
nc = nomenclature.Nomenclature(Path(P_irp) / 'definitions' ) # load nomenclature locally
nc.variable

# Convert RECC_DF to IamDataFrame: 
df_base = pyam.IamDataFrame(RECC_DF, model='ODYM-RECC v2.4', region='World')


"""Validate dataframe against nomenclature"""
nc.validate(df_base)
ExpFileName = 'RECC_to_IIASA_DB_export_' + Sect
df_base.to_excel(ExpFileName + '.xlsx') # exports to uploadable xlsx file, saved in current working directory.
#df_base.to_csv('RECC_to_IIASA_DB_export.csv')




### Sandbox
df_base.timeseries()


P1 = os.path.join(P_irp,'definitions','variables')
P2 = os.path.join(P_irp,'definitions','variables','types')
"""Assert that all yaml files in `path` can be parsed without errors"""
nomenclature.testing.assert_valid_yaml(P1)
nomenclature.testing.assert_valid_yaml(P2)

df_base.unit_mapping
df_base.filter(variable='Emissions|Kyoto Gases', region='World').plot()    


             
                 
#
#
#