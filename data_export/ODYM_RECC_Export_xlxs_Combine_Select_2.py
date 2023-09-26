# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for checking and plotting.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select_2.xlxs

This script takes results from different single (or few region) model runs to 
compile single-region results or aggregate into smaller world regions (like Europe or Asia).

"""
# Import required libraries:
import os
import openpyxl
import numpy as np
import uuid
import RECC_Paths # Import path file
import copy

def get_RECC_resfile_pos(Label,region,Resultsheet):
    # Find the index for the given Label
    idx = 1    
    while True:
        if Resultsheet.cell(idx, 1).value == Label and Resultsheet.cell(idx, 3).value == region:
            break # that gives us the right index to read the Label from the result table.
        idx += 1 
    return idx

# Create UUID of script run
Current_UUID = str(uuid.uuid4())

# Read from default location:
CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select_2.xlsx')    
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
    
# Definitions/Specifications
Model_id      = CF[CS].cell(1,2).value
Model_date    = CF[CS].cell(1,4).value
outpath       = CF[CS].cell(1,6).value
fn_add        = CF[CS].cell(1,8).value

# Prepare result workbook
RB = openpyxl.Workbook() # Export other model results, calibration values, flags, etc.
cs = RB.active
cs.title = 'Cover'
cs.cell(row=2, column=2).value = Model_id
cs.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
cs.cell(row=4, column=2).value = Model_date
rs = RB.create_sheet('Results')

# Read specs and variable matchings
scen = [] # list of target scenarios
offs = [] # list of offsets for the different scenarios rel. to starting position
secs = [] # list of matching results folders for the different scenarios


# Read scenarios
r = 4
while True:
    if CF[CS].cell(r,1).value is None:
        break
    if CF[CS].cell(r,1).value == 1: # Only if the SELECT flag in col. A is set to 1.
        scen.append(CF[CS].cell(r,2).value)
        offs.append(CF[CS].cell(r,5).value)
        secs.append([])
        for sl in range(0,20):
            secs[-1].append(CF[CS].cell(r,6+sl).value)
    r += 1
    
# Move to parameter list:
while True:
    if CF[CS].cell(r,2).value == 'Target indicator label':
        break    
    r += 1
r += 1    

# Read indicator list:
ti = []    # target indicator
ri = []    # RECC indicator
tu = []    # target unit
ru = []    # RECC unit
cf = []    # Conv. factor
sl = []    # sector list
rr = []    # RECC regional resolution

while True:
    if CF[CS].cell(r,2).value is None:
        break
    ti.append(CF[CS].cell(r,2).value)
    ri.append(CF[CS].cell(r,3).value)
    tu.append(CF[CS].cell(r,4).value)
    ru.append(CF[CS].cell(r,5).value)
    cf.append(CF[CS].cell(r,6).value)
    sl.append(CF[CS].cell(r,7).value)
    rr.append(CF[CS].cell(r,8).value)
    r += 1

nos = len(scen) # number of scenarios
noi = len(ti)   # number of indicators

# split sector labels at '+' for different sectors
sL = [i.split('+') for i in sl]

# Move to region aggregation definition:
while True:
    if CF[CS].cell(r,2).value == 'AggRegion':
        break    
    r += 1
r += 1   

# Region List
Ar = []    # Aggregated (target) regions for export
Dr = []    # RECC detailed regions

r0 = r

while True:
    if CF[CS].cell(r,2).value is None:
        break
    Ar.append(CF[CS].cell(r,2).value)
    c = 3
    Dr.append([])
    while True:
        if CF[CS].cell(r,c).value is None:
            break
        Dr[r-r0].append(CF[CS].cell(r,c).value)
        c += 1    
    r += 1

nor = len(Ar)

# iterate over scenarios, parse results:
Res = np.zeros((nor*nos*noi,46+7)) # main result array 46 years, one blank column, and 6 columns for cumulative results

Folders = list(set([item for sublist in secs for item in sublist]))

# Extract, aggregated, and reformat results
for rf in Folders:
    if rf is not None: # if a folder is given, extract all indicators and write to corresponding position in result array:
        print('Reading data from ' + rf)
        RECC_ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,rf)) if filename.startswith('ODYM_RECC_ModelResults_')]
        RECC_RF      = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,rf,RECC_ResFile[0]))
        RECC_RS      = RECC_RF['Model_Results']
        parts = rf.split('__')
        region = parts[0]
        sector = parts[3]
        # look for where to put data from this result folder:
        for s in range(0,nos): # iterate over all selected scenarios
            for i in range(0,noi): # for all indicators
                print('Reading data for ' + ti[i])
                for r in range(0,nor): # for all regions
                    if region == Ar[r] or region in Dr[r]: # the current result file region is or is part of the current target region
                        targetpos = r*nos*noi + i * nos + s # position in Res array: outer index: region, middle index: indicator, inner index: scenario
                        if sector in sL[i]: # if current sector is part of target sector for indicator
                            if rf in secs[s]: # if current folder is in list for currect scenario --> extract results!
                                if rr[i] == 'Aggregate': # use indicator for aggregate region label and add to results:
                                    idx = get_RECC_resfile_pos(ri[i],region,RECC_RS) # position of that indicator for that region in the RECC result file
                                    for t in range(0,46): # read and add values
                                        Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[i] # Value from Excel to array
                                if rr[i] == 'Aggregate_from_Single': # use indicator for aggregate region label and add to results:
                                    for sir in range(0,len(Dr[r])):
                                        idx = get_RECC_resfile_pos(ri[i],Dr[r][sir],RECC_RS) # position of that indicator for that region in the RECC result file
                                        for t in range(0,46): # read and add values
                                            Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[i] # Value from Excel to array
        
# Determine cumulative quantities:
Res[:,47] = Res[:,5:36].sum(axis=1) # Cum. 2020-2050
Res[:,48] = Res[:,5:46].sum(axis=1) # Cum. 2020-2060
Res[:,49] = Res[:,5:15].sum(axis=1) # Cum. 2020-2029
Res[:,50] = Res[:,15:25].sum(axis=1) # Cum. 2030-2039
Res[:,51] = Res[:,25:35].sum(axis=1) # Cum. 2040-2049
Res[:,52] = Res[:,35:45].sum(axis=1) # Cum. 2050-2059

# Export results to xlsx
# Define column labels
rs.cell(row=1, column=1).value = 'Model id'
rs.cell(row=1, column=2).value = 'Region'
rs.cell(row=1, column=3).value = 'Variable'
rs.cell(row=1, column=4).value = 'Scenario'
rs.cell(row=1, column=5).value = 'Sectors'
rs.cell(row=1, column=6).value = 'Unit'
for t in range(0,46):
    rs.cell(row=1, column=7+t).value = 2015 + t
rs.cell(row=1, column=46+7).value = 'No data'
rs.cell(row=1, column=47+7).value = 'Cum. 2020-2050 (incl.)'
rs.cell(row=1, column=48+7).value = 'Cum. 2020-2060 (incl.)'
rs.cell(row=1, column=49+7).value = 'Cum. 2020-2029'
rs.cell(row=1, column=50+7).value = 'Cum. 2030-2039'
rs.cell(row=1, column=51+7).value = 'Cum. 2040-2040'
rs.cell(row=1, column=52+7).value = 'Cum. 2050-2059'
for t in range(1,60):
    rs.cell(row=1, column=t).font = openpyxl.styles.Font(bold=True)
# fill labels
for m in range(0,nor*nos*noi):
     r = m // (noi*nos)
     x = m %  (noi*nos)
     i = x // nos
     s = x %  nos
     rs.cell(row=m+2, column=1).value = Model_id 
     rs.cell(row=m+2, column=2).value = Ar[r]
     rs.cell(row=m+2, column=3).value = ti[i]
     rs.cell(row=m+2, column=4).value = scen[s]
     rs.cell(row=m+2, column=5).value = sl[i]
     rs.cell(row=m+2, column=6).value = tu[i]
# fill data    
for m in range(0,nor*nos*noi):
    for n in range(0,46+7):
        rs.cell(row=2+m, column=7+n).value = Res[m,n]
        
# Special: All regions to global aggregate, if region is outer index and all regions add up
region_no = nos*noi
Res_r = Res.reshape((nor,nos*noi,46+7)) # reshape to region as separate dimension
Res_r_agg = Res_r.sum(axis=0) # sum over regions
start_ind = copy.deepcopy(m)
# fill labels
for m in range(start_ind,start_ind+nos*noi):
     i = (m-start_ind) // nos
     s = (m-start_ind) %  nos
     rs.cell(row=m+2, column=1).value = Model_id 
     rs.cell(row=m+2, column=2).value = 'Global'
     rs.cell(row=m+2, column=3).value = ti[i]
     rs.cell(row=m+2, column=4).value = scen[s]
     rs.cell(row=m+2, column=5).value = sl[i]
     rs.cell(row=m+2, column=6).value = tu[i]
# fill data    
for m in range(start_ind,start_ind+nos*noi):
    for n in range(0,46+7):
        rs.cell(row=2+m, column=7+n).value = Res_r_agg[m-start_ind,n]        

# Save exported results
RB.save(os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '.xlsx')) 
                 
#
#
#
#
#
#