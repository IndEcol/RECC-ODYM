# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for checking and plotting.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select.xlxs

This script takes results from different single (or few region) model runs to 
compile single-region results or aggregate into smaller world regions (like Europe or Asia).

Export time series and cumulative results to separate sheets, better for subsequent analysis with pandas/pyplot etc.

Aggregate different indicators into a single one, like different building types

Extract single-region indicators from aggregate region model runs

Documentation and how to in RECCv2.5_EXPORT_Combine_Select.xlxs
"""

# Import required libraries:
import os
import openpyxl

import numpy as np
import uuid
import RECC_Paths # Import path file
import copy

def get_RECC_resfile_pos(Label,region,Resultsheet):
    # Find the index for the given Label
    idx = 1    
    while True:
        if Resultsheet.cell(idx, 1).value == Label and Resultsheet.cell(idx, 3).value == region:
            break # that gives us the right index to read the Label from the result table.
        idx += 1 
    return idx

# Create UUID of script run
Current_UUID = str(uuid.uuid4())

# Read from default location:
CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')    
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
    
# Definitions/Specifications
Model_id      = CF['Cover'].cell(6,4).value
Model_date    = CF['Cover'].cell(7,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value
glob_agg      = CF[CS].cell(1,6).value

# Prepare result workbook
RB = openpyxl.Workbook() # Export other model results, calibration values, flags, etc.
cs = RB.active
cs.title = 'Cover'
cs.cell(row=2, column=2).value = Model_id
cs.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
cs.cell(row=4, column=2).value = Model_date
resS = RB.create_sheet('Results')
rsCS = RB.create_sheet('Results_Cumulative')

# Read specs and variable matchings
scen = [] # list of target scenarios
offs = [] # list of offsets for the different scenarios rel. to starting position
secs = [] # list of matching results folders for the different scenarios


# Read scenarios
r = 4
while True:
    if CF[CS].cell(r,1).value is None:
        break
    if CF[CS].cell(r,1).value == 1: # Only if the SELECT flag in col. A is set to 1.
        scen.append(CF[CS].cell(r,2).value)
        offs.append(CF[CS].cell(r,5).value)
        secs.append([])
        for sl in range(0,100): # up to 100 result folders for each scenario
            secs[-1].append(CF[CS].cell(r,6+sl).value)
    r += 1
    
# Move to parameter list:
while True:
    if CF[CS].cell(r,2).value == 'Target indicator label':
        break    
    r += 1
r += 1    

# Read indicator list:
tif= []    # target indicator full list
ri = []    # RECC indicator list
tu = []    # target unit
tuc= []    # target unit cumulative
ru = []    # RECC unit
cf = []    # Conv. factor
sl = []    # sector list
rr = []    # RECC regional resolution

while True:
    if CF[CS].cell(r,2).value is None:
        break
    tif.append(CF[CS].cell(r,2).value)
    ri.append(CF[CS].cell(r,3).value)
    tu.append(CF[CS].cell(r,4).value)
    tuc.append(CF[CS].cell(r,9).value)
    ru.append(CF[CS].cell(r,5).value)
    cf.append(CF[CS].cell(r,6).value)
    sl.append(CF[CS].cell(r,7).value)
    rr.append(CF[CS].cell(r,8).value)
    r += 1

nos = len(scen) # number of scenarios
ti  = list(set(tif)) # only unique indicators
noi = len(ti)   # number of indicators
noif= len(tif) # number of source indicators

# sort target units into same order as target indicators
tu  = [tu[tif.index(z)]  for z in ti]
tuc = [tuc[tif.index(z)] for z in ti]

# split sector labels at '+' for different sectors
sL = [i.split('+') for i in sl]

# Move to region aggregation definition:
while True:
    if CF[CS].cell(r,2).value == 'AggRegion':
        break    
    r += 1
r += 1   

# Region List
Ar = []    # Aggregated (target) regions for export
Dr = []    # RECC detailed regions

r0 = r

while True:
    if CF[CS].cell(r,2).value is None:
        break
    Ar.append(CF[CS].cell(r,2).value)
    c = 3
    Dr.append([])
    while True:
        if CF[CS].cell(r,c).value is None:
            break
        Dr[r-r0].append(CF[CS].cell(r,c).value)
        c += 1    
    r += 1

nor = len(Ar)

# iterate over scenarios, parse results:
Res  = np.zeros((nor*nos*noi,46)) # main result array 46 years
ResC = np.zeros((nor*nos*noi,6))  # main results, 6 columns for cumulative results

Folders = list(set([item for sublist in secs for item in sublist]))

# Extract, aggregated, and reformat results
for rf in Folders:
    if rf is not None: # if a folder is given, extract all indicators and write to corresponding position in result array:
        print('Reading data from ' + rf)
        RECC_ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,rf)) if filename.startswith('ODYM_RECC_ModelResults_')]
        RECC_RF      = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,rf,RECC_ResFile[0]))
        RECC_RS      = RECC_RF['Model_Results']
        parts = rf.split('__')
        region = parts[0]
        sector = parts[3]
        # look for where to put data from this result folder:
        for s in range(0,nos): # iterate over all selected scenarios
            for j in range(0,noif): # for all source indicators
                i = ti.index(tif[j]) # target position of indicator
                print('Reading source data for ' + ti[i])
                for r in range(0,nor): # for all regions
                    if region == Ar[r] or region in Dr[r]: # the current result file region is or is part of the current target region
                        targetpos = r*nos*noi + i * nos + s # position in Res array: outer index: region, middle index: indicator, inner index: scenario
                        if sector in sL[j]: # if current sector is part of source sector(s) for target indicator
                            if rf in secs[s]: # if current folder is in list for currect scenario --> extract results!
                                if rr[j] == 'Aggregate': # use indicator for aggregate region label and add to results:
                                    try:
                                        idx = get_RECC_resfile_pos(ri[j],region,RECC_RS) # position of that indicator for that region in the RECC result file
                                    except: # no such indicator for this region
                                        break
                                    for t in range(0,46): # read and add values
                                        Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array
                                if rr[j] == 'Aggregate_from_Single': # use indicator for aggregate region label and add to results:
                                    for sir in range(0,len(Dr[r])):
                                        try:
                                            idx = get_RECC_resfile_pos(ri[j],Dr[r][sir],RECC_RS) # position of that indicator for that region in the RECC result file
                                        except: # no such indicator for this region
                                            break
                                        for t in range(0,46): # read and add values
                                            Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array
                                if rr[j] == 'Extract from aggregate':
                                    try:
                                        idx = get_RECC_resfile_pos(ri[j],Ar[r],RECC_RS) # position of that indicator for that region in the RECC result file
                                    except: # no such indicator for this region
                                        break
                                    for t in range(0,46): # read and add values
                                        Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array                                    
        
# Determine cumulative quantities:
ResC[:,0] = Res[:,5:36].sum(axis=1) # Cum. 2020-2050
ResC[:,1] = Res[:,5:46].sum(axis=1) # Cum. 2020-2060
ResC[:,2] = Res[:,5:15].sum(axis=1) # Cum. 2020-2029
ResC[:,3] = Res[:,15:25].sum(axis=1) # Cum. 2030-2039
ResC[:,4] = Res[:,25:35].sum(axis=1) # Cum. 2040-2049
ResC[:,5] = Res[:,35:45].sum(axis=1) # Cum. 2050-2059

# Export results to xlsx
# Define column labels
for rs in [resS,rsCS]:
    rs.cell(row=1, column=1).value = 'Model id'
    rs.cell(row=1, column=2).value = 'Region'
    rs.cell(row=1, column=3).value = 'Indicator'
    rs.cell(row=1, column=4).value = 'Scenario'
    rs.cell(row=1, column=5).value = 'Sectors'
    rs.cell(row=1, column=6).value = 'Unit'
    
for t in range(0,46):
    resS.cell(row=1, column=7+t).value = 2015 + t
rsCS.cell(row=1, column=7).value = 'Cum. 2020-2050 (incl.)'
rsCS.cell(row=1, column=8).value = 'Cum. 2020-2060 (incl.)'
rsCS.cell(row=1, column=9).value = 'Cum. 2020-2029'
rsCS.cell(row=1, column=10).value = 'Cum. 2030-2039'
rsCS.cell(row=1, column=11).value = 'Cum. 2040-2049'
rsCS.cell(row=1, column=12).value = 'Cum. 2050-2059'
for t in range(1,53):
    resS.cell(row=1, column=t).font = openpyxl.styles.Font(bold=True)
for t in range(1,13):
    resS.cell(row=1, column=t).font = openpyxl.styles.Font(bold=True)
    
# fill labels
for m in range(0,nor*nos*noi):
     r = m // (noi*nos)
     x = m %  (noi*nos)
     i = x // nos
     s = x %  nos
     for rs in [resS,rsCS]:
         rs.cell(row=m+2, column=1).value = Model_id 
         rs.cell(row=m+2, column=2).value = Ar[r]
         rs.cell(row=m+2, column=3).value = ti[i]
         rs.cell(row=m+2, column=4).value = scen[s]
         rs.cell(row=m+2, column=5).value = sl[i]
     resS.cell(row=m+2, column=6).value   = tu[i]
     rsCS.cell(row=m+2, column=6).value   = tuc[i]
# fill data    
for m in range(0,nor*nos*noi):
    for n in range(0,46):
        resS.cell(row=2+m, column=7+n).value = Res[m,n]
    for n in range(0,6):
        rsCS.cell(row=2+m, column=7+n).value = ResC[m,n]
        
        
# Special: All regions to global aggregate, if region is outer index and all regions add up
if glob_agg    == 'True':
    region_no  = nos*noi
    Res_r      = Res.reshape((nor,nos*noi,46)) # reshape to region as separate dimension
    Res_r_agg  = Res_r.sum(axis=0) # sum over regions
    Res_rC     = ResC.reshape((nor,nos*noi,6)) # reshape to region as separate dimension
    Res_rC_agg = Res_rC.sum(axis=0) # sum over regions

    start_ind = copy.deepcopy(m)+1
    # fill labels
    for m in range(start_ind,start_ind+nos*noi):
         i = (m-start_ind) // nos
         s = (m-start_ind) %  nos
         for rs in [resS,rsCS]:
             rs.cell(row=m+2, column=1).value = Model_id 
             rs.cell(row=m+2, column=2).value = 'Global'
             rs.cell(row=m+2, column=3).value = ti[i]
             rs.cell(row=m+2, column=4).value = scen[s]
             rs.cell(row=m+2, column=5).value = sl[i]
         resS.cell(row=m+2, column=6).value   = tu[i]
         rsCS.cell(row=m+2, column=6).value   = tuc[i]
    # fill data    
    for m in range(start_ind,start_ind+nos*noi):
        for n in range(0,46):
            resS.cell(row=2+m, column=7+n).value = Res_r_agg[m-start_ind,n]        
        for n in range(0,6):
            rsCS.cell(row=2+m, column=7+n).value = Res_rC_agg[m-start_ind,n]        
            
# Save exported results
RB.save(os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '_sep.xlsx')) 
                 
#
#
#
#
#
#