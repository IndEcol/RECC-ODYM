# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for parsing and checking with the CRAFT team at BoKu.

See https://github.com/iiasa/irp-internal-workflow for naming conventions.

"""
# Import required libraries:
import os
import openpyxl
import numpy as np
import pandas as pd
import uuid
from pathlib import Path
import RECC_Paths # Import path file

def get_RECC_resfile_pos(Label,region,Resultsheet):
    # Find the index for the given Label
    idx = 1    
    while True:
        if Resultsheet.cell(idx, 1).value == Label and Resultsheet.cell(idx, 3).value == region:
            break # that gives us the right index to read the Label from the result table.
        idx += 1 
    return idx

# Create UUID of script run
Current_UUID = str(uuid.uuid4())

# Read from default location:
CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select_1.xlsx')    
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
    
# Definitions/Specifications
Model_id      = CF[CS].cell(1,2).value
Model_date    = CF[CS].cell(1,4).value
outpath       = CF[CS].cell(1,6).value


# Prepare result workbook
RB = openpyxl.Workbook() # Export other model results, calibration values, flags, etc.
cs = RB.active
cs.title = 'Cover'
cs.cell(row=2, column=2).value = Model_id
cs.cell(row=2, column=2).font = openpyxl.styles.Font(bold=True)
cs.cell(row=4, column=2).value = Model_date
rs = RB.create_sheet('Results')

# Read specs and variable matchings
scen = [] # list of target scenarios
offs = [] # list of offsets for the different scenarios rel. to starting position
secs = [[],[],[],[],[],[],[],[],[],[]] # list of matching results folders for sectors 1, 2, and 3 - 10
secS = ['sector1','sector2','sector3','sector4','sector5','sector6','sector7','sector8','sector9','sector10']

secL = [] # Sector labels
for i in range(0,10):
    secL.append(CF[CS].cell(3,6+i).value)

# Read scenarios
r = 4
while True:
    if CF[CS].cell(r,1).value is None:
        break
    if CF[CS].cell(r,1).value == 1: # Only if the SELECT flag in col. A is set to 1.
        scen.append(CF[CS].cell(r,2).value)
        offs.append(CF[CS].cell(r,5).value)
        for sl in range(0,10):
            secs[0].append(CF[CS].cell(r,6+sl).value)
    r += 1
    
# Move to parameter list:
while True:
    if CF[CS].cell(r,2).value == 'Target indicator label':
        break    
    r += 1
r += 1    

# Read indicator list:
ti = []    # target indicator
ri = []    # RECC indicator
tu = []    # target unit
ru = []    # RECC unit
cf = []    # Conv. factor
sl = []    # sector list
tr = []    # target region
rr = []    # RECC regions

r0 = r

while True:
    if CF[CS].cell(r,2).value is None:
        break
    ti.append(CF[CS].cell(r,2).value)
    ri.append(CF[CS].cell(r,3).value)
    tu.append(CF[CS].cell(r,4).value)
    ru.append(CF[CS].cell(r,5).value)
    cf.append(CF[CS].cell(r,6).value)
    sl.append(CF[CS].cell(r,7).value)
    tr.append(CF[CS].cell(r,8).value)
    c = 9
    rr.append([])
    while True:
        if CF[CS].cell(r,c).value is None:
            break
        rr[r-r0].append(CF[CS].cell(r,c).value)
        c += 1
    r += 1

nos = len(scen) # number of scenarios
noi = len(ti)   # number of indicators

# Create sector list with actual labels, not placeholders
sL = sl.copy()
for i in range(0,7): # Find all sectorX instances and replace them with their actual sector labels
    for j in range(0,10):
        if secS[j] in sL[i]:
            sL[i] = sL[i].replace(secS[j],secL[j])

# iterate over scenarios, parse results:
Res = np.zeros((nos*noi,46+7)) # main result array 46 years, one blank column, and 6 columns for cumulative results

for S in range(0,nos): # iterate over all selected scenarios
    for s in range(0,3): # iterate over all sectors 
        rf = secs[s][S]# result folder
        if rf is not None: # if a folder is given, extract all indicators and write to corresponding position in result array:
            print('Reading data from ' + rf)
            RECC_ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,rf)) if filename.startswith('ODYM_RECC_ModelResults_')]
            RECC_RF      = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,rf,RECC_ResFile[0]))
            RECC_RS      = RECC_RF['Model_Results']
            for ind in range(0,noi): # for all indicators
                print('Reading data for ' + ti[ind])
                for reg in range(0,len(rr[ind])): # for all regions to be aggreated
                    idx = get_RECC_resfile_pos(ri[ind],rr[ind][reg],RECC_RS) # position of that indicator for that region in the RECC result file
                    if secS[s] in sl[ind]: # if the current sector is in the sector list for this indicator
                        targetpos = ind * nos + S # position in Res array: outer index: indicator, inner index: scenario
                        for t in range(0,46): # read and add values
                            Res[targetpos,t] += RECC_RS.cell(idx+offs[S],t+8).value * cf[ind] # Value from Excel to array

# Determine cumulative quantities:
Res[:,47] = Res[:,5:36].sum(axis=1) # Cum. 2020-2050
Res[:,48] = Res[:,5:46].sum(axis=1) # Cum. 2020-2060
Res[:,49] = Res[:,5:16].sum(axis=1) # Cum. 2020-2030
Res[:,50] = Res[:,15:26].sum(axis=1) # Cum. 2030-2040
Res[:,51] = Res[:,25:36].sum(axis=1) # Cum. 2040-2050
Res[:,52] = Res[:,35:46].sum(axis=1) # Cum. 2050-2060

# Export results to xlsx
# Define column labels
rs.cell(row=1, column=1).value = 'Model id'
rs.cell(row=1, column=2).value = 'Variable'
rs.cell(row=1, column=3).value = 'Scenario'
rs.cell(row=1, column=4).value = 'Region'
rs.cell(row=1, column=5).value = 'Sectors'
rs.cell(row=1, column=6).value = 'Unit'
for t in range(0,46):
    rs.cell(row=1, column=7+t).value = 2015 + t
rs.cell(row=1, column=47+7).value = 'Cum. 2020-2050'
rs.cell(row=1, column=48+7).value = 'Cum. 2020-2060'
rs.cell(row=1, column=49+7).value = 'Cum. 2020-2030'
rs.cell(row=1, column=50+7).value = 'Cum. 2030-2040'
rs.cell(row=1, column=51+7).value = 'Cum. 2040-2050'
rs.cell(row=1, column=52+7).value = 'Cum. 2050-2060'
for t in range(1,60):
    rs.cell(row=1, column=t).font = openpyxl.styles.Font(bold=True)
# fill labels
for m in range(0,nos*noi):
     rs.cell(row=m+2, column=1).value = Model_id 
     rs.cell(row=m+2, column=2).value = ti[m // nos]
     rs.cell(row=m+2, column=3).value = scen[m % nos]
     rs.cell(row=m+2, column=4).value = tr[m // nos]
     rs.cell(row=m+2, column=5).value = sL[m // nos]
     rs.cell(row=m+2, column=6).value = tu[m // nos]
# fill data    
for m in range(0,nos*noi):
    for n in range(0,46+7):
        rs.cell(row=2+m, column=7+n).value = Res[m,n]

# Save exported results
RB.save(os.path.join(RECC_Paths.export_path,outpath,'Results_RECCv2.5_CRAFT_Coupling.xlsx')) 
                 
#
#
#
#
#
#
