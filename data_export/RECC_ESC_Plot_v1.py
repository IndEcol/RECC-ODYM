# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 13:54:19 2023

@author: spauliuk
"""
import os
import plotnine
import openpyxl
from plotnine import *
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
import RECC_Paths # Import path file


def get_RECC_resfile_pos(Label,region,Resultsheet):
    # Find the index for the given Label
    idx = 1    
    while True:
        if Resultsheet.cell(idx, 1).value == Label and Resultsheet.cell(idx, 3).value == region:
            break # that gives us the right index to read the Label from the result table.
        idx += 1 
    return idx

# Definitions/Specifications
CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')   
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
Model_id      = CF['Cover'].cell(6,4).value
Model_date    = CF['Cover'].cell(7,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value
glob_agg      = CF[CS].cell(1,6).value

# Read specs and variable matchings
scen = [] # list of target scenarios
offs = [] # list of offsets for the different scenarios rel. to starting position
secs = [] # list of matching results folders for the different scenarios

# Read scenarios
r = 4
while True:
    if CF[CS].cell(r,1).value is None:
        break
    if CF[CS].cell(r,1).value == 1: # Only if the SELECT flag in col. A is set to 1.
        scen.append(CF[CS].cell(r,2).value)
        offs.append(CF[CS].cell(r,5).value)
        secs.append([])
        for sl in range(0,20):
            secs[-1].append(CF[CS].cell(r,6+sl).value)
    r += 1
    
# Move to parameter list:
while True:
    if CF[CS].cell(r,2).value == 'Target indicator label':
        break    
    r += 1
r += 1    

# Read indicator list:
tif= []    # target indicator
ri = []    # RECC indicator
tu = []    # target unit
ru = []    # RECC unit
cf = []    # Conv. factor
sl = []    # sector list
rr = []    # RECC regional resolution

while True:
    if CF[CS].cell(r,2).value is None:
        break
    tif.append(CF[CS].cell(r,2).value)
    ri.append(CF[CS].cell(r,3).value)
    tu.append(CF[CS].cell(r,4).value)
    ru.append(CF[CS].cell(r,5).value)
    cf.append(CF[CS].cell(r,6).value)
    sl.append(CF[CS].cell(r,7).value)
    rr.append(CF[CS].cell(r,8).value)
    r += 1

nos = len(scen) # number of scenarios
ti  = list(set(tif)) # only unique indicators
noi = len(ti)   # number of indicators
noif= len(tif) # number of source indicators

# split sector labels at '+' for different sectors
sL = [i.split('+') for i in sl]

# Move to region aggregation definition:
while True:
    if CF[CS].cell(r,2).value == 'AggRegion':
        break    
    r += 1
r += 1   

# Region List
Ar = []    # Aggregated (target) regions for export
Dr = []    # RECC detailed regions

r0 = r

while True:
    if CF[CS].cell(r,2).value is None:
        break
    Ar.append(CF[CS].cell(r,2).value)
    c = 3
    Dr.append([])
    while True:
        if CF[CS].cell(r,c).value is None:
            break
        Dr[r-r0].append(CF[CS].cell(r,c).value)
        c += 1    
    r += 1

nor = len(Ar)

# iterate over scenarios, parse results:
Res  = np.zeros((nor*nos*noi,46)) # main result array 46 years
ResC = np.zeros((nor*nos*noi,6))  # main results, 6 columns for cumulative results

Folders = list(set([item for sublist in secs for item in sublist]))

# Extract, aggregated, and reformat results
for rf in Folders:
    if rf is not None: # if a folder is given, extract all indicators and write to corresponding position in result array:
        print('Reading data from ' + rf)
        RECC_ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,rf)) if filename.startswith('ODYM_RECC_ModelResults_')]
        RECC_RF      = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,rf,RECC_ResFile[0]))
        RECC_RS      = RECC_RF['Model_Results']
        parts = rf.split('__')
        region = parts[0]
        sector = parts[3]
        # look for where to put data from this result folder:
        for s in range(0,nos): # iterate over all selected scenarios
            for j in range(0,noif): # for all source indicators
                i = ti.index(tif[j]) # target position of indicator
                print('Reading data for ' + ti[i])
                for r in range(0,nor): # for all regions
                    if region == Ar[r] or region in Dr[r]: # the current result file region is or is part of the current target region
                        targetpos = r*nos*noi + i * nos + s # position in Res array: outer index: region, middle index: indicator, inner index: scenario
                        if sector in sL[i]: # if current sector is part of target sector for indicator
                            if rf in secs[s]: # if current folder is in list for currect scenario --> extract results!
                                if rr[j] == 'Aggregate': # use indicator for aggregate region label and add to results:
                                    try:
                                        idx = get_RECC_resfile_pos(ri[j],region,RECC_RS) # position of that indicator for that region in the RECC result file
                                    except: # no such indicator for this region
                                        break
                                    for t in range(0,46): # read and add values
                                        Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array
                                if rr[j] == 'Aggregate_from_Single': # use indicator for aggregate region label and add to results:
                                    for sir in range(0,len(Dr[r])):
                                        try:
                                            idx = get_RECC_resfile_pos(ri[j],Dr[r][sir],RECC_RS) # position of that indicator for that region in the RECC result file
                                        except: # no such indicator for this region
                                            break
                                        for t in range(0,46): # read and add values
                                            Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array
                                if rr[j] == 'Extract from aggregate':
                                    try:
                                        idx = get_RECC_resfile_pos(ri[j],Ar[r],RECC_RS) # position of that indicator for that region in the RECC result file
                                    except: # no such indicator for this region
                                        break
                                    for t in range(0,46): # read and add values
                                        Res[targetpos,t] += RECC_RS.cell(idx+offs[s],t+8).value * cf[j] # Value from Excel to array                                    
        
# Special: All regions to global aggregate, if region is outer index and all regions add up
if glob_agg    == 'True':
    region_no  = nos*noi
    Res_r      = Res.reshape((nor,nos*noi,46)) # reshape to region as separate dimension
    Res_r_agg  = Res_r.sum(axis=0) # sum over regions
    Res_rC     = ResC.reshape((nor,nos*noi,6)) # reshape to region as separate dimension
    Res_rC_agg = Res_rC.sum(axis=0) # sum over regions

r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define ESC plot':
        break    
    r += 1
r += 1

ctitles = []
ctypes  = []
cregs   = []
cscens  = []

while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ctitles.append(CF[CS].cell(r,2).value)
    ctypes.append(CF[CS].cell(r,3).value)
    cregs.append(CF[CS].cell(r,4).value)
    cscens.append(CF[CS].cell(r,5).value)
    r += 1

# determine ESC indicators and plot ESC cascades
# find ESC parameters in Res array: outer index: Ar, middle index: ti, inner indicator: scen

for c in range(0,len(ctitles)):
    if ctypes[c] == 'version_1':
        if cscens[c] == 'All':
            nocs = nos # number of cascade scenarios nocs = number of scenarios nos
        else:
            nocs = len(cscens[c].split(';'))
        if cregs[c] == 'Global':
            rind = -1
        else:
            rind = Ar.index(cregs[c])
        # Define data container
        esc_data = np.zeros((6,46,nocs)) # 6 decoupling indices, 46 years, nocs scenarios
        
        # Decoupling 1: Lower stock levels
        # Decoupling 2: Operational energy per stock, etc.
        edx  = ti.index('Energy cons., use phase, res+non-res buildings')
        rebx = ti.index('In-use stock, res. buildings')
        nrbx = ti.index('In-use stock, nonres. buildings')
        matm = ti.index('Final consumption of materials')
        rec1 = ti.index('ReUse of materials in products, construction grade steel')
        rec2 = ti.index('ReUse of materials in products, concrete')
        rec3 = ti.index('ReUse of materials in products, wood and wood products')
        rec4 = ti.index('Secondary construction steel')
        ghg1 = ti.index('GHG emissions, res. buildings, use phase')
        ghg2 = ti.index('GHG emissions, non-res. buildings, use phase')
        ghg3 = ti.index('GHG emissions, res+non-res buildings, energy supply')
        maf1 = ti.index('Material footprint, metal ores, system-wide')
        maf2 = ti.index('Material footprint, non-metallic minerals, system-wide')
        maf3 = ti.index('Material footprint, biomass (dry weight), system-wide')

                
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
            
        if rind == -1:
            for sc in range(0,nocs):
                targetpos_edx  = edx  * nos + scen.index(cscenss[sc]) # position in Res array: outer index: region, middle index: indicator, inner index: scenario
                targetpos_rebx = rebx * nos + scen.index(cscenss[sc])
                targetpos_nrbx = nrbx * nos + scen.index(cscenss[sc])
                targetpos_matm = matm * nos + scen.index(cscenss[sc])
                targetpos_rec1 = rec1 * nos + scen.index(cscenss[sc])
                targetpos_rec2 = rec2 * nos + scen.index(cscenss[sc])
                targetpos_rec3 = rec3 * nos + scen.index(cscenss[sc])
                targetpos_rec4 = rec4 * nos + scen.index(cscenss[sc])
                targetpos_ghg1 = ghg1 * nos + scen.index(cscenss[sc])
                targetpos_ghg2 = ghg1 * nos + scen.index(cscenss[sc])
                targetpos_ghg3 = ghg1 * nos + scen.index(cscenss[sc])
                targetpos_maf1 = maf1 * nos + scen.index(cscenss[sc])
                targetpos_maf2 = maf2 * nos + scen.index(cscenss[sc])
                targetpos_maf3 = maf3 * nos + scen.index(cscenss[sc])
                esc_data[0,:,sc] = Res_r_agg[targetpos_rebx,:] + Res_r_agg[targetpos_nrbx,:]
                esc_data[1,:,sc] = Res_r_agg[targetpos_edx,:] / (Res_r_agg[targetpos_rebx,:] + Res_r_agg[targetpos_nrbx,:])   
                esc_data[2,:,sc] = Res_r_agg[targetpos_matm,:] / (Res_r_agg[targetpos_rebx,:] + Res_r_agg[targetpos_nrbx,:])  
                esc_data[3,:,sc] = (Res_r_agg[targetpos_rec1,:] + Res_r_agg[targetpos_rec2,:] + Res_r_agg[targetpos_rec3,:] + Res_r_agg[targetpos_rec4,:]) / Res_r_agg[targetpos_matm,:]
                esc_data[4,:,sc] = (Res_r_agg[targetpos_maf1,:] + Res_r_agg[targetpos_maf2,:] + Res_r_agg[targetpos_maf3,:]) / Res_r_agg[targetpos_matm,:]
                esc_data[5,:,sc] = (Res_r_agg[targetpos_ghg1,:] + Res_r_agg[targetpos_ghg2,:] + Res_r_agg[targetpos_ghg3,:]) / Res_r_agg[targetpos_edx,:]
        else:
            for sc in range(0,nocs):
                targetpos_edx  = rind*nos*noi + edx  * nos + scen.index(cscenss[sc]) # position in Res array: outer index: region, middle index: indicator, inner index: scenario
                targetpos_rebx = rind*nos*noi + rebx * nos + scen.index(cscenss[sc])
                targetpos_nrbx = rind*nos*noi + nrbx * nos + scen.index(cscenss[sc])
                targetpos_matm = rind*nos*noi + matm * nos + scen.index(cscenss[sc])
                targetpos_rec1 = rind*nos*noi + rec1 * nos + scen.index(cscenss[sc])
                targetpos_rec2 = rind*nos*noi + rec2 * nos + scen.index(cscenss[sc])
                targetpos_rec3 = rind*nos*noi + rec3 * nos + scen.index(cscenss[sc])
                targetpos_rec4 = rind*nos*noi + rec4 * nos + scen.index(cscenss[sc])
                targetpos_ghg1 = rind*nos*noi + ghg1 * nos + scen.index(cscenss[sc])
                targetpos_ghg2 = rind*nos*noi + ghg2 * nos + scen.index(cscenss[sc])
                targetpos_ghg3 = rind*nos*noi + ghg3 * nos + scen.index(cscenss[sc])
                targetpos_maf1 = rind*nos*noi + maf1 * nos + scen.index(cscenss[sc])
                targetpos_maf2 = rind*nos*noi + maf2 * nos + scen.index(cscenss[sc])
                targetpos_maf3 = rind*nos*noi + maf3 * nos + scen.index(cscenss[sc])
                esc_data[0,:,sc] = Res[targetpos_rebx,:] + Res[targetpos_nrbx,:]
                esc_data[1,:,sc] = Res[targetpos_edx,:] / (Res[targetpos_rebx,:] + Res[targetpos_nrbx,:])    
                esc_data[2,:,sc] = Res[targetpos_matm,:] / (Res[targetpos_rebx,:] + Res[targetpos_nrbx,:])    
                esc_data[3,:,sc] = (Res[targetpos_rec1,:] + Res[targetpos_rec2,:] + Res[targetpos_rec3,:] + Res[targetpos_rec4,:]) / Res[targetpos_matm,:]
                esc_data[4,:,sc] = (Res[targetpos_maf1,:] + Res[targetpos_maf2,:] + Res[targetpos_maf3,:]) / Res[targetpos_matm,:]
                esc_data[5,:,sc] = (Res[targetpos_ghg1,:] + Res[targetpos_ghg2,:] + Res[targetpos_ghg3,:]) / Res[targetpos_edx,:]
        # normalize esc data so that all time series are in relation to SSP2 (maximum):
        for mm in range(0,46):
            esc_data[0,mm,:] = esc_data[0,mm,:] / esc_data[0,mm,:].max()

        # normalize the esc_data
        esc_data[:,0,:] = 0 # start year is 2016, 2015 data are not considered.
        esc_data_Divisor = np.einsum('cs,t->cts',esc_data[:,1,:],np.ones(46))
        esc_data_n = np.divide(esc_data, esc_data_Divisor, out=np.zeros_like(esc_data_Divisor), where=esc_data_Divisor!=0)
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=6 , figsize=(21, 3))        
        fig.suptitle('Energy service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        axs[0].plot(np.arange(2016,2061), esc_data_n[0,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[0,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[0].set_title('(1) Stock per service')
        
        axs[1].plot(np.arange(2016,2061), esc_data_n[1,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[1,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[1].set_title('(2) Operational energy per stock')
        
        axs[2].plot(np.arange(2016,2061), esc_data_n[2,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[2,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[2].set_title('(3) Build-up material per stock')
        
        axs[3].plot(np.arange(2016,2061), esc_data_n[3,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[3,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[3].set_title('(4) Circlar material use rate')
        
        axs[4].plot(np.arange(2016,2061), esc_data_n[4,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[4,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[4].set_title('(5) material footprint per final consumption')        
        
        axs[5].plot(np.arange(2016,2061), esc_data_n[5,1::,:], linewidth = 1.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[5,1::,:], linewidth = 1.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[5].set_title('(6) GHG emissions per energy use')
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=1, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '.png'), dpi=150, bbox_inches='tight')



#
#
#
# The end.
#
#
#