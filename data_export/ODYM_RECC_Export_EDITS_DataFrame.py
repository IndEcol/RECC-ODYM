# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 15:42:03 2021

@author: spauliuk

This script takes a number of RECC scenarios (as defined in list), 
loads a number of results and then compiles them into an excel workbook/csv 
file for parsing and checking in the IIASA EDITS database.

See https://github.com/iiasa/irp-internal-workflow for naming conventions.

"""
# Import required libraries:
import os
import openpyxl
import numpy as np
import pandas as pd
import uuid
from pathlib import Path
import RECC_Paths # Import path file
pE = RECC_Paths.EDITS_path

def get_RECC_resfile_pos(Label,Resultsheet):
    # Find the index for the given Label
    idx = 1    
    while True:
        if Resultsheet.cell(idx, 1).value == Label:
            break # that gives us the right index to read the Label from the result table.
        idx += 1 
    return idx

# Create UUID of script run
Current_UUID = str(uuid.uuid4())
RECC_Paths.results_path_save = Path(RECC_Paths.results_path_eval) / ('RECC_Results_' + Current_UUID)
if not os.path.exists(RECC_Paths.results_path_save): # Create scrip run results directory.
    os.makedirs(RECC_Paths.results_path_save)

# Definitions
SpecPath      = os.path.join(pE,'EDITS_EXPORT_RECCv2.5.xlsx')
SpecFile      = openpyxl.load_workbook(SpecPath)
SpecSheet     = SpecFile['RECC_Export_EDITS']

ResFileP      = os.path.join(pE,'OUTPUTS - RECCv2.5_templateV4.xlsx')
ResFile       = openpyxl.load_workbook(ResFileP)
RS            = ResFile['data']

# Read specs and variable matchings
scen = [] # list of scenarios
rebf = [] # list of matching reb results folders
nrbf = [] # list of matching nrb results folders
offs = [] # list of offsets for the different scenarios rel. to starting position

for m in range(0,4):
    scen.append(SpecSheet.cell(m+3,2).value)
    rebf.append(SpecSheet.cell(m+3,6).value)
    nrbf.append(SpecSheet.cell(m+3,7).value)
    offs.append(SpecSheet.cell(m+3,5).value)
    
# iterate over scenarios, parse and write results:
outline = 2    
for m in range(10,1000):
    if SpecSheet.cell(m,2).value is None:
        break
    Edits_var = SpecSheet.cell(m,2).value
    Edits_uni = SpecSheet.cell(m,3).value
    RECC_var  = SpecSheet.cell(m,5).value
    Sector    = SpecSheet.cell(m,4).value
    Conv_Fact = SpecSheet.cell(m,7).value
    # check the four different scenarios:
    if Sector == 'reb':
        RECC_FList = rebf
    if Sector == 'nrb':
        RECC_FList = nrbf        
    for r in range(0,4): # loop over the four scenarios
        RECC_ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,RECC_FList[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
        RECC_Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,RECC_FList[r],RECC_ResFile[0]))
        RECC_Resultsheet2 = RECC_Resultfile2['Model_Results']
        idx = get_RECC_resfile_pos(RECC_var,RECC_Resultsheet2)
        # Write results to export file
        RS.cell(outline,1).value = 'RECC v2.5'
        RS.cell(outline,2).value = scen[r]
        RS.cell(outline,3).value = RECC_Resultsheet2.cell(idx+offs[r],3).value # Region
        RS.cell(outline,4).value = Edits_var # Variable
        RS.cell(outline,5).value = Edits_uni # Unit
        # write values
        print(idx+offs[r])
        for t in range(0,45):
            RS.cell(outline,t+6).value = RECC_Resultsheet2.cell(idx+offs[r],t+9).value * Conv_Fact # Value
        outline +=1
        

# Save exported results
ResFile.save(os.path.join(pE,'OUTPUTS - RECCv2.5_templateV4.xlsx')) 
                 
#
#
#
#
#
#
