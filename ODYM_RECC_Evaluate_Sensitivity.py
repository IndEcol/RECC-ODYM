# -*- coding: utf-8 -*-
"""
Created on Wed Oct 17 10:37:00 2018

@author: spauliuk
"""
def main(RegionalScope,FolderList,SectorString,Current_UUID):
        
    import openpyxl
    import numpy as np
    import matplotlib.pyplot as plt  
    import pylab
    import RECC_Paths # Import path file    
    import os
    
    RECC_Paths.results_path_save = os.path.join(RECC_Paths.results_path_eval,'RECC_Results_' + Current_UUID)
    
    PlotExpResolution = 150 # dpi 150 for overview or 500 for paper
    
    #Sensitivity analysis folder order, by default, all strategies are off, one by one is implemented each at a time.
    ## Vehicles:
    #Baseline (no RE)
    #FabYieldImprovement
    #FabScrapDiversion
    #EoL_RR_Improvement
    #ChangeMaterialComposition
    #ReduceMaterialContent
    #ReUse_Materials
    #LifeTimeExtension
    #CarSharing
    #RideSharing
    #NoRecycling
    
    ## Buildings:
    #Baseline (no RE)
    #FabYieldImprovement
    #FabScrapDiversion
    #EoL_RR_Improvement
    #ChangeMaterialComposition
    #ReduceMaterialContent
    #ReUse_Materials
    #LifeTimeExtension
    #MoreIntenseUse
    #NoRecycling
    
    # Sensitivity plots
    
    NS = 3 # SSP
    NR = 2 # RCP
    
    if SectorString == 'pav':
        NE      = 11 # no of sensitivities for cascade
        LWE     = ['Higher yield, manuf. efficiency','Fab scrap diversion','EoL_RR_Improvement','Material substitution','ReduceMaterialContent','ReUse_Materials','LifeTimeExtension','CarSharing','RideSharing','No recycling']
        Offset1 = 9.9
        
    if SectorString == 'reb':
        NE      = 10 # no of Res. eff. scenarios for cascade
        LWE     = ['Higher yield, manuf. efficiency','Fab scrap diversion','EoL_RR_Improvement','Material substitution','ReduceMaterialContent','ReUse_Materials','LifeTimeExtension','More intense use','No recycling']
        Offset1 = 8.9
        
    if SectorString == 'nrb':
        NE      = 10 # no of Res. eff. scenarios for cascade
        LWE     = ['Higher yield, manuf. efficiency','Fab scrap diversion','EoL_RR_Improvement','Material substitution','ReduceMaterialContent','ReUse_Materials','LifeTimeExtension','More intense use','No recycling']
        Offset1 = 8.9   
        
    CumEms_Sens2050      = np.zeros((NS,NR,NE)) # SSP-Scenario x RES scenario
    CumEms_Sens2060      = np.zeros((NS,NR,NE)) # SSP-Scenario x RES scenario
    AnnEms2030_Sens      = np.zeros((NS,NR,NE)) # SSP-Scenario x RES scenario
    AnnEms2050_Sens      = np.zeros((NS,NR,NE)) # SSP-Scenario x RES scenario
    AvgDecadalEms        = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RES scenario x 4 decades
    # for use phase di emissions:
    UseCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalUseEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    # for material-related emissions:
    MatCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalMatEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    # for manufacturing-related emissions:
    ManCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalManEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    # for forestry and wood waste related emissions:
    ForCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalForEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060 
    # for recycling credit:
    RecCreditCum2050 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditCum2060 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditAnn2030 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditAnn2050 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    RecCreditAvgDec  = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    
    # get result items:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[0])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[0],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    # Find the index for sysem-wide emissions, the recycling credit and others:
    swe = 1    
    while True:
        if Resultsheet2.cell(swe+1, 1).value == 'GHG emissions, system-wide _3579di':
            break # that gives us the right index to read the recycling credit from the result table.
        swe += 1        
    rci = 1
    while True:
        if Resultsheet2.cell(rci+1, 1).value == 'GHG emissions, recycling credits':
            break # that gives us the right index to read the recycling credit from the result table.
        rci += 1
    mci = 1
    while True:
        if Resultsheet2.cell(mci+1, 1).value == 'GHG emissions, material cycle industries and their energy supply _3di_9di':
            break # that gives us the right index to read the recycling credit from the result table.
        mci += 1

    up1i = 1
    while True:
        if Resultsheet2.cell(up1i+1, 1).value == 'GHG emissions, use phase _7d':
            break # that gives us the right index from the result table.
        up1i += 1  
    up2i = 1
    while True:
        if Resultsheet2.cell(up2i+1, 1).value == 'GHG emissions, use phase scope 2 (electricity) _7i':
            break # that gives us the right index from the result table.
        up2i += 1  
    up3i = 1
    while True:
        if Resultsheet2.cell(up3i+1, 1).value == 'GHG emissions, use phase other indirect (non-el.) _7i':
            break # that gives us the right index from the result table.
        up3i += 1  

    mfi = 1
    while True:
        if Resultsheet2.cell(mfi+1, 1).value == 'GHG emissions, manufacturing _5i, all':
            break # that gives us the right index from the result table.
        mfi += 1 
    fci = 1
    while True:
        if Resultsheet2.cell(fci+1, 1).value == 'GHG emissions, energy recovery from waste wood (biogenic C plus energy substitution within System)':
            break # that gives us the right index from the result table.
        fci += 1 
    wci = 1
    while True:
        if Resultsheet2.cell(wci+1, 1).value == 'GHG sequestration by forests (w. neg. sign)':
            break # that gives us the right index from the result table.
        wci += 1 
    
    for r in range(0,NE): # RE scenario                
        # import system-wide GHG and material-related emissions
        ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
        Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[r],ResFile[0]))
        Resultsheet2 = Resultfile2['Model_Results']
        # system-wide emissions results
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time
                    CumEms_Sens2050[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value
                for t in range(0,45): # time
                    CumEms_Sens2060[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value    
                AnnEms2030_Sens[s,c,r]   = Resultsheet2.cell(swe+ 2*s +c+1,23).value
                AnnEms2050_Sens[s,c,r]   = Resultsheet2.cell(swe+ 2*s +c+1,43).value
                AvgDecadalEms[s,c,r,0]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(13,23)])/10
                AvgDecadalEms[s,c,r,1]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(23,33)])/10
                AvgDecadalEms[s,c,r,2]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(33,43)])/10
                AvgDecadalEms[s,c,r,3]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(43,53)])/10
        # Use phase results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    UseCumEms2050[s,c,r] += Resultsheet2.cell(up1i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up2i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up3i+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    UseCumEms2060[s,c,r] += Resultsheet2.cell(up1i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up2i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up3i+ 2*s +c+1,t+9).value                    
                UseAnnEms2030[s,c,r]      = Resultsheet2.cell(up1i+ 2*s +c+1,23).value  + Resultsheet2.cell(up2i+ 2*s +c+1,23).value  + Resultsheet2.cell(up3i+ 2*s +c+1,23).value  
                UseAnnEms2050[s,c,r]      = Resultsheet2.cell(up1i+ 2*s +c+1,43).value  + Resultsheet2.cell(up2i+ 2*s +c+1,43).value  + Resultsheet2.cell(up3i+ 2*s +c+1,43).value  
                AvgDecadalUseEms[s,c,r,0] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalUseEms[s,c,r,1] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalUseEms[s,c,r,2] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalUseEms[s,c,r,3] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(43,53)])/10          
        # Material results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    MatCumEms2050[s,c,r] += Resultsheet2.cell(mci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    MatCumEms2060[s,c,r] += Resultsheet2.cell(mci+ 2*s +c+1,t+9).value                    
                MatAnnEms2030[s,c,r]      = Resultsheet2.cell(mci+ 2*s +c+1,23).value
                MatAnnEms2050[s,c,r]      = Resultsheet2.cell(mci+ 2*s +c+1,43).value
                AvgDecadalMatEms[s,c,r,0] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalMatEms[s,c,r,1] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalMatEms[s,c,r,2] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalMatEms[s,c,r,3] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(43,53)])/10    
        # Manufacturing results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    ManCumEms2050[s,c,r] += Resultsheet2.cell(mfi+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    ManCumEms2060[s,c,r] += Resultsheet2.cell(mfi+ 2*s +c+1,t+9).value                    
                ManAnnEms2030[s,c,r]      = Resultsheet2.cell(mfi+ 2*s +c+1,23).value
                ManAnnEms2050[s,c,r]      = Resultsheet2.cell(mfi+ 2*s +c+1,43).value
                AvgDecadalManEms[s,c,r,0] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalManEms[s,c,r,1] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalManEms[s,c,r,2] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalManEms[s,c,r,3] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(43,53)])/10                 
        # Forestry results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    ForCumEms2050[s,c,r] += Resultsheet2.cell(fci+ 2*s +c+1,t+9).value + Resultsheet2.cell(wci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    ForCumEms2060[s,c,r] += Resultsheet2.cell(fci+ 2*s +c+1,t+9).value + Resultsheet2.cell(wci+ 2*s +c+1,t+9).value                    
                ForAnnEms2030[s,c,r]      = Resultsheet2.cell(fci+ 2*s +c+1,23).value  + Resultsheet2.cell(wci+ 2*s +c+1,23).value
                ForAnnEms2050[s,c,r]      = Resultsheet2.cell(fci+ 2*s +c+1,43).value  + Resultsheet2.cell(wci+ 2*s +c+1,43).value
                AvgDecadalForEms[s,c,r,0] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalForEms[s,c,r,1] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalForEms[s,c,r,2] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalForEms[s,c,r,3] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(43,53)])/10              
        # recycling credit
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    RecCreditCum2050[s,c,r]+= Resultsheet2.cell(rci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    RecCreditCum2060[s,c,r]+= Resultsheet2.cell(rci+ 2*s +c+1,t+9).value
                RecCreditAnn2030[s,c,r]     = Resultsheet2.cell(rci+ 2*s +c+1,23).value
                RecCreditAnn2050[s,c,r]     = Resultsheet2.cell(rci+ 2*s +c+1,43).value
                RecCreditAvgDec[s,c,r,0]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(13,23)])/10
                RecCreditAvgDec[s,c,r,1]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(23,33)])/10
                RecCreditAvgDec[s,c,r,2]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(33,43)])/10
                RecCreditAvgDec[s,c,r,3]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(43,53)])/10                       
    
    ### Tornado plot for sensitivity
            
    MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.1)) # select 12 colors from the 'Paired' color map.            
    
    Scens = ['LED','SSP1','SSP2']
    Rcens  = ['Base','RCP2_6']
    Titles = ['Ann_2030_GHG_Sens','Ann_2050_GHG_Sens','Cum_2050_GHG_Sens']
    
    for npp in range(0,3): # three different variables plotted
        for m in range(0,NS): # SSP
            for c in range(0,NR): # RCP
                # Fill Data container with indices SSP x RES
                #2030 emissions
                if npp == 0:
                    Data = AnnEms2030_Sens[:,c,1::]-np.einsum('S,n->Sn',AnnEms2030_Sens[:,c,0],np.ones(NE-1))
                    Base = AnnEms2030_Sens[:,c,0]
                #2050 emissions
                if npp == 1:
                    Data = AnnEms2050_Sens[:,c,1::]-np.einsum('S,n->Sn',AnnEms2050_Sens[:,c,0],np.ones(NE-1))
                    Base = AnnEms2050_Sens[:,c,0]
                #2050 cum. emissions
                if npp == 2:
                    Data = CumEms_Sens2050[:,c,1::]-np.einsum('S,n->Sn',CumEms_Sens2050[:,c,0],np.ones(NE-1))
                    Base = CumEms_Sens2050[:,c,0]  
                                        
                # plot results
                fig  = plt.figure(figsize=(5,NE))
                ax1  = plt.axes([0.08,0.08,0.85,0.9])
                
                Poss = np.arange(NE-1,0,-1)
                ax1.barh(Poss,Data[m,:], color=MyColorCycle, lw=0.4)
        
                # plot text and labels
                for mm in range(0,NE-1):
                    plt.text(Data[m,:].min()*0.9, Offset1-mm, LWE[mm] + ': ' + ("%3.0f" % Data[m,mm]),fontsize=14,fontweight='bold')          
                
                plt.text(Data[m,:].min()*0.59, 10.5, 'Baseline: ' + ("%3.0f" % Base[m]) + r' Mt CO$_2$-eq/yr.',fontsize=14,fontweight='bold')
        
                plt.title(RegionalScope + '_' + SectorString + '_' + Titles[npp] + '_' + Scens[m] + '_' + Rcens[c], fontsize = 18)
                plt.ylabel('RE strategies.', fontsize = 18)
                plt.xlabel(r'Mt of CO$_2$-eq.', fontsize = 14)
                #plt.xticks([0.25,1.25,2.25,3.25,4.25,5.25])
                plt.yticks([])
                #ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
                #plt_lgd  = plt.legend(handles = ProxyHandlesList,labels = LWE,shadow = False, prop={'size':16},ncol=1, loc = 'upper right' ,bbox_to_anchor=(1.91, 1)) 
                plt.axis([Data[m,:].min() -15, Data[m,:].max() +80, 0.3, 11.1])
            
                plt.show()
                fig_name = RegionalScope + '_' + SectorString + '_' + Titles[npp] + '_' + Scens[m] + '_' + Rcens[c] + '.png'
                fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = PlotExpResolution, bbox_inches='tight')             
      
    return CumEms_Sens2050, CumEms_Sens2060, AnnEms2030_Sens, AnnEms2050_Sens, AvgDecadalEms, UseCumEms2050, UseCumEms2060, UseAnnEms2030, UseAnnEms2050, AvgDecadalUseEms, MatCumEms2050, MatCumEms2060, MatAnnEms2030, MatAnnEms2050, AvgDecadalMatEms, ManCumEms2050, ManCumEms2060, ManAnnEms2030, ManAnnEms2050, AvgDecadalManEms, ForCumEms2050, ForCumEms2060, ForAnnEms2030, ForAnnEms2050, AvgDecadalForEms, RecCreditCum2050, RecCreditCum2060, RecCreditAnn2030, RecCreditAnn2050, RecCreditAvgDec

# code for script to be run as standalone function
if __name__ == "__main__":
    main()
