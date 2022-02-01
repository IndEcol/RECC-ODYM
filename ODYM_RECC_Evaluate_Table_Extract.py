# -*- coding: utf-8 -*-
"""
Created on Wed Oct 17 10:37:00 2018

@author: spauliuk
"""
def main(RegionalScope,ThreeSectoList,Current_UUID):
    
    import openpyxl
    import numpy as np
    import matplotlib.pyplot as plt  
    import os
    import RECC_Paths # Import path file   
    
    RECC_Paths.results_path_save = os.path.join(RECC_Paths.results_path_eval,'RECC_Results_' + Current_UUID)
    
    GHG_Table_Overview   = np.zeros((4,6,2)) # Table GHG overview format 4 scopes, 6 times, 2 RCP

    # No ME scenario
    # Find result fild with common start but unique UUID in name:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,ThreeSectoList[0])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,ThreeSectoList[0],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    # Find the index for the recycling credit and others:
    tci = 1
    while True:
        if Resultsheet2.cell(tci+1, 1).value == 'GHG emissions, system-wide _3579di':
            break # that gives us the right index to read the recycling credit from the result table.
        tci += 1
    
    rci = 1
    while True:
        if Resultsheet2.cell(rci+1, 1).value == 'GHG emissions, recycling credits':
            break # that gives us the right index to read the recycling credit from the result table.
        rci += 1
    mci = 1
    while True:
        if Resultsheet2.cell(mci+1, 1).value == 'GHG emissions, material cycle industries and their energy supply _3di_9di':
            break # that gives us the right index to read the recycling credit from the result table.
        mci += 1
        
    for nnn in range(0,2):
        GHG_Table_Overview[0,0,nnn] = Resultsheet2.cell(tci +3 + nnn,9).value
        GHG_Table_Overview[0,1,nnn] = Resultsheet2.cell(tci +3 + nnn,13).value
        GHG_Table_Overview[0,2,nnn] = Resultsheet2.cell(tci +3 + nnn,23).value
        GHG_Table_Overview[0,3,nnn] = Resultsheet2.cell(tci +3 + nnn,33).value
        GHG_Table_Overview[0,4,nnn] = Resultsheet2.cell(tci +3 + nnn,43).value
        GHG_Table_Overview[0,5,nnn] = Resultsheet2.cell(tci +3 + nnn,53).value
        
        GHG_Table_Overview[2,0,nnn] = Resultsheet2.cell(mci +3 + nnn,9).value
        GHG_Table_Overview[2,1,nnn] = Resultsheet2.cell(mci +3 + nnn,13).value
        GHG_Table_Overview[2,2,nnn] = Resultsheet2.cell(mci +3 + nnn,23).value
        GHG_Table_Overview[2,3,nnn] = Resultsheet2.cell(mci +3 + nnn,33).value
        GHG_Table_Overview[2,4,nnn] = Resultsheet2.cell(mci +3 + nnn,43).value
        GHG_Table_Overview[2,5,nnn] = Resultsheet2.cell(mci +3 + nnn,53).value
                
    # Full ME scenario
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,ThreeSectoList[-1])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,ThreeSectoList[-1],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    # Find the index for the recycling credit and others:
    tci = 1
    while True:
        if Resultsheet2.cell(tci+1, 1).value == 'GHG emissions, system-wide _3579di':
            break # that gives us the right index to read the recycling credit from the result table.
        tci += 1
    
    rci = 1
    while True:
        if Resultsheet2.cell(rci+1, 1).value == 'GHG emissions, recycling credits':
            break # that gives us the right index to read the recycling credit from the result table.
        rci += 1
    mci = 1
    while True:
        if Resultsheet2.cell(mci+1, 1).value == 'GHG emissions, material cycle industries and their energy supply _3di_9di':
            break # that gives us the right index to read the recycling credit from the result table.
        mci += 1
        
    for nnn in range(0,2):        
        GHG_Table_Overview[1,0,nnn] = Resultsheet2.cell(tci +3 + nnn,9).value
        GHG_Table_Overview[1,1,nnn] = Resultsheet2.cell(tci +3 + nnn,13).value
        GHG_Table_Overview[1,2,nnn] = Resultsheet2.cell(tci +3 + nnn,23).value
        GHG_Table_Overview[1,3,nnn] = Resultsheet2.cell(tci +3 + nnn,33).value
        GHG_Table_Overview[1,4,nnn] = Resultsheet2.cell(tci +3 + nnn,43).value
        GHG_Table_Overview[1,5,nnn] = Resultsheet2.cell(tci +3 + nnn,53).value
        
        GHG_Table_Overview[3,0,nnn] = Resultsheet2.cell(mci +3 + nnn,9).value
        GHG_Table_Overview[3,1,nnn] = Resultsheet2.cell(mci +3 + nnn,13).value
        GHG_Table_Overview[3,2,nnn] = Resultsheet2.cell(mci +3 + nnn,23).value
        GHG_Table_Overview[3,3,nnn] = Resultsheet2.cell(mci +3 + nnn,33).value
        GHG_Table_Overview[3,4,nnn] = Resultsheet2.cell(mci +3 + nnn,43).value
        GHG_Table_Overview[3,5,nnn] = Resultsheet2.cell(mci +3 + nnn,53).value
        

    return GHG_Table_Overview

# code for script to be run as standalone function
if __name__ == "__main__":
    main()
