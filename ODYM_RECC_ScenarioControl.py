# -*- coding: utf-8 -*-
"""
Created on Jan 5th, 2020, as copy of RECC_ScenarioControl_V2_2.py

@author: spauliuk
"""

"""

File RECC_ScenarioControl.py

Script that modifies the RECC config file to run a list of scenarios and executes RECC main script for each scenario config.

"""

# Import required libraries:
import os
import openpyxl

import RECC_Paths # Import path file
import ODYM_RECC_Main

#ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
ScenarioSetting = 'Test'
#ScenarioSetting = 'Buildings_Global_Config_list'
#ScenarioSetting = 'Buildings_Global_EDITS'
#ScenarioSetting = 'Glob_BLd_MediumWood'
#ScenarioSetting = 'CRAFT_Coupling_Config_list'
#ScenarioSetting = 'pav_reb_Config_list_all'

# open scenario sheet
ModelConfigListFile  = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_ModelConfig_List.xlsx'))
ModelConfigListSheet = ModelConfigListFile[ScenarioSetting]
SheetName = 'Buildings_Global'
#Read control lines and execute main model script
ResultFolders = []
Row = 3
# search for script config list entry
while ModelConfigListSheet.cell(Row+1, 3).value != None:
    RegionalScope = ModelConfigListSheet.cell(Row+1, 3).value
    print(RegionalScope)
    Config = {}
    for m in range(3,28):
        Config[ModelConfigListSheet.cell(3, m+1).value] = ModelConfigListSheet.cell(Row+1, m+1).value
    Row += 1
    # rewrite RECC model config
    mywb = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))
    
    sheet = mywb.get_sheet_by_name('Cover')
    sheet['D4'] = SheetName
    sheet = mywb.get_sheet_by_name(SheetName)
    sheet['D7']   = RegionalScope
    sheet['G21']  = Config['RegionSelect']
    sheet['G28']  = Config['Products'] # manufacturing sectors have the same indices as the products.
    sheet['G29']  = Config['Sectors']
    sheet['G30']  = Config['Products']
    sheet['G34']  = Config['NonresidentialBuildings']
    sheet['G49']  = Config['Regions32goods']
    # The indices below need to be updated when new parameters are added to the parameter list
    sheet['D203'] = Config['Logging_Verbosity']
    sheet['D204'] = Config['Include_REStrategy_FabYieldImprovement']
    sheet['D205'] = Config['Include_REStrategy_FabScrapDiversion']
    sheet['D206'] = Config['Include_REStrategy_EoL_RR_Improvement']
    sheet['D207'] = Config['ScrapExport']
    sheet['D208'] = Config['ScrapExportRecyclingCredit']
    sheet['D209'] = Config['IncludeRecycling']
    sheet['D210'] = Config['Include_REStrategy_MaterialSubstitution']
    sheet['D211'] = Config['Include_REStrategy_UsingLessMaterialByDesign']
    sheet['D212'] = Config['Include_REStrategy_ReUse']
    sheet['D213'] = Config['Include_REStrategy_LifeTimeExtension']
    sheet['D214'] = Config['Include_REStrategy_MoreIntenseUse']
    sheet['D215'] = Config['Include_REStrategy_CarSharing']
    sheet['D216'] = Config['Include_REStrategy_RideSharing']
    sheet['D217'] = Config['Include_REStrategy_ModalSplit']
    sheet['D218'] = Config['SectorSelect']
    sheet['D219'] = Config['Include_Renovation_reb']
    sheet['D220'] = Config['Include_Renovation_nrb']
    sheet['D221'] = Config['No_EE_Improvements']
    sheet['D240'] = Config['PlotResolution']
    
    mywb.save(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))

    # run the ODYM-RECC model
    OutputDict = ODYM_RECC_Main.main()
    ResultFolders.append(OutputDict['Name_Scenario'])

# Export ResultFolders:
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'ResultFolders'
Fr = 3
for Fname in ResultFolders:
    ws1.cell(row=Fr+1, column=4).value = Fname 
    Fr +=1
book.save(os.path.join(RECC_Paths.results_path,'ResultFolders.xlsx'))   
#
#
