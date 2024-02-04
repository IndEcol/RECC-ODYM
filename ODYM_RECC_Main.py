# -*- coding: utf-8 -*-
"""
Created on February 21, 2020, as copy of ODYM_RECC_V2_3.py

@authors: spauliuk
"""

"""
File ODYM_RECC_Main.py

Contains the ODYM-RECC model v 2.5 for the resource efficiency climate change mitigation nexus
Model version 2_5: Global coverage of five sectors.
passenger vehicles, residential buildings, non-residential buildings, electricity generation, and appliances.

dependencies:
    numpy >= 1.9
    scipy >= 0.14

"""
def main():
    # Import required libraries:
    import os
    import sys
    import logging as log
    import openpyxl
    import numpy as np
    import time
    import datetime
    #import scipy.io
    import pandas as pd
    import shutil   
    import uuid
    import matplotlib.pyplot as plt   
    from matplotlib.lines import Line2D
    import importlib
    import getpass
    from copy import deepcopy
    from tqdm import tqdm
    import scipy.stats
    #from scipy.interpolate import interp1d
    from scipy.interpolate import make_interp_spline
    import pylab
    import pickle
    
    import RECC_Paths # Import path file
    log.getLogger('matplotlib.font_manager').disabled = True    # required for preventing debugging messages in some console versions
    #import re
    __version__ = str('2.5')
    ##################################
    #    Section 1)  Initialize      #
    ##################################
    # add ODYM module directory to system path
    sys.path.insert(0, os.path.join(os.path.join(RECC_Paths.odym_path,'odym'),'modules'))
    ### 1.1.) Read main script parameters
    # Mylog.info('### 1.1 - Read main script parameters')
    ProjectSpecs_Name_ConFile = 'RECC_Config.xlsx'
    Model_Configfile = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), data_only=True)
    ScriptConfig = {'Model Setting': Model_Configfile['Cover'].cell(4,4).value}
    Model_Configsheet = Model_Configfile[ScriptConfig['Model Setting']]
    #Read debug modus:   
    DebugCounter = 0
    while Model_Configsheet.cell(DebugCounter+1, 3).value != 'Logging_Verbosity':
        DebugCounter += 1
    ScriptConfig['Logging_Verbosity'] = Model_Configsheet.cell(DebugCounter+1,4).value # Read loggin verbosity once entry was reached.    
    # Extract user name from main file
    ProjectSpecs_User_Name     = getpass.getuser()
    
    # import packages whose location is now on the system path:    
    import ODYM_Classes as msc # import the ODYM class file
    importlib.reload(msc)
    import ODYM_Functions as msf  # import the ODYM function file
    importlib.reload(msf)
    import dynamic_stock_model as dsm # import the dynamic stock model library
    importlib.reload(dsm)
    
    Name_Script        = Model_Configsheet.cell(6,4).value
    if Name_Script != 'ODYM_RECC_Main':  # Name of this script must equal the specified name in the Excel config file
        raise AssertionError('Fatal: The name of the current script does not match to the sript name specfied in the project configuration file. Exiting the script.')
    # the model will terminate if the name of the script that is run is not identical to the script name specified in the config file.
    Name_Scenario            = Model_Configsheet.cell(7,4).value # Regional scope as torso for scenario name
    StartTime                = datetime.datetime.now()
    TimeString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day) + '__' + str(StartTime.hour) + '_' + str(StartTime.minute) + '_' + str(StartTime.second)
    #DateString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day)
    ProjectSpecs_Path_Result = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString )
    
    if not os.path.exists(ProjectSpecs_Path_Result): # Create model run results directory.
        os.makedirs(ProjectSpecs_Path_Result)
    # Initialize logger
    if ScriptConfig['Logging_Verbosity'] == 'DEBUG':
        log_verbosity = eval("log.DEBUG")  
    log_filename = Name_Scenario + '__' + TimeString + '.md'
    [Mylog, console_log, file_log] = msf.function_logger(log_filename, ProjectSpecs_Path_Result,
                                                         log_verbosity, log_verbosity)
    # log header and general information
    Time_Start = time.time()
    ScriptConfig['Current_UUID'] = str(uuid.uuid4())
    Mylog.info('# Simulation from ' + time.asctime())
    Mylog.info('Unique ID of scenario run: ' + ScriptConfig['Current_UUID'])
    
    ### 1.2) Read model control parameters
    Mylog.info('### 1.2 - Read model control parameters')
    #Read control and selection parameters into dictionary
    ScriptConfig = msf.ParseModelControl(Model_Configsheet,ScriptConfig)
    
    Mylog.info('Script: ' + Name_Script + '.py')
    Mylog.info('Model script version: ' + __version__)
    Mylog.info('Model functions version: ' + msf.__version__())
    Mylog.info('Model classes version: ' + msc.__version__())
    Mylog.info('Current User: ' + ProjectSpecs_User_Name)
    Mylog.info('Current Scenario: ' + Name_Scenario)
    Mylog.info(ScriptConfig['Description'])
    Mylog.debug('----\n')
    
    ### 1.3) Organize model output folder and logger
    Mylog.info('### 1.3 Organize model output folder and logger')
    #Copy Config file and model script into that folder
    shutil.copy(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), os.path.join(ProjectSpecs_Path_Result, ProjectSpecs_Name_ConFile))
    #shutil.copy(Name_Script + '.py'      , os.path.join(ProjectSpecs_Path_Result, Name_Script + '.py'))
    
    ######################################################
    #     Section 2a) Read classifications and data      #
    ######################################################
    Mylog.info('## 2 - Read classification items and define all classifications')
    ### 2.1) # Read model run config data
    Mylog.info('### 2.1 - Read model run config data')
    # Note: This part reads the items directly from the Exel master,
    # will be replaced by reading them from version-managed csv file.
    class_filename       = str(ScriptConfig['Version of master classification']) + '.xlsx'
    Classfile            = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,class_filename), data_only=True)
    Classsheet           = Classfile['MAIN_Table']
    MasterClassification = msf.ParseClassificationFile_Main(Classsheet,Mylog)
        
    Mylog.info('Read and parse config table, including the model index table, from model config sheet.')
    IT_Aspects,IT_Description,IT_Dimension,IT_Classification,IT_Selector,IT_IndexLetter,PL_Names,PL_Description,PL_Version,PL_IndexStructure,PL_IndexMatch,PL_IndexLayer,PL_SubFolder,PL_ProxyCode,PL_ProcMethod,PL_UpdateOverwrite,PrL_Number,PrL_Name,PrL_Comment,PrL_Type,ScriptConfig = msf.ParseConfigFile(Model_Configsheet,ScriptConfig,Mylog)    
    
    Mylog.info('Define model classifications and select items for model classifications according to information provided by config file.')
    ModelClassification  = {} # Dict of model classifications
    for m in range(0,len(IT_Aspects)):
        ModelClassification[IT_Aspects[m]] = deepcopy(MasterClassification[IT_Classification[m]])
        EvalString = msf.EvalItemSelectString(IT_Selector[m],len(ModelClassification[IT_Aspects[m]].Items))
        if EvalString.find(':') > -1: # range of items is taken
            RangeStart = int(EvalString[0:EvalString.find(':')])
            RangeStop  = int(EvalString[EvalString.find(':')+1::])
            ModelClassification[IT_Aspects[m]].Items = ModelClassification[IT_Aspects[m]].Items[RangeStart:RangeStop]           
        elif EvalString.find('[') > -1: # selected items are taken
            ModelClassification[IT_Aspects[m]].Items = [ModelClassification[IT_Aspects[m]].Items[i] for i in eval(EvalString)]
        elif EvalString == 'all':
            None
        else:
            Mylog.error('Item select error for aspect ' + IT_Aspects[m] + ' were found in datafile.')
            break
        
    ### 2.2) # Define model index table and parameter dictionary
    Mylog.info('### 2.2 - Define model index table and parameter dictionary')
    Model_Time_Start = int(min(ModelClassification['Time'].Items))
    Model_Time_End   = int(max(ModelClassification['Time'].Items))
    Model_Duration   = Model_Time_End - Model_Time_Start + 1
    
    Mylog.info('Define index table dataframe.')
    IndexTable = pd.DataFrame({'Aspect'        : IT_Aspects,  # 'Time' and 'Element' must be present!
                               'Description'   : IT_Description,
                               'Dimension'     : IT_Dimension,
                               'Classification': [ModelClassification[Aspect] for Aspect in IT_Aspects],
                               'IndexLetter'   : IT_IndexLetter})  # Unique one letter (upper or lower case) indices to be used later for calculations.
    
    # Default indexing of IndexTable, other indices are produced on the fly
    IndexTable.set_index('Aspect', inplace=True)
    
    # Add indexSize to IndexTable:
    IndexTable['IndexSize'] = pd.Series([len(IndexTable.Classification[i].Items) for i in range(0, len(IndexTable.IndexLetter))],
                                        index=IndexTable.index)
    
    # list of the classifications used for each indexletter
    IndexTable_ClassificationNames = [IndexTable.Classification[i].Name for i in range(0, len(IndexTable.IndexLetter))]
    
    # 2.3) Define shortcuts for the most important index sizes:
    Nt = len(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items)
    Ne = len(IndexTable.Classification[IndexTable.index.get_loc('Element')].Items)
    Nc = len(IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items)
    Nr = len(IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items)
    Nl = len(IndexTable.Classification[IndexTable.index.get_loc('Region11')].Items)
    No = len(IndexTable.Classification[IndexTable.index.get_loc('Region1')].Items)
    NG = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('G')].Items)
    Ng = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items)
    Np = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items)
    NB = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items)
    NN = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items) # varies: 24 for region-specific nrb and 4 for aggregated global resolution.
    NI = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items)
    Na = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items)
    #NA = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('A')].Items)
    NS = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items)
    NR = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    Nw = len(IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items)
    Nm = len(IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items)
    NP = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('P')].Items)    
    NX = len(IndexTable.Classification[IndexTable.index.get_loc('Extensions')].Items)
    Nx = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('x')].Items)   
    Nn = len(IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items)
    NV = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('V')].Items)
    Ns = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('s')].Items)
    #NT = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items)
    NL = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items)
    NO = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items)    
    NM = len(IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items)
    #IndexTable.loc['t']['Classification'].Items # get classification items
    
    SwitchTime = Nc-Nt+1 # Index of first model year (2016)
    # 2.4) Read model data and parameters.
    Mylog.info('Read model data and parameters.')
    
    ParFileName = os.path.join(RECC_Paths.data_path,'RECC_ParameterDict_' + ScriptConfig['RegionalScope'] + '.dat')
    try: # Load Pickle parameter dict to save processing time
        ParFileObject = open(ParFileName,'rb')  
        ParameterDict = pickle.load(ParFileObject)
        Mylog.info('Read model data and parameters from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
        #for individual parameters load new data if specified accordingly in config file
        mo_start = 0 # set mo for re-reading a certain parameter
        mo_reading_true = 0 
        for mo in range(mo_start,len(PL_Names)):
            #mo = 76 # set mo for re-reading a certain parameter
            #ParPath = os.path.join(os.path.abspath(os.path.join(ProjectSpecs_Path_Main, '.')), 'ODYM_RECC_Database', PL_Version[mo])             
            if PL_UpdateOverwrite[mo] == 'True': # new data is supposed to be used to replace data loaded from parameter dict
                mo_reading_true += 1
                if mo_reading_true == 1:
                    Mylog.info('Updating and overwriting parameter data in pickled parameter dict for selected parameters as specified in config file:')
                if PL_SubFolder[mo] == 'default': # path is not in subfolder but in main data directory
                    ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
                else: # parameter file is in subfolder, add this to path
                    ParPath = os.path.join(RECC_Paths.data_path, PL_SubFolder[mo], PL_Names[mo] + '_' + PL_Version[mo])
                Mylog.info('Reading parameter ' + PL_Names[mo] + ' and overwriting values in pickled parameter dict')
                #MetaData, Values = msf.ReadParameter(ParPath = ParPath,ThisPar = PL_Names[mo], ThisParIx = PL_IndexStructure[mo], IndexMatch = PL_IndexMatch[mo], ThisParLayerSel = PL_IndexLayer[mo], MasterClassification,IndexTable,IndexTable_ClassificationNames,ScriptConfig,Mylog) # Do not change order of parameters handed over to function!
                # Do not change order of parameters handed over to function!
                MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                                     PL_IndexLayer[mo], PL_ProcMethod[mo], MasterClassification, IndexTable,
                                                     IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
                ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                            UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                            Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                            Unit=MetaData['Dataset_Unit'])
                Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
                Mylog.info('_')        
        Mylog.info('Reading of parameters finished.')
        Mylog.info(str(mo_reading_true) + ' parameter file(s) read additionally and overwritten in pickled parameter dict.')
        if mo_reading_true > 0: #if new parameter values were added to parameter dict from previous run
            CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
            Mylog.info('New parameter reading sequence UUID: ' + CheckKey)
            Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
            ParameterDict['Checkkey'] = CheckKey
            # Save to pickle file for next model run
            ParFileObject = open(ParFileName,'wb') 
            pickle.dump(ParameterDict,ParFileObject)
        else: #if no new parameter data was read
            Mylog.info('Model data and parameters were read from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
        ParFileObject.close()      
    except:
        msf.check_dataset(RECC_Paths.data_path,PL_Names,PL_Version,PL_SubFolder,Mylog)
        ParameterDict = {}
        mo_start = 0 # set mo for re-reading a certain parameter
        for mo in range(mo_start,len(PL_Names)):
            #mo = 76 # set mo for re-reading a certain parameter
            #ParPath = os.path.join(os.path.abspath(os.path.join(ProjectSpecs_Path_Main, '.')), 'ODYM_RECC_Database', PL_Version[mo])
            if PL_SubFolder[mo] == 'default': # path is not in subfolder
                ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
            else: # parameter file is in subfolder, add this to path
                ParPath = os.path.join(RECC_Paths.data_path, PL_SubFolder[mo], PL_Names[mo] + '_' + PL_Version[mo])
            Mylog.info('Reading parameter ' + PL_Names[mo])
            #MetaData, Values = msf.ReadParameter(ParPath = ParPath,ThisPar = PL_Names[mo], ThisParIx = PL_IndexStructure[mo], IndexMatch = PL_IndexMatch[mo], ThisParLayerSel = PL_IndexLayer[mo], MasterClassification,IndexTable,IndexTable_ClassificationNames,ScriptConfig,Mylog) # Do not change order of parameters handed over to function!
            # Do not change order of parameters handed over to function!
            MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                                 PL_IndexLayer[mo], PL_ProcMethod[mo], MasterClassification, IndexTable,
                                                 IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
            ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                        UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                        Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                        Unit=MetaData['Dataset_Unit'])
            Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
            Mylog.info('_')
        Mylog.info('Reading of parameters finished.')
        CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
        Mylog.info('Current parameter reading sequence UUID: ' + CheckKey)
        Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
        ParameterDict['Checkkey'] = CheckKey
        # Save to pickle file for next model run
        ParFileObject = open(ParFileName,'wb') 
        pickle.dump(ParameterDict,ParFileObject)   
        ParFileObject.close()
        
    Mylog.info('_')
    Mylog.info('_')
    
    ########################################################
    #     Section 2b) Automatic assignment of proxies      #
    ########################################################
    # 0) obtain specific indices and positions:
    LEDindex        = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('LED')
    SSP1index       = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('SSP1')
    SSP2index       = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('SSP2')
    
    # a) Currently not used 
    # b) Set reference population dataset.
    ParameterDict['2_P_Population_Reference'] = msc.Parameter(Name='2_P_Population_Reference', ID='2_P_Population_Reference',
                                                UUID=ParameterDict[ScriptConfig['Population_Reference']].UUID, P_Res=None, MetaData=ParameterDict[ScriptConfig['Population_Reference']].MetaData,
                                                Indices=ParameterDict[ScriptConfig['Population_Reference']].Indices, Values=ParameterDict[ScriptConfig['Population_Reference']].Values, Uncert=None,
                                                Unit=ParameterDict[ScriptConfig['Population_Reference']].MetaData['Dataset_Unit'])
    
    # c) assign proxy data to model parameters
    for mo in range(0,len(PL_Names)):
        ProxyCommands = PL_ProxyCode[mo].split('_')
        if ProxyCommands[0] == 'replicate': # replicate the values of another regional index for the focus region index r:
            IndexStructure_Old = ParameterDict[PL_Names[mo]].Indices
            IndexStructure_New = IndexStructure_Old.replace(ProxyCommands[1],'r')
            ParameterDict[PL_Names[mo]].Indices = IndexStructure_New # note down the focus region aspect in the parameter definition
            # Assing new values with specified proxy mapper: 
            ParameterDict[PL_Names[mo]].Values = np.einsum(IndexStructure_Old+',r'+ProxyCommands[1]+'->'+IndexStructure_New,ParameterDict[PL_Names[mo]].Values,ParameterDict[ScriptConfig['r_'+ProxyCommands[1]+'_Mapping']].Values)
        if ProxyCommands[0] == 'scale': # scale down the values of another regional index for the focus region index r according to another indicator (usually population).
            if ProxyCommands[1] == 'P-startyear': # scale with population in the start year, which has NO aspects t and S.
                IndexStructure_Old = ParameterDict[PL_Names[mo]].Indices
                IndexStructure_New = IndexStructure_Old.replace(ProxyCommands[2],'r')
                ParameterDict[PL_Names[mo]].Indices = IndexStructure_New # note down the focus region aspect in the parameter definition
                NewRegIndLength = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc(ProxyCommands[2])].Items)
                # Calculate population shares of actual (high-resolution) model region in proxy (aggregate) region:
                Pop_r_Agg = np.einsum('r,r'+ProxyCommands[2]+'->'+ProxyCommands[2],ParameterDict['2_P_Population_Reference'].Values[0,0,:,0],ParameterDict[ScriptConfig['r_'+ProxyCommands[2]+'_Mapping']].Values) # population in r summed up to aggregate region (like d)
                Pop_r_rep = np.einsum('r,'+ProxyCommands[2]+'->r'+ProxyCommands[2],ParameterDict['2_P_Population_Reference'].Values[0,0,:,0],np.ones((NewRegIndLength)))
                Pop_r_rep_filter = np.multiply(Pop_r_rep,ParameterDict[ScriptConfig['r_'+ProxyCommands[2]+'_Mapping']].Values)
                Pop_rep_d = np.einsum(ProxyCommands[2]+',r->r'+ProxyCommands[2],Pop_r_Agg,np.ones((Nr))) # Replicate the aggregated population values across high-resolution regional dimension
                P_Share_Proxy = np.divide(Pop_r_rep_filter,Pop_rep_d) # Calculate share of disaggregated in aggregated population for full regional resolution r
                P_Share_Proxy[np.isnan(P_Share_Proxy)] = 0
                ParameterDict[PL_Names[mo]].Values = np.einsum(IndexStructure_Old+',r'+ProxyCommands[2]+'->'+IndexStructure_New,ParameterDict[PL_Names[mo]].Values,P_Share_Proxy)        
            if ProxyCommands[1] == 'P': # scale with population, which has aspects t and S, so these must be present in the affected parameter!
                IndexStructure_Old = ParameterDict[PL_Names[mo]].Indices
                IndexStructure_New = IndexStructure_Old.replace(ProxyCommands[2],'r')
                ParameterDict[PL_Names[mo]].Indices = IndexStructure_New # note down the focus region aspect in the parameter definition
                NewRegIndLength = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc(ProxyCommands[2])].Items)
                # Calculate population shares of actual (high-resolution) model region in proxy (aggregate) region:
                Pop_r_Agg = np.einsum('trS,r'+ProxyCommands[2]+'->t'+ProxyCommands[2]+'S',ParameterDict['2_P_Population_Reference'].Values[0,:,:,:],ParameterDict[ScriptConfig['r_'+ProxyCommands[2]+'_Mapping']].Values) # population in r summed up to aggregate region (like d)
                Pop_r_rep = np.einsum('trS,'+ProxyCommands[2]+'->tr'+ProxyCommands[2]+'S',ParameterDict['2_P_Population_Reference'].Values[0,:,:,:],np.ones((NewRegIndLength)))
                Pop_r_rep_filter = np.multiply(Pop_r_rep,np.einsum('r'+ProxyCommands[2]+',tS->trdS',ParameterDict[ScriptConfig['r_'+ProxyCommands[2]+'_Mapping']].Values,np.ones((Nt,NS))))
                Pop_rep_d = np.einsum('t'+ProxyCommands[2]+'S,r->tr'+ProxyCommands[2]+'S',Pop_r_Agg,np.ones((Nr))) # Replicate the aggregated population values across high-resolution regional dimension
                P_Share_Proxy = np.divide(Pop_r_rep_filter,Pop_rep_d) # Calculate share of disaggregated in aggregated population for full regional resolution r
                P_Share_Proxy[np.isnan(P_Share_Proxy)] = 0
                ParameterDict[PL_Names[mo]].Values = np.einsum(IndexStructure_Old+',tr'+ProxyCommands[2]+'S->'+IndexStructure_New,ParameterDict[PL_Names[mo]].Values,P_Share_Proxy)        
                
    ##############################################################
    #     Section 3)  Interpolate missing parameter values:      #
    ##############################################################
    SectorList      = eval(ScriptConfig['SectorSelect'])
    if 'nrb' in SectorList and 'nrbg' in SectorList:
        raise AssertionError('Fatal: Non-residential buildings are included both globally (nrbg) and for individual regions (nrb). Double-counting. Exiting the script, check config file.')    
    
    # index location and range of pass. vehs. in product list.
    try:
        Sector_pav_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('passenger vehicles')
    except:
        Sector_pav_loc  = np.nan
    try:
        Sector_pav_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items]
    except:
        if 'pav' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect p must also be selected for aspect g. Exiting the script.')
        else:
            Sector_pav_rge = []
    # index location and range of res. builds. in product list.
    try:
        Sector_reb_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('residential buildings')
    except:
        Sector_reb_loc  = np.nan
    try:
        Sector_reb_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items]
    except:
        if 'reb' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect B must also be selected for aspect g. Exiting the script.')
        else:
            Sector_reb_rge = []
    # index location and range of nonres. builds. in product list.
    if 'nrb' in SectorList:
        try:
            Sector_nrb_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('nonresidential buildings r')
        except:
            Sector_nrb_loc  = np.nan
        try:
            Sector_nrb_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items]
        except:
            Sector_nrb_rge  = []
    else:
        Sector_nrb_rge  = []
    # index location and range of nonres. builds. global in product list.    
    if 'nrbg' in SectorList:       
        try:
            Sector_nrbg_loc = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('nonresidential buildings g')
        except:
            Sector_nrbg_loc = np.nan
        try:
            Sector_nrbg_rge = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items]
        except:
            Sector_nrbg_rge = []
    else:
        Sector_nrbg_rge = []        
     # index location and range of industry in product list.    
    try:
        Sector_ind_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('industry')
    except:
        Sector_ind_loc  = np.nan
    try:
        Sector_ind_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items]
    except:
        if 'ind' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect I must also be selected for aspect g. Exiting the script.')
        else:
            Sector_ind_rge = []
    # index location and range of appliances in product list.    
    try:
        Sector_app_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('appliances')
    except:
        Sector_app_loc  = np.nan
    try:
        Sector_app_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items]
    except:
        if 'app' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect a must also be selected for aspect g. Exiting the script.')
        else:
            Sector_app_rge = []
        
    Cement_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('cement')
    Concrete_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('concrete')
    ConcrAgg_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('concrete aggregates')
    Wood_loc      = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('wood and wood products')
    WroughtAl_loc = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('wrought Al')
    CastAl_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('cast Al')
    Copper_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('copper electric grade')
    Plastics_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('plastics')
    Zinc_loc      = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('zinc')
    PrimCastAl_loc= IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of cast Al, primary')
    PrimWrAl_loc  = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of wrought Al, primary')
    EffCastAl_loc = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of cast Al, efficient')
    EffWrAl_loc   = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of wrought Al, efficient')
    PrimCGSteel_loc  = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of construction grade steel, primary')
    PrimASteel_loc   = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of automotive steel, primary')
    PrimSSteel_loc   = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of stainless steel, primary')
    PrimCastIron_loc = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of cast iron, primary')
    H2CGSteel_loc    = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of construction grade steel, H2')
    H2ASteel_loc     = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of automotive steel, H2')
    H2SSteel_loc     = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of stainless steel, H2')
    H2CastIron_loc   = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of cast iron, H2')
    PaperProd_loc    = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of paper and cardboard, primary')
    WoodProd_loc     = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items.index('production of wood and wood products, primary')
    Woodwaste_loc = IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items.index('used construction wood')
    Electric_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('electricity')
    NatuGas_loc   = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('natural gas')
    WoodFuel_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('fuel wood')
    Hydrogen_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('hydrogen')
    all_loc       = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('all')
    Carbon_loc    = IndexTable.Classification[IndexTable.index.get_loc('Element')].Items.index('C')
    ClimPolScen   = IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items.index('RCP2.6')
    CO2_loc       = IndexTable.Classification[IndexTable.index.get_loc('Extensions')].Items.index('CO2 emissions per main output')
    GWP100_loc    = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('GWP100')
    Land_loc      = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Land occupation (LOP)')
    Water_loc     = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Water consumption potential (WCP)')
    dynGWP100_loc = IndexTable.Classification[IndexTable.index.get_loc('Cumulative env. pressure')].Items.index('dynGWP100')
    AllMat_loc    = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Raw material input (RMI), all materials')
    FosFuel_loc   = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Raw material input (RMI), fossil fuels')
    MetOres_loc   = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Raw material input (RMI), metal ores')
    nMetOres_loc  = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Raw material input (RMI), non-metallic minerals')
    Biomass_loc   = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Raw material input (RMI), biomass')
    PrimEn_loc    = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('primary energy')
    Heating_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Heating')
    Cooling_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Cooling')
    DomstHW_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('DHW')
    Service_Drivg = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Driving')
    Bio_El_loc    = IndexTable.Classification[IndexTable.index.get_loc('Industry')].Items.index('bio powerplant')
    Bio_El_CS_loc = IndexTable.Classification[IndexTable.index.get_loc('Industry')].Items.index('biomass power plant with CCS')
    
    Service_Reb   = np.array([Heating_loc,Cooling_loc,DomstHW_loc])
    Ind_2015      = 115 #index of year 2015
    #Ind_2017      = 117 #index of year 2017
    Ind_2020      = 120 #index of year 2020
    IsClose_Remainder_Small = 1e-15 
    IsClose_Remainder_Large = 1e-7 
    DPI_RES        = ScriptConfig['Plot4Max'] # 100 for overview or 500 for paper plots, defined in ModelConfig_List
    
    # Determine location of the indices of individual sectors in the region-specific list and in the list of all goods
    # indices of sectors with same regional scope in complete goods list
    Sector_11reg_rge    = Sector_ind_rge
    Sector_1reg_rge     = Sector_app_rge + Sector_nrbg_rge
    #indices of individual end-use sectors within regionally separated product lists, check with classification master file!
    Sector_ind_rge_reg  = np.arange(0,18,1)
    Sector_app_rge_reg  = np.arange(0,12,1)
    Sector_nrbg_rge_reg = np.arange(12,16,1)
    
    OutputDict      = {}  # Dictionary with output variables for entire model run, to export checks and analyses.
    
    # 1a) Currently not used
    
    # 1b) Currently not used
        
    # 1c) Currently not used
    
    # 1d) Split concrete in building archetypes into cement and aggregates but keep concrete separately
    ParameterDict['3_MC_BuildingArchetypes'].Values[:,:,Cement_loc]   = ParameterDict['3_MC_BuildingArchetypes'].Values[:,:,Cement_loc] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_BuildingArchetypes'].Values[:,:,Concrete_loc].copy()
    ParameterDict['3_MC_BuildingArchetypes'].Values[:,:,ConcrAgg_loc] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_BuildingArchetypes'].Values[:,:,Concrete_loc].copy()
    
    ParameterDict['3_MC_NonResBuildingArchetypes'].Values[:,:,Cement_loc]   = ParameterDict['3_MC_NonResBuildingArchetypes'].Values[:,:,Cement_loc] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_NonResBuildingArchetypes'].Values[:,:,Concrete_loc].copy()
    ParameterDict['3_MC_NonResBuildingArchetypes'].Values[:,:,ConcrAgg_loc] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_NonResBuildingArchetypes'].Values[:,:,Concrete_loc].copy()
        
    # 1e) Compile parameter for building energy conversion efficiency:
    ParameterDict['4_TC_ResidentialEnergyEfficiency'] = msc.Parameter(Name='4_TC_ResidentialEnergyEfficiency', ID='4_TC_ResidentialEnergyEfficiency',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values                                   = np.einsum('VRrn,tS->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency_Default'].Values[:,:,:,:,0],np.ones((Nt,NS)))
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Heating_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[Heating_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Cooling_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Cooling'].Values[Cooling_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[DomstHW_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[DomstHW_loc,:,:,Electric_loc,:,:] / 100
    
    ParameterDict['4_TC_NonResidentialEnergyEfficiency'] = msc.Parameter(Name='4_TC_NonResidentialEnergyEfficiency', ID='4_TC_NonResidentialEnergyEfficiency',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values                                   = np.einsum('VRrn,tS->VRrntS',ParameterDict['4_TC_NonResEnergyEfficiency_Default'].Values[:,:,:,:,0],np.ones((Nt,NS)))
    ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[Heating_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_NonResEnergyEfficiency_Scenario_Heating'].Values[Heating_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[Cooling_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_NonResEnergyEfficiency_Scenario_Cooling'].Values[Cooling_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[DomstHW_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_NonResEnergyEfficiency_Scenario_Heating'].Values[DomstHW_loc,:,:,Electric_loc,:,:] / 100
    
    # 1f) Derive energy supply multipliers for buildings for future age-cohorts
    # From energy carrier split and conversion efficiency, the multipliers converting 1 MJ of final building energy demand into different energy carriers are determined.
    # For details around the ancillary quantity anc, see the model documentation.
    Divisor = ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values #VRrntS
    Anc = np.divide(np.einsum('VRrnt,S->VRrntS',ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values, np.ones(NS)), Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    
    # Define energy carrier split for useful energy
    ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'] = msc.Parameter(Name='3_SHA_EnergyCarrierSplit_Buildings_uf', ID='3_SHA_EnergyCarrierSplit_Buildings_uf',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'].Values = np.divide(Anc, np.einsum('VRrtS,n->VRrntS',np.einsum('VRrntS->VRrtS',Anc),np.ones(Nn)), out=np.zeros_like(Divisor), where=Divisor!=0)
    
    Divisor = ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values #VRrntS
    Anc = np.divide(np.einsum('VRrnt,S->VRrntS',ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings'].Values, np.ones(NS)), Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    
    # Define energy carrier split for useful energy
    ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings_uf'] = msc.Parameter(Name='3_SHA_EnergyCarrierSplit_NonResBuildings_uf', ID='3_SHA_EnergyCarrierSplit_NonResBuildings_uf',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings_uf'].Values = np.divide(Anc, np.einsum('VRrtS,n->VRrntS',np.einsum('VRrntS->VRrtS',Anc),np.ones(Nn)), out=np.zeros_like(Divisor), where=Divisor!=0)
    
    # 2a) Determine future energy intensity and material composition of vehicles by mixing archetypes:
    # Check if RE strategies are active and set implementation curves to 2016 value if not.
    if 'pav' in SectorList:
        if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
            ParameterDict['3_SHA_LightWeighting_Vehicles'].Values  = np.einsum('prS,t->prtS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values[:,:,0,:],np.ones((Nt)))
        DownSizingBuffer = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.copy()
        if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True': # consider lightweighting trough UsingLessMaterialByDesign (segment shift)
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 1:
                        ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,:,nnS] = np.einsum('s,t->st',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,0,nnS],np.ones((Nt)))
        else: # no lightweighting trough UsingLessMaterialByDesign.
            # for regions not selected above (high-income regions): Set downsizing to 2016 levels.
            ParameterDict['3_SHA_DownSizing_Vehicles'].Values = np.einsum('srS,t->srtS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,:,0,:],np.ones((Nt)))
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 1:
                        ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,:,nnS]  = DownSizingBuffer[:,nnr,:,nnS].copy()
        ParameterDict['3_MC_RECC_Vehicles_RECC'] = msc.Parameter(Name='3_MC_RECC_Vehicles_RECC', ID='3_MC_RECC_Vehicles_RECC',
                                                   UUID=None, P_Res=None, MetaData=None,
                                                   Indices='cmprS', Values=np.zeros((Nc,Nm,Np,Nr,NS)), Uncert=None,
                                                   Unit='kg/unit')
        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[0:SwitchTime-1,:,:,:,:] = np.einsum('cmpr,S->cmprS',ParameterDict['3_MC_RECC_Vehicles'].Values[0:SwitchTime-1,:,:,:],np.ones(NS))
        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchTime-1::,:,:,:,:] = \
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[24,28,32,36,40,44],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[25,29,33,37,41,45],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[26,30,34,38,42,46],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[27,31,35,39,43,47],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[0 ,4 ,8 ,12,16,20],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[1 ,5 ,9 ,13,17,21],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[2 ,6 ,10,14,18,22],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[3 ,7 ,11,15,19,23],:]))
        ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchTime-1::,:,Service_Drivg,:,:,:] = \
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[24,28,32,36,40,44],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[25,29,33,37,41,45],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[26,30,34,38,42,46],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[27,31,35,39,43,47],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[0 ,4 ,8 ,12,16,20],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[1 ,5 ,9 ,13,17,21],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[2 ,6 ,10,14,18,22],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[3 ,7 ,11,15,19,23],:]))
    
    # 2b) Determine future energy intensity and material composition of residential buildings by mixing archetypes:
    # Expand building light-weighting split to all building types:
    if 'reb' in SectorList:
        ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('B,rtSR->BrtSR',np.ones(NB),ParameterDict['3_SHA_LightWeighting_Buildings'].Values[Sector_reb_loc,:,:,:,:]).copy()
        if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
            ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('BrSR,t->BrtSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values[:,:,0,:,:],np.ones((Nt)))
        if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'False': # no lightweighting trough UsingLessMaterialByDesign.
            ParameterDict['3_SHA_DownSizing_Buildings'].Values = np.einsum('urSR,t->urtSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values[:,:,0,:,:],np.ones((Nt)))
        ParameterDict['3_MC_RECC_Buildings_RECC'] = msc.Parameter(Name='3_MC_RECC_Buildings_RECC', ID='3_MC_RECC_Buildings_RECC',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='cmBrS', Values=np.zeros((Nc,Nm,NB,Nr,NS,NR)), Uncert=None,
                                                Unit='kg/m2')
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,:,:,:,:,:] = np.einsum('cmBr,SR->cmBrSR',ParameterDict['3_MC_RECC_Buildings'].Values[0:115,:,:,:],np.ones((NS,NR)))
        # Split concrete into cement and aggregates for historic age-cohorts (for future age-cohorts, this is done already for the archetypes).
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,Cement_loc,:,:,:,:]   = ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]       * ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,Concrete_loc,:,:,:,:].copy()
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,ConcrAgg_loc,:,:,:,:] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,Concrete_loc,:,:,:,:].copy()
        # Mix future archetypes for material composition
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[115::,:,:,:,:,:] = \
        np.einsum('BrcSR,BmrcSR->cmBrSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcSR,Brm->BmrcSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[87,88,89,90,91,92,93,94,95,96,97,98,99],:,:])) +\
        np.einsum('BrcSR,BmrcSR->cmBrSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcSR,Brm->BmrcSR',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[61,62,63,64,65,66,67,68,69,70,71,72,73],:,:])) +\
        np.einsum('BrcSR,BmrcSR->cmBrSR',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcSR,Brm->BmrcSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[74,75,76,77,78,79,80,81,82,83,84,85,86],:,:])) +\
        np.einsum('BrcSR,BmrcSR->cmBrSR',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcSR,Brm->BmrcSR',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[48,49,50,51,52,53,54,55,56,57,58,59,60],:,:]))
        # Replicate values for Al, Cu, plastics for future age-cohorts as the archetypes don't have such information.
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[115::,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc,Zinc_loc],:,:,:,:] = np.einsum('mBr,cSR->cmBrSR',ParameterDict['3_MC_RECC_Buildings'].Values[110,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc,Zinc_loc],:,:].copy(),np.ones((Nt,NS,NR)))
        
        ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values = np.einsum('cBVnrS,R->cBVnrSR',ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values,np.ones(NR)) # replicate for both RCP scenarios
        ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[115::,:,:,:,:,:,:] = \
        np.einsum('BrcSR,BnrVcSR->cBVnrSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcSR,BrVn->BnrVcSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[87,88,89,90,91,92,93,94,95,96,97,98,99],:,:,:])) +\
        np.einsum('BrcSR,BnrVcSR->cBVnrSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcSR,BrVn->BnrVcSR',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[61,62,63,64,65,66,67,68,69,70,71,72,73],:,:,:])) +\
        np.einsum('BrcSR,BnrVcSR->cBVnrSR',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcSR,BrVn->BnrVcSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[74,75,76,77,78,79,80,81,82,83,84,85,86],:,:,:])) +\
        np.einsum('BrcSR,BnrVcSR->cBVnrSR',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcSR,BrVn->BnrVcSR',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[48,49,50,51,52,53,54,55,56,57,58,59,60],:,:,:]))
        # The archetypes report useful energy for 'all' energy carriers together! Must be split into different energy carriers.
        # Will happen below as energy carrier split and final-to-useful conversion efficieny is RCP-scenario dependent.
        
        # Define time-dependent final energy parameter:
        ParameterDict['3_EI_Products_UsePhase_resbuildings_t'] = msc.Parameter(Name='3_EI_Products_UsePhase_resbuildings_t', ID='3_EI_Products_UsePhase_resbuildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='cBVnrt', Values=np.zeros((Nc,NB,NV,Nn,Nr,Nt)), Uncert=None,
                                                    Unit='MJ/m2/yr')
        ParameterDict['3_MC_RECC_Buildings_t'] = msc.Parameter(Name='3_MC_RECC_Buildings_t', ID='3_MC_RECC_Buildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='mBrctSR', Values=np.zeros((Nm,NB,Nr,Nc,Nt,NS,NR)), Uncert=None,
                                                    Unit='kg/m2')    
    
    # 2c) Determine future energy intensity and material composition of nonresidential buildings by mixing archetypes:
    if 'nrb' in SectorList:
        # Expand building light-weighting split to all building types:
        ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values = np.einsum('N,rtSR->NrtSR',np.ones(NN),ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values[Sector_nrb_loc,:,:,:,:]).copy()
        if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
            ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values = np.einsum('NrSR,t->NrtSR',ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values[:,:,0,:,:],np.ones((Nt)))
        if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'False': # no lightweighting trough UsingLessMaterialByDesign.
            ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values = np.einsum('urSR,t->urtSR',ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values[:,:,0,:,:],np.ones((Nt)))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'] = msc.Parameter(Name='3_MC_RECC_NonResBuildings_RECC', ID='3_MC_RECC_NonResBuildings_RECC',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='cmNrSR', Values=np.zeros((Nc,Nm,NN,Nr,NS,NR)), Uncert=None,
                                                    Unit='kg/m2')
        #For 3_MC: copy over historic age-cohorts first
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,:,:,:,:] = np.einsum('cmNr,SR->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,:,:],np.ones((NS,NR)))
        # Split concrete into cement and aggregates for historic age-cohorts (for future age-cohorts, this is done already for the archetypes).
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,Cement_loc,:,:,:,:]    = ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,Cement_loc,:,:,:,:] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,Concrete_loc,:,:,:,:].copy()
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,ConcrAgg_loc,:,:,:,:]  = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,Concrete_loc,:,:,:,:].copy()
        #For 3_MC: Replicate standard type data for other types as proxy type, as only for those, empirical 3_MC data were compiled.
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[0,1,2,3],:,:,:]     = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,1,:], np.ones((NS,NR)),np.ones(4))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[4,5,6,7],:,:,:]     = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,5,:], np.ones((NS,NR)),np.ones(4))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[8,9,10,11],:,:,:]   = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,9,:], np.ones((NS,NR)),np.ones(4))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[12,13,14,15],:,:,:] = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,13,:],np.ones((NS,NR)),np.ones(4))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[16,17,18,19],:,:,:] = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,17,:],np.ones((NS,NR)),np.ones(4))
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[0:115,:,[20,21,22,23],:,:,:] = np.einsum('cmr,SR,N->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[0:115,:,21,:],np.ones((NS,NR)),np.ones(4))
        # Mix future archetypes for material composition
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[115::,:,:,:,:,:] = \
        np.einsum('NrcSR,NmrcSR->cmNrSR',ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values,     np.einsum('urcSR,Nrm->NmrcSR',ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,    ParameterDict['3_MC_NonResBuildingArchetypes'].Values[[110,114,106,102,158,162,154,150,174,178,170,166,190,194,186,182,126,130,122,118,142,146,138,134],:,:])) +\
        np.einsum('NrcSR,NmrcSR->cmNrSR',ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values,     np.einsum('urcSR,Nrm->NmrcSR',1 - ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,ParameterDict['3_MC_NonResBuildingArchetypes'].Values[[109,113,105,101,157,161,153,149,173,177,169,165,189,193,185,181,125,129,121,117,141,145,137,133],:,:])) +\
        np.einsum('NrcSR,NmrcSR->cmNrSR',1 - ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values, np.einsum('urcSR,Nrm->NmrcSR',ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,    ParameterDict['3_MC_NonResBuildingArchetypes'].Values[[111,115,107,103,159,163,155,151,175,179,171,167,191,195,187,183,127,131,123,119,143,147,139,135],:,:])) +\
        np.einsum('NrcSR,NmrcSR->cmNrSR',1 - ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values, np.einsum('urcSR,Nrm->NmrcSR',1 - ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,ParameterDict['3_MC_NonResBuildingArchetypes'].Values[[108,112,104,100,156,160,152,148,172,176,168,164,188,192,184,180,124,128,120,116,140,144,136,132],:,:]))
        # Replicate values for Al, Cu, plastics for future age-cohorts as the archetypes don't have such information.
        ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[115::,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc,Zinc_loc],:,:,:,:] = np.einsum('mNr,cSR->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings'].Values[110,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc,Zinc_loc],:,:].copy(),np.ones((Nt,NS,NR)))
        
        # For 3_EI_Products_UsePhase_nonresbuildings: Replicate SSP1 values to LED and SSP2 scenarios (scenario aspect is irrelevant anyway because these are historic data only)
        ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values = np.einsum('cNVnrS,R->cNVnrSR',ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values,np.ones(NR)) # replicate for both RCP scenarios
        ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[:,:,:,:,:,:,:]     = np.einsum('cNVnrR,S->cNVnrSR',ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[:,:,:,:,:,1,:],np.ones((NS)))
        # Mix future archetypes
        ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[115::,:,:,:,:,:,:] = \
        np.einsum('NrcSR,NnrVcSR->cNVnrSR',ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values,     np.einsum('urcSR,NrVn->NnrVcSR',ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,    ParameterDict['3_EI_NonResBuildingArchetypes'].Values[[110,114,106,102,158,162,154,150,174,178,170,166,190,194,186,182,126,130,122,118,142,146,138,134],:,:,:])) +\
        np.einsum('NrcSR,NnrVcSR->cNVnrSR',ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values,     np.einsum('urcSR,NrVn->NnrVcSR',1 - ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,ParameterDict['3_EI_NonResBuildingArchetypes'].Values[[109,113,105,101,157,161,153,149,173,177,169,165,189,193,185,181,125,129,121,117,141,145,137,133],:,:,:])) +\
        np.einsum('NrcSR,NnrVcSR->cNVnrSR',1 - ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values, np.einsum('urcSR,NrVn->NnrVcSR',ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,    ParameterDict['3_EI_NonResBuildingArchetypes'].Values[[111,115,107,103,159,163,155,151,175,179,171,167,191,195,187,183,127,131,123,119,143,147,139,135],:,:,:])) +\
        np.einsum('NrcSR,NnrVcSR->cNVnrSR',1 - ParameterDict['3_SHA_LightWeighting_NonResBuildings'].Values, np.einsum('urcSR,NrVn->NnrVcSR',1 - ParameterDict['3_SHA_DownSizing_NonResBuildings'].Values,ParameterDict['3_EI_NonResBuildingArchetypes'].Values[[108,112,104,100,156,160,152,148,172,176,168,164,188,192,184,180,124,128,120,116,140,144,136,132],:,:,:]))
        ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'] = msc.Parameter(Name='3_EI_Products_UsePhase_nonresbuildings_t', ID='3_EI_Products_UsePhase_nonresbuildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='cNVnrt', Values=np.zeros((Nc,NN,NV,Nn,Nr,Nt)), Uncert=None,
                                                    Unit='MJ/m2/yr')
        ParameterDict['3_MC_RECC_NonResBuildings_t'] = msc.Parameter(Name='3_MC_RECC_NonResBuildings_t', ID='3_MC_RECC_NonResBuildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='mNrctSR', Values=np.zeros((Nm,NN,Nr,Nc,Nt,NS,NR)), Uncert=None,
                                                    Unit='kg/m2')      
    
    if 'nrbg' in SectorList:
        # Split concrete into cement and aggregates:
        # Cement for buildings remains, as this item refers to cement in mortar, screed, and plaster. Cement in concrete is calculated as ParameterDict['3_MC_CementContentConcrete'].Values * concrete and added here. 
        # Concrete aggregates (0.87*concrete) are considered as well.
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Cement_loc,:]   = ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Cement_loc,:] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:].copy()
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[ConcrAgg_loc,:] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:].copy()
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:] = 0
        
    # 3) Currently not in use.
    
    # 4) Fabrication yield and fabrication scrap diversion:
    # Extrapolate 2050-2060 as 2015 values
    index = PL_Names.index('4_PY_Manufacturing')
    ParameterDict[PL_Names[index]].Values[:,:,:,:,1::,:] = np.einsum('t,mwgFr->mwgFtr',np.ones(45),ParameterDict[PL_Names[index]].Values[:,:,:,:,0,:])
    if ScriptConfig['Include_REStrategy_FabScrapDiversion'] == 'False':
        ParameterDict['6_PR_FabricationScrapDiversion'].Values = np.zeros((Nm,Nw,No,NS))
    
    # 5) EoL RR, apply world average to all regions
    ParameterDict['4_PY_EoL_RecoveryRate'].Values = np.einsum('gmwW,r->grmwW',ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,0,:,:,:],np.ones((Nr)))
    
    # 6) Energy carrier split of vehicles, replicate fixed values for all regions and age-cohorts etc.
    ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles'].Values = np.einsum('pn,crVS->cprVnS',ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles'].Values[115,:,0,Service_Drivg,:,SSP1index].copy(),np.ones((Nc,Nr,NV,NS)))
    
    # 7) RE strategy potentials for individual countries are replicated from global average:
    ParameterDict['6_PR_ReUse_Bld'].Values                      = np.einsum('mB,r->mBr',ParameterDict['6_PR_ReUse_Bld'].Values[:,:,0],np.ones(Nr))
    ParameterDict['6_PR_ReUse_nonresBld'].Values                = np.einsum('mN,r->mNr',ParameterDict['6_PR_ReUse_nonresBld'].Values[:,:,0],np.ones(Nr))
    ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values = np.einsum('pS,r->prS',ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values[:,0,:],np.ones(Nr))
    ParameterDict['6_PR_EoL_RR_Improvement'].Values             = np.einsum('gmwW,r->grmwW',ParameterDict['6_PR_EoL_RR_Improvement'].Values[:,0,:,:,:],np.ones(Nr))
    
    # 8) Define a multi-regional RE strategy and building renovation scaleup parameter
    ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'] = msc.Parameter(Name='3_SHA_RECC_REStrategyScaleUp_r', ID='3_SHA_RECC_REStrategyScaleUp_r',
                                                      UUID=None, P_Res=None, MetaData=None,
                                                      Indices='trSR', Values=np.zeros((Nt,Nr,NS,NR)), Uncert=None,
                                                      Unit='1')
    ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values        = np.einsum('RtS,r->trSR',ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[:,0,:,:],np.ones(Nr)).copy()
    ParameterDict['3_SHA_BuildingRenovationScaleUp_r'] = msc.Parameter(Name='3_SHA_BuildingRenovationScaleUp_r', ID='3_SHA_BuildingRenovationScaleUp_r',
                                                      UUID=None, P_Res=None, MetaData=None,
                                                      Indices='trSR', Values=np.zeros((Nt,Nr,NS,NR)), Uncert=None,
                                                      Unit='1')
    ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values        = np.einsum('RtS,r->trSR',ParameterDict['3_SHA_BuildingRenovationScaleUp'].Values[:,0,:,:],np.ones(Nr)).copy()
    
    # 9) Currently not used    
    
    # 10) Set future vehicle reuse to 2015 levels if strategy is not included:
    # (To reflect that reuse is already happening to some extent.)
    if ScriptConfig['Include_REStrategy_ReUse'] == 'False':
        ParameterDict['6_PR_ReUse_Veh'].Values       = np.einsum('mprS,t->mprtS',ParameterDict['6_PR_ReUse_Veh'].Values[:,:,:,1,:],np.ones(Nt)) # stay at current levels, which are > 0.
        ParameterDict['6_PR_ReUse_Bld'].Values       = np.zeros(ParameterDict['6_PR_ReUse_Bld'].Values.shape) # set to zero, which corresponds to current levels.
        ParameterDict['6_PR_ReUse_nonresBld'].Values = np.zeros(ParameterDict['6_PR_ReUse_nonresBld'].Values.shape) # set to zero, which corresponds to current levels.
        
    # 11) MODEL CALIBRATION
    # Calibrate vehicle kilometrage: No longer used! VKM is now calibrated in scenario target table process to deliver correct pC stock number for 2015.
    #### ParameterDict['3_IO_Vehicles_UsePhase'].Values[3,:,:,:]                             = ParameterDict['3_IO_Vehicles_UsePhase'].Values[3,:,:,:]                           * np.einsum('r,tS->rtS',ParameterDict['6_PR_Calibration'].Values[0,:],np.ones((Nt,NS)))
    # Calibrate vehicle fuel consumption, cgVnrS   
    if 'pav' in SectorList:
        ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,Service_Drivg,:,:,:]        = ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,Service_Drivg,:,:,:]      * np.einsum('r,cpnS->cpnrS',ParameterDict['6_PR_Calibration'].Values[1,:],np.ones((115,Np,Nn,NS)))
    # Calibrate res. building energy consumption
    if 'reb' in SectorList:
        ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:,:]      = ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:,:]    * np.einsum('r,cBVnSR->cBVnrSR',ParameterDict['6_PR_Calibration'].Values[2,:],np.ones((115,NB,3,Nn,NS,NR)))
    # Calibrate nonres. building energy consumption
    if 'nrb' in SectorList:
        ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:,:]   = ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:,:] * np.einsum('r,cNVnSR->cNVnrSR',ParameterDict['6_PR_Calibration'].Values[3,:],np.ones((115,NN,3,Nn,NS,NR)))
    
    # 12) No recycling scenario (counterfactual reference)
    if ScriptConfig['IncludeRecycling'] == 'False': # no recycling and remelting
        ParameterDict['4_PY_EoL_RecoveryRate'].Values            = np.zeros(ParameterDict['4_PY_EoL_RecoveryRate'].Values.shape)
        ParameterDict['4_PY_MaterialProductionRemelting'].Values = np.zeros(ParameterDict['4_PY_MaterialProductionRemelting'].Values.shape)
        
    # 13) Currently not used.
      
    # 14) Define parameter for future vehicle stock:
    # a) calculated passenger vehicle stock
    ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_passvehicles', ID='2_S_RECC_FinalProducts_Future_passvehicles',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='StGr', Values=np.zeros((NS,Nt,NG,Nr)), Uncert=None,
                                                Unit='cars per person')
    # b) calculated vehicle kilometrage
    ParameterDict['3_IO_Vehicles_UsePhase_eff'] = msc.Parameter(Name='3_IO_Vehicles_UsePhase_eff', ID='3_IO_Vehicles_UsePhase_eff',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VrtS', Values=np.zeros((NV,Nr,Nt,NS)), Uncert=None,
                                                Unit='km per vehicle')
    
    # 15) Define parameter for future building stock:
    # actual future res and nonres building stock
    ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_resbuildings_act', ID='2_S_RECC_FinalProducts_Future_resbuildings_act',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='StGr', Values=np.zeros((NS,Nt,NG,Nr)), Uncert=None,
                                                Unit='m2 per person')
    ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_NonResBuildings_act', ID='2_S_RECC_FinalProducts_Future_NonResBuildings_act',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='GrtS', Values=np.zeros((NG,Nr,Nt,NS)), Uncert=None,
                                                Unit='m2 per person')    
    # 3_IO changing over time:
    ParameterDict['3_IO_Buildings_UsePhase'] = msc.Parameter(Name='3_IO_Buildings_UsePhase', ID='3_IO_Buildings_UsePhase',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='tcBVrS', Values=np.zeros((Nt,Nc,NB,NV,Nr,NS)), Uncert=None,
                                                Unit='1')
    # Historic age-cohorts:
    # ParameterDict['3_IO_Buildings_UsePhase_Historic'] is a combination of climate and socioeconomic 3_IO determinants.
    # We single out the former and keep them constant and let the socioeconomic factors change according to the '3_IO_Buildings_UsePhase_Future_...' parameters.
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Heating_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Heating_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating,np.ones(Nt))
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW     = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,DomstHW_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,DomstHW_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW,np.ones(Nt))
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Cooling_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Cooling_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling,np.ones(Nt))
    # Correct for if some of the corrections lead to IO > 1 which may be the case when hist. IO data are incomplete and thus set to 1 already.
    ParameterDict['3_IO_Buildings_UsePhase'].Values[ParameterDict['3_IO_Buildings_UsePhase'].Values > 1] = 1
    # Future age-cohorts:
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Heating_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,DomstHW_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Cooling_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    
    # expand 3_IO parameter for nonres buildings with 1 for post 2015 years:
    ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[SwitchTime::,:,:,:,:] = 1
    
    # 16) Currently not used.
    
    # 17) No energy efficiency improvements (counterfactual reference)
    # Freeze type split and archetypes at 2015/17/2020 levels:
    if ScriptConfig['No_EE_Improvements'] == 'True':
        SwitchIndex = Ind_2015 # choose between Ind_2015/17/20 (for continuation from 2015/17/20 onwards.
        if 'pav' in SectorList:
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 0: # don't consider type split reset for cases, where type split change involves larger vehicles.
                        ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchIndex::,:,:,:,nnr,nnS]    = np.einsum('pVn,c->cpVn',ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchIndex,:,:,:,nnr,nnS],np.ones(Nc-SwitchIndex))
                        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchIndex::,:,:,nnr,nnS]  = np.einsum('mp,c->cmp',ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchIndex,:,:,nnr,nnS],np.ones(Nc-SwitchIndex))
            ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[:,:,:,:,4::]                           = np.einsum('GrRp,t->GrRpt',ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[:,:,:,:,4],np.ones(Nt-4)) # index 4 is year 2020.
        if 'reb' in SectorList:
            ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchIndex::,:,:,:,:,:,:]  = np.einsum('BVnrSR,c->cBVnrSR',ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchIndex,:,:,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_MC_RECC_Buildings_RECC'].Values[SwitchIndex::,:,:,:,:,:]               = np.einsum('mBrSR,c->cmBrSR',ParameterDict['3_MC_RECC_Buildings_RECC'].Values[SwitchIndex,:,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,:,4::,:]                            = np.einsum('BrS,t->BrtS',ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
            ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4::,:]                 = np.einsum('VRrnS,t->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
        if 'nrb' in SectorList:            
            ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[SwitchIndex::,:,:,:,:,:,:] = np.einsum('NVnrSR,c->cNVnrSR',ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[SwitchIndex,:,:,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[SwitchIndex::,:,:,:,:,:]         = np.einsum('mNrSR,c->cmNrSR',ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[SwitchIndex,:,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[:,:,4::,:]                      = np.einsum('NrS,t->NrtS',ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
            ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[:,:,:,:,4::,:]              = np.einsum('VRrnS,t->VRrntS',ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[:,:,:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
    
    # 18) Currently not used.
    
    # 19) Make sure that all share parameters are non-negative and add up to 100%:
    # not necessary for res. buildings as data fulfil constraints.
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values[ParameterDict['3_SHA_TypeSplit_Buildings'].Values < 0] = 0
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values = ParameterDict['3_SHA_TypeSplit_Buildings'].Values / np.einsum('rtS,B->BrtS',ParameterDict['3_SHA_TypeSplit_Buildings'].Values.sum(axis=0),np.ones(NB))
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values[np.isnan(ParameterDict['3_SHA_TypeSplit_Buildings'].Values)] = 0
    ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values < 0] = 0
    ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values = ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values / np.einsum('rtSR,B->BrtSR',ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.sum(axis=0),np.ones(NN))
    ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[np.isnan(ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values)] = 0
    
    # 20) Extrapolate appliances beyond 2050:
    for noS in range(0,NS):
        for noR in range(0,NR):
            for noa in range(0,Na):
                if ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,140,noS,noR,noa] != 0:
                    growthrate = (ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,150,noS,noR,noa]/ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,140,noS,noR,noa]-1)/10
                else:
                    growthrate = 0
                for noT in range(151,161):
                    ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,noT,noS,noR,noa] = ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,150,noS,noR,noa] * np.power(1+growthrate,noT-150)
        
    # 21) Currently not used
        
    # 22) calculate Stocks on 1. Jan 2016:    
    pC_AgeCohortHist           = np.zeros((NG,Nr))
    #pC_FutureStock             = np.zeros((NS,NG,Nr))
    # a) from historic data:
    Stocks_2016_passvehicles   = ParameterDict['2_S_RECC_FinalProducts_2015_passvehicles'].Values[0,:,:,:].sum(axis=0)
    pCStocks_2016_passvehicles = np.einsum('pr,r->rp',Stocks_2016_passvehicles,1/ParameterDict['2_P_Population_Reference'].Values[0,0,:,1]) 
    Stocks_2016_resbuildings   = ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:].sum(axis=0)
    pCStocks_2016_resbuildings = np.einsum('Br,r->rB',Stocks_2016_resbuildings,1/ParameterDict['2_P_Population_Reference'].Values[0,0,:,1]) 
    Stocks_2016_nresbuildings  = ParameterDict['2_S_RECC_FinalProducts_2015_nonresbuildings'].Values[0,:,:,:].sum(axis=0)
    pCStocks_2016_nresbuildings= np.einsum('Nr,r->rN',Stocks_2016_nresbuildings,1/ParameterDict['2_P_Population_Reference'].Values[0,0,:,1]) 
    if 'pav' in SectorList:
        pC_AgeCohortHist[Sector_pav_loc, :] = pCStocks_2016_passvehicles.sum(axis =1)
    if 'reb' in SectorList:
        pC_AgeCohortHist[Sector_reb_loc, :] = pCStocks_2016_resbuildings.sum(axis =1)
    if 'nrb' in SectorList:    
        pC_AgeCohortHist[Sector_nrb_loc, :] = pCStocks_2016_nresbuildings.sum(axis =1)
    OutputDict['pC_AgeCohortHist']   = pC_AgeCohortHist.copy()
    # b) from future stock curves:
    #This is done for the individual sector calculations below.
    
    # c) Total 2015 material stock, all in Mt!
    if 'pav' in SectorList:
        TotalMaterialStock_2015_pav = np.einsum('cgr,cmgr->mgr',ParameterDict['2_S_RECC_FinalProducts_2015_passvehicles'].Values[0,:,:,:],ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[:,:,:,:,0])/1000
    if 'reb' in SectorList:
        TotalMaterialStock_2015_reb = np.einsum('cgr,cmgr->mgr',ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:],ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,0,0])/1000
    if 'nrb' in SectorList:
        TotalMaterialStock_2015_nrb = np.einsum('cgr,cmgr->mgr',ParameterDict['2_S_RECC_FinalProducts_2015_nonresbuildings'].Values[0,:,:,:],ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[:,:,:,:,0,0])/1000
        
    # 23) Material and process-dependent electricity mix (with aluminium electricity mix)
    # reshape electricity mix: oRit->PRit
    Par_ElectricityMix_P  = np.einsum('RIt,P->PRIt', ParameterDict['4_SHA_ElectricityMix_World'].Values[0,:,:,:], np.ones(NP) )
    if ScriptConfig['Include_AluminiumElectricityMix'] == 'True':
        # overwright aluminium energy mix
        for mP in [PrimCastAl_loc,PrimWrAl_loc,EffCastAl_loc,EffWrAl_loc]:
            Par_ElectricityMix_P[mP,:,:,:] = np.einsum('I,Rt->RIt',ParameterDict['4_SHA_ElectricityMix_World_Alu'].Values[0,0,:,0],np.ones((NR,Nt)))
            
    # Time dependent extensions. Versions: o-version for generic energy use, material-version for primary production. 
    # This is needed because aluminium might have a separate electricity mix, depending on ScriptConfig
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'] = msc.Parameter(Name='4_PE_ProcessExtensions_EnergyCarriers_MJ_o', ID='4_PE_ProcessExtensions_EnergyCarriers_MJ_o',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='nxotR', Values=np.zeros((Nn,Nx,No,Nt,NR)), Uncert=None,
                                                Unit='impact-eq/MJ')
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'] = msc.Parameter(Name='4_PE_ProcessExtensions_EnergyCarriers_MJ_material', ID='4_PE_ProcessExtensions_EnergyCarriers_MJ_material',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='mnxotR', Values=np.zeros((Nm,Nn,Nx,No,Nt,NR)), Uncert=None,
                                                Unit='impact-eq/MJ') # For energy demand of material production
    
    # replicate 2015 values. The current dataset contains only the 2015 initial value. 
    # Formula calculates impact / kg * kg /MJ = impact / MJ
    # Electricity is added separately in the next step
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values         = np.einsum('nxo,n,tR->nxotR',  ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,:,0],ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values,np.ones((Nt,NR)))
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values = np.einsum('nxo,n,tRm->mnxotR',ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,:,0],ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values,np.ones((Nt,NR,Nm)))
    # Replicate 2015 values for 4_PE_ProcessExtensions_Industry
    ParameterDict['4_PE_ProcessExtensions_Industry'].Values = np.einsum('Ixo,t->Ixot',ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,:,0],np.ones((Nt)))
    # add electricity calculated from electriciy mix
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[Electric_loc,:,:,:,:] = np.einsum('oRIt,Ixot->xotR',
                                                                                                        ParameterDict['4_SHA_ElectricityMix_World'].Values,             # MJ industry/MJ el        
                                                                                                        ParameterDict['4_PE_ProcessExtensions_Industry'].Values/3.6)    # impact/MJ industry
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,Electric_loc,:,:,:,:] = np.einsum('oRIt,Ixot,m->mxotR',
                                                                                                        ParameterDict['4_SHA_ElectricityMix_World'].Values,             # MJ industry/MJ el        
                                                                                                        ParameterDict['4_PE_ProcessExtensions_Industry'].Values/3.6,    # impact/MJ industry
                                                                                                        np.ones((Nm)) )
    if ScriptConfig['Include_AluminiumElectricityMix'] == 'True':
        # overwrite aluminium energy mix
        ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[[WroughtAl_loc,CastAl_loc],Electric_loc,:,:,:,:] = np.einsum('oRIt,Ixot,m->mxotR',
                                                                                                            ParameterDict['4_SHA_ElectricityMix_World_Alu'].Values,             # MJ industry/MJ el        
                                                                                                            ParameterDict['4_PE_ProcessExtensions_Industry'].Values/3.6,    # impact/MJ industry
                                                                                                            np.ones((2)) )
    # Now the same, but with extended regional scope for later use
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'] = msc.Parameter(Name='4_PE_ProcessExtensions_EnergyCarriers_MJ_r', ID='4_PE_ProcessExtensions_EnergyCarriers_MJ_r',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='nxrtR', Values=np.zeros((Nn,Nx,Nr,Nt,NR)), Uncert=None,
                                                Unit='[impact unit]/MJ')
    # replicate 2015 values. In current dataset is only initial value. Electricity is added separately in the next step
    # Formula calculates impact / kg * kg /MJ = impact / MJ
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values = np.einsum('nx,n,rtR->nxrtR',ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,0,0],ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values,np.ones((Nr,Nt,NR)))
    
    # add electricity calculated from electriciy mix
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,:] = np.einsum('rRIt,Ixt->xrtR',
                                                                                                        ParameterDict['4_SHA_ElectricityMix'].Values,                            # MJ industry/MJ el        
                                                                                                        ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,:]/3.6)    # impact/MJ industry
    
    # emission factors
    EmissionFactorElectricity_r = ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,GWP100_loc,:,:,:]
    EmissionFactorElectricity_o = ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[Electric_loc,GWP100_loc,0,:,:]
    
    # diagnostic
    impact_elect =  np.einsum('I,Ix->x',
                        ParameterDict['4_SHA_ElectricityMix_World'].Values[0,0,:,0],              # MJ industry/MJ el        
                        ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,0]/3.6,     # impact/MJ industry 
                         )  # Impact/MJ el
        
    # 24) Computing residuals(=process emissions) for material production processes   
    ParameterDict['4_PE_ProcessExtensions_Residual'] = msc.Parameter(Name='4_PE_ProcessExtensions_Residual', ID='4_PE_ProcessExtensions_Residual',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='Pxt', Values=np.zeros((NP,Nx,Nt)), Uncert=None,
                                                Unit='[impact unit]/kg')    
    # Fuel contributions
    fuel_production = np.einsum('nx,n,Pn->Px',
                         ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,0,0], # impact/kg fuel        
                         ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values,             # kg fuel/MJ fuel 
                         ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,0,0]            # MJ fuel/kg mat 
                         )  # Impact/kg mat
    # Direct contributions
    direct_impact = np.einsum('Xn,xX,Pn->Px',
                         ParameterDict['6_PR_DirectEmissions'].Values,             # impact/MJ fuel  
                         ParameterDict['6_MIP_CharacterisationFactors'].Values,
                         ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,0,0]  # MJ fuel/kg mat 
                         )  # Impact/kg mat
    # Electricity generation
    elec_production = np.einsum('P,Pi,ix->Px',
                        ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,Electric_loc,0,0],    # MJ el/kg mat
                        Par_ElectricityMix_P[:,0,:,0],              # MJ industry/MJ el        
                        ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,0]/3.6        # impact/MJ industry 
                         )  # Impact/kg mat
    # compute residuals
    residuals = ParameterDict['4_PE_ProcessExtensions_Materials'].Values[:,:,0,0] - fuel_production - direct_impact - elec_production   # Index_ Px, Unit: [impact unit]/kg mat
    # remove primary energy from residuals, as accounted for separately, only process emisisons here:
    residuals[:,PrimEn_loc] = 0
    # remove fossil fuels from residuals, as accounted for separately, only process emisisons here:
    residuals[:,FosFuel_loc] = 0    
    # remove biomass from residuals, as accounted for separately, only process emisisons here. Keep only biomass input for wood and paper production:
    mfilter = np.zeros((NM))    
    mfilter[WoodProd_loc] = 1
    mfilter[PaperProd_loc] = 1
    residuals[:,Biomass_loc] = residuals[:,Biomass_loc]  * mfilter
     
    # RCP production processes represent technologies that are not out there yet. 
    # They are modelled as the same production steps, with different energy inputs. 
    # Hence, the residual extensions (i.e. the process impacts) are the same as the Baseline technology.
    residuals[EffCastAl_loc,:]  = residuals[PrimCastAl_loc,:]
    residuals[EffWrAl_loc,:]    = residuals[PrimWrAl_loc,:]
    residuals[H2CGSteel_loc,:]  = residuals[PrimCGSteel_loc,:]
    residuals[H2ASteel_loc,:]   = residuals[PrimASteel_loc,:]
    residuals[H2CastIron_loc,:] = residuals[PrimCastIron_loc,:]
    residuals[H2SSteel_loc,:]   = residuals[PrimSSteel_loc,:]    
    ParameterDict['4_PE_ProcessExtensions_Residual'].Values[:,:,0] = residuals
    # Replicate residuals over time
    ParameterDict['4_PE_ProcessExtensions_Residual'].Values = np.einsum('Px,t->Pxt',ParameterDict['4_PE_ProcessExtensions_Residual'].Values[:,:,0],np.ones((Nt)))
    
    # 25) Dynamic primary production intensity for diagnostic purposes
    MaterialProductionIntensity_P = \
        np.einsum('Pxt,R->PxtR',
                  ParameterDict['4_PE_ProcessExtensions_Residual'].Values,
                  np.ones((NR))) + \
        np.einsum('Xn,xX,Pnt,R->PxtR',
                   ParameterDict['6_PR_DirectEmissions'].Values[:,:],              
                   ParameterDict['6_MIP_CharacterisationFactors'].Values,
                   ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0],
                   np.ones((NR))) +\
        np.einsum('nxtR,Pnt->PxtR',
                   ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,0,:,:],              
                   ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0])  
    MaterialProductionIntensity_m = np.einsum('PxtR,tmRP->mxtR',
                  MaterialProductionIntensity_P,
                  ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,:,:])
    MaterialProductionIntensityGHG_m = MaterialProductionIntensity_m[:,GWP100_loc,:,:]
    
        
    ##########################################################
    #    Section 4) Initialize dynamic MFA model for RECC    #
    ##########################################################
    Mylog.info('Initialize dynamic MFA model for RECC')
    Mylog.info('Define RECC system and processes.')
    
    #Define arrays for result export:
    Impacts_System_13579di               = np.zeros((Nx,Nt,NS,NR))
    Impacts_System_3579di                = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_7d                  = np.zeros((Nx,Nt,NS,NR))
    Impacts_OtherThanUsePhaseDirect      = np.zeros((Nx,Nt,NS,NR))
    Impacts_Materials_3di_9di            = np.zeros((Nx,Nt,NS,NR)) # all processes and their energy supply chains except for manufacturing and use phase
    Impacts_Vehicles_Direct              = np.zeros((Nx,Nt,Nr,NS,NR)) # use phase only
    Impacts_ReBuildgs_Direct             = np.zeros((Nx,Nt,Nr,NS,NR)) # use phase only
    Impacts_NRBuildgs_Direct             = np.zeros((Nx,Nt,Nr,NS,NR)) # use phase only
    Impacts_NRBuildgs_Direct_g           = np.zeros((Nx,Nt,NS,NR)) # use phase only
    Impacts_Vehicles_indir               = np.zeros((Nx,Nt,NS,NR)) # energy supply only
    Impacts_AllBuildings_indir           = np.zeros((Nx,Nt,NS,NR)) # energy supply only
    Impacts_Manufact_5di_all             = np.zeros((Nx,Nt,NS,NR))
    Impacts_WasteMgt_9di_all             = np.zeros((Nx,Nt,NS,NR))
    Impacts_PrimaryMaterial_3di          = np.zeros((Nx,Nt,NS,NR))
    Impacts_PrimaryMaterial_3di_m        = np.zeros((Nx,Nt,Nm,NS,NR))
    Impacts_SecondaryMetal_di_m          = np.zeros((Nx,Nt,Nm,NS,NR))
    Impacts_UsePhase_7i_Scope2_El        = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_7i_OtherIndir       = np.zeros((Nx,Nt,NS,NR))
    Impacts_MaterialCycle_5di_9di        = np.zeros((Nx,Nt,NS,NR))
    Impacts_RecyclingCredit              = np.zeros((Nx,Nt,NS,NR))
    Impacts_ForestCO2Uptake              = np.zeros((Nx,Nt,NS,NR))
    Impacts_ForestCO2Uptake_r            = np.zeros((Nx,Nt,Nr,NS,NR))
    Impacts_EnergyRecoveryWasteWood      = np.zeros((Nx,Nt,NS,NR))
    Impacts_ByEnergyCarrier_UsePhase_d   = np.zeros((Nx,Nt,Nr,Nn,NS,NR))
    Impacts_ByEnergyCarrier_UsePhase_i   = np.zeros((Nx,Nt,Nr,Nn,NS,NR))
    Impacts_Energy_Supply_All            = np.zeros((Nx,Nt,NS,NR)) # impacts of all energy carriers for all processes. 
    
    dynGWP_System_3579di                 = np.zeros((NS,NR)) # dynGWP100 of entire system
    dynGWP_WoodCycle                     = np.zeros((NS,NR)) # dynGWP100 of wood use: forest uptake + wood-related emissions from waste mgt. Pos sign for flow from system to environment.
    
    Material_Inflow                  = np.zeros((Nt,Ng,Nm,NS,NR))
    Scrap_Outflow                    = np.zeros((Nt,Nw,NS,NR))
    PrimaryProduction                = np.zeros((Nt,Nm,NS,NR))
    SecondaryProduct                 = np.zeros((Nt,Nm,NS,NR))
    SecondaryExport                  = np.zeros((Nt,Nm,NS,NR))
    SecondaryProduct_EoL_Pot         = np.zeros((Nt,Nm,NS,NR)) # Secondary material from EoL material flows only, part of F_9_12, for reporting only
    RenovationMaterialInflow_7       = np.zeros((Nt,Nm,NS,NR))
    Element_Material_Composition     = np.zeros((Nt,Nm,Ne,NS,NR))
    Element_Material_Composition_raw = np.zeros((Nt,Nm,Ne,NS,NR))
    Element_Material_Composition_con = np.zeros((Nt,Nm,Ne,NS,NR))
    Manufacturing_Output             = np.zeros((Nt,Ng,Nm,NS,NR))
    StockMatch_2015                  = np.zeros((NG,Nr))
    NegInflowFlags                   = np.zeros((NS,NR))
    #NegInflowFlags_After2020         = np.zeros((NS,NR))
    FabricationScrap                 = np.zeros((Nt,Nw,NS,NR))
    EnergyCons_UP_Vh                 = np.zeros((Nt,NS,NR))
    EnergyCons_UP_Bd                 = np.zeros((Nt,NS,NR))
    EnergyCons_Mn                    = np.zeros((Nt,NS,NR))
    EnergyCons_Wm                    = np.zeros((Nt,NS,NR))
    EnergyCons_PP                    = np.zeros((Nt,NS,NR))
    EnergyCons_PP_m                  = np.zeros((Nt,Nm,NS,NR))
    EnergyCons_UP_serv_pav           = np.zeros((Nt,Nr,NV,NS,NR))
    EnergyCons_UP_serv_reb           = np.zeros((Nt,Nr,NV,NS,NR))
    EnergyCons_UP_serv_nrb           = np.zeros((Nt,Nr,NV,NS,NR))
    EnergyCons_UP_total              = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_UP_reb                = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_UP_nrb                = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_total                 = np.zeros((Nt,Nn,NS,NR))
    StockCurves_Totl                 = np.zeros((Nt,NG,NS,NR))
    StockCurves_Prod                 = np.zeros((Nt,Ng,NS,NR))
    StockCurves_Mat                  = np.zeros((Nt,Nm,NS,NR))
    Inflow_Prod                      = np.zeros((Nt,Ng,NS,NR))
    Inflow_Prod_r                    = np.zeros((Nt,Nr,Ng,NS,NR))
    Outflow_Prod                     = np.zeros((Nt,Ng,NS,NR))
    Outflow_Prod_r                   = np.zeros((Nt,Nr,Ng,NS,NR))
    EoL_Products_for_WasteMgt        = np.zeros((Nt,Ng,NS,NR))
    Outflow_Materials_Usephase_all   = np.zeros((Nt,Nm,NS,NR))
    Outflow_Products_Usephase_all    = np.zeros((Nt,Ng,NS,NR))
    WasteMgtLosses_To_Landfill       = np.zeros((Nt,Ne,NS,NR))
    Population                       = np.zeros((Nt,Nr,NS,NR))
    pCStocksCurves                   = np.zeros((Nt,NG,Nr,NS,NR))
    Passenger_km                     = np.zeros((Nt,NS,NR))
    Vehicle_km                       = np.zeros((Nt,NS,NR))
    Service_IO_ResBuildings          = np.zeros((Nt,NV,NS,NR))
    Service_IO_NonResBuildings       = np.zeros((Nt,NV,NS,NR))
    ReUse_Materials                  = np.zeros((Nt,Nm,NS,NR))
    Carbon_IndustrialRoundwood_bld   = np.zeros((Nt,Nr,NS,NR)) # Industrial roundwood, hard and softwood, for processing into structural wood elements for residential and non-residential buildings. Unit: Mt/yr of C.
    Carbon_Fuelwood_bld              = np.zeros((Nt,NS,NR)) # Fuelwood, hard and softwood, for use in building heating and hot water only (no cooking fuel).
    Carbon_Fuelwood_el               = np.zeros((Nt,NS,NR)) # Fuelwood, hard and softwood, for use in electricity generation.
    Carbon_Fuelwood_release          = np.zeros((Nt,NS,NR)) # Total wood C outflow from fuelwood, in form of CO2.
    Carbon_Wood_Inflow               = np.zeros((Nt,Nr,NS,NR))
    Carbon_Wood_Outflow              = np.zeros((Nt,Nr,NS,NR))
    Carbon_Wood_Stock                = np.zeros((Nt,Nr,NS,NR))
    Cement_Inflow                    = np.zeros((Nt,Nr,NS,NR))
    Vehicle_FuelEff                  = np.zeros((Nt,Np,Nr,NS,NR))
    ResBuildng_EnergyCons            = np.zeros((Nt,NB,Nr,NS,NR))
    GWP_bio_Credit                   = np.zeros((Nt,NS,NR))
    EnergySubst_WtE_EL               = np.zeros((Nt,NS,NR))
    EnergySubst_WtE_NG               = np.zeros((Nt,NS,NR))
    FuelWoodSubst_WoodWaste          = np.zeros((Nt,NS,NR))
    BiogenicCO2WasteCombustion       = np.zeros((Nt,NS,NR))
    SysVar_RoundwoodConstruc_c_1_2_r = np.zeros((Nt,Nr,NS,NR))
    SysVar_WoodWasteIncineration     = np.zeros((Nt,Nr,Nw,Ne,NS,NR))
    SysVar_EoLCascEntry              = np.zeros((Nt,Nr,NS,NR))
    SysVar_CascadeRelease            = np.zeros((Nt,Nr,Nw,Ne,NS,NR))
    SysVar_WoodWaste_Gas_El          = np.zeros((Nt,Nr,NS,NR))
    WoodCascadingInflow              = np.zeros((Nt,Nr,NS,NR))
    WoodCascadingStock               = np.zeros((Nt,Nr,NS,NR))
    Stock_2020_pav                   = np.zeros((Nt,Nr,NS,NR))
    Stock_2020_reb                   = np.zeros((Nt,Nr,NS,NR))
    Stock_2020_nrb                   = np.zeros((Nt,Nr,NS,NR))
    NegInflowFlags                   = np.zeros((NG,NS,NR))
    time_dsm                         = np.arange(0,Nc,1) # time array of [0:Nc) needed for some sectors
    
    ExitFlags = {} # Exit flags for individual model runs
    #  Examples for testing
    #mS = 1
    #mR = 1
    
    # Select and loop over scenarios
    for mS in range(0,NS):
        for mR in range(0,NR):
    
            SName = IndexTable.loc['Scenario'].Classification.Items[mS]
            RName = IndexTable.loc['Scenario_RCP'].Classification.Items[mR]
            Mylog.info('_')
            Mylog.info('Computing RECC model for SSP scenario ' + SName + ' and RE scenario ' + RName + '.')
            
            # Initialize MFA system
            RECC_System = msc.MFAsystem(Name='RECC_SingleScenario',
                                        Geogr_Scope='19 regions + 1 single country', #IndexTableR.Classification[IndexTableR.set_index('IndexLetter').index.get_loc('r')].Items,
                                        Unit='Mt',
                                        ProcessList=[],
                                        FlowDict={},
                                        StockDict={},
                                        ParameterDict=ParameterDict,
                                        Time_Start=Model_Time_Start,
                                        Time_End=Model_Time_End,
                                        IndexTable=IndexTable,
                                        Elements=IndexTable.loc['Element'].Classification.Items,
                                        Graphical=None)
                                  
            # Check Validity of index tables:
            # returns true if dimensions are OK and time index is present and element list is not empty
            RECC_System.IndexTableCheck()
            
            # Add processes to system
            for m in range(0, len(PrL_Number)):
                RECC_System.ProcessList.append(msc.Process(Name = PrL_Name[m], ID   = PrL_Number[m]))
                
            # Define system variables: Flows.    
            RECC_System.FlowDict['F_0_1']     = msc.Flow(Name='CO2 uptake', P_Start=0, P_End=1,
                                                     Indices='t,r,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_1_2']     = msc.Flow(Name='harvested wood', P_Start=1, P_End=2,
                                                     Indices='t,r,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_2_3']     = msc.Flow(Name='timber consumed by sawmills', P_Start=2, P_End=3,
                                                     Indices='t,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_2_7']     = msc.Flow(Name='wood fuel use', P_Start=2, P_End=7,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None) # flow is directly routed to use phase.
    
            RECC_System.FlowDict['F_7_0']     = msc.Flow(Name='wood fuel use direct emissions', P_Start=7, P_End=0,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_0_3']     = msc.Flow(Name='ore input', P_Start=0, P_End=3,
                                                     Indices='t,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_3_4']     = msc.Flow(Name='primary material production' , P_Start = 3, P_End = 4, 
                                                     Indices = 't,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_3_10']    = msc.Flow(Name='primary material production waste' , P_Start = 3, P_End = 10, 
                                                     Indices = 't,r,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_4_5']     = msc.Flow(Name='primary material consumption' , P_Start = 4, P_End = 5, 
                                                     Indices = 't,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_5_6']     = msc.Flow(Name='manufacturing output' , P_Start = 5, P_End = 6, 
                                                     Indices = 't,o,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                
            RECC_System.FlowDict['F_6_7']     = msc.Flow(Name='final consumption', P_Start=6, P_End=7,
                                                     Indices='t,r,g,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_6_7_Nl']  = msc.Flow(Name='final consumption Nl', P_Start=6, P_End=7,
                                                     Indices='t,l,L,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_6_7_No']  = msc.Flow(Name='final consumption No', P_Start=6, P_End=7,
                                                     Indices='t,o,O,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_7_8']     = msc.Flow(Name='EoL products' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_7_8_Nl']  = msc.Flow(Name='EoL products Nl' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
        
            RECC_System.FlowDict['F_7_8_No']  = msc.Flow(Name='EoL products No' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_0']     = msc.Flow(Name='obsolete stock formation' , P_Start = 8, P_End = 0, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_8_0_Nl']  = msc.Flow(Name='obsolete stock formation Nl' , P_Start = 8, P_End = 0, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_0_No']  = msc.Flow(Name='obsolete stock formation No' , P_Start = 8, P_End = 0, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_9']     = msc.Flow(Name='waste mgt. input' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_9_Nl']  = msc.Flow(Name='waste mgt. input Nl' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_9_No']  = msc.Flow(Name='waste mgt. input No' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17']    = msc.Flow(Name='product re-use in' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17_Nl'] = msc.Flow(Name='product re-use in Nl' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17_No'] = msc.Flow(Name='product re-use in No' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                    
            RECC_System.FlowDict['F_17_6']    = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_17_6_Nl'] = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
     
            RECC_System.FlowDict['F_17_6_No'] = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                    
            RECC_System.FlowDict['F_9_10']    = msc.Flow(Name='old scrap' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,r,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_10_Nl'] = msc.Flow(Name='old scrap Nl' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,l,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_10_No'] = msc.Flow(Name='old scrap No' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                    
            RECC_System.FlowDict['F_5_10']    = msc.Flow(Name='new scrap' , P_Start = 5, P_End = 10, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_10_9']    = msc.Flow(Name='scrap use' , P_Start = 10, P_End = 9, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_10_9w']   = msc.Flow(Name='wood waste use' , P_Start = 10, P_End = 9, 
                                                     Indices = 't,r,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_12']    = msc.Flow(Name='secondary material production' , P_Start = 9, P_End = 12, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_10_12']   = msc.Flow(Name='fabscrapdiversion' , P_Start = 10, P_End = 12, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)        
            
            RECC_System.FlowDict['F_12_5']    = msc.Flow(Name='secondary material consumption' , P_Start = 12, P_End = 5, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_12_0']    = msc.Flow(Name='excess secondary material' , P_Start = 12, P_End = 0, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_0']     = msc.Flow(Name='waste mgt. and remelting losses' , P_Start = 9, P_End = 0, 
                                                     Indices = 't,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            # Define system variables: Stocks.
            RECC_System.StockDict['dS_0']     = msc.Stock(Name='System environment stock change', P_Res=0, Type=1,
                                                     Indices = 't,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_1t']    = msc.Stock(Name='Forestry stock change, timber', P_Res=1, Type=1,
                                                     Indices = 't,r,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_1t']     = msc.Stock(Name='Forestry carbon stock, timber', P_Res=1, Type=0,
                                                     Indices = 't,c,r,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_1f']    = msc.Stock(Name='Forestry stock change, fuel wood', P_Res=1, Type=1,
                                                     Indices = 't,r,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_1f']     = msc.Stock(Name='Forestry carbon stock, fuel wood', P_Res=1, Type=0,
                                                     Indices = 't,c,r,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)        
            
            RECC_System.StockDict['S_7']      = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_7_Nl']   = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_7_No']   = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_7']     = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_7_Nl']  = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_7_No']  = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_9']     = msc.Stock(Name='Wood cascading buffer', P_Res=9, Type=0,
                                                     Indices = 't,c,r,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_9']    = msc.Stock(Name='Wood cascading buffer change', P_Res=9, Type=1,
                                                     Indices = 't,c,r,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_10']     = msc.Stock(Name='Fabrication scrap buffer', P_Res=10, Type=0,
                                                     Indices = 't,c,o,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_10']    = msc.Stock(Name='Fabrication scrap buffer change', P_Res=10, Type=1,
                                                     Indices = 't,o,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_10w']    = msc.Stock(Name='Wood waste buffer', P_Res=10, Type=0,
                                                     Indices = 't,c,r,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_10w']   = msc.Stock(Name='Wood waste buffer change', P_Res=10, Type=1,
                                                     Indices = 't,r,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_12']     = msc.Stock(Name='secondary material buffer', P_Res=12, Type=0,
                                                     Indices = 't,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_12']    = msc.Stock(Name='Secondary material buffer change', P_Res=12, Type=1,
                                                     Indices = 't,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.Initialize_StockValues() # Assign empty arrays to stocks according to dimensions.
            RECC_System.Initialize_FlowValues()  # Assign empty arrays to flows according to dimensions.
            
            ##########################################################
            #    Section 5) Solve dynamic MFA model for RECC         #
            ##########################################################
            Mylog.info('Solve dynamic MFA model for the RECC project for all sectors chosen.')
            # THIS IS WHERE WE EXPAND FROM THE FORMAL MODEL STRUCTURE AND CODE WHATEVER IS NECESSARY TO SOLVE THE MODEL EQUATIONS.
            SwitchTime=Nc - Model_Duration +1 # Year when future modelling horizon starts: 1.1.2016        
    
            # Determine empty result containers for all sectors
            Stock_Detail_UsePhase_p     = np.zeros((Nt,Nc,Np,Nr)) # index structure: tcpr. Unit: million items.
            Stock_2020_decline_p        = np.zeros((Nt,Np,Nr))    # index structure: tpr.  Unit: million items.
            Outflow_Detail_UsePhase_p   = np.zeros((Nt,Nc,Np,Nr)) # index structure: tcpr. Unit: million items.
            Inflow_Detail_UsePhase_p    = np.zeros((Nt,Np,Nr))    # index structure: tpr.  Unit: million items.
            
            Stock_Detail_UsePhase_B     = np.zeros((Nt,Nc,NB,Nr)) # index structure: tcBr. Unit: million m².
            Stock_2020_decline_B        = np.zeros((Nt,NB,Nr))    # index structure: tBr.  Unit: million m².
            Stock_2020_agestruct_B      = np.zeros((Nc,NB,Nr))    # index structure: cBr.  Unit: million m².
            Outflow_Detail_UsePhase_B   = np.zeros((Nt,Nc,NB,Nr)) # index structure: tcBr. Unit: million m².
            Inflow_Detail_UsePhase_B    = np.zeros((Nt,NB,Nr))    # index structure: tBr.  Unit: million m².
        
            Stock_Detail_UsePhase_N     = np.zeros((Nt,Nc,NN,Nr)) # index structure: tcNr. Unit: million m².
            Stock_2020_decline_N        = np.zeros((Nt,NN,Nr))    # index structure: tNr.  Unit: million m².
            Outflow_Detail_UsePhase_N   = np.zeros((Nt,Nc,NN,Nr)) # index structure: tcNr. Unit: million m².
            Inflow_Detail_UsePhase_N    = np.zeros((Nt,NN,Nr))    # index structure: tNr.  Unit: million m². 
            
            Stock_Detail_UsePhase_Ng    = np.zeros((Nt,Nc,NN,No)) # index structure: tcNo. Unit: million m².
            Outflow_Detail_UsePhase_Ng  = np.zeros((Nt,Nc,NN,No)) # index structure: tcNo. Unit: million m².
            Inflow_Detail_UsePhase_Ng   = np.zeros((Nt,NN,No))    # index structure: tNo.  Unit: million m².         
            
            Stock_Detail_UsePhase_I     = np.zeros((Nt,Nc,NI,Nl)) # index structure: tcIL. Unit: GW.
            Outflow_Detail_UsePhase_I   = np.zeros((Nt,Nc,NI,Nl)) # index structure: tcIl. Unit: GW.
            Inflow_Detail_UsePhase_I    = np.zeros((Nt,NI,Nl))    # index structure: tIl.  Unit: GW.
        
            Stock_Detail_UsePhase_a     = np.zeros((Nt,Nc,Na,No)) # index structure: tcao. Unit: # of items (1).
            Outflow_Detail_UsePhase_a   = np.zeros((Nt,Nc,Na,No)) # index structure: tcao. Unit: # of items (1).
            Inflow_Detail_UsePhase_a    = np.zeros((Nt,Na,No))    # index structure: tao.  Unit: # of items (1).    
            
            F_6_7_ren                   = np.zeros((Nt,Nc,Nr,Ng,Nm,Ne)) # Indices='t,c,r,g,m,e', # inflow of renovation material, Mt/yr
            F_6_7_new                   = np.zeros((Nt,Nr,Ng,Nm,Ne))    # Indices='t,r,g,m,e',   # inflow of material in new products, Mt/yr
        
            # Sector: Passenger vehicles
            if 'pav' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, passenger vehicles.')
                # 1) Determine kilometrage endogenously and apply stock-driven model
                SF_Array                    = np.zeros((Nc,Nc,Np,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                
                #Get historic stock at end of 2015 by age-cohort, and covert unit to Vehicles: million.
                TotalStock_UsePhase_Hist_cpr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_passvehicles'].Values[0,:,:,:]
                
                # Determine total future stock, product level. Units: Vehicles: million.
                # Option implemented: By service curve.
                Total_Service_pav_tr_pC                     = np.einsum('rt->tr',RECC_System.ParameterDict['1_F_Function_Future'].Values[Sector_pav_loc,:,:,mS])
                if ScriptConfig['Include_REStrategy_CarSharing'] == 'False': # set carsharing to zero.
                    RECC_System.ParameterDict['6_PR_CarSharingShare'].Values  = np.zeros(RECC_System.ParameterDict['6_PR_CarSharingShare'].Values.shape)
                if ScriptConfig['Include_REStrategy_RideSharing'] == 'False': # set ride-sharing to zero.                
                    RECC_System.ParameterDict['6_PR_RideSharingShare'].Values = np.zeros(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values.shape)
    
                # i) Calculate pc stocks in the four subdivisions: CaS, RiS, CaS+RiS, none:
                Total_Vehicle_km_pav_tr_pC          = np.zeros((Nt,Nr))
                TotalStockCurves_UsePhase_p_pC_test = np.zeros((Nt,Nr)) 
                for nrr in range(0,Nr):
                    for ntt in range(0,Nt):   
                        # convert abs. change in vehicle OR to relative change:
                        MIP_RideSharing_Occupancy_rel = (RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] + RECC_System.ParameterDict['6_MIP_RideSharing_Occupancy'].Values[mS,nrr]) / RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS]
                        # calculate future stock levels:
                        s0        = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(1 - RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS])
                        s_CaS     = (RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(1 - RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  * (RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr])
                        s_RiS     = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  / (MIP_RideSharing_Occupancy_rel)
                        s_CaS_RiS = (RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  / (MIP_RideSharing_Occupancy_rel / RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr])
    
                        s_total   = s0.copy() + s_CaS.copy() + s_RiS.copy() + s_CaS_RiS.copy()
                        TotalStockCurves_UsePhase_p_pC_test[ntt,nrr] = s_total.copy()
                        TotalStockCurves_UsePhase_p_pC_test[np.isnan(TotalStockCurves_UsePhase_p_pC_test)] = 0 # ignore drive technologies where there is no stock.
    
                        # ii) Calculate average vehicle kilometrage and average occupancy rate:
                        vkm       = ((s0 + s_RiS) + (s_CaS + s_CaS_RiS) / RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr]) * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS].copy() / s_total
                        Total_Vehicle_km_pav_tr_pC[ntt,nrr] = vkm.copy()
                        # Overwrite predefined values by internally calculated vehicle-km:
                        RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[Service_Drivg,nrr,ntt,mS] = Total_Vehicle_km_pav_tr_pC[ntt,nrr]
                        #ocr       = Total_Service_pav_tr_pC[ntt,nrr] / (s_total * vkm)
                        
                RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[np.isnan(RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values)] = 0
                # iii) Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.         
                # This needs to be made sure during the scenario framing process! Here, only the accounting and model equations to convert PKM to VKM and stock are executed, no further checks are made.
                TotalStockCurves_UsePhase_p_pC   = TotalStockCurves_UsePhase_p_pC_test.copy()
                TotalStockCurves_UsePhase_p      = np.einsum('tr,tr->tr',TotalStockCurves_UsePhase_p_pC, RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS])
                RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'].Values[mS,:,Sector_pav_loc,:] = TotalStockCurves_UsePhase_p_pC.copy() 
                
                # iv) adjust vehicle lifetime to new effective value to reflect impact of car-sharing:
                # First, replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_p = np.einsum('c,pr->prc',np.ones((Nc)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_passvehicles'].Values)
                # Second, adjust lifetime if car-sharing is present
                if ScriptConfig['Include_REStrategy_CarSharing'] == 'True': # adjust lifetime of future age-cohorts.
                    for npp in range(0,Np):
                        for nrr in range(0,Nr):
                            for ntt in range(0,Nt):
                                Par_RECC_ProductLifetime_p[npp,nrr,ntt+SwitchTime-1] = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS]/100 + RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] * RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr]/100) * Par_RECC_ProductLifetime_p[npp,nrr,ntt+SwitchTime-1]
                            
                # Include_REStrategy_LifeTimeExtension: Product lifetime extension.
                # Third, change lifetime of future age-cohorts according to lifetime extension parameter
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                    Par_RECC_ProductLifetime_p[:,:,SwitchTime -1::] = np.einsum('crp,prc->prc',1 + np.einsum('cr,pr->crp',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values[:,:,mS]),Par_RECC_ProductLifetime_p[:,:,SwitchTime -1::])
                
                # 2) Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                for p in tqdm(range(0, Np), unit=' vehicles types'):
                    for r in range(0, Nr):
                        LifeTimes = Par_RECC_ProductLifetime_p[p, r, :]
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.5 * LifeTimes} # flat decline: obsolescence of 16 % in 3 years around mean lifetime.
                        SF_Array[:, :, p, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                        np.fill_diagonal(SF_Array[:, :, p, r],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
        
                # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
                for r in range(0,Nr):   
                    FutureStock                 = np.zeros((Nc))
                    FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_p[1::, r].copy() # Future total stock
                    InitialStock                = TotalStock_UsePhase_Hist_cpr[:,:,r].copy()
                    InitialStocksum             = InitialStock.sum()
                    StockMatch_2015[Sector_pav_loc,r] = TotalStockCurves_UsePhase_p[0, r]/InitialStocksum
                    SFArrayCombined             = SF_Array[:,:,:,r]
                    TypeSplit                   = np.zeros((Nc,Np))
                    TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[Sector_pav_loc,r,mR,:,1::].transpose() # indices: cp
                    
                    RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                    Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                    
                    # For a single-vehicle dynamic LCA, use this code block below to replace actual flows by single vehicle with fixed lifetime:
    #                VType = 4 # 0 for gasoline, 4 for BEV
    #                VLT   = 15 # fixed lifetime in years
    #                VTin  = 120 # year index of production, 120 stands for 2020.
    #                Var_I = np.zeros((Nc,Np))
    #                Var_I[VTin,VType] = 1
    #                Var_S = np.zeros((Nc,Nc,Np))
    #                Var_S[VTin:VTin+VLT,VTin,VType] = 1
    #                Var_O = np.zeros((Nc,Nc,Np))
    #                Var_O[VTin+VLT,VTin,VType] = 1
    #                InitialStock = np.zeros((Nc,Np))
                    
                    # Below, the results are added with += because the different commodity groups (buildings, vehicles) are calculated separately
                    # to introduce the type split for each, but using the product resolution of the full model with all sectors.
                    Stock_Detail_UsePhase_p[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                    Stock_Detail_UsePhase_p[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcpr
                    Stock_2020_decline_p[1::,:,r]        += Var_S[SwitchTime::,0:Ind_2020+1,:].copy().sum(axis=1) # tpr
                    Outflow_Detail_UsePhase_p[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcpr
                    Inflow_Detail_UsePhase_p[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tpr
                    Inflow_Prod_r[1::,r,Sector_pav_rge,mS,mR] = Var_I[SwitchTime::,:].copy()
                    # Check for negative inflows:
                    if IFlags.sum() != 0:
                        NegInflowFlags[Sector_pav_loc,mS,mR] = 1 # flag this scenario                
                    
                # Here so far: Units: Vehicles: million. for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_pav_loc,mS,mR] = TotalStockCurves_UsePhase_p.sum(axis =1).copy()
                StockCurves_Prod[:,Sector_pav_rge,mS,mR] = np.einsum('tcpr->tp',Stock_Detail_UsePhase_p).copy()
                pCStocksCurves[:,Sector_pav_loc,:,mS,mR] = TotalStockCurves_UsePhase_p_pC.copy()
                Population[:,:,mS,mR]                    = RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]
                Inflow_Prod[:,Sector_pav_rge,mS,mR]      = np.einsum('tpr->tp',Inflow_Detail_UsePhase_p).copy()
                Inflow_Prod_r[:,:,Sector_pav_rge,mS,mR]  = np.einsum('tpr->trp',Inflow_Detail_UsePhase_p).copy()
                Outflow_Prod[:,Sector_pav_rge,mS,mR]     = np.einsum('tcpr->tp',Outflow_Detail_UsePhase_p).copy()
                Outflow_Prod_r[:,:,Sector_pav_rge,mS,mR]     = np.einsum('tcpr->trp',Outflow_Detail_UsePhase_p).copy()
    
            # Sector: Residential buildings
            if 'reb' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, residential buildings.')
                # 1) Determine total stock and apply stock-driven model
                SF_Array                    = np.zeros((Nc,Nc,NB,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                
                #Get historic stock at end of 2015 by age-cohort, and covert unit to Buildings: million m2.
                TotalStock_UsePhase_Hist_cBr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:]
                
                # Determine total future stock, product level. Units: Buildings: million m2.
                TotalStockCurves_UsePhase_B_pC_test = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[mS,:,Sector_reb_loc,:]
                    
                # 2) Include (or not) the RE strategies for the use phase:
                # Include_REStrategy_MoreIntenseUse:
                if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True': 
                    # Calculate counter-factual scenario: X% decrease of stock levels by 2050 compared to scenario reference. X coded in parameter ..._MIUPotential
                    if SName != 'LED':
                        RemainingFraction = 1-RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_MIUPotential'].Values[Sector_reb_loc,0,mS] / 100
                        #clamped_spline = make_interp_spline(np.arange(0,Nt,1), MIURamp, bc_type=([(2, 0)], [(1, 0)]))
                        clamped_spline = make_interp_spline([0,2,Nt-5,Nt], [1,1,RemainingFraction,RemainingFraction], bc_type=([(2, 0)], [(1, 0)]))
                        MIURamp_Spline = clamped_spline(np.arange(0,Nt,1))
                        MIURamp_Spline[MIURamp_Spline>1]=1
                        MIURamp_Spline[MIURamp_Spline<RemainingFraction]=RemainingFraction
                        
                        TotalStockCurves_UsePhase_B_pC_test    = TotalStockCurves_UsePhase_B_pC_test * np.einsum('t,r->tr',MIURamp_Spline,np.ones((Nr)))
                # Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.
                TotalStockCurves_UsePhase_B_pC_LED_ref = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[LEDindex,:,Sector_reb_loc,:]
                TotalStockCurves_UsePhase_B_pC         = np.maximum(TotalStockCurves_UsePhase_B_pC_test,TotalStockCurves_UsePhase_B_pC_LED_ref)
                TotalStockCurves_UsePhase_B            = np.einsum('tr,tr->tr',TotalStockCurves_UsePhase_B_pC,RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]) 
                RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mS,:,Sector_reb_loc,:] = TotalStockCurves_UsePhase_B_pC.copy()
            
                # Include_REStrategy_LifeTimeExtension: Product lifetime extension.
                # First, replicate lifetimes for post 2020 age-cohorts from 2020 values as the parameter file only specifies values up to 2020:
                RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[:,:,120::] = np.einsum('c,Br->Brc',np.ones((41)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[:,:,120]).copy()
                Par_RECC_ProductLifetime_B = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values.copy()
                # Second, change lifetime of future age-cohorts according to lifetime extension parameter
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                    # option: future age-cohorts only, used in ODYM-RECC v2.2:
                    #Par_RECC_ProductLifetime_B[:,:,SwitchTime -1::] = np.einsum('crB,Brc->Brc',1 + np.einsum('cr,Br->crB',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_LifeTimeExtension_resbuildings'].Values[:,:,mS]),Par_RECC_ProductLifetime_B[:,:,SwitchTime -1::])
                    # option: all age-cohorts, used from ODYM-RECC v2.3 onwards, which leads to:
                    if ScriptConfig['Include_Renovation_reb'] == 'True' and ScriptConfig['No_EE_Improvements'] == 'False': 
                        # Increase lifetime of all res. buildings instantaneously:
                        Par_RECC_ProductLifetime_B = np.einsum('Brc,Brc->Brc',np.einsum('Br,c->Brc',1 + RECC_System.ParameterDict['6_PR_LifeTimeExtension_resbuildings'].Values[:,:,mS],np.ones(Nc)),Par_RECC_ProductLifetime_B)
                    else:
                        # gradual increase of lifetime by age-cohort, including historic age-cohorts, starting from 0:
                        for B in range(0, NB):
                            for r in range(0, Nr):
                                LTE_Pot = RECC_System.ParameterDict['6_PR_LifeTimeExtension_resbuildings'].Values[B,r,mS]
                                LTE_Rampupcurve = np.zeros(Nc)
                                try:
                                    LTE_Rampupcurve[0:SwitchTime] = np.arange(0,LTE_Pot,LTE_Pot/SwitchTime)
                                except:
                                    None # LTE_Pot = 0, no LTE
                                LTE_Rampupcurve[SwitchTime::] = LTE_Pot
                                Par_RECC_ProductLifetime_B[B,r,:] = np.einsum('c,c->c',1 + LTE_Rampupcurve,Par_RECC_ProductLifetime_B[B,r,:])
    
                # 3) Dynamic stock model, with lifetime depending on age-cohort.
                # Build pdf array from lifetime distribution: Probability of survival.
                for B in tqdm(range(0, NB), unit=' res. building types'):
                    for r in range(0, Nr):
                        LifeTimes = Par_RECC_ProductLifetime_B[B, r, :]
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
                        SF_Array[:, :, B, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                        np.fill_diagonal(SF_Array[:, :, B, r],1) # no outflows from current year, 
                        # this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
        
                # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
                for r in range(0,Nr):   
                    FutureStock                 = np.zeros((Nc))
                    FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_B[1::, r].copy()# Future total stock
                    InitialStock                = TotalStock_UsePhase_Hist_cBr[:,:,r].copy()
                    InitialStocksum             = InitialStock.sum()
                    StockMatch_2015[Sector_reb_loc,r] = TotalStockCurves_UsePhase_B[0, r]/InitialStocksum
                    SFArrayCombined             = SF_Array[:,:,:,r]
                    TypeSplit                   = np.zeros((Nc,NB))
                    TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,r,1::,mR,mS].transpose() # indices: Bc
                    
                    RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                    Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                    
                    # Below, the results are added with += because the different commodity groups (buildings, vehicles) are calculated separately
                    # to introduce the type split for each, but using the product resolution of the full model with all sectors.
                    Stock_Detail_UsePhase_B[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                    Stock_Detail_UsePhase_B[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcBr
                    Stock_2020_decline_B[1::,:,r]        += Var_S[SwitchTime::,0:Ind_2020+1,:].copy().sum(axis=1) # tBr
                    Stock_2020_agestruct_B[:,:,r]        += Var_S[Ind_2020,:,:].copy() # cBr
                    Outflow_Detail_UsePhase_B[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcBr
                    Inflow_Detail_UsePhase_B[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tBr
                    # Check for negative inflows:
                    if IFlags.sum() != 0:
                        NegInflowFlags[Sector_reb_loc,mS,mR] = 1 # flag this scenario
        
                # Here so far: Units: Buildings: million m². for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_reb_loc,mS,mR] = TotalStockCurves_UsePhase_B.sum(axis =1).copy()
                StockCurves_Prod[:,Sector_reb_rge,mS,mR] = np.einsum('tcBr->tB',Stock_Detail_UsePhase_B).copy()
                pCStocksCurves[:,Sector_reb_loc,:,mS,mR] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mS,:,Sector_reb_loc,:].copy()
                Population[:,:,mS,mR]                    = RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]
                Inflow_Prod[:,Sector_reb_rge,mS,mR]      = np.einsum('tBr->tB',Inflow_Detail_UsePhase_B).copy()
                Inflow_Prod_r[:,:,Sector_reb_rge,mS,mR]  = np.einsum('tBr->trB',Inflow_Detail_UsePhase_B).copy()
                Outflow_Prod[:,Sector_reb_rge,mS,mR]     = np.einsum('tcBr->tB',Outflow_Detail_UsePhase_B).copy()
                Outflow_Prod_r[:,:,Sector_reb_rge,mS,mR] = np.einsum('tcpr->trp',Outflow_Detail_UsePhase_B).copy()
                
                # Include renovation of reb:
                RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS,mR] = np.einsum('cmBr,t->mBrct',RECC_System.ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,mS,mR],np.ones(Nt)) # mBrctSR
                if ScriptConfig['Include_Renovation_reb'] == 'True' and ScriptConfig['No_EE_Improvements'] == 'False': 
                    RenPot_E   = np.einsum('rcB,rB->rcB',RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_ResBuildings'].Values[:,0:SwitchTime,:],RECC_System.ParameterDict['3_SHA_EnergySavingsPot_Renovation_ResBuildings'].Values[:,mS,:]) # Unit: 1
                    RenPot_E_t = np.einsum('tr,rcB->trcB',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RenPot_E) # Unit: 1, Defined as share of stock crB that is renovated by year t * energy saving potential
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cBVnr,trcB->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:SwitchTime,:,:,:,:,mS,mR],(np.ones((Nt,Nr,Nc-Nt+1,NB))-RenPot_E_t)) # cBVnrt
                    # Add renovation material intensity to building material intensity:
                    RenPot_M_t = np.einsum('tr,rcB->trcB',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_ResBuildings'].Values[:,0:SwitchTime,:]) # Unit: 1, Defined as share of stock crB that is renovated by year t
                    MC_Ren = RECC_System.ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,mS,mR]*RECC_System.ParameterDict['3_MC_RECC_Buildings_Renovation_Relative'].Values + RECC_System.ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values
                    RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,0:SwitchTime,:,mS,mR] += np.einsum('cmBr,trcB->mBrct',MC_Ren[0:SwitchTime,:,:,:],RenPot_M_t)
                else:
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cBVnr,trcB->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:SwitchTime,:,:,:,:,mS,mR],np.ones((Nt,Nr,Nc-Nt+1,NB))) # cBVnrt
                # Add values for future age-cohorts, convert from useful to final energy, expand from 'all' to specific energy carriers
                RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime-1::,:,:,:,:,:]   = np.einsum('Vrnt,Vrnt,cBVr,t->cBVnrt',ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,mR,:,:,:,mS],ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'].Values[:,mR,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchTime-1::,:,:,all_loc,:,mS,mR],np.ones(Nt))
                # Split energy into different carriers for historic age-cohorts. Here, a shift of heating systems is assumed irrespective of the renovation activities. The energy carrier split is time dependent but not age-cohort dependent.
                RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:]     = np.einsum('Vrnt,cBVrt->cBVnrt',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:], RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,all_loc,:,:]) 
                 
                
            # Sector: Nonresidential buildings, by region
            if 'nrb' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, nonresidential buildings.')
                # 1) Determine total stock and apply stock-driven model
                SF_Array                    = np.zeros((Nc,Nc,NN,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                
                #Get historic stock at end of 2015 by age-cohort, and covert unit to nonres Buildings: million m2.
                TotalStock_UsePhase_Hist_cNr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_nonresbuildings'].Values[0,:,:,:]
                
                # Determine total future stock, product level. Units: nonres Buildings: million m2.
                TotalStockCurves_UsePhase_N_pC_test = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings'].Values[Sector_nrb_loc,:,:,mS]
                    
                # 2) Include (or not) the RE strategies for the use phase:
                # Include_REStrategy_MoreIntenseUse:
                if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True': 
                    # Calculate counter-factual scenario: X% decrease of stock levels by 2050 compared to scenario reference. X coded in parameter ..._MIUPotential
                    if SName != 'LED':
                        RemainingFraction = 1-RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_nonresbuildings_MIUPotential'].Values[Sector_nrb_loc,0,mS] / 100
                        #clamped_spline = make_interp_spline(np.arange(0,Nt,1), MIURamp, bc_type=([(2, 0)], [(1, 0)]))
                        clamped_spline = make_interp_spline([0,2,Nt-5,Nt], [1,1,RemainingFraction,RemainingFraction], bc_type=([(2, 0)], [(1, 0)]))
                        MIURamp_Spline = clamped_spline(np.arange(0,Nt,1))
                        MIURamp_Spline[MIURamp_Spline>1]=1
                        MIURamp_Spline[MIURamp_Spline<RemainingFraction]=RemainingFraction
                        
                        TotalStockCurves_UsePhase_N_pC_test    = TotalStockCurves_UsePhase_N_pC_test * np.einsum('t,r->rt',MIURamp_Spline,np.ones((Nr)))
                # Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.
                TotalStockCurves_UsePhase_N_pC_LED_ref = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings'].Values[Sector_nrb_loc,:,:,LEDindex]
                TotalStockCurves_UsePhase_N_pC         = np.maximum(TotalStockCurves_UsePhase_N_pC_test,TotalStockCurves_UsePhase_N_pC_LED_ref)
                TotalStockCurves_UsePhase_N            = np.einsum('rt,tr->tr',TotalStockCurves_UsePhase_N_pC,RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]) 
                RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'].Values[Sector_nrb_loc,:,:,mS] = TotalStockCurves_UsePhase_N_pC.copy()
            
                # Include_REStrategy_LifeTimeExtension: Product lifetime extension.
                Par_RECC_ProductLifetime_N = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_NonResbuildings'].Values.copy()
                # Second, change lifetime of future age-cohorts according to lifetime extension parameter
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                    # option: all age-cohorts, used from ODYM-RECC v2.3 onwards, which leads to:
                    if ScriptConfig['Include_Renovation_nrb'] == 'True' and ScriptConfig['No_EE_Improvements'] == 'False': 
                        # Increase lifetime of all nonres. buildings instantaneously:
                        Par_RECC_ProductLifetime_N = np.einsum('Nrc,Nrc->Nrc',np.einsum('Nr,c->Nrc',1 + RECC_System.ParameterDict['6_PR_LifeTimeExtension_nonresbuildings'].Values[:,:],np.ones(Nc)),Par_RECC_ProductLifetime_N)
                    else:
                        # gradual increase of lifetime by age-cohort, including historic age-cohorts, starting from 0:
                        for N in range(0, NN):
                            for r in range(0, Nr):
                                LTE_Pot = RECC_System.ParameterDict['6_PR_LifeTimeExtension_nonresbuildings'].Values[N,r]
                                LTE_Rampupcurve = np.zeros(Nc)
                                try:
                                    LTE_Rampupcurve[0:SwitchTime] = np.arange(0,LTE_Pot,LTE_Pot/SwitchTime)
                                except:
                                    None # LTE_Pot = 0, no LTE
                                LTE_Rampupcurve[SwitchTime::] = LTE_Pot
                                Par_RECC_ProductLifetime_N[N,r,:] = np.einsum('c,c->c',1 + LTE_Rampupcurve,Par_RECC_ProductLifetime_N[N,r,:])
      
                # 3) Dynamic stock model, with lifetime depending on age-cohort.
                # Build pdf array from lifetime distribution: Probability of survival.
                for N in tqdm(range(0, NN), unit=' nonres. building types'):
                    for r in range(0, Nr):
                        LifeTimes = Par_RECC_ProductLifetime_N[N, r, :]
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
                        SF_Array[:, :, N, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                        np.fill_diagonal(SF_Array[:, :, N, r],1) # no outflows from current year, 
                        # this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
        
                # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
                for r in range(0,Nr):   
                    FutureStock                 = np.zeros((Nc))
                    FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_N[1::, r].copy()# Future total stock
                    InitialStock                = TotalStock_UsePhase_Hist_cNr[:,:,r].copy()
                    InitialStocksum             = InitialStock.sum()
                    StockMatch_2015[Sector_nrb_loc,r] = TotalStockCurves_UsePhase_N[0, r]/InitialStocksum
                    SFArrayCombined             = SF_Array[:,:,:,r]
                    TypeSplit                   = np.zeros((Nc,NN))
                    TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[:,r,1::,mR,mS].transpose() # indices: Nc
                    
                    RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                    Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                    
                    # Below, the results are added with += because the different commodity groups (buildings, vehicles) are calculated separately
                    # to introduce the type split for each, but using the product resolution of the full model with all sectors.
                    Stock_Detail_UsePhase_N[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                    Stock_Detail_UsePhase_N[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcNr
                    Stock_2020_decline_N[1::,:,r]        += Var_S[SwitchTime::,0:Ind_2020+1,:].copy().sum(axis=1) # tNr
                    Outflow_Detail_UsePhase_N[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcNr
                    Inflow_Detail_UsePhase_N[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tNr
                    # Check for negative inflows:
                    if IFlags.sum() != 0:
                        NegInflowFlags[Sector_nrb_loc,mS,mR] = 1 # flag this scenario
        
                # Here so far: Units: Buildings: million m2. for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_nrb_loc,mS,mR] = TotalStockCurves_UsePhase_N.sum(axis =1).copy()
                StockCurves_Prod[:,Sector_nrb_rge,mS,mR] = np.einsum('tcNr->tN',Stock_Detail_UsePhase_N).copy()
                pCStocksCurves[:,Sector_nrb_loc,:,mS,mR] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'].Values[Sector_nrb_loc,:,:,mS].transpose().copy()
                Population[:,:,mS,mR]                    = RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]
                Inflow_Prod[:,Sector_nrb_rge,mS,mR]      = np.einsum('tNr->tN',Inflow_Detail_UsePhase_N).copy()
                Inflow_Prod_r[:,:,Sector_nrb_rge,mS,mR]  = np.einsum('tNr->trN',Inflow_Detail_UsePhase_N).copy()
                Outflow_Prod[:,Sector_nrb_rge,mS,mR]     = np.einsum('tcNr->tN',Outflow_Detail_UsePhase_N).copy()
                Outflow_Prod_r[:,:,Sector_nrb_rge,mS,mR] = np.einsum('tcpr->trp',Outflow_Detail_UsePhase_N).copy()
    
                # Include renovation of nrb:
                RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_t'].Values[:,:,:,:,:,mS,mR] = np.einsum('cmNr,t->mNrct',RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[:,:,:,:,mS,mR],np.ones(Nt)) # mNrctSR
                if ScriptConfig['Include_Renovation_nrb'] == 'True' and ScriptConfig['No_EE_Improvements'] == 'False': 
                    RenPot_E   = np.einsum('rcN,rN->rcN',RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_NonResBuildings'].Values[:,0:SwitchTime,:],RECC_System.ParameterDict['3_SHA_EnergySavingsPot_Renovation_NonResBuildings'].Values[:,mS,:]) # Unit: 1
                    RenPot_E_t = np.einsum('tr,rcN->trcN',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RenPot_E) # Unit: 1
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cNVnr,trcN->cNVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:SwitchTime,:,:,:,:,mS,mR],(np.ones((Nt,Nr,Nc-Nt+1,NN))-RenPot_E_t)) # cNVnrt
                    # Add renovation material intensity to building material intensity:
                    RenPot_M_t = np.einsum('tr,rcN->trcN',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_NonResBuildings'].Values[:,0:SwitchTime,:]) # Unit: 1, Defined as share of stock crN that is renovated by year t
                    MC_Ren = RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_Renovation_Absolute'].Values
                    RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_t'].Values[:,:,:,0:SwitchTime,:,mS,mR] += np.einsum('cmNr,trcN->mNrct',MC_Ren[0:SwitchTime,:,:,:],RenPot_M_t)
                else:
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cNVnr,trcN->cNVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:SwitchTime,:,:,:,:,mS,mR],(np.ones((Nt,Nr,Nc-Nt+1,NN)))) # cNVnrt
                # Add values for future age-cohorts, convert from useful to final energy, expand from 'all' to specific energy carriers
                RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[SwitchTime-1::,:,:,:,:,:]   = np.einsum('Vrnt,Vrnt,cNVr,t->cNVnrt',ParameterDict['4_TC_NonResidentialEnergyEfficiency'].Values[:,mR,:,:,:,mS],ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings_uf'].Values[:,mR,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[SwitchTime-1::,:,:,all_loc,:,mS,mR],np.ones(Nt))
                # Split energy into different carriers for historic age-cohorts. Here, a shift of heating systems is assumed irrespective of the renovation activities. The energy carrier split is time dependent but not age-cohort dependent.
                RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[0:SwitchTime,:,:,:,:,:]     = np.einsum('Vrnt,cNVrt->cNVnrt',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings'].Values[:,mR,:,:,:], RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[0:SwitchTime,:,:,all_loc,:,:]) 
                
            # Sector: Nonresidential buildings, global total
            if 'nrbg' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, nonresidential buildings.')
                SF_Array = np.zeros((Nc,Nc,NN,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.            
                s_nrbg   = RECC_System.ParameterDict['2_S_RECC_FinalProducts_nonresbuildings_g'].Values[:,:]  ### dimensions: 'Nt'
    
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[:,:,SwitchTime -1::] = np.einsum('Not,Not->Not',RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[:,:,SwitchTime -1::],1 + np.einsum('No,ot->Not',RECC_System.ParameterDict['6_PR_LifeTimeExtension_nonresbuildings_g'].Values[:,:],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS]))
    
                for N in tqdm(range(0, NN), unit='Mm²'):
                    for o in range(0, No):
                        # First, replicate lifetimes for all age-cohorts
                        LifeTimes_nrbg = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[N,o,:] # Dimensions: 'Noc'
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes_nrbg, 
                              'StdDev': 0.3 * LifeTimes_nrbg}
                # Compute evolution of nrbg in-use stock and related flows with stock-driven model
       
                        RECC_dsm_nrbg            = dsm.DynamicStockModel(time_dsm, s = s_nrbg[N,:].copy(), lt = lt)     
                        SF_Array[:, :, N, o]     = RECC_dsm_nrbg.compute_sf().copy()                             
                        np.fill_diagonal(SF_Array[:, :, N, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                        RECC_dsm_nrbg.sf         = SF_Array[:, :, N, o].copy()
                        
                        nrbg_sc, nrbg_oc, nrbg_i = RECC_dsm_nrbg.compute_stock_driven_model(NegativeInflowCorrect = False) # Unit: Mm²
                        
                        Stock_Detail_UsePhase_Ng[:,:,N,o]        = nrbg_sc[SwitchTime-1::,:].copy() # index structure: tcNo. Unit: million m².
                        Outflow_Detail_UsePhase_Ng[1::,:,N,o]    = nrbg_oc[SwitchTime::,:].copy()   # index structure: tcNo. Unit: million m².
                        Inflow_Detail_UsePhase_Ng[1::,N,o]       = nrbg_i[SwitchTime::].copy()      # index structure: tNo.  Unit: million m².     
                        
                # Here so far: Units: Buildings: million m2. for stocks, Mm² for flows.
                StockCurves_Totl[:,Sector_nrbg_loc,mS,mR] = np.einsum('tcNo->t', Stock_Detail_UsePhase_Ng).copy()
                StockCurves_Prod[:,Sector_nrbg_rge,mS,mR] = np.einsum('tcNo->tN',Stock_Detail_UsePhase_Ng).copy()
                pCStocksCurves[:,Sector_nrbg_loc,:,mS,mR] = 0  # pC stocks are not considered for this sector/dataset
                Inflow_Prod[:,Sector_nrbg_rge,mS,mR]      = np.einsum('tNo->tN',Inflow_Detail_UsePhase_Ng).copy()
                Outflow_Prod[:,Sector_nrbg_rge,mS,mR]     = np.einsum('tcNo->tN',Outflow_Detail_UsePhase_Ng).copy()                    
                
            # Sector: Industry, 11 region and global coverage, will be calculated separately and waste will be added to wast mgt. inflow for 1st region.
            if 'ind' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, industry.')
                # 1) Determine total stock and apply stock-driven model
                
                SF_Array                    = np.zeros((Nc,Nc,NI,Nl)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                i_Inflow_ind = RECC_System.ParameterDict['1_F_RECC_FinalProducts_industry'].Values[:,:,:,:,:]                     ### dimensions: rSRpt of TotalFutureInflow_UsePhase_ind
    
                # set lifetime parameter
                # First, Simply replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_ind = np.einsum('cl,I->Ilc',np.ones((Nc,Nl)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_industry'].Values)
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    Par_RECC_ProductLifetime_ind[:,:,SwitchTime -1::] = np.einsum('Ilt,Ilt->Ilt',Par_RECC_ProductLifetime_ind[:,:,SwitchTime -1::],1 + np.einsum('Il,t->Ilt',RECC_System.ParameterDict['6_PR_LifeTimeExtension_industry'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,0,:,mS]))
                    
                # Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                RECC_dsm_ind_s_c        = np.zeros((Nl,NS,NR,NI,Nc,Nc))
                RECC_dsm_ind_s_c_o_c    = np.zeros((Nl,NS,NR,NI,Nc,Nc))
                RECC_dsm_ind_o          = np.zeros((Nl,NS,NR,NI,Nc))
                Inflow                  = np.zeros((Nt,Nl,NI))
                
                TotalStockCurves_UsePhase_I = np.zeros((Nt,NI,Nl))
    
                for I in tqdm(range(0, NI), unit=' EGT types'):
                    for l in range(0, Nl):
                                LifeTimes = Par_RECC_ProductLifetime_ind[I, l, :]
                            
                                lt = {'Type'  : 'Normal',
                                      'Mean'  : LifeTimes,
                                      'StdDev': 0.3 * LifeTimes}
                # Compute inflow-driven model
                                RECC_dsm_ind                         = dsm.DynamicStockModel(time_dsm , i = i_Inflow_ind[l,mS,mR,I,:].copy()  , lt = lt)
                                SF_Array[:, :, I, l]                 = dsm.DynamicStockModel(time_dsm , i = Inflow  , lt = lt).compute_sf().copy()  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                                np.fill_diagonal(SF_Array[:, :, I, l],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                                # Those parts of the stock remain in use instead.
                                
                                RECC_dsm_ind.sf                      = SF_Array[:, :, I, l].copy()
                                RECC_dsm_ind_s_c[l,mS,mR,I,:,:]      = RECC_dsm_ind.compute_s_c_inflow_driven()
                                RECC_dsm_ind_s_c_o_c[l,mS,mR,I,:,:]  = RECC_dsm_ind.compute_o_c_from_s_c()
                                RECC_dsm_ind_o[l,mS,mR,I,:]          = RECC_dsm_ind.compute_outflow_total()
     
                                Stock_Detail_UsePhase_I[:,:,I,l]     = RECC_dsm_ind_s_c[l,mS,mR,I,SwitchTime-1::,:]
                                Outflow_Detail_UsePhase_I[:,:,I,l]   = RECC_dsm_ind_s_c_o_c[l,mS,mR,I,SwitchTime-1::,:]
                                Outflow_Detail_UsePhase_I[0,:,I,l]   = 0 # no flow calculation in first year
                                Inflow_Detail_UsePhase_I[:,I,l]      = i_Inflow_ind[l,mS,mR,I,SwitchTime-1::] # index structure: tIl
                                Inflow_Detail_UsePhase_I[0,I,l]      = 0 # no flow calculation in first year
            
            
                TotalStockCurves_UsePhase_I[:,:,:] = Stock_Detail_UsePhase_I[:,:,:,:].sum(axis=1) 
                                
            # Here so far: Units: Electricity: GW. for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_ind_loc,mS,mR] = TotalStockCurves_UsePhase_I[:,:,:].sum(axis=1).sum(axis=1).copy()
                StockCurves_Prod[:,Sector_ind_rge,mS,mR] = TotalStockCurves_UsePhase_I[:,:,:].sum(axis=2).copy()
                Inflow_Prod[:,Sector_ind_rge,mS,mR]      = np.einsum('tIl->tI',Inflow_Detail_UsePhase_I).copy()
                Outflow_Prod[:,Sector_ind_rge,mS,mR]     = np.einsum('tcIl->tI',Outflow_Detail_UsePhase_I).copy()                      
               
                
            # Sector: Appliances, global coverage, will be calculated separately and waste will be added to wast mgt. inflow for 1st region.
            if 'app' in SectorList:            
                Mylog.info('Calculate inflows and outflows for use phase, appliances.')
                
                SF_Array     = np.zeros((Nc,Nc,Na,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                i_Inflow_app = RECC_System.ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[:,:,:,:,:] # dimensions:ocSRa
            
                # set lifetime parameter
                # First, Simply replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_app = np.einsum('co,a->aoc',np.ones((Nc,No)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_appliances'].Values)
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    Par_RECC_ProductLifetime_app[:,:,SwitchTime -1::] = np.einsum('aot,aot->aot',Par_RECC_ProductLifetime_app[:,:,SwitchTime -1::],1 + np.einsum('ao,ot->aot',RECC_System.ParameterDict['6_PR_LifeTimeExtension_appliances'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS]))
                    
                # Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                RECC_dsm_app_s_c        = np.zeros((No,NS,NR,Na,Nc,Nc))
                RECC_dsm_app_s_c_o_c    = np.zeros((No,NS,NR,Na,Nc,Nc))
                RECC_dsm_app_o          = np.zeros((No,NS,NR,Na,Nc))
                            
                TotalStockCurves_UsePhase_a = np.zeros((Nt,Na,Nl))
                
                for a in tqdm(range(0, Na), unit=' App types'):
                    for o in range(0, No):
                        LifeTimes = Par_RECC_ProductLifetime_app[a, o, :]
                    
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
      
                # Compute inflow-driven model     
                        RECC_dsm_app                         = dsm.DynamicStockModel(time_dsm , i = i_Inflow_app[o,:,mS,mR,a].copy()  , lt = lt)  
                        SF_Array[:, :, a, o]                 = dsm.DynamicStockModel(time_dsm , i = i_Inflow_app[o,:,mS,mR,a].copy()  , lt = lt).compute_sf().copy() 
                        np.fill_diagonal(SF_Array[:, :, a, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                                
                        RECC_dsm_app.sf                      = SF_Array[:, :, a, o].copy()
                        RECC_dsm_app_s_c[o,mS,mR,a,:,:]      = RECC_dsm_app.compute_s_c_inflow_driven()
                        RECC_dsm_app_s_c_o_c[o,mS,mR,a,:,:]  = RECC_dsm_app.compute_o_c_from_s_c()
                        RECC_dsm_app_o[o,mS,mR,a,:]          = RECC_dsm_app.compute_outflow_total()
    
                        Stock_Detail_UsePhase_a[:,:,a,o]     = RECC_dsm_app_s_c[o,mS,mR,a,SwitchTime-1::,:]
                        Outflow_Detail_UsePhase_a[:,:,a,o]   = RECC_dsm_app_s_c_o_c[o,mS,mR,a,SwitchTime-1::,:]
                        Outflow_Detail_UsePhase_a[0,:,a,o]   = 0 # no flow calculation in first year
                        Inflow_Detail_UsePhase_a[:,a,o]      = i_Inflow_app[o,SwitchTime-1::,mS,mR,a] # index structure: tIl
                        Inflow_Detail_UsePhase_a[0,a,o]      = 0 # no flow calculation in first year
    
                TotalStockCurves_UsePhase_a[:,:,:]           = Stock_Detail_UsePhase_a[:,:,:,:].sum(axis=1) 
                                
            # Here so far: Units: 1 (# items). for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_app_loc,mS,mR]     = TotalStockCurves_UsePhase_a[:,:,:].sum(axis=1).sum(axis=1).copy()
                StockCurves_Prod[:,Sector_app_rge,mS,mR]     = TotalStockCurves_UsePhase_a[:,:,:].sum(axis=2).copy()
                Inflow_Prod[:,Sector_app_rge,mS,mR]          = np.einsum('tIl->tI',Inflow_Detail_UsePhase_a).copy()
                Outflow_Prod[:,Sector_app_rge,mS,mR]         = np.einsum('tcIl->tI',Outflow_Detail_UsePhase_a).copy()   
    
            # Archive 2015 pC stock values for future curves:
            pC_FutureStock_2015             = np.zeros((NS,NG,Nr))
            # b) from scenario curves:
            for mSS in range(0,NS):
                if 'pav' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_pav_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'].Values[mSS,0,Sector_pav_loc,:]
                if 'reb' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_reb_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mSS,0,Sector_reb_loc,:]
                if 'nrb' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_nrb_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'].Values[Sector_nrb_loc,:,0,mSS]
            OutputDict['pC_FutureStock_2015']   = pC_FutureStock_2015.copy()
    
            
            # ABOVE: each sector separate, individual regional resolution. BELOW: all sectors together, global total.
            
            # Prepare parameters:        
            # include light-weighting in future MC parameter, cmgr
            Par_RECC_MC_Nr = np.zeros((Nc,Nm,Ng,Nr,NS,NR,Nt))  # Unit: vehicles: kg/item, buildings: kg/m².
            if 'pav' in SectorList:
                Par_RECC_MC_Nr[:,:,Sector_pav_rge,:,mS,mR,:]      = np.einsum('cmpr,t->pcmrt',RECC_System.ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[:,:,:,:,mS],np.ones(Nt))
            if 'reb' in SectorList:
                Par_RECC_MC_Nr[:,:,Sector_reb_rge,:,mS,mR,:]      = np.einsum('mBrct->Bcmrt',RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS,mR])
            if 'nrb' in SectorList: 
                Par_RECC_MC_Nr[:,:,Sector_nrb_rge,:,mS,mR,:]      = np.einsum('mNrct->Ncmrt',RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_t'].Values[:,:,:,:,:,mS,mR])
            Par_RECC_MC_Nl = np.zeros((Nc,Nm,NL,Nl,NS))          # for electricity generation technologies in kt/GW
            Par_RECC_MC_Nl[:,:,Sector_ind_rge_reg,:,mS]        = np.einsum('lc,Im->Icml',np.ones((Nl,Nc)), RECC_System.ParameterDict['3_MC_RECC_industry'].Values[:,:])       #3_MC_RECC_industry has dimensions Im
            Par_RECC_MC_No = np.zeros((Nc,Nm,NO,No,NS))          # for appliances in g/unit, nonres. buildings in kg/m²
            Par_RECC_MC_No[:,:,Sector_app_rge_reg,:,mS]        = np.einsum('c,oOm->Ocmo',np.ones((Nc)), RECC_System.ParameterDict['3_MC_RECC_appliances'].Values[:,:,:])      #3_MC_RECC_appliances has dimensions oam
            if 'nrbg' in SectorList:
                Par_RECC_MC_No[:,:,Sector_nrbg_rge_reg,:,mS]       = np.einsum('c,o,mN->Ncmo',np.ones((Nc)), np.ones((No)), RECC_System.ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[:,:])  #3_MC_RECC_Nonresbuildings_g has dimensions mN
            # Units: Vehicles: kg/unit, Buildings: kg/m2  
            
            # historic element composition of materials:
            Par_Element_Composition_of_Materials_m   = np.zeros((Nc,Nm,Ne)) # Unit: 1. Aspects: cme, produced in age-cohort c. Applies to new manufactured goods.
            Par_Element_Composition_of_Materials_m[0:Nc-Nt+1,:,:] = np.einsum('c,me->cme',np.ones(Nc-Nt+1),RECC_System.ParameterDict['3_MC_Elements_Materials_ExistingStock'].Values)
            # For future age-cohorts, the total is known but the element breakdown of this parameter will be updated year by year in the loop below.
            Par_Element_Composition_of_Materials_m[:,:,0] = 1 # element 0 is 'all', for which the mass share is always 100%.
            
            # future element composition of materials inflow use phase (mix new and reused products)
            Par_Element_Composition_of_Materials_c   = np.zeros((Nt,Nm,Ne)) # cme, produced in age-cohort c. Applies to new manufactured goods.
            
            # Element composition of material in the use phase
            Par_Element_Composition_of_Materials_u   = Par_Element_Composition_of_Materials_m.copy() # cme
            
            # Manufacturing yield and other improvements
            # Reduce cement content by up to percentage indicate in 3_SHA_CementContentReduction parameter:
            if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True':
                Par_RECC_MC_Nr[115::,Cement_loc,:,:,mS,mR,:] = Par_RECC_MC_Nr[115::,Cement_loc,:,:,mS,mR,:] * (1 - RECC_System.ParameterDict['3_SHA_CementContentReduction'].Values[Cement_loc] * np.einsum('oc,gt->cgot',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS],np.ones((Ng,Nt)))).copy()
    
            Par_FabYieldLoss = np.einsum('mwggto->mwgto',RECC_System.ParameterDict['4_PY_Manufacturing'].Values) # take diagonal of product = manufacturing process
            
            # Consider Fabrication yield improvement and reduction in cement content of concrete and plaster
            if ScriptConfig['Include_REStrategy_FabYieldImprovement'] == 'True':
                Par_FabYieldImprovement = np.einsum('w,tmgo->mwgto',np.ones((Nw)),np.einsum('ot,mgo->tmgo',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS],RECC_System.ParameterDict['6_PR_FabricationYieldImprovement'].Values[:,:,:,mS]))
            else:
                Par_FabYieldImprovement = 0              
                
            Par_FabYieldLoss_Raster = Par_FabYieldLoss > 0    
            Par_FabYieldLoss        = Par_FabYieldLoss - Par_FabYieldLoss_Raster * Par_FabYieldImprovement #mwgto
            Par_FabYieldLoss_total  = np.einsum('mwgto->mgto',Par_FabYieldLoss)
            Divisor                 = 1-Par_FabYieldLoss_total
            Par_FabYield_total_inv  = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # mgto
            # Determine total element composition of products (c: age-cohort), needs to be updated for future age-cohorts, is done below after material cycle computation.
            Par_3_MC_Stock_ByElement_Nr = np.einsum('cmgrt,cme->tcrgme',Par_RECC_MC_Nr[:,:,:,:,mS,mR,:],Par_Element_Composition_of_Materials_m) # Unit: vehicles: kg/item, buildings: kg/m².
            Par_3_MC_Stock_ByElement_Nl = np.einsum('cmLl,cme->clLme',Par_RECC_MC_Nl[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: ind: kt/GW
            Par_3_MC_Stock_ByElement_No = np.einsum('cmOo,cme->coOme',Par_RECC_MC_No[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: app: g/unit, nrbg: kg/m²
            # Consider EoL recovery rate improvement:
            if ScriptConfig['Include_REStrategy_EoL_RR_Improvement'] == 'True':
                Par_RECC_EoL_RR =  np.einsum('t,grmw->trmgw',np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,:,:,:,0] *0.01) \
                + np.einsum('tr,grmw->trmgw',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_EoL_RR_Improvement'].Values[:,:,:,:,0]*0.01)
                Par_RECC_WoodWaste_Cascading = np.einsum('t,wmWr->twmWr',np.ones((Nt)),RECC_System.ParameterDict['4_PY_WoodCascading'].Values * 0.01) \
                + np.einsum('tr,wmWr->twmWr',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_WoodCascading_Improvement'].Values*0.01)
            else:    
                Par_RECC_EoL_RR =  np.einsum('t,grmw->trmgw',np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,:,:,:,0] *0.01)
                Par_RECC_WoodWaste_Cascading = np.einsum('t,wmWr->twmWr',np.ones((Nt)),RECC_System.ParameterDict['4_PY_WoodCascading'].Values * 0.01)
            
            # For regional dimension 11 and 1
            Par_RECC_EoL_RR_Nl = np.einsum('l,t,Lmw->tlmLw',np.ones((Nl)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_11reg_rge,0,:,:,0] *0.01)
            Par_RECC_EoL_RR_No = np.einsum('o,t,Omw->tomOw',np.ones((No)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_1reg_rge,0,:,:,0] *0.01)
            
            # Calculate reuse factor
            # For vehicles, scenarios already included in target table output!
            ReUseFactor_tmprS = np.einsum('mprtS->tmprS',RECC_System.ParameterDict['6_PR_ReUse_Veh'].Values/100)
            # For Buildings, scenarios obtained from RES scaleup curves
            ReUseFactor_tmBrS = np.einsum('tmBr,S->tmBrS',np.einsum('tr,mBr->tmBr',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_ReUse_Bld'].Values),np.ones((NS)))
            ReUseFactor_tmNrS = np.einsum('tmNr,S->tmNrS',np.einsum('tr,mNr->tmNr',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_ReUse_nonresBld'].Values),np.ones((NS)))
            
            Mylog.info('Translate total flows into individual materials and elements, for 2015 and historic age-cohorts.')
            if 'pav' in SectorList:
                # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
                # Stock elemental composition, will be updated for future years:
                RECC_System.StockDict['S_7'].Values[:,:,:,Sector_pav_rge,:,:] = \
                np.einsum('tcrpme,tcpr->tcrpme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_pav_rge,:,:],Stock_Detail_UsePhase_p)/1000   # Indices='t,c,r,p,m,e'
                # Outflow, 'all' elements only:
                RECC_System.FlowDict['F_7_8'].Values[:,:,:,Sector_pav_rge,:,0] = \
                np.einsum('pcmrt,tcpr->ptcrm',Par_RECC_MC_Nr[:,:,Sector_pav_rge,:,mS,mR,:],Outflow_Detail_UsePhase_p)/1000 # all elements, Indices='t,c,r,p,m'
                # Inflow as mass balance, to account for renovation material inflows to other age-cohorts than the current one (t=c).
                RECC_System.FlowDict['F_6_7'].Values[1::,:,Sector_pav_rge,:,0]   = \
                np.einsum('ptcrm->ptrm',np.diff(RECC_System.StockDict['S_7'].Values[:,:,:,Sector_pav_rge,:,0],1,axis=1)) + np.einsum('ptcrm->ptrm',RECC_System.FlowDict['F_7_8'].Values[1::,:,:,Sector_pav_rge,:,0])
                # inflow of materials in new products, for checking:
                for mmt in range(0,Nt):
                    F_6_7_new[mmt,:,Sector_pav_rge,:,0] = np.einsum('pr,pmr->prm',Inflow_Detail_UsePhase_p[mmt,:,:],Par_RECC_MC_Nr[SwitchTime+mmt-1,:,Sector_pav_rge,:,mS,mR,mmt])/1000
                # Check_pav = (RECC_System.FlowDict['F_6_7'].Values[1::,0,Sector_pav_rge,:,0] - F_6_7_new[1::,0,Sector_pav_rge,:,0]).sum() # must be 0.
                    
            if 'reb' in SectorList:        
                # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
                # Stock elemental composition, historic for each element and for future years: 'all' elements only
                RECC_System.StockDict['S_7'].Values[:,:,:,Sector_reb_rge,:,:] = \
                np.einsum('tcrBme,tcBr->tcrBme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,:],Stock_Detail_UsePhase_B)/1000   # Indices='t,c,r,B,m'
                # Outflow, 'all' elements only:
                RECC_System.FlowDict['F_7_8'].Values[:,:,:,Sector_reb_rge,:,0] = \
                np.einsum('Btcrm,tcBr->Btcrm',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,0],Outflow_Detail_UsePhase_B)/1000 # all elements, Indices='t,c,r,B,m'
                # Inflow of renovation material as stock multiplied with change in material composition:
                F_6_7_ren[1::,:,:,Sector_reb_rge,:,0]  = np.einsum('tcBr,Btcrm->Btcrm',Stock_Detail_UsePhase_B[1::,:,:,:],np.diff(Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,0],1,axis=1))/1000
                # inflow of materials in new products
                for mmt in range(0,Nt):
                    F_6_7_new[mmt,:,Sector_reb_rge,:,0] = np.einsum('Br,Brm->Brm',Inflow_Detail_UsePhase_B[mmt,:,:],Par_3_MC_Stock_ByElement_Nr[mmt,SwitchTime+mmt-1,:,Sector_reb_rge,:,0])/1000
                # Check_reb = (RECC_System.FlowDict['F_6_7'].Values[1::,0,Sector_reb_rge,:,0] - F_6_7_new[1::,0,Sector_reb_rge,:,0] - F_6_7_ren[1::,:,0,Sector_reb_rge,:,0].sum(axis=2)) # must be 0.
                RECC_System.FlowDict['F_6_7'].Values[:,:,Sector_reb_rge,:,0]   = np.einsum('Btrm->Btrm',F_6_7_new[:,:,Sector_reb_rge,:,0]) + np.einsum('Btcrm->Btrm',F_6_7_ren[:,:,:,Sector_reb_rge,:,0])
                
            if 'nrb' in SectorList:
                # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
                # Stock elemental composition, historic for each element and for future years: 'all' elements only
                RECC_System.StockDict['S_7'].Values[:,:,:,Sector_nrb_rge,:,:] = \
                np.einsum('tcrNme,tcNr->tcrNme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_nrb_rge,:,:],Stock_Detail_UsePhase_N)/1000   # Indices='t,c,r,N,m,e'
                # Outflow, 'all' elements only:
                RECC_System.FlowDict['F_7_8'].Values[:,:,:,Sector_nrb_rge,:,0] = \
                np.einsum('Ntcrm,tcNr->Ntcrm',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_nrb_rge,:,0],Outflow_Detail_UsePhase_N)/1000 # all elements, Indices='t,c,r,N,m'
                # Inflow of renovation material as stock multiplied with change in material composition:
                F_6_7_ren[1::,:,:,Sector_nrb_rge,:,0]  = np.einsum('tcNr,Ntcrm->Ntcrm',Stock_Detail_UsePhase_N[1::,:,:,:],np.diff(Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_nrb_rge,:,0],1,axis=1))/1000
                # inflow of materials in new products
                for mmt in range(0,Nt):
                    F_6_7_new[mmt,:,Sector_nrb_rge,:,0] = np.einsum('Nr,Nrm->Nrm',Inflow_Detail_UsePhase_N[mmt,:,:],Par_3_MC_Stock_ByElement_Nr[mmt,SwitchTime+mmt-1,:,Sector_nrb_rge,:,0])/1000
                # Check_nrb = (RECC_System.FlowDict['F_6_7'].Values[1::,0,Sector_nrb_rge,:,0] - F_6_7_new[1::,0,Sector_nrb_rge,:,0] - F_6_7_ren[1::,:,0,Sector_nrb_rge,:,0].sum(axis=2)) # must be 0.
                RECC_System.FlowDict['F_6_7'].Values[:,:,Sector_nrb_rge,:,0]   = np.einsum('Ntrm->Ntrm',F_6_7_new[:,:,Sector_nrb_rge,:,0]) + np.einsum('Ntcrm->Ntrm',F_6_7_ren[:,:,:,Sector_nrb_rge,:,0])
                            
            # 1_Nl_No) Inflow, outflow and stock first year for Nl and No regional aggregation and Sector I and a
            RECC_System.FlowDict['F_6_7_Nl'].Values[0,:,Sector_ind_rge_reg,:,:]   = \
            np.einsum('Ilme,Il ->Ilme',Par_3_MC_Stock_ByElement_Nl[SwitchTime-1,:,Sector_ind_rge_reg,:,:],Inflow_Detail_UsePhase_I[0,:,:])/1000 # all elements, Indices='t,l,I,m,e'  
            RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_app_rge_reg,:,:]   = \
            np.einsum('aome,ao->aome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_app_rge_reg,:,:],Inflow_Detail_UsePhase_a[0,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'  
            if 'nrbg' in SectorList:
                RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_nrbg_rge_reg,:,:]   = \
                np.einsum('Nome,No->Nome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_nrbg_rge_reg,:,:],Inflow_Detail_UsePhase_Ng[0,:,:])/1000 # all elements, Indices='t,o,N,m,e'  
        
            RECC_System.FlowDict['F_7_8_Nl'].Values[0,0:SwitchTime,:,Sector_ind_rge_reg,:,:] = \
            np.einsum('clIme,cpl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:SwitchTime,:,Sector_ind_rge_reg,:,:],Outflow_Detail_UsePhase_I[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,l,I,m,e'
            RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_app_rge_reg,:,:] = \
            np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_app_rge_reg,:,:],Outflow_Detail_UsePhase_a[0,0:SwitchTime,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'
            if 'nrbg' in SectorList:
                RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_nrbg_rge_reg,:,:] = \
                np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_nrbg_rge_reg,:,:],Outflow_Detail_UsePhase_Ng[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,o,N,m,e'
        
            RECC_System.StockDict['S_7_Nl'].Values[0,0:SwitchTime,:,Sector_ind_rge_reg,:,:] = \
            np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:SwitchTime,:,Sector_ind_rge_reg,:,:],Stock_Detail_UsePhase_I[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,l,p,m,e'
            RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_app_rge_reg,:,:] = \
            np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_app_rge_reg,:,:],Stock_Detail_UsePhase_a[0,0:SwitchTime,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'
            if 'nrbg' in SectorList:
                RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_nrbg_rge_reg,:,:] = \
                np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_nrbg_rge_reg,:,:],Stock_Detail_UsePhase_Ng[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,o,N,m,e'
    
            # 2) Inflow, future years, all elements only
            RECC_System.FlowDict['F_6_7_Nl'].Values[1::,:,Sector_ind_rge_reg,:,0]   = \
            np.einsum('Itlm,tIl->Itlm',Par_3_MC_Stock_ByElement_Nl[SwitchTime::,:,Sector_ind_rge_reg,:,0],Inflow_Detail_UsePhase_I[1::,:,:])/1000 # all elements, Indices='t,l,I,m'  
            RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_app_rge_reg,:,0]   = \
            np.einsum('atom,tao->atom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_app_rge_reg,:,0],Inflow_Detail_UsePhase_a[1::,:,:])/1000000000000 # all elements, Indices='t,o,a,m'  
            if 'nrbg' in SectorList:
                RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_nrbg_rge_reg,:,0]   = \
                np.einsum('Ntom,tNo->Ntom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_nrbg_rge_reg,:,0],Inflow_Detail_UsePhase_Ng[1::,:,:])/1000 # all elements, Indices='t,o,N,m'              
            #Units so far: Mt/yr
            
            Mylog.info('Calculate material stocks and flows, material cycles, determine elemental composition.')
            # Units: Mt and Mt/yr.
            # This calculation is done year-by-year, and the elemental composition of the materials is in part determined by the scrap flow metal composition
            
            for t in tqdm(range(1,Nt), unit=' years'):  # 1: 2016
                CohortOffset = t +Nc -Nt # index of current age-cohort.   
                # First, before going down to the material layer, we consider obsolete stock formation and re-use.
                
                # 1) Convert use phase outflow to system variables.
                # Split flows into materials and chemical elements.
                # Calculate use phase outflow and obsolete stock formation
                # ObsStockFormation = ObsStockFormationFactor(t,g,r) * Outflow_Detail_UsePhase(t,c,g,r), currently not implemented. 
                
                if 'pav' in SectorList:
                    RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,:] = \
                    np.einsum('pcrme,cpr->pcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_pav_rge,:,:],Outflow_Detail_UsePhase_p[t,0:CohortOffset,:,:])/1000 # All elements.
                if 'reb' in SectorList:
                    RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,:],Outflow_Detail_UsePhase_B[t,0:CohortOffset,:,:])/1000 # All elements.
                if 'nrb' in SectorList:
                    RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,:] = \
                    np.einsum('Ncrme,cNr->Ncrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_nrb_rge,:,:],Outflow_Detail_UsePhase_N[t,0:CohortOffset,:,:])/1000 # All elements.
    
                # 1_Nl_No)
                RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:] = \
                np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:CohortOffset,:,Sector_ind_rge_reg,:,:],Outflow_Detail_UsePhase_I[t,0:CohortOffset,:,:])/1000 # All elements.
                RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:] = \
                np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_app_rge_reg,:,:],Outflow_Detail_UsePhase_a[t,0:CohortOffset,:,:])/1000000000000 # All elements.
                if 'nrbg' in SectorList:
                    RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:] = \
                    np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_nrbg_rge_reg,:,:],Outflow_Detail_UsePhase_Ng[t,0:CohortOffset,:,:])/1000 # All elements.
    
                # RECC_System.FlowDict['F_8_0'].Values = MatContent * ObsStockFormation. Currently 0, already defined.
                            
                # 2) Consider re-use of materials in product groups (via components), as ReUseFactor(m,g,r,R,t) * RECC_System.FlowDict['F_7_8'].Values(t,c,r,g,m,e)
                # Distribute material for re-use onto product groups
                if 'pav' in SectorList:
                    ReUsePotential_Materials_t_m_Veh = np.einsum('mpr,pcrm->m',ReUseFactor_tmprS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8'].Values[t,:,:,Sector_pav_rge,:,0]) # in Mt
                if 'reb' in SectorList:
                    ReUsePotential_Materials_t_m_Bld = np.einsum('mBr,Bcrm->m',ReUseFactor_tmBrS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8'].Values[t,:,:,Sector_reb_rge,:,0]) # in Mt
                if 'nrb' in SectorList:
                    ReUsePotential_Materials_t_m_NRB = np.einsum('mNr,Ncrm->m',ReUseFactor_tmNrS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8'].Values[t,:,:,Sector_nrb_rge,:,0]) # in Mt
                # in the future, re-use will be a region-to-region parameter depicting, e.g., the export of used vehicles from the EU to Africa.
                # check whether inflow is big enough for potential to be used, correct otherwise:
                for mmm in range(0,Nm):
                    # Vehicles
                    if 'pav' in SectorList:
                        if RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_pav_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_Veh[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_Veh[mmm] > 0:
                                ReUsePotential_Materials_t_m_Veh[mmm] = RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_pav_rge,mmm,0].sum()
                    # residential buildings
                    if 'reb' in SectorList:
                        if RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_reb_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_Bld[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_Bld[mmm] > 0:
                                ReUsePotential_Materials_t_m_Bld[mmm] = RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_reb_rge,mmm,0].sum()
                    # nonresidential buildings
                    if 'nrb' in SectorList:
                        if RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_nrb_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_NRB[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_NRB[mmm] > 0:
                                ReUsePotential_Materials_t_m_NRB[mmm] = RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_nrb_rge,mmm,0].sum()
                    
                # Vehicles
                if 'pav' in SectorList:
                    Divisor = np.einsum('m,crp->pcrm',np.einsum('pcrm->m',RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,0]),np.ones((CohortOffset,Nr,Np)))
                    MassShareVeh = np.divide(RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: pcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,:] = \
                    np.einsum('cme,pcrm->pcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,pcrm->pcrm',ReUsePotential_Materials_t_m_Veh,MassShareVeh))  # All elements.
                # residential Buildings
                if 'reb' in SectorList:
                    Divisor = np.einsum('m,crB->Bcrm',np.einsum('Bcrm->m',RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0]),np.ones((CohortOffset,Nr,NB)))
                    MassShareBld = np.divide(RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_Bld,MassShareBld))  # All elements.
                # nonresidential Buildings
                if 'nrb' in SectorList:
                    Divisor = np.einsum('m,crN->Ncrm',np.einsum('Ncrm->m',RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,0]),np.ones((CohortOffset,Nr,NN)))
                    MassShareNRB = np.divide(RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Ncrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,:] = \
                    np.einsum('cme,Ncrm->Ncrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Ncrm->Ncrm',ReUsePotential_Materials_t_m_NRB,MassShareNRB))  # All elements.
                
                # reused material mapped to final consumption region and good, proportional to final consumption breakdown into products and regions.
                # can be replaced by region-by-region reuse parameter.             
                Divisor = np.einsum('m,rg->rgm',np.einsum('rgm->m',RECC_System.FlowDict['F_6_7'].Values[t,:,:,:,0]),np.ones((Nr,Ng)))
                InvMass = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
                RECC_System.FlowDict['F_17_6'].Values[t,0:CohortOffset,:,:,:,:] = \
                np.einsum('cme,rgm->crgme',np.einsum('crgme->cme',RECC_System.FlowDict['F_8_17'].Values[t,0:CohortOffset,:,:,:,:]),\
                RECC_System.FlowDict['F_6_7'].Values[t,:,:,:,0]*InvMass)
                
                # 3) calculate inflow waste mgt as EoL products - obsolete stock formation - re-use
                RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:]           = np.einsum('crgme->rgme',RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,:,:,:]    - RECC_System.FlowDict['F_8_0'].Values[t,0:CohortOffset,:,:,:,:]    - RECC_System.FlowDict['F_8_17'].Values[t,0:CohortOffset,:,:,:,:])
                if len(Sector_11reg_rge) > 0:
                    RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:]    = np.einsum('clLme->lLme',RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_0_Nl'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_17_Nl'].Values[t,0:CohortOffset,:,:,:,:])
                if len(Sector_1reg_rge) > 0:
                    RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:]    = np.einsum('coOme->oOme',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_0_No'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,:,:,:])
            
                # 4) EoL products to postconsumer scrap: trwe. Add Waste mgt. losses.
                RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]            = np.einsum('rmgw,rgme->rwe',Par_RECC_EoL_RR[t,:,:,:,:],RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    
                if len(Sector_11reg_rge) > 0:                    
                    RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:]     = np.einsum('lmLw,lLme->lwe',Par_RECC_EoL_RR_Nl[t,:,:,:,:],RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:])    
                if len(Sector_1reg_rge) > 0:            
                    RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:]     = np.einsum('omOw,oOme->owe',Par_RECC_EoL_RR_No[t,:,:,:,:],RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:])    
    
                # 5) Add re-use flow to inflow and calculate manufacturing output as final consumption - re-use, in Mt/yr, all elements, trgme, element composition not yet known.
                RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]                   = np.einsum('rgm->gm',RECC_System.FlowDict['F_6_7'].Values[t,:,:,:,0])    - np.einsum('crgm->gm',RECC_System.FlowDict['F_17_6'].Values[t,:,:,:,:,0])     # global total
                if len(Sector_11reg_rge) > 0:            
                    RECC_System.FlowDict['F_5_6'].Values[t,0,Sector_11reg_rge,:,0]    = np.einsum('lLm->Lm',RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,0]) - np.einsum('clLm->Lm',RECC_System.FlowDict['F_17_6_Nl'].Values[t,:,:,:,:,0])  # global total
                if len(Sector_1reg_rge) > 0:            
                    RECC_System.FlowDict['F_5_6'].Values[t,0,Sector_1reg_rge,:,0]     = np.einsum('oOm->Om',RECC_System.FlowDict['F_6_7_No'].Values[t,:,:,:,0]) - np.einsum('coOm->Om',RECC_System.FlowDict['F_17_6_No'].Values[t,:,:,:,:,0])  # global total
                Manufacturing_Output[t,:,:,mS,mR]                                 = RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0].copy()
                
                # 6) Calculate total manufacturing input and primary production, all elements, element composition not yet known.
                # Add fabrication scrap diversion, new scrap and calculate remelting.
                #Manufacturing_Input_m_ref    = np.einsum('mg,gm->m', Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                #Manufacturing_Input_gm_ref   = np.einsum('mg,gm->gm',Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                # Same as above, but the variables below will be adjusted for diverted fab scrap and uses subsequently:
                Manufacturing_Input_m_adj    = np.einsum('mg,gm->m', Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                Manufacturing_Input_gm_adj   = np.einsum('mg,gm->gm',Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                # split manufacturing material input into different products g:
                Manufacturing_Input_Split_gm = np.einsum('gm,m->gm', Manufacturing_Input_gm_adj, np.divide(1, Manufacturing_Input_m_adj, out=np.zeros_like(Manufacturing_Input_m_adj), where=Manufacturing_Input_m_adj!=0))
                            
                # 7) Determine available fabrication scrap diversion and secondary material, total and by element.
                # Determine fabscrapdiversionpotential: (for steel scrap only acc. to parameter)
                Fabscrapdiversionpotential_twm                     = np.einsum('wm,ow->wm',np.einsum('o,mwo->wm',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,t,mS],RECC_System.ParameterDict['6_PR_FabricationScrapDiversion'].Values[:,:,:,mS]),RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,0]).copy()
                # break down fabscrapdiversionpotential to e:
                NewScrapElemShares                                 = msf.TableWithFlowsToShares(RECC_System.StockDict['S_10'].Values[t-1,t-1,0,:,1::],axis=1) # element composition of fab scrap
                Fabscrapdiversionpotential_twme                    = np.zeros((Nt,Nw,Nm,Ne))
                Fabscrapdiversionpotential_twme[t,:,:,0]           = Fabscrapdiversionpotential_twm # total mass (all chem. elements)
                Fabscrapdiversionpotential_twme[t,:,:,1::]         = np.einsum('wm,we->wme',Fabscrapdiversionpotential_twm,NewScrapElemShares) # other chemical elements
                Fabscrapdiversionpotential_tme                     = np.einsum('wme->me',Fabscrapdiversionpotential_twme[t,:,:,:])
                RECC_System.FlowDict['F_10_12'].Values[t,:,:,:]    = np.einsum('wme,o->ome',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No))
                RECC_System.FlowDict['F_10_9'].Values[t,:,:,:]     = np.einsum('owe->owe',np.einsum('we,o->owe',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:].sum(axis=0) + RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:].sum(axis=0) + RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:].sum(axis=0),np.ones(No)) + RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:].copy()) - np.einsum('wme,o->owe',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No)).copy()
                RECC_System.FlowDict['F_10_9'].Values[t,:,Woodwaste_loc,Carbon_loc] = 0
                RECC_System.FlowDict['F_10_9w'].Values[t,:,:,:]    = RECC_System.StockDict['S_10w'].Values[t-1,t-1,:,:,:].copy()
                RECC_System.FlowDict['F_10_9w'].Values[t,:,Woodwaste_loc,Carbon_loc] += RECC_System.FlowDict['F_9_10'].Values[t,:,Woodwaste_loc,Carbon_loc]            
                RECC_System.FlowDict['F_9_12'].Values[t,:,:,:]     = np.einsum('owe,wmePo->ome',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:],RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[:,:,:,:,0,:])
                RECC_System.FlowDict['F_9_12'].Values[t,:,:,0]     = np.einsum('ome->om',RECC_System.FlowDict['F_9_12'].Values[t,:,:,1::])
                # Calculate cascade input from EoL for results:
                SysVar_EoLCascEntry[t,:,mS,mR]                    += np.einsum('r,r->r',Par_RECC_WoodWaste_Cascading[t,Woodwaste_loc,Wood_loc,Woodwaste_loc,:],RECC_System.FlowDict['F_9_10'].Values[t,:,Woodwaste_loc,Carbon_loc])
    
                # 8) MARKET BALANCE for secondary materials:
    
                # a) Use diverted fab scrap first, if possible:
                DivFabScrap_to_Manuf = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if Fabscrapdiversionpotential_tme[mat,0] > 0:
                        if Fabscrapdiversionpotential_tme[mat,0] > Manufacturing_Input_m_adj[mat]:
                            DivFabScrap_to_Manuf[mat,:]            = (Fabscrapdiversionpotential_tme[mat,:] * Manufacturing_Input_m_adj[mat] / Fabscrapdiversionpotential_tme[mat,0]).copy()
                        else:
                            DivFabScrap_to_Manuf[mat,:]            = Fabscrapdiversionpotential_tme[mat,:].copy()
                Non_DivFabScrap                                    = Fabscrapdiversionpotential_tme - DivFabScrap_to_Manuf
                RemainingManufactInputDemand_1                     = Manufacturing_Input_m_adj - DivFabScrap_to_Manuf[:,0]    
                
                # b) Use secondary material, if possible:            
                SecondaryMaterialUse = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0] > 0:
                        if RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0] > RemainingManufactInputDemand_1[mat]:
                            SecondaryMaterialUse[mat,:]            = (RECC_System.FlowDict['F_9_12'].Values[t,0,mat,:] * RemainingManufactInputDemand_1[mat] / RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0]).copy()
                        else:
                            SecondaryMaterialUse[mat,:]            = RECC_System.FlowDict['F_9_12'].Values[t,0,mat,:].copy()
                Non_UsedSecMaterial                                = RECC_System.FlowDict['F_9_12'].Values[t,0,:,:] - SecondaryMaterialUse
                RemainingManufactInputDemand_2                     = RemainingManufactInputDemand_1 - SecondaryMaterialUse[:,0]
                
                # c) Use stock-piled secondary material, if available and if possible:
                RECC_System.StockDict['S_12'].Values[t,0,:,:]      = RECC_System.StockDict['S_12'].Values[t-1,0,:,:] # age stock pile by 1 year
                StockPileSecondaryMaterialUse = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if RECC_System.StockDict['S_12'].Values[t,0,mat,0] > 0:
                        if RECC_System.StockDict['S_12'].Values[t,0,mat,0] > RemainingManufactInputDemand_2[mat]:
                            StockPileSecondaryMaterialUse[mat,:]   = (RECC_System.StockDict['S_12'].Values[t,0,mat,:] * RemainingManufactInputDemand_2[mat] / RECC_System.StockDict['S_12'].Values[t,0,mat,0]).copy()
                        else:
                            StockPileSecondaryMaterialUse[mat,:]   = RECC_System.StockDict['S_12'].Values[t,0,mat,:].copy()
                RECC_System.StockDict['S_12'].Values[t,0,:,:]      = RECC_System.StockDict['S_12'].Values[t,0,:,:] - StockPileSecondaryMaterialUse.copy()
                PrimaryProductionDemand = RemainingManufactInputDemand_2 - StockPileSecondaryMaterialUse[:,0]
                
                # d) Convert internal calculations to system variables:
                RECC_System.FlowDict['F_12_5'].Values[t,0,:,:]     = DivFabScrap_to_Manuf + SecondaryMaterialUse + StockPileSecondaryMaterialUse
                RECC_System.FlowDict['F_4_5'].Values[t,:,:]        = np.einsum('m,me->me',PrimaryProductionDemand,RECC_System.ParameterDict['3_MC_Elements_Materials_Primary'].Values)
                if ScriptConfig['ScrapExport'] == 'True':
                    RECC_System.FlowDict['F_12_0'].Values[t,0,:,:] = Non_DivFabScrap.copy() + Non_UsedSecMaterial.copy() 
                else:
                    RECC_System.StockDict['S_12'].Values[t,0,:,:] += Non_DivFabScrap.copy() + Non_UsedSecMaterial.copy()
             
                # e) Element composition of material flows:         
                Manufacturing_Input_me_final                       = RECC_System.FlowDict['F_4_5'].Values[t,:,:] + RECC_System.FlowDict['F_12_5'].Values[t,0,:,:]
                Manufacturing_Input_gme_final                      = np.einsum('gm,me->gme',Manufacturing_Input_Split_gm,Manufacturing_Input_me_final)
                Element_Material_Composition_Manufacturing         = msf.DetermineElementComposition_All_Oth(Manufacturing_Input_me_final)
                Element_Material_Composition_raw[t,:,:,mS,mR]      = Element_Material_Composition_Manufacturing.copy()
                
                Element_Material_Composition[t,:,:,mS,mR]          = Element_Material_Composition_Manufacturing.copy()
                Par_Element_Composition_of_Materials_m[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
                Par_Element_Composition_of_Materials_u[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
    
                # End of 8)MARKET BALANCE for secondary material.
    
                # 9) Primary production (wood is calulated separately)
                RECC_System.FlowDict['F_3_4'].Values[t,:,:]        = RECC_System.FlowDict['F_4_5'].Values[t,:,:].copy()
                RECC_System.FlowDict['F_0_3'].Values[t,:,:]        = RECC_System.FlowDict['F_3_4'].Values[t,:,:].copy() 
                RECC_System.FlowDict['F_0_3'].Values[t,Wood_loc,Carbon_loc] = 0
    
                # Total roundwood and sawmill losses for construction wood production, by region
                # First, estimate regional split of construction wood use by final demand:
                WoodUse_Divisor = np.einsum(',r->r',RECC_System.FlowDict['F_6_7'].Values[t,:,:,Wood_loc,0].sum(),np.ones(Nr))
                Par_Carbon_Timber_ByRegion_Rel            = np.divide(RECC_System.FlowDict['F_6_7'].Values[t,:,:,Wood_loc,0].sum(axis=1), WoodUse_Divisor, out=np.zeros_like(WoodUse_Divisor), where=WoodUse_Divisor!=0)  # aspects: [r]        
                # Second, calculate carbon in industrial roundwood for structural timber in buildings by region, in Mt of C,
                # as well as waste flows (trimmings, wood shavings) to be send to the waste management industries:
                if Par_Carbon_Timber_ByRegion_Rel.sum() > 0:
                    SysVar_RoundwoodConstruc_c_1_2_r[t,:,mS,mR]     = np.einsum('r,r,r->r', 1 / ParameterDict['4_PY_TimberRoundWood'].Values[:,t,mS], Par_Carbon_Timber_ByRegion_Rel, np.einsum(',r->r',RECC_System.FlowDict['F_3_4'].Values[t,Wood_loc,Carbon_loc],np.ones(Nr)))
                RECC_System.FlowDict['F_1_2'].Values[t,:,Carbon_loc] = SysVar_RoundwoodConstruc_c_1_2_r[t,:,mS,mR]
                RECC_System.FlowDict['F_2_3'].Values[t,Wood_loc,Carbon_loc] = SysVar_RoundwoodConstruc_c_1_2_r[t,:,mS,mR].sum()
                RECC_System.FlowDict['F_3_10'].Values[t,:,Woodwaste_loc,Carbon_loc] = Par_Carbon_Timber_ByRegion_Rel * (RECC_System.FlowDict['F_2_3'].Values[t,Wood_loc,Carbon_loc] - RECC_System.FlowDict['F_3_4'].Values[t,Wood_loc,Carbon_loc])
        
                # 10) Calculate manufacturing output, at global level only
                RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,:]    = np.einsum('me,gm->gme',Element_Material_Composition_Manufacturing,RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0])
            
                # 10a) Calculate material composition of product consumption
                Throughput_FinalGoods_me                           = RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,:].sum(axis =0) + np.einsum('crgme->me',RECC_System.FlowDict['F_17_6'].Values[t,0:CohortOffset,:,:,:,:])
                Element_Material_Composition_cons                  = msf.DetermineElementComposition_All_Oth(Throughput_FinalGoods_me)
                Element_Material_Composition_con[t,:,:,mS,mR]      = Element_Material_Composition_cons.copy()
                
                Par_Element_Composition_of_Materials_c[t,:,:]      = Element_Material_Composition_cons.copy()
                Par_3_MC_Stock_ByElement_Nl[CohortOffset,:,:,:,:]  = np.einsum('mLl,me->lLme',Par_RECC_MC_Nl[CohortOffset,:,:,:,mS],Par_Element_Composition_of_Materials_c[t,:,:]) # clLme
                Par_3_MC_Stock_ByElement_No[CohortOffset,:,:,:,:]  = np.einsum('mOo,me->oOme',Par_RECC_MC_No[CohortOffset,:,:,:,mS],Par_Element_Composition_of_Materials_c[t,:,:]) # cOome                    
                
                # 11) Calculate manufacturing scrap 
                RECC_System.FlowDict['F_5_10'].Values[t,0,:,:]     = np.einsum('gme,mwg->we',Manufacturing_Input_gme_final,Par_FabYieldLoss[:,:,:,t,0]) 
                # Fabrication scrap, to be recycled next year:
                RECC_System.StockDict['S_10'].Values[t,t,:,:,:]    = RECC_System.FlowDict['F_5_10'].Values[t,:,:,:].copy()
                # Remove wood waste, which is treated separately:
                RECC_System.StockDict['S_10'].Values[t,t,:,Woodwaste_loc,Carbon_loc] = 0
                # Wood waste, to be recycled next year:
                RECC_System.StockDict['S_10w'].Values[t,t,:,:,:]   = RECC_System.FlowDict['F_3_10'].Values[t,:,:,:].copy()
                RECC_System.StockDict['S_10w'].Values[t,t,:,Woodwaste_loc,Carbon_loc]   += np.einsum('r,x->r',Par_Carbon_Timber_ByRegion_Rel,RECC_System.FlowDict['F_5_10'].Values[t,:,Woodwaste_loc,Carbon_loc].copy())
            
                # 12) Calculate element composition of final consumption and latest age-cohort in in-use stock
                if 'pav' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_pav_rge,:,:] = Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_pav_rge,:,:]
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_pav_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[CohortOffset,:,Sector_pav_rge,:,mS,mR,t])
                    RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_pav_rge,:,:]   = \
                    np.einsum('prme,pr->prme',Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_pav_rge,:,:],Inflow_Detail_UsePhase_p[t,:,:])/1000 # all elements, Indices='t,r,p,m,e'
                    RECC_System.StockDict['S_7'].Values[t,0:CohortOffset+1,:,Sector_pav_rge,:,:] = \
                    np.einsum('pcrme,cpr->pcrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset+1,:,Sector_pav_rge,:,:],Stock_Detail_UsePhase_p[t,0:CohortOffset+1,:,:])/1000 # All elements.
    
                if 'reb' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_reb_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[SwitchTime+t-1,:,Sector_reb_rge,:,mS,mR,t])
                    # Determine element breakdown of inflow and renovation material
                    RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_reb_rge,:,:]   = \
                    np.einsum('me,Brm->Brme',Par_Element_Composition_of_Materials_c[t,:,:],RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_reb_rge,:,0]) # all elements, Indices='t,r,B,m,e'
                    F_6_7_ren[t,:,:,Sector_reb_rge,:,:] = np.einsum('me,Bcrm->Bcrme',Par_Element_Composition_of_Materials_c[t,:,:],F_6_7_ren[t,:,:,Sector_reb_rge,:,0]) # all elements, Indices='t,c,r,B,m,e' (c is age-cohort where material flows)
                    # Value of Par_3_MC_Stock_ByElement_Nr for current year and all previous age-cohorts c < t need to be updated, as they change due to the adding of recycling materials in the current year. 
                    # Determine the element material composition at the end of last year, as weighting factor for existing stock
                    Divisor  = np.einsum('Bcrm,e->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,0],np.ones(Ne))
                    Par_ElementComposition_LastYear = np.divide(Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,:],Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) #Bcrme
                    # Compile all materials present in stock broken down by element:
                    # Here, The materials present in stock consist of the current products in stock * their element composition of last year plus the inflow of renovation material with this years material production element composition.
                    StockMat = F_6_7_ren[t,0:CohortOffset,:,Sector_reb_rge,:,:] + np.einsum('Bcrm,Bcrme->Bcrme',RECC_System.StockDict['S_7'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0] - F_6_7_ren[t,0:CohortOffset,:,Sector_reb_rge,:,0],Par_ElementComposition_LastYear)
                    Divisor  = np.einsum('Bcrm,e->Bcrme',StockMat[:,:,:,:,0],np.ones(Ne))
                    # Caculate product element composition of latest age-cohort from total materials by element:
                    Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_reb_rge,:,:]  = np.einsum('Bcmr,Bcrme->Bcrme',Par_RECC_MC_Nr[0:CohortOffset,:,Sector_reb_rge,:,mS,mR,t],np.divide(StockMat,Divisor, out=np.zeros_like(Divisor), where=Divisor!=0))
                    # Update stock: break down material into elements:                
                    RECC_System.StockDict['S_7'].Values[t,0:CohortOffset +1,:,Sector_reb_rge,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset +1,:,Sector_reb_rge,:,:],Stock_Detail_UsePhase_B[t,0:CohortOffset +1,:,:])/1000
                    
                if 'nrb' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_nrb_rge,:,:]   = np.einsum('me,Nmr->Nrme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[SwitchTime+t-1,:,Sector_nrb_rge,:,mS,mR,t])
                    # Determine element breakdown of inflow and renovation material
                    RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_nrb_rge,:,:]   = \
                    np.einsum('me,Nrm->Nrme',Par_Element_Composition_of_Materials_c[t,:,:],RECC_System.FlowDict['F_6_7'].Values[t,:,Sector_nrb_rge,:,0]) # all elements, Indices='t,r,N,m,e'
                    F_6_7_ren[t,:,:,Sector_nrb_rge,:,:] = np.einsum('me,Ncrm->Ncrme',Par_Element_Composition_of_Materials_c[t,:,:],F_6_7_ren[t,:,:,Sector_nrb_rge,:,0]) # all elements, Indices='t,c,r,N,m,e' (c is age-cohort where material flows)
                    # Value of Par_3_MC_Stock_ByElement_Nr for current year and all previous age-cohorts c < t need to be updated, as they change due to the adding of recycling materials in the current year. 
                    # Determine the element material composition at the end of last year, as weighting factor for existing stock
                    Divisor  = np.einsum('Ncrm,e->Ncrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_nrb_rge,:,0],np.ones(Ne))
                    Par_ElementComposition_LastYear = np.divide(Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_nrb_rge,:,:],Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) #Ncrme
                    # Compile all materials present in stock broken down by element:
                    # Here, The materials present in stock consist of the current products in stock * their element composition of last year plus the inflow of renovation material with this years material production element composition.
                    StockMat = F_6_7_ren[t,0:CohortOffset,:,Sector_nrb_rge,:,:] + np.einsum('Ncrm,Ncrme->Ncrme',RECC_System.StockDict['S_7'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,0] - F_6_7_ren[t,0:CohortOffset,:,Sector_nrb_rge,:,0],Par_ElementComposition_LastYear)
                    Divisor  = np.einsum('Ncrm,e->Ncrme',StockMat[:,:,:,:,0],np.ones(Ne))
                    # Caculate product element composition of latest age-cohort from total materials by element:
                    Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_nrb_rge,:,:]  = np.einsum('Ncmr,Ncrme->Ncrme',Par_RECC_MC_Nr[0:CohortOffset,:,Sector_nrb_rge,:,mS,mR,t],np.divide(StockMat,Divisor, out=np.zeros_like(Divisor), where=Divisor!=0))
                    # Update stock: break down material into elements:                
                    RECC_System.StockDict['S_7'].Values[t,0:CohortOffset +1,:,Sector_nrb_rge,:,:] = \
                    np.einsum('Ncrme,cNr->Ncrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset +1,:,Sector_nrb_rge,:,:],Stock_Detail_UsePhase_N[t,0:CohortOffset +1,:,:])/1000
                    
                RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,:]   = \
                np.einsum('lIme,Il->lIme',Par_3_MC_Stock_ByElement_Nl[CohortOffset,:,:,:,:],Inflow_Detail_UsePhase_I[t,:,:])/1000 # all elements, Indices='t,l,I,m,e'                
                RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge_reg,:,:]   = \
                np.einsum('aome,ao->aome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_app_rge_reg,:,:], Inflow_Detail_UsePhase_a[t,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'                
                
                if 'nrbg' in SectorList: # results for global region o are on position 0 of r index.
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_nrbg_rge_reg,:,:]   = \
                    np.einsum('Nome,No->Nome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_nrbg_rge_reg,:,:],Inflow_Detail_UsePhase_Ng[t,:,:])/1000 # all elements, Indices='t,o,N,m,e'                
                    
                RECC_System.StockDict['S_7_Nl'].Values[t,0:CohortOffset +1,:,Sector_ind_rge_reg,:,:] = \
                np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:CohortOffset +1,:,Sector_ind_rge_reg,:,:],Stock_Detail_UsePhase_I[t,0:CohortOffset +1,:,:])/1000 # All elements. In Mt
                RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_app_rge_reg,:,:] = \
                np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_app_rge_reg,:,:],Stock_Detail_UsePhase_a[t,0:CohortOffset +1,:,:])/1000000000000 # All elements.In Mt
                if 'nrbg' in SectorList: # results for global region o are on position 0 of r index.
                    RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_nrbg_rge_reg,:,:] = \
                    np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_nrbg_rge_reg,:,:],Stock_Detail_UsePhase_Ng[t,0:CohortOffset +1,:,:])/1000 # All elements.In Mt
                  
                # 13) Calculate wood cascading
                # Add cascading material to cascading stock:
                RECC_System.StockDict['S_9'].Values[t,t,:,:,:] = np.einsum('r,rwe->rwe',Par_RECC_WoodWaste_Cascading[t,Woodwaste_loc,Wood_loc,Woodwaste_loc,:],RECC_System.FlowDict['F_10_9w'].Values[t,:,:,:])
                WoodCascadingInflow[t,:,mS,mR]                 = (44/12) * (1 / ParameterDict['3_MC_CO2FromWoodCombustion'].Values[CO2_loc,Wood_loc]) * np.einsum('r,r->r',    Par_RECC_WoodWaste_Cascading[t,Woodwaste_loc,Wood_loc,Woodwaste_loc,:],RECC_System.FlowDict['F_10_9w'].Values[t,:,Woodwaste_loc,Carbon_loc])
                # Carry forward cascading stock:
                RECC_System.StockDict['S_9'].Values[t,0:t,:,:,:] = RECC_System.StockDict['S_9'].Values[t-1,0:t,:,:,:]
                clt = int(ParameterDict['3_LT_Wood_Cascade'].Values) # cascading lifetime (fixed)
                if t > clt: # need to be far enough in the future to release cascading stocks from 2016 and thereafter
                    casc_release = RECC_System.StockDict['S_9'].Values[t,t-clt,:,:,:].copy()
                    RECC_System.StockDict['S_9'].Values[t,t-clt,:,:,:] = 0
                    SysVar_WoodWasteIncineration[t,:,:,:,mS,mR] += casc_release
                    SysVar_CascadeRelease[t,:,:,:,mS,mR]        += casc_release # record cascade release separately
                # Send wood material to final combustion, both from the current year (no cascading) and after cascading
                SysVar_WoodWasteIncineration[t,:,:,:,mS,mR] += np.einsum('r,rwe->rwe',1 - Par_RECC_WoodWaste_Cascading[t,Woodwaste_loc,Wood_loc,Woodwaste_loc,:],RECC_System.FlowDict['F_10_9w'].Values[t,:,:,:])
                # SysVar_WoodWasteIncineration contains carbon flows whose related CO2 emissions are already accounted for as use phase direct emissions. Calculate to determine system-wide C release after wood use.
                
                # 14) Calculate waste mgt. losses.
                RECC_System.FlowDict['F_9_0'].Values[t,:]          = np.einsum('rgme->e',RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    - np.einsum('rwe->e',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]) \
                                                                   + np.einsum('lLme->e',RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:]) - np.einsum('lwe->e',RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:]) \
                                                                   + np.einsum('oOme->e',RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:]) - np.einsum('owe->e',RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:]) \
                                                                   + np.einsum('rwe->e',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:])      - np.einsum('ome->e',RECC_System.FlowDict['F_9_12'].Values[t,:,:,:]) \
                                                                   + np.einsum('rwe->e',SysVar_WoodWasteIncineration[t,:,:,:,mS,mR])
                
                # 15) Calculate stock changes
                RECC_System.StockDict['dS_7'].Values[t,:,:,:,:,:]     = RECC_System.StockDict['S_7'].Values[t,:,:,:,:,:]    - RECC_System.StockDict['S_7'].Values[t-1,:,:,:,:,:]
                RECC_System.StockDict['dS_7_Nl'].Values[t,:,:,:,:,:]  = RECC_System.StockDict['S_7_Nl'].Values[t,:,:,:,:,:] - RECC_System.StockDict['S_7_Nl'].Values[t-1,:,:,:,:,:]
                RECC_System.StockDict['dS_7_No'].Values[t,:,:,:,:,:]  = RECC_System.StockDict['S_7_No'].Values[t,:,:,:,:,:] - RECC_System.StockDict['S_7_No'].Values[t-1,:,:,:,:,:]            
                RECC_System.StockDict['dS_9'].Values[t,:,:,:,:]       = RECC_System.StockDict['S_9'].Values[t,:,:,:,:]      - RECC_System.StockDict['S_9'].Values[t-1,:,:,:,:]
                RECC_System.StockDict['dS_10'].Values[t,:,:,:]        = RECC_System.StockDict['S_10'].Values[t,t,:,:,:]     - RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:]
                RECC_System.StockDict['dS_10w'].Values[t,:,:,:]       = RECC_System.StockDict['S_10w'].Values[t,t,:,:,:]    - RECC_System.StockDict['S_10w'].Values[t-1,t-1,:,:,:]
                RECC_System.StockDict['dS_12'].Values[t,:,:,:]        = RECC_System.StockDict['S_12'].Values[t,:,:,:]       - RECC_System.StockDict['S_12'].Values[t-1,:,:,:]
                RECC_System.StockDict['dS_0'].Values[t,:]             = RECC_System.FlowDict['F_9_0'].Values[t,:]           + np.einsum('rme->e',RECC_System.FlowDict['F_12_0'].Values[t,:,:,:]) + np.einsum('crgme->e',RECC_System.FlowDict['F_8_0'].Values[t,:,:,:,:,:]) - np.einsum('me->e',RECC_System.FlowDict['F_0_3'].Values[t,:,:])
                
            # Diagnostics:
            # Tbd.
                
            ##########################################################
            #    Section 6) Post-process RECC model solution         #
            ##########################################################            
            # All GHG and material flows/indicators in Mt/yr, all energy flows in TJ/yr.
            # All energy flows are _final_ energy unless otherwise indicated by variable name.
                        
            # A) Calculate intensity of operation, by sector
            # Hop over to save computation time:
            # SysVar_StockServiceProvision_UsePhase_pav = np.einsum('Vrt,tcpr->tcprV',  RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],     Stock_Detail_UsePhase_p)
            # SysVar_StockServiceProvision_UsePhase_reb = np.einsum('tcBVr,tcBr->tcBrV',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],     Stock_Detail_UsePhase_B)
            # SysVar_StockServiceProvision_UsePhase_nrb = np.einsum('cNVr,tcNr->tcNrV', RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS], Stock_Detail_UsePhase_N)
            # Unit: million km/yr for vehicles, million m2 for buildings by three use types: heating, cooling, and DHW.
            # Aggreated computation of building service for export:
            SysVar_StockServiceProvision_UsePhase_reb_agg = np.einsum('tcBVr,tcBr->tV',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],     Stock_Detail_UsePhase_B)
            SysVar_StockServiceProvision_UsePhase_nrb_agg = np.einsum('cNVr,tcNr->tV' ,RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS], Stock_Detail_UsePhase_N)
            
            # B) Calculate total operational energy use, by sector
            # Removed to save computation time and memory.
            
            # C) Translate 'all' energy carriers to specific ones, use phase, by sector
            SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all      = np.zeros((Nt,Nn,Nr))
            if 'pav' in SectorList:
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav  = np.einsum('cprVn,cpVr,Vrt,tcpr->trpn',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles' ].Values[:,:,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[:,:,:,-1,:,mS],RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],Stock_Detail_UsePhase_p, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByService_pav        = np.einsum('cprVn,cpVr,Vrt,tcpr->trV', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles' ].Values[:,:,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[:,:,:,-1,:,mS],RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],Stock_Detail_UsePhase_p, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trpn->tnr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            else:
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav  = np.zeros((Nt,Nr,Np,Nn))
                SysVar_EnergyDemand_UsePhase_ByService_pav        = np.zeros((Nt,Nr,NV))
            if 'reb' in SectorList:
                # 3_EI_Products_UsePhase_resbuildings_t contains EI for final energy (F_16_7) for both historic and future age cohorts (conversion happended above).
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb  = np.einsum('cBVnrt,tcBVr,tcBr->trBn', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByService_reb        = np.einsum('cBVnrt,tcBVr,tcBr->trV',  ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trBn->tnr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            else:
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb  = np.zeros((Nt,Nr,NB,Nn))
                SysVar_EnergyDemand_UsePhase_ByService_reb        = np.zeros((Nt,Nr,NV))
            if 'nrb' in SectorList: 
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb  = np.einsum('cNVnrt,cNVr,tcNr->trNn', ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values,RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS],Stock_Detail_UsePhase_N, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByService_nrb        = np.einsum('cNVnrt,cNVr,tcNr->trV',  ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values,RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS],Stock_Detail_UsePhase_N, optimize = True)
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trNn->tnr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)
            else:
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb  = np.zeros((Nt,Nr,NN,Nn))
                SysVar_EnergyDemand_UsePhase_ByService_nrb        = np.zeros((Nt,Nr,NV))
            if 'nrbg' in SectorList:     
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg = np.zeros((Nt,No,NN,Nn)) # Not yet quantified!
            else:
                SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg = np.zeros((Nt,No,NN,Nn))
                
            # D) Calculate energy demand of the other industries, all in TJ/yr.
            SysVar_EnergyDemand_PrimaryProd    = 1000 * np.einsum('Pnt,tmP,tm->tmPn',RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0],RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_3_4'].Values[:,:,0])
            # Since we have no separate hydrogen production and no impacts from H2 production, loop back the energy demand of H2 production and add it to the parameter:
            SysVar_EnergyDemand_PrimaryProd   += np.einsum('tmPY,Ynt->tmPn',SysVar_EnergyDemand_PrimaryProd,RECC_System.ParameterDict['4_EI_HydrogenEnergyDemand'].Values)
            SysVar_EnergyDemand_Manufacturing  = np.zeros((Nt,Nn))
            if 'pav' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('pn,tpr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_pav_rge,:,110,-1],Inflow_Detail_UsePhase_p)        # conversion factor: 1, as MJ/item     = TJ/Million items.
            if 'reb' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('Bn,tBr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_reb_rge,:,110,-1],Inflow_Detail_UsePhase_B)        # conversion factor: 1, as MJ/m²       = TJ/Million m².
            if 'nrb' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('Nn,tNr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_nrb_rge,:,110,-1],Inflow_Detail_UsePhase_N)        # conversion factor: 1, as MJ/m²       = TJ/Million m².
            if 'nrbg' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('Nn,tNr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_nrbg_rge,:,110,-1],Inflow_Detail_UsePhase_N)        # conversion factor: 1, as MJ/m²      = TJ/Million m².                
            if 'ind' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('In,tIr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_ind_rge,:,110,-1],Inflow_Detail_UsePhase_I) * 1e-6 # conversion factor: 1e-6, as TJ/GW    = 10e-6 MJ/GW. 
            if 'app' in SectorList:
                SysVar_EnergyDemand_Manufacturing += np.einsum('an,tar->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_app_rge,:,110,-1],Inflow_Detail_UsePhase_a) * 1e-6 # conversion factor: 1e-6, as TJ/item  = 10e-6 MJ/items. 
            SysVar_EnergyDemand_WasteMgt       = 1000 * (np.einsum('wn,trw->tn',RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10'].Values[:,:,:,0]) +\
                                                        np.einsum('wn,trw->tn', RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10_Nl'].Values[:,:,:,0]) +\
                                                        np.einsum('wn,trw->tn', RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10_No'].Values[:,:,:,0]))
            SysVar_EnergyDemand_Remelting      = 1000 * np.einsum('mn,trm->tn', RECC_System.ParameterDict['4_EI_RemeltingEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_12'].Values[:,:,:,0])
            SysVar_EnergyDemand_Remelting_m    = 1000 * np.einsum('mn,trm->tnm',RECC_System.ParameterDict['4_EI_RemeltingEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_12'].Values[:,:,:,0])
            
            # Calculate total energy demand by individual energy carrier
            SysVar_TotalEnergyDemand_16_all = np.einsum('trpn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav) + np.einsum('trBn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) \
            + np.einsum('trNn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb) + np.einsum('toNn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg) \
            + np.einsum('tmPn->tn',SysVar_EnergyDemand_PrimaryProd) + SysVar_EnergyDemand_Manufacturing + SysVar_EnergyDemand_WasteMgt + SysVar_EnergyDemand_Remelting
            SysVar_EnergySupply_15_16 = SysVar_TotalEnergyDemand_16_all.copy() # Supply from process 15, will be corrected for WtE below
            
            # E) Calculate biomass-related carbon flows and stocks in waste management and forestry.
            # Wood material lost in waste mgt
            Wood_dry_perCarbon                                     = 1 / (RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44)
            SysVar_Energy_FuelWood_2_7_tr                          = np.einsum('trp->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,WoodFuel_loc]) + np.einsum('trB->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,WoodFuel_loc]) + np.einsum('trN->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,WoodFuel_loc])
            SysVar_Carbon_FuelWood_2_7_tr                          = SysVar_Energy_FuelWood_2_7_tr / RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] / 1000 # carbon only! Mt/yr of C.
            # Fuel wood demand from use phase that is met with available waste wood, and remaining wood waste. carbon only! Mt/yr of C.
            # CO2 ems. from this substitution flow are already accounted for in use phase direct emissions.
            # Remaining wood waste that is not used as fuel wood goes to WtE:
            SysVar_WoodWasteFuelWoodSubst_tr                       = np.minimum(SysVar_Carbon_FuelWood_2_7_tr,SysVar_WoodWasteIncineration[:,:,Woodwaste_loc,Carbon_loc,mS,mR]) # Mt/yr of C.
            FuelWoodSubst_WoodWaste[:,mS,mR]                       = SysVar_WoodWasteFuelWoodSubst_tr.sum(axis=1) # in Mt C / yr        
            # Energy Recovery: in TJ/yr, carbon, CO2 and generated heat and electricity that substitutes demand and reduces energy flow from process 16.
            C_WtE_tr                                               = SysVar_WoodWasteIncineration[:,:,Woodwaste_loc,Carbon_loc,mS,mR] - SysVar_WoodWasteFuelWoodSubst_tr # waste to energy (WtE) carbon flow in Mt C/yr
            SysVar_WoodWaste_Gas_El[:,:,mS,mR]                     = C_WtE_tr
            SysVar_WtE_CO2_xt                                      = np.zeros((Nx,Nt)) # Extensions for 9_16 energy supply
            SysVar_WtE_CO2_xt[GWP100_loc,:]                        = 44/12 * np.einsum('tr->t',C_WtE_tr) # in Mt CO2/yr
            # Determine use phase nat. gas substitution in buildings
            EnergySupplyReduction_NatGas_tr                        = np.minimum(SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all[:,NatuGas_loc,:],C_WtE_tr * RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] * 0.001) # in TJ/yr
            EnergySubst_WtE_NG[:,mS,mR]                            = EnergySupplyReduction_NatGas_tr.sum(axis = 1) # in TJ/yr
            ExcessWoodwasteEnergy_tr                               = C_WtE_tr * RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] * 0.001 - EnergySupplyReduction_NatGas_tr
            # Convert back to carbon and determine electricity generation:
            C_ExcessWoodwasteEnergy_tr                             = ExcessWoodwasteEnergy_tr / (RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] * 0.001)
            EnergySupplyReduction_Electr_tr                        = 1000 * RECC_System.ParameterDict['4_PE_ElectricityFromWoodCombustion'].Values[Woodwaste_loc,0,0] * Wood_dry_perCarbon * C_ExcessWoodwasteEnergy_tr # in TJ/yr
            EnergySubst_WtE_EL[:,mS,mR]                            = EnergySupplyReduction_Electr_tr.sum(axis=1) # in TJ/yr
            # Reduce fuel wood and electricity supply. Only affects external energy supply from process 15, not emissions, as these are accounted for via the process-wise energy demand.
            SysVar_EnergySupply_15_16[:,WoodFuel_loc]             -= 1000 * SysVar_WoodWasteFuelWoodSubst_tr.sum(axis=1) * RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc]
            SysVar_EnergySupply_15_16[:,NatuGas_loc]              -= EnergySubst_WtE_NG[:,mS,mR]         
            SysVar_EnergySupply_15_16[:,Electric_loc]             -= EnergySubst_WtE_EL[:,mS,mR] 
            SysVar_EnergySupply_9_16                               = SysVar_TotalEnergyDemand_16_all - SysVar_EnergySupply_15_16
            # Calculate total energy demand for all energy carriers together. ONLY CORRECT if 'all' energy carriers are at last location -1!
            SysVar_EnergySupply_15_16[:,-1]                        = SysVar_EnergySupply_15_16[:,0:-1].sum(axis=1)
            SysVar_EnergySupply_9_16[:,-1]                         = SysVar_EnergySupply_9_16[:,0:-1].sum(axis=1)
            SysVar_EnergySupply_9_16_NG_trn                        = np.zeros((Nt,Nr,Nn))
            SysVar_EnergySupply_9_16_NG_trn[:,:,NatuGas_loc]       = EnergySupplyReduction_NatGas_tr
            SysVar_EnergySupply_9_16_El_trn                        = np.zeros((Nt,Nr,Nn))
            SysVar_EnergySupply_9_16_El_trn[:,:,Electric_loc]      = EnergySupplyReduction_Electr_tr
            SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all[:,-1] = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all[:,0:-1].sum(axis=1)
            # Unit: TJ/yr.        
            SysVar_EnergySavings_WasteToEnergy = np.zeros((Nt,Nn))
            SysVar_EnergySavings_WasteToEnergy[:,Electric_loc]     = EnergySubst_WtE_EL[:,mS,mR].copy()
            SysVar_EnergySavings_WasteToEnergy[:,NatuGas_loc]      = EnergySubst_WtE_NG[:,mS,mR].copy()
        
            # b) energy and carbon in fuel wood in TJ/yr (energy) and Mt/yr (carbon)
            SysVar_Energy_FuelWood_2_7_tr                         -= 1000 * SysVar_WoodWasteFuelWoodSubst_tr * RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] # in TJ/yr
            SysVar_Carbon_FuelWood_2_7_net_tr                      = SysVar_Carbon_FuelWood_2_7_tr - SysVar_WoodWasteFuelWoodSubst_tr 
            SysVar_Carbon_FuelWood_1_2_tr                          = SysVar_Carbon_FuelWood_2_7_net_tr        
            # For export, in Mt C/yr
            Carbon_Fuelwood_bld[:,mS,mR]                           = SysVar_Carbon_FuelWood_1_2_tr.sum(axis = 1)
            Carbon_Fuelwood_el[:,mS,mR]                            = 0.001 * np.einsum('I,It->t', ParameterDict['4_PE_Carbon_for_Electricity_Generation'].Values[:,Electric_loc],np.einsum('It,t->It', ParameterDict['4_SHA_ElectricityMix_World'].Values[0,mR,:,:],SysVar_EnergySupply_15_16[:,Electric_loc]))
            
            # c) carbon for energy use (Mt/yr)
            RECC_System.FlowDict['F_2_7'].Values[:,Carbon_loc]     = SysVar_Carbon_FuelWood_2_7_net_tr.sum(axis = 1) 
            RECC_System.FlowDict['F_7_0'].Values                   = RECC_System.FlowDict['F_2_7'].Values.copy() # This is for mass balance only. The emissions from burning fuel wood in the use phase are accounted for by the direct emissions already.
            RECC_System.FlowDict['F_1_2'].Values[:,:,Carbon_loc]  += SysVar_Carbon_FuelWood_1_2_tr # aspects: tr, in Mt C/yr
            
            # d) Quantify wood carbon stock change (through harvested wood and its regrowth only, no soil carbon balance, albedo, etc.)
            if ScriptConfig['ForestryModel'] == 'GrowthCurve':
                for mmr in range(0,Nr):
                    # only the forest carbon pool change relative to the baseline is quantified, not the total forest carbon stock.
                    Forest_GrowthTable_timber = np.zeros((Nc,Nc))
                    np.fill_diagonal(Forest_GrowthTable_timber, -1* SysVar_RoundwoodConstruc_c_1_2_r[:,mmr,mS,mR])
                    Forest_GrowthTable_fuelwd = np.zeros((Nt,Nt))
                    np.fill_diagonal(Forest_GrowthTable_fuelwd, -1* SysVar_Carbon_FuelWood_1_2_tr[:,mmr])
                    # We also take into account for the regrowth of forest attributable to all wood present in the 2015 stock.
                    RegrowthCurve_Timber = scipy.stats.norm.cdf(np.arange(0,Nc,1),RECC_System.ParameterDict['3_LT_ForestRotationPeriod_Timber'  ].Values[Wood_loc]    /2, RECC_System.ParameterDict['3_LT_ForestRotationPeriod_Timber'  ].Values[Wood_loc]    /4)
                    RegrowthCurve_FuelWo = scipy.stats.norm.cdf(np.arange(0,Nt,1),RECC_System.ParameterDict['3_LT_ForestRotationPeriod_FuelWood'].Values[WoodFuel_loc]/2, RECC_System.ParameterDict['3_LT_ForestRotationPeriod_FuelWood'].Values[WoodFuel_loc]/4)
                    # Scale regrowth curves and insert them into growth tables:
                    for nnc in range(0,Nc):
                        Forest_GrowthTable_timber[nnc+1::,nnc] = Forest_GrowthTable_timber[nnc,nnc] * (1 - RegrowthCurve_Timber[0:Nc-nnc-1])
                    for nnt in range(0,Nt):
                        Forest_GrowthTable_fuelwd[nnt+1::,nnt] = Forest_GrowthTable_fuelwd[nnt,nnt] * (1 - RegrowthCurve_FuelWo[0:Nt-nnt-1])
            
                    # Assign growth table values to stock and stock changes in process 1:
                    # only the forest carbon pool change relative to the baseline is quantified, not the total forest carbon stock.    
                    RECC_System.StockDict['S_1t'].Values[:,:,mmr,Carbon_loc]              = Forest_GrowthTable_timber[SwitchTime-1::,:]
                    RECC_System.StockDict['S_1f'].Values[:,SwitchTime-1::,mmr,Carbon_loc] = Forest_GrowthTable_fuelwd 
                                
                    RECC_System.StockDict['dS_1t'].Values[:,mmr,Carbon_loc]   = Forest_GrowthTable_timber.sum(axis=1)[SwitchTime-1::]
                    RECC_System.StockDict['dS_1t'].Values[1::,mmr,Carbon_loc] = np.diff(Forest_GrowthTable_timber.sum(axis=1))[SwitchTime-1::]
                    RECC_System.StockDict['dS_1f'].Values[:,mmr,Carbon_loc]   = Forest_GrowthTable_fuelwd.sum(axis=1).copy()
                    RECC_System.StockDict['dS_1f'].Values[1::,mmr,Carbon_loc] = np.diff(Forest_GrowthTable_fuelwd.sum(axis=1)).copy()
        
            if ScriptConfig['ForestryModel'] == 'CarbonNeutral': # Default option   
                # only the forest carbon pool change relative to the baseline is quantified, not the total forest carbon stock.
                RECC_System.StockDict['S_1t'].Values[:,:,:,Carbon_loc]              = 0
                RECC_System.StockDict['S_1f'].Values[:,SwitchTime-1::,:,Carbon_loc] = 0 
                            
                RECC_System.StockDict['dS_1t'].Values[:,:,Carbon_loc]   = 0
                RECC_System.StockDict['dS_1t'].Values[1::,:,Carbon_loc] = 0
                RECC_System.StockDict['dS_1f'].Values[:,:,Carbon_loc]   = 0
                RECC_System.StockDict['dS_1f'].Values[1::,:,Carbon_loc] = 0
        
            RECC_System.FlowDict['F_0_1'].Values[:,:,Carbon_loc] = RECC_System.FlowDict['F_1_2'].Values[:,:,Carbon_loc] + RECC_System.StockDict['dS_1t'].Values[:,:,Carbon_loc] + RECC_System.StockDict['dS_1f'].Values[:,:,Carbon_loc]
            # For the growth curve method, this flow has a large negative initial value due to the boundary conditions. The value for year 0 is set to 0 in the results of the calculations using this system variable, as year 0 is for initialisation only.
            
            # F) Check whether flow value arrays match their indices, etc.
            RECC_System.Consistency_Check() 
        
            # G) Determine Mass Balance
            # Commment out to save computation time:
            #BalAbs = -1 # means that mass bal. computation was commented out to save computation time.
            Bal = RECC_System.MassBalance()
            BalAbs = np.abs(Bal).sum()
            Mylog.info('Total mass balance deviation (np.abs(Bal).sum() for socioeconomic scenario ' + SName + ' and RE scenario ' + RName + ': ' + str(BalAbs) + ' Mt.')                    
    
            # H) Calculate direct combustion-related GHG emissions in the use phase (resulting from the combustion of energy carriers in processes)
            # Unit: Mt/yr. 1 kg/MJ = 1kt/TJ
            SysExt_DirectImpacts_UsePhase_Vehicles      = 0.001 * np.einsum('Xn,xX,trpn->xtrp',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            SysExt_DirectImpacts_UsePhase_Buildings     = 0.001 * np.einsum('Xn,xX,trBn->xtrB',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            SysExt_DirectImpacts_UsePhase_NRBuildgs     = 0.001 * np.einsum('Xn,xX,trNn->xtrN',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)
            SysExt_DirectImpacts_UsePhase_NRBuildgs_g   = 0.001 * np.einsum('Xn,xX,toNn->xtoN',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg)
            SysExt_DirectImpacts_PrimaryProd            = 0.001 * np.einsum('Xn,xX,tmPn->xtm' ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_PrimaryProd)
            SysExt_DirectImpacts_Manufacturing          = 0.001 * np.einsum('Xn,xX,tn->xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_Manufacturing)
            SysExt_DirectImpacts_WasteMgt               = 0.001 * np.einsum('Xn,xX,tn->xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_WasteMgt) + SysVar_WtE_CO2_xt # include WtE emissions
            SysExt_DirectImpacts_Remelting              = 0.001 * np.einsum('Xn,xX,tn->xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_Remelting)
            SysExt_DirectImpacts_Remelting_m            = 0.001 * np.einsum('Xn,xX,tnm->xtm'  ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_Remelting_m)
    
            # I) Calculate process emissions in industry.
            # Double-counting is avoided since there are no process extensions for concrete, only for cement and aggregates.
            # Because of different material production technolgies, aspects m and P don't have a 1-to-1 correspondance anymore. 
            # 4_SHA_MaterialsTechnologyShare_debugging model this 1-to-1.
            # Material production emissions have now multiple contributions: residuals[ now called process ], energy carriers (and electricity), direct emissions. 
            # Direct contributions has been computet already in SysExt_DirectEmissions_PrimaryProd. 
            # Energy Carriers will be computed in the indirect contributions below.
            # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land.
            SysExt_ProcessImpacts_PrimaryProd_mP         = np.einsum('Pxt,tmP,tm->xPmt',ParameterDict['4_PE_ProcessExtensions_Residual'].Values,RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_3_4'].Values[:,:,0])       
            SysExt_ProcessImpacts_PrimaryProd_P          = np.einsum('Pxt,tmP,tm->xPt', ParameterDict['4_PE_ProcessExtensions_Residual'].Values,RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_3_4'].Values[:,:,0])       
            SysExt_ProcessImpacts_PrimaryProd_m          = np.einsum('Pxt,tmP,tm->xtm', ParameterDict['4_PE_ProcessExtensions_Residual'].Values,RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_3_4'].Values[:,:,0])       
            SysExt_ProcessImpacts_PrimaryProd            = np.einsum('Pxt,tmP,tm->xt',  ParameterDict['4_PE_ProcessExtensions_Residual'].Values,RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_3_4'].Values[:,:,0])       
            
            # J) Calculate emissions from energy supply
            # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land.        
            SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings      = 0.001 * np.einsum('nxrt,trBn->xt',RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) \
                                                                           + 0.001 * np.einsum('nxrt,trNn->xt',RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb) \
                                                                           + 0.001 * np.einsum('nxot,toNn->xt',RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg)
            if 'reb' in SectorList or 'nrb' in SectorList: # only if buildings are covered                                                                   
                SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings -= 0.001 * np.einsum('nxrt,trn->xt', RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergySupply_9_16_NG_trn + SysVar_EnergySupply_9_16_El_trn) # subtract impacts that do not occur because some energy is supplied from WtE.
            
            SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_EL   = 0.001 * np.einsum('xrt,trB->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,Electric_loc]) \
                                                                           + 0.001 * np.einsum('xrt,trN->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,Electric_loc]) \
                                                                           + 0.001 * np.einsum('xot,toN->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[Electric_loc,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg[:,:,:,Electric_loc])
            if 'reb' in SectorList or 'nrb' in SectorList: # only if buildings are covered                                                                       
                SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_EL -= 0.001 * np.einsum('nxrt,trn->xt', RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergySupply_9_16_El_trn) # subtract impacts that do not occur because some energy is supplied from WtE.    
            
            SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_Ot   = SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings - SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_EL
            
            SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles          = 0.001 * np.einsum('nxrt,trpn->xt',RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_EL       = 0.001 * np.einsum('xrt,trp->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,Electric_loc])   # electricity only
            SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_Ot       = SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles - SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_EL
            
            if Nr > 1:
                SysExt_IndirectImpacts_EnergySupply_PrimaryProd            = 0.001 * np.einsum('mnxot,tmPn->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,:,:,:,:,mR],SysVar_EnergyDemand_PrimaryProd)
                SysExt_IndirectImpacts_EnergySupply_PrimaryProd_m          = 0.001 * np.einsum('mnxot,tmPn->xtm', RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,:,:,:,:,mR],SysVar_EnergyDemand_PrimaryProd)
                SysExt_IndirectImpacts_EnergySupply_Manufacturing          = 0.001 * np.einsum('nxot,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergyDemand_Manufacturing)
                SysExt_IndirectImpacts_EnergySupply_WasteMgt               = 0.001 * np.einsum('nxot,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergyDemand_WasteMgt)
                SysExt_IndirectImpacts_EnergySupply_Remelting              = 0.001 * np.einsum('nxot,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergyDemand_Remelting)
                SysExt_IndirectImpacts_EnergySupply_Remelting_m            = 0.001 * np.einsum('nxot,tnm->xtm',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergyDemand_Remelting_m)
                SysExt_IndirectImpacts_EnergySupply_All                    = 0.001 * np.einsum('nxot,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_o'].Values[:,:,:,:,mR],SysVar_EnergySupply_15_16) + SysVar_WtE_CO2_xt
            else:
                SysExt_IndirectImpacts_EnergySupply_PrimaryProd            = 0.001 * np.einsum('mnxt,tmPn->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,:,:,0,:,mR],SysVar_EnergyDemand_PrimaryProd)
                SysExt_IndirectImpacts_EnergySupply_PrimaryProd_m          = 0.001 * np.einsum('mnxt,tmPn->xtm', RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,:,:,0,:,mR],SysVar_EnergyDemand_PrimaryProd)
                SysExt_IndirectImpacts_EnergySupply_Manufacturing          = 0.001 * np.einsum('nxt,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR],SysVar_EnergyDemand_Manufacturing)
                SysExt_IndirectImpacts_EnergySupply_WasteMgt               = 0.001 * np.einsum('nxt,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR],SysVar_EnergyDemand_WasteMgt)
                SysExt_IndirectImpacts_EnergySupply_Remelting              = 0.001 * np.einsum('nxt,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR],SysVar_EnergyDemand_Remelting)
                SysExt_IndirectImpacts_EnergySupply_Remelting_m            = 0.001 * np.einsum('nxt,tnm->xtm',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR],SysVar_EnergyDemand_Remelting_m)
                SysExt_IndirectImpacts_EnergySupply_All                    = 0.001 * np.einsum('nxt,tn->xt',     RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR],SysVar_EnergySupply_15_16) + SysVar_WtE_CO2_xt
            
            # Calculate emissions by energy carrier:
            SysExt_DirectImpacts_UsePhase_Vehicles_n                       = 0.001 * np.einsum('Xn,xX,trpn->xtrn',  RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            SysExt_DirectImpacts_UsePhase_ResBuildings_n                   = 0.001 * np.einsum('Xn,xX,trBn->xtrn',  RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            SysExt_IndirectImpacts_EnergySupply_UsePhase_ResBuildings_n    = 0.001 * np.einsum('nxrt,trBn->xtrn',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) 
            SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_n        = 0.001 * np.einsum('nxrt,trpn->xtrn',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            
            # K) Calculate emissions benefits  
            if ScriptConfig['ScrapExportRecyclingCredit'] == 'True':
                SysExt_EnergyDemand_RecyclingCredit                 = -1 * 1000  * np.einsum('Pnt,tmP,tm->tmn',RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0],RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_12_0'].Values[:,-1,:,0]) # in TJ/yr
                SysExt_DirectImpacts_RecyclingCredit                = -1 * 0.001 * np.einsum('Xn,xX,tmn->xt'  ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysExt_EnergyDemand_RecyclingCredit) # in Mt CO2-eq/yr
                SysExt_ProcessImpacts_RecyclingCredit               = -1 *         np.einsum('Pxt,tmP,tm->xt' ,RECC_System.ParameterDict['4_PE_ProcessExtensions_Residual'].Values,RECC_System.ParameterDict['4_SHA_MaterialsTechnologyShare'].Values[0,:,:,mR,:],RECC_System.FlowDict['F_12_0'].Values[:,-1,:,0]) # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land.
                SysExt_IndirectImpacts_EnergySupply_RecyclingCredit = -1 * 0.001 * np.einsum('mnxt,tmn->xt'   ,RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_Materials'].Values[:,:,:,0,:,mR],SysExt_EnergyDemand_RecyclingCredit) # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land. 
            else:
                SysExt_EnergyDemand_RecyclingCredit                 = np.zeros((Nt,Nm,Nn))
                SysExt_DirectImpacts_RecyclingCredit                = np.zeros((Nx,Nt))
                SysExt_ProcessImpacts_RecyclingCredit               = np.zeros((Nx,Nt))
                SysExt_IndirectImpacts_EnergySupply_RecyclingCredit = np.zeros((Nx,Nt))
                
            # L) GWP_bio calculation, using the original RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values
            # and NOT the extended lifetime.
            SysVar_GHGEms_GWP_bio_r = np.zeros((NX,Nt,Nr))
            SysVar_GHGEms_GWP_bio_o = np.zeros((NX,Nt))
            
            if 'reb' in SectorList:
                for ntt in range(0,Nt):
                    for nrr in range(0,Nr):
                        for nBB in range(0,NB):
                            #total carbon in wood (carbon content ca. 0.5)
                            mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7'].Values[ntt,nrr,Sector_reb_rge[nBB],9,0]
                            GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[nBB,nrr,ntt+SwitchTime-1]))]
                            SysVar_GHGEms_GWP_bio_r[0,ntt,nrr] += 44/12 * mass_C * GWP_bio # convert from C to CO2
            if 'nrb' in SectorList:
                for ntt in range(0,Nt):
                    for nrr in range(0,Nr):
                        for nNN in range(0,NN):
                            #total carbon in wood (carbon content ca. 0.5)
                            mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7'].Values[ntt,nrr,Sector_nrb_rge[nNN],9,0]
                            GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_NonResbuildings'].Values[nNN,nrr,ntt+SwitchTime-1]))]
                            SysVar_GHGEms_GWP_bio_r[0,ntt,nrr] += 44/12 * mass_C * GWP_bio # convert from C to CO2     
            if 'nrbg' in SectorList:
                for ntt in range(0,Nt):
                    for nNN in range(0,NN):
                        #total carbon in wood (carbon content ca. 0.5)
                        mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7_No'].Values[ntt,0,Sector_nrbg_rge_reg[nNN],9,0]
                        GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[nNN,0,ntt+SwitchTime-1]))]
                        SysVar_GHGEms_GWP_bio_o[0,ntt] += 44/12 * mass_C * GWP_bio # convert from C to CO2
            SysVar_GHGEms_GWP_bio = np.einsum('Xtr->Xt',SysVar_GHGEms_GWP_bio_r) + SysVar_GHGEms_GWP_bio_o
            
            SysExt_CO2UptakeImpacts_Forests_r = np.zeros((Nx,Nt,Nr,Nm)) # Move CO2 uptake to GWP100 indicator at midpoint level, negative for forest growth/sequestration!
            SysExt_CO2UptakeImpacts_Forests_r[GWP100_loc,:,:,Wood_loc] = -1 * 44/12 * RECC_System.FlowDict['F_0_1'].Values[:,:,Carbon_loc] # negative sign because emissions are measured in X_0 direction.
            SysExt_CO2UptakeImpacts_Forests_r[:,0,:,:] = 0
            SysExt_CO2UptakeImpacts_Forests   = np.einsum('xtrm->xt',SysExt_CO2UptakeImpacts_Forests_r)
            
            # For the GrowthCurve Method, F_0_1 has a large negative initial value due to the boundary conditions. The value for year 0 is set to 0 in the results of the calculations using this system variable, as year 0 is for initialisation only.
            
            # M) Calculate pressure indicators of system, by process group, with recycling credits separate.
            # Number indicates the process number of the ODYM-RECC system definition
            # 'd' behind the number indicates direct, 'i' indirect emissions of that process.
            # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land. 
            SysExt_Impacts_UsePhase_7d             = np.einsum('xtrB->xt',SysExt_DirectImpacts_UsePhase_Buildings) + np.einsum('xtrN->xt',SysExt_DirectImpacts_UsePhase_NRBuildgs) + np.einsum('xtoN->xt',SysExt_DirectImpacts_UsePhase_NRBuildgs_g) + np.einsum('xtrp->xt',SysExt_DirectImpacts_UsePhase_Vehicles)
            SysExt_Impacts_UsePhase_7i_Scope2_El   = SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_EL + SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_EL
            SysExt_Impacts_UsePhase_7i_OtherIndir  = SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings_Ot + SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_Ot
            SysExt_Impacts_PrimaryMaterial_3di     = np.einsum('xtm->xt',SysExt_DirectImpacts_PrimaryProd) + SysExt_ProcessImpacts_PrimaryProd + SysExt_IndirectImpacts_EnergySupply_PrimaryProd
            SysExt_Impacts_PrimaryMaterial_3di_m   = SysExt_DirectImpacts_PrimaryProd + SysExt_ProcessImpacts_PrimaryProd_m + SysExt_IndirectImpacts_EnergySupply_PrimaryProd_m
            SysExt_Impacts_Manufacturing_5di       = SysExt_DirectImpacts_Manufacturing + SysExt_IndirectImpacts_EnergySupply_Manufacturing
            SysExt_Impacts_WasteMgtRemelting_9di   = SysExt_DirectImpacts_WasteMgt + SysExt_DirectImpacts_Remelting + SysExt_IndirectImpacts_EnergySupply_WasteMgt + SysExt_IndirectImpacts_EnergySupply_Remelting
            SysExt_Impacts_MaterialCycle_5di_9di   = SysExt_Impacts_Manufacturing_5di + SysExt_Impacts_WasteMgtRemelting_9di
            SysExt_Impacts_RecyclingCredit         = SysExt_DirectImpacts_RecyclingCredit + SysExt_ProcessImpacts_RecyclingCredit + SysExt_IndirectImpacts_EnergySupply_RecyclingCredit
            # Calculate total env. pressure of system
            SysExt_Impacts_OtherThanUsePhaseDirect = SysExt_Impacts_UsePhase_7i_Scope2_El + SysExt_Impacts_UsePhase_7i_OtherIndir + SysExt_Impacts_PrimaryMaterial_3di + SysExt_Impacts_MaterialCycle_5di_9di
            # Compute total emissions of all processes in the system.
            SysExt_TotalImpacts_13579di            = SysExt_Impacts_UsePhase_7d + SysExt_Impacts_OtherThanUsePhaseDirect + SysExt_CO2UptakeImpacts_Forests # include forest C sequestration
            SysExt_TotalImpacts_3579di             = SysExt_Impacts_UsePhase_7d + SysExt_Impacts_OtherThanUsePhaseDirect # no forest C sequestration
            SysExt_Impacts_Materials_3di_9di       = SysExt_Impacts_PrimaryMaterial_3di + SysExt_Impacts_WasteMgtRemelting_9di
            
            # N) Calculate other indicators
            # Secondary material from EoL material flows only, part of F_9_12, for reporting only:
            SecondaryProduct_EoL_Potential = np.einsum('twe,wmeP->tme',RECC_System.FlowDict['F_9_10'].Values.sum(axis=1),RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[:,:,:,:,0,0])
            SecondaryProduct_EoL_Potential[:,:,0] = np.einsum('tme->tm',SecondaryProduct_EoL_Potential[:,:,1::])
            
            # O) Compile results
            # Unit: 1 billion (1e9) impact units: Mt CO2-eq, Mt of material, km³ of water, 1000 km² of land. 
            Impacts_System_13579di[:,:,mS,mR]                 = SysExt_TotalImpacts_13579di[:,:].copy()
            Impacts_System_3579di[:,:,mS,mR]                  = SysExt_TotalImpacts_3579di[:,:].copy()
            Impacts_UsePhase_7d[:,:,mS,mR]                    = SysExt_Impacts_UsePhase_7d[:,:].copy()
            Impacts_UsePhase_7i_Scope2_El[:,:,mS,mR]          = SysExt_Impacts_UsePhase_7i_Scope2_El[:,:].copy()
            Impacts_UsePhase_7i_OtherIndir[:,:,mS,mR]         = SysExt_Impacts_UsePhase_7i_OtherIndir[:,:].copy()
            Impacts_MaterialCycle_5di_9di[:,:,mS,mR]          = SysExt_Impacts_MaterialCycle_5di_9di[:,:].copy()
            Impacts_RecyclingCredit[:,:,mS,mR]                = SysExt_Impacts_RecyclingCredit[GWP100_loc,:].copy()
            Impacts_ForestCO2Uptake[:,:,mS,mR]                = np.einsum('xtrm->xt', SysExt_CO2UptakeImpacts_Forests_r)[:,:].copy()
            Impacts_ForestCO2Uptake_r[:,:,:,mS,mR]            = np.einsum('xtrm->xtr',SysExt_CO2UptakeImpacts_Forests_r)[:,:,:].copy()
            Impacts_EnergyRecoveryWasteWood[:,:,mS,mR]        = SysVar_WtE_CO2_xt.copy()
            Impacts_OtherThanUsePhaseDirect[:,:,mS,mR]        = SysExt_Impacts_OtherThanUsePhaseDirect[:,:].copy() # all non use-phase processes
            Impacts_Materials_3di_9di[:,:,mS,mR]              = SysExt_Impacts_Materials_3di_9di[:,:].copy()
            Impacts_Vehicles_Direct[:,:,:,mS,mR]              = np.einsum('xtrp->xtr',SysExt_DirectImpacts_UsePhase_Vehicles)[:,:,:].copy()
            Impacts_ReBuildgs_Direct[:,:,:,mS,mR]             = np.einsum('xtrB->xtr',SysExt_DirectImpacts_UsePhase_Buildings)[:,:,:].copy()
            Impacts_NRBuildgs_Direct[:,:,:,mS,mR]             = np.einsum('xtrN->xtr',SysExt_DirectImpacts_UsePhase_NRBuildgs)[:,:,:].copy()
            Impacts_NRBuildgs_Direct_g[:,:,mS,mR]             = np.einsum('xtoN->xt' ,SysExt_DirectImpacts_UsePhase_NRBuildgs_g)[:,:].copy()
            Impacts_PrimaryMaterial_3di[:,:,mS,mR]            = SysExt_Impacts_PrimaryMaterial_3di[:,:].copy()
            Impacts_PrimaryMaterial_3di_m[:,:,:,mS,mR]        = SysExt_Impacts_PrimaryMaterial_3di_m[:,:,:].copy()
            Impacts_Manufact_5di_all[:,:,mS,mR]               = SysExt_Impacts_Manufacturing_5di[:,:].copy()
            Impacts_WasteMgt_9di_all[:,:,mS,mR]               = SysExt_Impacts_WasteMgtRemelting_9di[:,:].copy()
            Impacts_Energy_Supply_All[:,:,mS,mR]              = SysExt_IndirectImpacts_EnergySupply_All[:,:].copy() # For checking the material footprint, energy impacts on GWP are accounted for elsewhere. Emissions from energy use are NOT included here (combustion emissions).
            Carbon_Fuelwood_release[:,mS,mR]                  = Carbon_Fuelwood_bld[:,mS,mR] + Carbon_Fuelwood_el[:,mS,mR]
            # other emissions breakdown
            Impacts_SecondaryMetal_di_m[:,:,:,mS,mR]          = (SysExt_DirectImpacts_Remelting_m + SysExt_IndirectImpacts_EnergySupply_Remelting_m)[:,:,:].copy()
            Impacts_Vehicles_indir[:,:,mS,mR]                 = SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles[:,:].copy()
            Impacts_AllBuildings_indir[:,:,mS,mR]             = SysExt_IndirectImpacts_EnergySupply_UsePhase_AllBuildings[:,:].copy()
            #Impacts_ByEnergyCarrier_UsePhase_d[:,:,:,:,mS,mR] = (SysExt_DirectImpacts_UsePhase_Vehicles_n + SysExt_DirectImpacts_UsePhase_ResBuildings_n)[:,:,:,:].copy()
            #Impacts_ByEnergyCarrier_UsePhase_i[:,:,:,:,mS,mR] = (SysExt_IndirectImpacts_EnergySupply_UsePhase_Vehicles_n + SysExt_IndirectImpacts_EnergySupply_UsePhase_ResBuildings_n)[:,:,:,:].copy()
            
            dynGWP_System_3579di[mS,mR]                       = np.einsum('t,t->',SysExt_TotalImpacts_3579di[GWP100_loc,:],RECC_System.ParameterDict['6_MIP_Cumulative_Pressure_Indicators'].Values[GWP100_loc,dynGWP100_loc,:])
            dynGWP_WoodCycle[mS,mR]                           = np.einsum('tr,t->',SysExt_CO2UptakeImpacts_Forests_r[GWP100_loc,:,:,Wood_loc],RECC_System.ParameterDict['6_MIP_Cumulative_Pressure_Indicators'].Values[GWP100_loc,dynGWP100_loc,:]) + np.einsum('tr,t->', 44/12 * SysVar_WoodWasteIncineration[:,:,Woodwaste_loc,Carbon_loc,mS,mR],RECC_System.ParameterDict['6_MIP_Cumulative_Pressure_Indicators'].Values[GWP100_loc,dynGWP100_loc,:])
            
            # Mass flows
            Material_Inflow[:,:,:,mS,mR]                = np.einsum('trgm->tgm',RECC_System.FlowDict['F_6_7'].Values[:,:,:,:,0]).copy()
            if 'ind' in SectorList:
                Material_Inflow[:,Sector_11reg_rge,:,mS,mR] = np.einsum('tlLm->Ltm',RECC_System.FlowDict['F_6_7_Nl'].Values[:,:,:,:,0]).copy()
            if 'app' in SectorList or 'nrbg' in SectorList:
                Material_Inflow[:,Sector_1reg_rge,:,mS,mR]  = np.einsum('toOm->Otm',RECC_System.FlowDict['F_6_7_No'].Values[:,:,:,:,0]).copy()            
            Scrap_Outflow[:,:,mS,mR]                    = np.einsum('trw->tw',RECC_System.FlowDict['F_9_10'].Values[:,:,:,0]).copy() + np.einsum('trw->tw',RECC_System.FlowDict['F_9_10_Nl'].Values[:,:,:,0]).copy() + np.einsum('trw->tw',RECC_System.FlowDict['F_9_10_No'].Values[:,:,:,0]).copy()
            PrimaryProduction[:,:,mS,mR]                = RECC_System.FlowDict['F_3_4'].Values[:,:,0].copy()
            SecondaryProduct[:,:,mS,mR]                 = RECC_System.FlowDict['F_9_12'].Values[:,0,:,0].copy()
            SecondaryExport[:,:,mS,mR]                  = RECC_System.FlowDict['F_12_0'].Values[:,0,:,0].copy() 
            SecondaryProduct_EoL_Pot[:,:,mS,mR]         = SecondaryProduct_EoL_Potential[:,:,0]
            RenovationMaterialInflow_7[:,:,mS,mR]       = np.einsum('tcrgm->tm',F_6_7_ren[:,:,:,:,:,0]).copy()
            FabricationScrap[:,:,mS,mR]                 = RECC_System.FlowDict['F_5_10'].Values[:,0,:,0].copy()
            ReUse_Materials[:,:,mS,mR]                  = np.einsum('tcrgm->tm',RECC_System.FlowDict['F_17_6'].Values[:,:,:,:,:,0]) + np.einsum('tclLm->tm',RECC_System.FlowDict['F_17_6_Nl'].Values[:,:,:,:,:,0]) + np.einsum('tcoOm->tm',RECC_System.FlowDict['F_17_6_No'].Values[:,:,:,:,:,0])
            Carbon_IndustrialRoundwood_bld[:,:,mS,mR]   = SysVar_RoundwoodConstruc_c_1_2_r[:,:,mS,mR].copy()
            Carbon_Wood_Inflow[:,:,mS,mR]               = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('trg->tr', RECC_System.FlowDict['F_6_7'].Values[:,:,:,Wood_loc,0]).copy())
            Carbon_Wood_Outflow[:,:,mS,mR]              = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('tcrg->tr',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,Wood_loc,0]).copy())
            Carbon_Wood_Stock[:,:,mS,mR]                = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('tcrg->tr',RECC_System.StockDict['S_7'].Values[:,:,:,:,Wood_loc,0]).copy())
            WoodCascadingStock[:,:,mS,mR]               = np.einsum('tcrw->tr',RECC_System.StockDict['S_9'].Values[:,:,:,:,Carbon_loc])
            Cement_Inflow[:,:,mS,mR]                    = np.einsum('trg->tr', RECC_System.FlowDict['F_6_7'].Values[:,:,:,Cement_loc,0]).copy()
            # Energy flows
            EnergyCons_UP_Vh[:,mS,mR]                   = np.einsum('trpn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav).copy()
            EnergyCons_UP_Bd[:,mS,mR]                   = np.einsum('trBn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb).copy() + np.einsum('trNn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb).copy() + np.einsum('toNn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg).copy()
            EnergyCons_Mn[:,mS,mR]                      = SysVar_EnergyDemand_Manufacturing.sum(axis =1).copy()
            EnergyCons_Wm[:,mS,mR]                      = SysVar_EnergyDemand_WasteMgt.sum(axis =1).copy() +  SysVar_EnergyDemand_Remelting.sum(axis =1).copy()
            EnergyCons_PP[:,mS,mR]                      = np.einsum('tmPn->t',SysVar_EnergyDemand_PrimaryProd)
            EnergyCons_PP_m[:,:,mS,mR]                  = np.einsum('tmPn->tm',SysVar_EnergyDemand_PrimaryProd)
            EnergyCons_UP_serv_pav[:,:,Service_Drivg,mS,mR] = SysVar_EnergyDemand_UsePhase_ByService_pav[:,:,Service_Drivg].copy()
            EnergyCons_UP_serv_reb[:,:,Heating_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,Heating_loc].copy()
            EnergyCons_UP_serv_reb[:,:,Cooling_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,Cooling_loc].copy()
            EnergyCons_UP_serv_reb[:,:,DomstHW_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,DomstHW_loc].copy()
            EnergyCons_UP_serv_nrb[:,:,Heating_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_nrb[:,:,Heating_loc].copy()
            EnergyCons_UP_serv_nrb[:,:,Cooling_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_nrb[:,:,Cooling_loc].copy()
            EnergyCons_UP_serv_nrb[:,:,DomstHW_loc,mS,mR]   = SysVar_EnergyDemand_UsePhase_ByService_nrb[:,:,DomstHW_loc].copy()
            EnergyCons_UP_total[:,:,mS,mR]              = np.einsum('tnr->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all)
            EnergyCons_UP_reb[:,:,mS,mR]                = np.einsum('trBn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            EnergyCons_UP_nrb[:,:,mS,mR]                = np.einsum('trNn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)
            EnergyCons_total[:,:,mS,mR]                 = SysVar_TotalEnergyDemand_16_all.copy()
            # Service flows
            Passenger_km[:,mS,mR]                       = np.einsum('rt,tr->t',RECC_System.ParameterDict['1_F_Function_Future'].Values[Sector_pav_loc,:,:,mS],RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]).copy()
            # Hop over to save memory:
            # Vehicle_km[:,mS,mR]                         = np.einsum('tcpr->t',SysVar_StockServiceProvision_UsePhase_pav[:,:,:,:,Service_Driving])
            Vehicle_km[:,mS,mR]                         = np.einsum('rt,tcpr->t', RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[Service_Drivg,:,:,mS],   Stock_Detail_UsePhase_p)
            Service_IO_ResBuildings[:,:,mS,mR]          = SysVar_StockServiceProvision_UsePhase_reb_agg.copy()
            Service_IO_NonResBuildings[:,:,mS,mR]       = SysVar_StockServiceProvision_UsePhase_nrb_agg.copy()
            # Parameters        
            Vehicle_FuelEff[:,:,:,mS,mR]                = np.einsum('tpnr->tpr',RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchTime-1::,:,Service_Drivg,:,:,mS])
            ResBuildng_EnergyCons[:,:,:,mS,mR]          = np.einsum('VtBnr->tBr',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchTime-1::,:,Service_Reb,:,:,mS,mR])
            GWP_bio_Credit[:,mS,mR]                     = SysVar_GHGEms_GWP_bio[0,:].copy()
            # Product flows
            EoL_Products_for_WasteMgt[:,:,mS,mR]        = np.einsum('trgm->tg', RECC_System.FlowDict['F_8_9'].Values[:,:,:,:,0]).copy()
            Outflow_Products_Usephase_all[:,:,mS,mR]    = np.einsum('tcrgm->tg',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,:,0]).copy()
            if 'ind' in SectorList:
                EoL_Products_for_WasteMgt[:,Sector_11reg_rge,mS,mR]        = np.einsum('tlLm->tL', RECC_System.FlowDict['F_8_9_Nl'].Values[:,:,:,:,0]).copy()
                Outflow_Products_Usephase_all[:,Sector_11reg_rge,mS,mR]    = np.einsum('tclLm->tL',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,:,0]).copy()   
            if 'app' in SectorList or 'nrbg' in SectorList:
                EoL_Products_for_WasteMgt[:,Sector_1reg_rge,mS,mR]         = np.einsum('toOm->tO', RECC_System.FlowDict['F_8_9_No'].Values[:,:,:,:,0]).copy()
                Outflow_Products_Usephase_all[:,Sector_1reg_rge,mS,mR]     = np.einsum('tcoOm->tO',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0]).copy()               
            Outflow_Materials_Usephase_all[:,:,mS,mR]   = np.einsum('tcrgm->tm',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,:,0]).copy() + np.einsum('tclLm->tm',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,:,0]).copy() + np.einsum('tcoOm->tm',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0]).copy()
            WasteMgtLosses_To_Landfill[:,:,mS,mR]       = RECC_System.FlowDict['F_9_0'].Values.copy()
            StockCurves_Mat[:,:,mS,mR]                  = np.einsum('tcrgm->tm',RECC_System.StockDict['S_7'].Values[:,:,:,:,:,0]).copy() + np.einsum('tclLm->tm',RECC_System.StockDict['S_7_Nl'].Values[:,:,:,:,:,0]).copy() + np.einsum('tcoOm->tm',RECC_System.StockDict['S_7_No'].Values[:,:,:,:,:,0]).copy()
            if 'pav' in SectorList:
                Stock_2020_pav[:,:,mS,mR]               = Stock_2020_decline_p.sum(axis=1)
            if 'reb' in SectorList:
                Stock_2020_reb[:,:,mS,mR]               = Stock_2020_decline_B.sum(axis=1)
            if 'nrb' in SectorList:
                Stock_2020_nrb[:,:,mS,mR]               = Stock_2020_decline_N.sum(axis=1)               
            
            # Extract calibration for SSP1:
            if mS == 1:
                E_Calib_Vehicles  = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,0:7])
                E_Calib_Buildings = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,0:7])
                E_Calib_NRBuildgs = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,0:7])
            
            # Determine exit flags            
            ExitFlags['Positive_Inflow_F6_7_R32_SSP_'  + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_6_7'].Values.min(),0, IsClose_Remainder_Small)
            ExitFlags['Positive_Outflow_F7_8_R32_SSP_' + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_7_8'].Values.min(),0, IsClose_Remainder_Small)  
            ExitFlags['Positive_Inflow_F8_9_R32_SSP_'  + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_8_9'].Values.min(),0, IsClose_Remainder_Small)
            
            # del RECC_System # Delete system when done, clear memory.
            
    ##############################################################
    #   Section 7) Export and plot results, save, and close      #
    ##############################################################
    Mylog.info('## 5 - Evaluate results, save, and close')
    Mylog.info('### 5.0 - Check data and results for boundary constraints and plausibility. Exit flags.')
    Mylog.info('Model input')          
    
    ExitFlags['3_SHA_LightWeighting_Vehicles_min']             = ParameterDict['3_SHA_LightWeighting_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_LightWeighting_Vehicles_max']             = ParameterDict['3_SHA_LightWeighting_Vehicles'].Values.max()/100 <= 1 # unit is % not 1!
    ExitFlags['3_SHA_DownSizing_Vehicles_min']                 = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_DownSizing_Vehicles_max']                 = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Vehicles_min']                  = ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_Vehicles_max']                  = ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Vehicles_sum']                  = np.isclose(ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.sum(),Nr*NR*Nt, IsClose_Remainder_Large)
    ExitFlags['3_SHA_TypeSplit_Buildings_min']                 = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_Buildings_max']                 = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Buildings_sum']                 = np.isclose(ParameterDict['3_SHA_TypeSplit_Buildings'].Values.sum(),Nr*Nt*NS*NR, IsClose_Remainder_Large)
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_min']           = ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_max']           = ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_sum']           = np.isclose(ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.sum(),Nr*Nt*NS*NR, IsClose_Remainder_Large)
    ExitFlags['LTE_Renovation_Consistency']                    = bool(ScriptConfig['Include_REStrategy_LifeTimeExtension']) & bool(ScriptConfig['Include_Renovation_reb']) & bool(ScriptConfig['Include_Renovation_nrb'])
    ExitFlags['Secondary_Material_Flows_Positive']             = SecondaryProduct.min() >= 0
    
    Mylog.info('Model exit flags:')
    for key in ExitFlags:
        Mylog.info(key + ': ' + str(ExitFlags[key]))
        
    Mylog.info('Model output')
               
    Mylog.info('### 5.1 - Create plots and include in logfiles')
    Mylog.info('Plot and export results')
    
    book = openpyxl.Workbook() # Model results in iamc style (row: specifier, columns: years)
    ws1 = book.active
    ws1.title = 'Cover'
    ws1.cell(row=3, column=2).value = 'ScriptConfig'
    ws1.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
    m = 4
    for x in sorted(ScriptConfig.keys()):
        ws1.cell(row=m, column=2).value = x
        ws1.cell(row=m, column=3).value = ScriptConfig[x]
        m +=1
    
    ws2 = book.create_sheet('Model_Results')
    ColLabels = ['Indicator','Unit','Region','System_location','RE scen','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws2.cell(row=1, column=m+1).value = ColLabels[m]
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws2.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    
    # GHG overview, bulk materials, material footprint
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[GWP100_loc,:,:,:],2,len(ColLabels),'GHG emissions, system-wide _3579di (excl. forests)','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_13579di[GWP100_loc,:,:,:],newrowoffset, len(ColLabels),'GHG emissions, system-wide _13579di (incl. forests)','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[AllMat_loc,:,:,:],newrowoffset,  len(ColLabels),'Material footprint, all materials, system-wide _3579di','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[FosFuel_loc,:,:,:],newrowoffset, len(ColLabels),'Material footprint, fossil fuels, system-wide _3579di','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[MetOres_loc,:,:,:],newrowoffset, len(ColLabels),'Material footprint, metal ores, system-wide _3579di','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[nMetOres_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, non-metallic minerals, system-wide _3579di','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[Biomass_loc,:,:,:],newrowoffset, len(ColLabels),'Material footprint, biomass (dry weight), system-wide _3579di','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[Land_loc,:,:,:],newrowoffset,    len(ColLabels),'Land footprint, system-wide _3579di','1000 km² of land',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_System_3579di[Water_loc,:,:,:],newrowoffset,   len(ColLabels),'Water footprint, system-wide _3579di','km³ of water',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, primary material production _3di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'Process and direct emissions in process 3 and related energy supply','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,8,:,:],newrowoffset,len(ColLabels),'Cement production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,0:4,:,:].sum(axis=1),newrowoffset, len(ColLabels),'Primary steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # final material consumption, fab scrap
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Material_Inflow[:,:,m,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Final consumption of materials: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,0:4,:,:]),newrowoffset,len(ColLabels),'Final consumption of materials: iron&steel (4 groups)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,4:6,:,:]),newrowoffset,len(ColLabels),'Final consumption of materials: aluminium (2 groups)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,[0,1,2,3,4,5,6,10],:,:]),newrowoffset,len(ColLabels),'Final consumption of metals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,[8,11,12,13],:,:]),newrowoffset,len(ColLabels),'Final consumption of non-metallic minerals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,[9],:,:]),newrowoffset,len(ColLabels),'Final consumption of biomaterials/wood (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,[7],:,:]),newrowoffset,len(ColLabels),'Final consumption of plastics (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    for m in range(0,Nw):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,FabricationScrap[:,m,:,:],newrowoffset,len(ColLabels),'Fabrication scrap: ' + IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_5_10','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # secondary materials
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,0:4,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,4:6,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,6,:,:],newrowoffset,len(ColLabels),'Secondary copper','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,[0,1,2,3,4,5,6,10],:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary production of metals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,[8,11,12,13],:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary production of non-metallic mineralic materials (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,9,:,:],newrowoffset,len(ColLabels),'Secondary production of biomaterials/wood (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,7,:,:],newrowoffset,len(ColLabels),'Secondary production of plastics (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_7d[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_OtherThanUsePhaseDirect[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, other than use phase direct: all industries and energy supply','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # GHG emissions, detail
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('trSR->tSR',Impacts_Vehicles_Direct[GWP100_loc,:,:,:,:]),newrowoffset,len(ColLabels),'GHG emissions, vehicles, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('trSR->tSR',Impacts_ReBuildgs_Direct[GWP100_loc,:,:,:,:]),newrowoffset,len(ColLabels),'GHG emissions, res. buildings, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('trSR->tSR',Impacts_NRBuildgs_Direct[GWP100_loc,:,:,:,:]),newrowoffset,len(ColLabels),'GHG emissions, non-res. buildings, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Vehicles_indir[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, vehicles, energy supply _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_AllBuildings_indir[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, res+non-res buildings, energy supply _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Manufact_5di_all[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, manufacturing _5i, all','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_5_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_WasteMgt_9di_all[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, waste mgt. and remelting _9di, all','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_9_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,4:6,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Primary Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,6,:,:],newrowoffset,len(ColLabels),'Primary Cu production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Materials_3di_9di[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, material cycle industries and their energy supply _3di_9di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_3_0 and related energy supply emissions','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # energy flows
    for nn in range(0,Nn):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_total[:,nn,:,:],newrowoffset,len(ColLabels),'energy supply, system-wide, excl. internal WtE: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'TJ / yr',ScriptConfig['RegionalScope'],'F_15_x','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for nn in range(0,Nn):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_total[:,nn,:,:],newrowoffset,len(ColLabels),'energy consumption, use phase: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'TJ / yr',ScriptConfig['RegionalScope'],'F_15_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_reb[:,nn,:,:],newrowoffset,len(ColLabels),'energy consumption, use phase, res. buildings: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'TJ / yr',ScriptConfig['RegionalScope'],'F_15_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_nrb[:,nn,:,:],newrowoffset,len(ColLabels),'energy consumption, use phase, nonres. buildings: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'TJ / yr',ScriptConfig['RegionalScope'],'F_15_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_Vh,newrowoffset,len(ColLabels),'Energy cons., use phase, vehicles','TJ/yr',ScriptConfig['RegionalScope'],'E_16_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_Bd,newrowoffset,len(ColLabels),'Energy cons., use phase, res+non-res buildings','TJ/yr',ScriptConfig['RegionalScope'],'E_16_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_Mn,newrowoffset,len(ColLabels),'Energy cons., manufacturing','TJ/yr',ScriptConfig['RegionalScope'],'E_16_5','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_Wm,newrowoffset,len(ColLabels),'Energy cons., waste mgt. and remelting','TJ/yr',ScriptConfig['RegionalScope'],'E_16_9','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_PP,newrowoffset,len(ColLabels),'Energy cons., primary material production (all materials and energy carriers)','TJ/yr',ScriptConfig['RegionalScope'],'E_16_3','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergySubst_WtE_EL,newrowoffset,len(ColLabels),'Electricity demand (system-wide) met by energy recovery from internal wood waste','TJ/yr',ScriptConfig['RegionalScope'],'E_9_16','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergySubst_WtE_NG,newrowoffset,len(ColLabels),'Natural gas demand (system-wide) met by energy recovery from internal wood waste','TJ/yr',ScriptConfig['RegionalScope'],'E_9_16','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,FuelWoodSubst_WoodWaste,newrowoffset,len(ColLabels),'Fuel wood demand (from use phase) met by internal wood waste','Mt C/ yr',ScriptConfig['RegionalScope'],'F_9_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for nn in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_PP_m[:,nn,:,:],newrowoffset,len(ColLabels),'Energy cons., primary production of '+ IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[nn]+', (all energy carriers)','TJ/yr',ScriptConfig['RegionalScope'],'E_16_3','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # material footprints
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_7d[FosFuel_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, fossil fuels, for use phase energy demand','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'for use phase energy demand','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[MetOres_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, metal ores, for primary material production','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'for primary material production','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[nMetOres_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, non-metallic minerals, for primary material production','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'for primary material production','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[Biomass_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, biomass, for primary material production (both energy supply and feedstock)','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'for primary material production','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[FosFuel_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, fossil fuels, for primary material production (both energy supply and feedstock)','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'for primary material production','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for nn in range(0, Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di_m[MetOres_loc,:,nn,:,:],newrowoffset,len(ColLabels),'Material footprint, metal ores only, production of primary _3di_' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[nn],'Mt/yr',ScriptConfig['RegionalScope'],'Env. extension of F_3_4','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Energy_Supply_All[FosFuel_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, fossil fuels, for the entire (system-wide) energy supply','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'Env. extension to F_16_all','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Energy_Supply_All[Biomass_loc,:,:,:],newrowoffset,len(ColLabels),'Material footprint, biomass/biofuel, for the entire (system-wide) energy supply','Mt/yr of raw materials',ScriptConfig['RegionalScope'],'Env. extension to F_16_all','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # primary energy input
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Energy_Supply_All[PrimEn_loc,:,:,:],newrowoffset,len(ColLabels),'Primary energy input, all energy carriers, for the entire (system-wide) energy supply','TJ/yr of primary energy',ScriptConfig['RegionalScope'],'Env. extension to F_16_all','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_7i_Scope2_El[PrimEn_loc,:,:,:] + Impacts_UsePhase_7i_OtherIndir[PrimEn_loc,:,:,:],newrowoffset,len(ColLabels),'Primary energy input, all energy carriers, for use phase energy demand','TJ/yr of primary energy',ScriptConfig['RegionalScope'],'Env. extension to F_16_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di[PrimEn_loc,:,:,:],newrowoffset,len(ColLabels),'Primary energy input, all energy carriers, for primary material production','TJ/yr of primary energy',ScriptConfig['RegionalScope'],'fEnv. extension to F_16_3','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # stocks
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_pav_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, pass. vehicles','million units',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Stock_2020_pav[:,mr,:,:],newrowoffset,len(ColLabels),'Stock curve of all pre 2021 age-cohorts, pass. vehs.','Vehicles: million, Buildings: million m²',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_reb_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, res. buildings','million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Stock_2020_reb[:,mr,:,:],newrowoffset,len(ColLabels),'Stock curve of all pre 2021 age-cohorts, res. blds.','Vehicles: million, Buildings: million m²',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_nrb_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Stock_2020_nrb[:,mr,:,:],newrowoffset,len(ColLabels),'Stock curve of all pre 2021 age-cohorts, non-res. blds.','Vehicles: million, Buildings: million m²',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'In-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million, Buildings: million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Mat[:,mm,:,:],newrowoffset,len(ColLabels),'In-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Mat.sum(axis=1),newrowoffset,len(ColLabels),'In-use stock, all materials','Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #per capita stocks per sector
    for mr in range(0,Nr):
        for mG in range(0,NG):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,pCStocksCurves[:,mG,mr,:,:],newrowoffset,len(ColLabels),'per capita in-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items[mG],'vehicles: cars per person, buildings: m2 per person',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part, per capita)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #population
    for mr in range(0,Nr):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Population[:,mr,:,:],newrowoffset,len(ColLabels),'Population','million',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'P (population)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #Using less material by design and material substitution shares
    # a) pass. vehicles:
    if 'pav' in SectorList:
        for mr in range(0,Nr):
            for ms in range(0,Ns): # vehicle downsizing parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[ms,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels),'Segment split of newly registered pass. vehicles, ' +IndexTable.Classification[IndexTable.index.get_loc('Car_segments')].Items[ms],'1',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
            for mp in range(0,len(Sector_pav_rge)): # vehicle lightweighting parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values[mp,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels), 'Share of light-weighted cars in newly registered ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)      
            for mp in range(0,len(Sector_pav_rge)): # export vehicle type split
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('Rt,S->tSR',ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[Sector_pav_loc,mr,:,mp,:],np.ones((NS))),newrowoffset,len(ColLabels), 'Type split of newly registered cars, ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)                      
    # b) res. buildings   
    if 'reb' in SectorList:     
        for mr in range(0,Nr): # building downsizing parameter is modified by script, result exported here.
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,ParameterDict['3_SHA_DownSizing_Buildings'].Values[0,mr,:,:,:],newrowoffset,len(ColLabels),'Share of newly built downsized res. buildings','%',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
            for mB in range(0,len(Sector_reb_rge)): # building lightweighting parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,ParameterDict['3_SHA_LightWeighting_Buildings'].Values[mB,mr,:,:,:],newrowoffset,len(ColLabels),'Share of newly built light-weighted ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_reb_rge[mB]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    #passenger and vehicle km, building service
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Passenger_km,newrowoffset,len(ColLabels),'passenger-km supplied by pass. vehicles','million km/yr',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Vehicle_km,newrowoffset,len(ColLabels),'vehicle-km driven by pass. vehicles','million km/yr',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_ResBuildings[:,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total heated floor space, res. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_ResBuildings[:,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total cooled floor space, res. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_NonResBuildings[:,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total heated floor space, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_NonResBuildings[:,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total cooled floor space, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Use phase indirect GHG, primary prodution GHG, material cycle and recycling credit
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase scope 2 (electricity) _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (electricity, for use phase energy)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_7i_OtherIndir[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase other indirect (non-el.) _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (other than el., for use phase energy)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_MaterialCycle_5di_9di[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, manufact, wast mgt., remelting and indirect _5di_9di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_9_0 + E_15_0 (part, for energy supply waste mgt.)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_RecyclingCredit[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, recycling credits','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'outside system','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_ForestCO2Uptake[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG sequestration by forests (w. neg. sign)','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'Process 1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_EnergyRecoveryWasteWood[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions from energy recovery from waste wood (WtE), does not include fuel wood use.','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'waste mgt. and energy supply','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Primary and secondary material production, if not included above already
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,0,:,:], newrowoffset,len(ColLabels),'Primary construction grade steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,1,:,:], newrowoffset,len(ColLabels),'Primary automotive steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,2,:,:], newrowoffset,len(ColLabels),'Primary stainless production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,3,:,:], newrowoffset,len(ColLabels),'Primary cast iron production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,4,:,:], newrowoffset,len(ColLabels),'Primary wrought Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,5,:,:], newrowoffset,len(ColLabels),'Primary cast Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,7,:,:], newrowoffset,len(ColLabels),'Primary plastics production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,9,:,:], newrowoffset,len(ColLabels),'Construction wood, structural, from industrial roundwood','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,10,:,:],newrowoffset,len(ColLabels),'Primary zinc production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,[0,1,2,3,4,5,6,10],:,:].sum(axis=1),newrowoffset,len(ColLabels),'Primary production of metals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,[8,11,12,13],:,:].sum(axis=1),newrowoffset,len(ColLabels),'Primary production of non-metallic mineral materials (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,0,:,:],  newrowoffset,len(ColLabels),'Secondary construction steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,1,:,:],  newrowoffset,len(ColLabels),'Secondary automotive steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,2,:,:],  newrowoffset,len(ColLabels),'Secondary stainless steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,3,:,:],  newrowoffset,len(ColLabels),'Secondary cast iron','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,4,:,:],  newrowoffset,len(ColLabels),'Secondary wrought Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,5,:,:],  newrowoffset,len(ColLabels),'Secondary cast Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,7,:,:],  newrowoffset,len(ColLabels),'Secondary plastics','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,9,:,:],  newrowoffset,len(ColLabels),'Recycled wood','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,10,:,:], newrowoffset,len(ColLabels),'Recycled zinc','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,11,:,:], newrowoffset,len(ColLabels),'Recycled concrete','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,0,:,:],  newrowoffset,len(ColLabels),'Potential for secondary construction steel from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,1,:,:],  newrowoffset,len(ColLabels),'Potential for secondary automotive steel from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,2,:,:],  newrowoffset,len(ColLabels),'Potential for secondary stainless steel from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,3,:,:],  newrowoffset,len(ColLabels),'Potential for secondary cast iron from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,4,:,:],  newrowoffset,len(ColLabels),'Potential for secondary wrought Al from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,5,:,:],  newrowoffset,len(ColLabels),'Potential for secondary cast Al from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,7,:,:],  newrowoffset,len(ColLabels),'Potential for secondary plastics from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,9,:,:],  newrowoffset,len(ColLabels),'Potential for recycled wood from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,10,:,:], newrowoffset,len(ColLabels),'Potential for recycled zinc from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct_EoL_Pot[:,11,:,:], newrowoffset,len(ColLabels),'Potential for recycled concrete from EoL products','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    # GHG of primary and secondary material production
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_PrimaryMaterial_3di_m[GWP100_loc,:,mm,:,:],newrowoffset,len(ColLabels),'GHG emissions, production of primary _3di_' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'Env. extension of F_3_4','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_SecondaryMetal_di_m[GWP100_loc,:,mm,:,:],newrowoffset,len(ColLabels),'GHG emissions, production of secondary _di_' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'E_9_0 (part) and associated em. in E_15_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # inflow and outflow of commodities
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'final consumption (use phase inflow), ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_pav_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all drive technologies together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod_r[:,mr,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'decommissioned buildings (use phase outflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod_r[:,mr,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'decommissioned buildings (use phase outflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod_r[:,mr,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'decommissioned buildings (use phase outflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        for mr in range(0,Nr):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod_r[:,mr,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'decommissioned buildings (use phase outflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'EoL products (use phase outflow), ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_pav_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all drive technologies together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    # Material reuse
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,ReUse_Materials[:,mm,:,:],newrowoffset,len(ColLabels),'ReUse of materials in products, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',ReUse_Materials[:,[0,1,2,3,4,5,6,10],:,:]),newrowoffset,len(ColLabels),'Reuse of metals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',ReUse_Materials[:,[8,11,12,13],:,:]),newrowoffset,len(ColLabels),'Reuse of non-metallic minerals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',ReUse_Materials[:,[9],:,:]),newrowoffset,len(ColLabels),'Reuse of biomaterials/wood (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',ReUse_Materials[:,[7],:,:]),newrowoffset,len(ColLabels),'Reuse of plastics (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)                    
    # carbon in wood inflow, stock, and outflow
    for mr in range(0,Nr):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_IndustrialRoundwood_bld[:,mr,:,:],newrowoffset,len(ColLabels),'Industrial roundwood, hard and softwood, for processing into structural wood elements for residential and non-residential buildings','Mt/yr of C (carbon)',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_1_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Fuelwood_bld,newrowoffset,len(ColLabels),'Fuelwood, hard and softwood, for use in building heating and hot water only (no cooking fuel).','Mt/yr of C (carbon)',ScriptConfig['RegionalScope'],'F_1_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Fuelwood_el, newrowoffset,len(ColLabels),'Fuelwood, hard and softwood, for use in electricity generation.','Mt/yr of C (carbon)',ScriptConfig['RegionalScope'],'F_1_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Fuelwood_release, newrowoffset,len(ColLabels),'Total wood C outflow from fuelwood, in form of CO2.','Mt/yr of C (carbon)',ScriptConfig['RegionalScope'],'F_x_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)            
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Inflow.sum(axis=1), newrowoffset,len(ColLabels),'Carbon in wood and wood products, final consumption/inflow','Mt/yr',ScriptConfig['RegionalScope'],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Outflow.sum(axis=1),newrowoffset,len(ColLabels),'Carbon in wood and wood products, EoL flows, outflow use phase','Mt/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Stock.sum(axis=1),  newrowoffset,len(ColLabels),'Carbon in wood and wood products, in-use stock','Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    for mr in range(0,Nr):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Inflow[:,mr,:,:], newrowoffset,len(ColLabels),'Carbon in wood and wood products, final consumption/inflow by region','Mt C/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Outflow[:,mr,:,:],newrowoffset,len(ColLabels),'Carbon in wood and wood products, EoL flows, outflow use phase by region','Mt C/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Stock[:,mr,:,:],  newrowoffset,len(ColLabels),'Carbon in wood and wood products, in-use stock by region','Mt C',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,WoodCascadingInflow[:,mr,:,:],newrowoffset,len(ColLabels),'Wood for cascading (inflow), by region','Mt/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part) (F_10_9w in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,SysVar_EoLCascEntry[:,mr,:,:],newrowoffset,len(ColLabels),'EoL wood for cascading (inflow), by region','Mt/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part) (F_10_9w in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,WoodCascadingStock[:,mr,:,:], newrowoffset,len(ColLabels),'Carbon in cascaded wood products, in-use stock by region','Mt C',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part) (S_9 in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,SysVar_WoodWasteIncineration[:,mr,Woodwaste_loc,Carbon_loc,:,:], newrowoffset,len(ColLabels),'Total carbon in wood waste for inc., by region, source of total biogenic CO2','Mt C / yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part) (S_9 in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,SysVar_CascadeRelease[:,mr,Woodwaste_loc,Carbon_loc,:,:], newrowoffset,len(ColLabels),'Carbon in outflow of cascading stock, by region','Mt C / yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part) (S_9 in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,SysVar_WoodWaste_Gas_El[:,mr,:,:], newrowoffset,len(ColLabels),'Carbon in wood waste for inc. (WtE), by region, excluding fuel wood subst.','Mt C / yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'S_7 (part) (S_9 in model code)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Inflow[:,mr,:,:], newrowoffset,len(ColLabels),'Cement, final consumption/inflow by region','Mt/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    # specific energy consumption of vehicles and buildings
    for mr in range(0,Nr):
        for mp in range(0,len(Sector_pav_rge)):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Vehicle_FuelEff[:,mp,mr,:,:],newrowoffset,len(ColLabels),'specific energy consumption, driving, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'MJ/km',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mr in range(0,Nr):
        for mB in range(0,len(Sector_reb_rge)):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,ResBuildng_EnergyCons[:,mB,mr,:,:],newrowoffset,len(ColLabels),'specific energy consumption, heating/cooling/DHW, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_reb_rge[mB]],'MJ/m2',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # specific energy consumption of vehicles and buildings
    for mr in range(0,Nr):
        # driving
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_pav[:,mr,Service_Drivg,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, pass. vehs., driving','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        # heating
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_reb[:,mr,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, res. bld. heating','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_nrb[:,mr,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, nonres. bld. heating','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        # cooling
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_reb[:,mr,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, res. bld. cooling','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_nrb[:,mr,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, nonres. bld. cooling','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        # domestic hot water (DHW)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_reb[:,mr,DomstHW_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, res. bld. DHW','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_serv_nrb[:,mr,DomstHW_loc,:,:],newrowoffset,len(ColLabels),'Total use pase energy consumption, nonres. bld. DHW','TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # GWP by energy carrier, vehicles and residential buildings
    # for mr in range(0,Nr):
    #     for mn in range(0,Nn):
    #         newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_ByEnergyCarrier_UsePhase_d[0,:,mr,mn,:,:] + Impacts_ByEnergyCarrier_UsePhase_i[0,:,mr,mn,:,:],newrowoffset,len(ColLabels),'GWP by energy carrier, use phase direct + indirect, all sectors covered by model run, ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[mn],'Mt/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'use phase and scope 2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # carbon credit from longterm biomass storage
    # newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_bio_Credit,newrowoffset,len(ColLabels),'GWP_bio_usephase','Mt / yr',ScriptConfig['RegionalScope'],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    # Excess secondary material export    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,0,:,:], newrowoffset,len(ColLabels),'Export of Secondary construction steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,1,:,:], newrowoffset,len(ColLabels),'Export of Secondary automotive steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,2,:,:], newrowoffset,len(ColLabels),'Export of Secondary stainless steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,3,:,:], newrowoffset,len(ColLabels),'Export of Secondary cast iron','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,4,:,:], newrowoffset,len(ColLabels),'Export of Secondary wrought Al','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,5,:,:], newrowoffset,len(ColLabels),'Export of Secondary cast Al','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,6,:,:], newrowoffset,len(ColLabels),'Export of Secondary copper','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,7,:,:], newrowoffset,len(ColLabels),'Export of Secondary plastics','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,8,:,:], newrowoffset,len(ColLabels),'Export of Secondary cement','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,9,:,:], newrowoffset,len(ColLabels),'Export of Recycled wood','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,10,:,:],newrowoffset,len(ColLabels),'Export of Recycled zinc','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,11,:,:],newrowoffset,len(ColLabels),'Export of Recycled concrete','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,12,:,:],newrowoffset,len(ColLabels),'Export of Secondary concrete aggregates','Mt / yr',ScriptConfig['RegionalScope'],'F_12_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # manufacturing output by materials
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,0,:,:].sum(axis=1), newrowoffset,len(ColLabels),'construction steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,1,:,:].sum(axis=1), newrowoffset,len(ColLabels),'automotive steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,2,:,:].sum(axis=1), newrowoffset,len(ColLabels),'stainless steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,3,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cast iron in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,4,:,:].sum(axis=1), newrowoffset,len(ColLabels),'wrought Al in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,5,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cast Al in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,6,:,:].sum(axis=1), newrowoffset,len(ColLabels),'copper in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,7,:,:].sum(axis=1), newrowoffset,len(ColLabels),'plastics in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,8,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cement in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,9,:,:].sum(axis=1), newrowoffset,len(ColLabels),'wood in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,10,:,:].sum(axis=1),newrowoffset,len(ColLabels),'zinc in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,11,:,:].sum(axis=1),newrowoffset,len(ColLabels),'concrete in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,12,:,:].sum(axis=1),newrowoffset,len(ColLabels),'concrete aggregates in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # postconsumer scrap
    for m in range(0,Nw):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Scrap_Outflow[:,m,:,:],newrowoffset,len(ColLabels),'Postconsumer scrap: ' + IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_9_10 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # EoL Products to waste mgt.
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EoL_Products_for_WasteMgt[:,mg,:,:],newrowoffset,len(ColLabels),'EoL Products to waste mgt., ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_8_9 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Outflow of products from use phase        
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Products_Usephase_all[:,mg,:,:],newrowoffset,len(ColLabels),'Outflow of products from use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Materials_Usephase_all[:,mm,:,:],newrowoffset,len(ColLabels),'Outflow of materials from use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',Outflow_Materials_Usephase_all[:,[0,1,2,3,4,5,6,10],:,:]),newrowoffset,len(ColLabels),'Use phase outflow of metals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',Outflow_Materials_Usephase_all[:,[8,11,12,13],:,:]),newrowoffset,len(ColLabels),'Use phase outflow of non-metallic minerals (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',Outflow_Materials_Usephase_all[:,[9],:,:]),newrowoffset,len(ColLabels),'Use phase outflow of biomaterials/wood (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tmSR->tSR',Outflow_Materials_Usephase_all[:,[7],:,:]),newrowoffset,len(ColLabels),'Use phase outflow of plastics (aggregate materials group)','Mt / yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)            
    # Losses from waste mgt.        
    for me in range(0,Ne):    
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,WasteMgtLosses_To_Landfill[:,me,:,:],newrowoffset,len(ColLabels),'Waste mgt and remelting losses, ' + IndexTable.Classification[IndexTable.index.get_loc('Element')].Items[me],'Mt/yr',ScriptConfig['RegionalScope'],'F_9_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Renovation material inflow:
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,RenovationMaterialInflow_7[:,mm,:,:],newrowoffset,len(ColLabels),'Inflow of renovation material into use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_6_7 (part: renovation inflow)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mr in range(0,Nr):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_ForestCO2Uptake_r[0,:,mr,:,:], newrowoffset,len(ColLabels),'CO2 uptake by forests by region (no trade, w. neg. sign)','Mt CO2/yr',IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items[mr],'F_0_1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    
        
    book2 = openpyxl.Workbook() # Export other model results, calibration values, flags, etc.
    wsx = book2.active
    wsx.title = 'Cover'
    wsx.cell(row=3, column=2).value = 'ScriptConfig'
    wsx.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
    m = 4
    for x in sorted(ScriptConfig.keys()):
        wsx.cell(row=m, column=2).value = x
        wsx.cell(row=m, column=3).value = ScriptConfig[x]
        m +=1       
        
    # Post calibration 2015 parameter values
    if 'pav' in SectorList:
        pav_Sheet = book2.create_sheet('passenger vehicles')
        pav_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
        pav_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)
        pav_Sheet.cell(2,2).value = 'region'
        pav_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,2).value = Rname
            pav_Sheet.cell(m+1,2).font  = openpyxl.styles.Font(bold=True)
            m+=1
        # pC stock values
        pav_Sheet.cell(2,3).value = '2015 per capita stock values, total (all segments and drive technologies), by model region. Unit: 1 (veh. per person).'
        pav_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_p_pC[0,m-2]
            m+=1
        # passenger-km
        pav_Sheet.cell(2,4).value = '2015 annual passenger kilometrage, by model region. Unit: km/yr.'
        pav_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,4).value = Total_Service_pav_tr_pC[0,m-2]
            m+=1
        # vehicle km
        pav_Sheet.cell(2,5).value = '2015 annual vehicle kilometrage, by model region. Unit: km/yr. Value for SSP1.'
        pav_Sheet.cell(2,5).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,5).value = ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[Service_Drivg,m-2,0,1]
            m+=1
        # vehicle occupancy rate
        pav_Sheet.cell(2,6).value = '2015 average vehicle occupancy rate, across all segments and drive technologies, by model region. Unit: km/yr. Value for SSP1.'
        pav_Sheet.cell(2,6).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,6).value = ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,m-2,0,1]
            m+=1
        # energy consumption, use phase
        pav_Sheet.cell(2,7).value = '2015 use phase energy consumption, across all segments and drive technologies, by model region. Unit: TJ/yr. Value for SSP1.'
        pav_Sheet.cell(2,7).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            pav_Sheet.cell(m+1,7).value = E_Calib_Vehicles[0,m-2]
            m+=1
            
    if 'reb' in SectorList:
        reb_Sheet = book2.create_sheet('residential buildings')
        reb_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
        reb_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)
        reb_Sheet.cell(2,2).value = 'region'
        reb_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            reb_Sheet.cell(m+1,2).value = Rname
            reb_Sheet.cell(m+1,2).font  = openpyxl.styles.Font(bold=True)        
            m+=1
        # pC stock values
        reb_Sheet.cell(2,3).value = '2015 per capita stock values, total (all building types and energy standards), by model region. Unit: m2 per person.'
        reb_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)      
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            reb_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_B_pC[0,m-2] # index structure is tr!
            m+=1
        # energy consumption, use phase
        reb_Sheet.cell(2,4).value = '2015 use phase energy consumption, across all building types and energy standards, by model region. Unit: TJ/yr. Value for SSP1.'
        reb_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)      
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            reb_Sheet.cell(m+1,4).value = E_Calib_Buildings[0,m-2]
            m+=1            
    
    if 'nrb' in SectorList:
        nrb_Sheet = book2.create_sheet('nonres. buildings')
        nrb_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
        nrb_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)  
        nrb_Sheet.cell(2,2).value = 'region'
        nrb_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            nrb_Sheet.cell(m+1,2).value = Rname
            m+=1
        # pC stock values
        nrb_Sheet.cell(2,3).value = '2015 per capita stock values, total (all building types and energy standards), by model region. Unit: m2 per person.'
        nrb_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            nrb_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_N_pC[m-2,0] # index structure is rt!
            m+=1
        # energy consumption, use phase
        nrb_Sheet.cell(2,4).value = '2015 use phase energy consumption, across all building types and energy standards, by model region. Unit: TJ/yr. Value for SSP1.'
        nrb_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
            nrb_Sheet.cell(m+1,4).value = E_Calib_NRBuildgs[0,m-2]
            m+=1     
            
    # Export time-less indicators, like dynGWP    
    ws3 = book2.create_sheet('Scenario_Indicators')
    Items_SSP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items
    Items_RCP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items
    for m in range(0,2*NS):
        ws3.cell(row=2, column=m+3).value = Items_SSP[m//2]
        ws3.cell(row=2, column=m+3).font  = openpyxl.styles.Font(bold=True)
        ws3.cell(row=3, column=m+3).value = Items_RCP[m%2]
        ws3.cell(row=3, column=m+3).font  = openpyxl.styles.Font(bold=True)
    
    ws3.cell(row=4, column=1).value = 'dynGWP_System_3579di'
    ws3.cell(row=4, column=1).font  = openpyxl.styles.Font(bold=True)
    ws3.cell(row=4, column=2).value = 'Mt CO_2 eq'
    ws3.cell(row=5, column=1).value = 'dynGWP_WoodCycle'
    ws3.cell(row=5, column=1).font  = openpyxl.styles.Font(bold=True)
    ws3.cell(row=5, column=2).value = 'Mt CO_2 eq'
    for m in range(0,NS):
        for n in range(0,NR):
            ws3.cell(row=4, column=3+NR*m+n).value = dynGWP_System_3579di[m,n]
            ws3.cell(row=5, column=3+NR*m+n).value = dynGWP_WoodCycle[m,n]
            
            
    # export emission factors. Unit: kgCO2 eq/MJ
    pd_xlsx_writer = pd.ExcelWriter(os.path.join(ProjectSpecs_Path_Result,'Extensions_'+ ScriptConfig['Current_UUID'] + '.xlsx'), engine="xlsxwriter")
    
    EF_bas = pd.DataFrame( data =EmissionFactorElectricity_r[:,:,0].transpose(), columns = IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items)
    EF_RCP = pd.DataFrame( data =EmissionFactorElectricity_r[:,:,1].transpose(), columns = IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items)
    
    EF_bas.to_excel(pd_xlsx_writer, sheet_name="EF_baseline")
    EF_RCP.to_excel(pd_xlsx_writer, sheet_name="EF_RCP")    
    
    # export residuals  
    pd_res = pd.DataFrame( 
        data = residuals[:,:], 
        columns = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items,
        index = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items)
    pd_res.to_excel(pd_xlsx_writer, sheet_name="EF_Residuals") 
    
    #pd_ecc = pd.DataFrame( # energy carriers contributions
    #     data = fuel_production[:,:], 
    #     columns = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items,
    #     index = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items)
    #pd_ecc.to_excel(pd_xlsx_writer, sheet_name="en_carr_contrib_fuels") 
    
    #pd_dir = pd.DataFrame( # direct contributions
    #     data = direct_impact[:,:], 
    #     columns = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items,
    #     index = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items)
    #pd_ecc.to_excel(pd_xlsx_writer, sheet_name="en_carr_contrib_direct") 
    
    #pd_el = pd.DataFrame( # electricity production contribution
    #     data = elec_production[:,:], 
    #     columns = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items,
    #     index = IndexTable.Classification[IndexTable.index.get_loc('MaterialProductionProcess')].Items)
    #pd_ecc.to_excel(pd_xlsx_writer, sheet_name="en_carr_contrib_electr") 
    
    ##############################
    # PLOT
    MyColorCycle = pylab.cm.Paired(np.arange(0,1,0.2))
    #linewidth = [1.2,2.4,1.2,1.2,1.2]
    linewidth  = [1.2,2,1.2]
    linewidth2 = [1.2,2,1.2]
    
    Figurecounter = 1
    LegendItems_SSP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items
    #LegendItems_RCP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items
    LegendItems_SSP_RE = ['LED, no EST', 'LED, 2°C ES', 'SSP1, no EST', 'SSP1, 2°C ES', 'SSP2, no EST', 'SSP2, 2°C ES']
    LegendItems_SSP_UP = ['Use Phase, SSP1, no EST', 'Rest of system GHG, SSP1, no EST, forestry excl.','Use Phase, SSP1, 2°C ES', 'Rest of system GHG, SSP1, 2°C ES, forestry excl.']
    ColorOrder         = [1,0,3]
    
    # policy baseline vs. RCP 2.6
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    for m in range(0,NS):
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),Impacts_System_13579di[0,:,m,0], linewidth = linewidth[m], color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.line((0, 0), 1, 1, fc=MyColorCycle[m,:]))
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),Impacts_System_13579di[0,:,m,1], linewidth = linewidth2[m], linestyle = '--', color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:]))     
    #plt_lgd  = plt.legend(reversed(ProxyHandlesList),reversed(LegendItems_SSP_RE),shadow = False, prop={'size':9}, loc = 'lower left')# 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.legend(LegendItems_SSP_RE,shadow = False, prop={'size':9}, loc = 'upper left')# 'upper right' ,bbox_to_anchor=(1.20, 1))    
    plt.ylabel('GHG emissions of system, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('System-wide emissions, by SSP scenario, '+ ScriptConfig['RegionalScope'] + '.', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2016, 2050, 0, ScriptConfig['Plot1Max']])
    plt.show()
    fig_name = 'GHG_Ems_Overview'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    for m in range(0,NS):
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),Impacts_PrimaryMaterial_3di[0,:,m,1])
    plt.legend(LegendItems_SSP,shadow = False, prop={'size':9}, loc = 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.ylabel('GHG emissions of primary material production, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('GHG primary materials, with EST', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2015, 2050, 0, ScriptConfig['Plot2Max']])
    plt.show()
    fig_name = 'GHG_PP_WithEST'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    # primary steel, no CP and 2°C combined:
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    for m in range(0,NS):
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),PrimaryProduction[:,0:3,m,0].sum(axis=1), linewidth = linewidth[m], color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.line((0, 0), 1, 1, fc=MyColorCycle[m,:]))
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),PrimaryProduction[:,0:3,m,1].sum(axis=1), linewidth = linewidth2[m], linestyle = '--', color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:]))     
    #plt_lgd  = plt.legend(reversed(ProxyHandlesList),reversed(LegendItems_SSP_RE),shadow = False, prop={'size':9}, loc = 'lower left')# 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.legend(LegendItems_SSP_RE,shadow = False, prop={'size':9}, loc = 'upper right')# 'upper right' ,bbox_to_anchor=(1.20, 1))    
    plt.ylabel('Primary steel production, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('Primary steel production, by SSP scenario, '+ ScriptConfig['RegionalScope'] + '.', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2017, 2050, 0, 0.15 * ScriptConfig['Plot2Max']])
    plt.show()
    fig_name = 'PSteel_Overview'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    #fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    # Cement production, no RE and RE combined:
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    for m in range(0,NS):
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),PrimaryProduction[:,8,m,0], linewidth = linewidth[m], color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.line((0, 0), 1, 1, fc=MyColorCycle[m,:]))
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1),PrimaryProduction[:,8,m,1], linewidth = linewidth2[m], linestyle = '--', color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:]))     
    #plt_lgd  = plt.legend(reversed(ProxyHandlesList),reversed(LegendItems_SSP_RE),shadow = False, prop={'size':9}, loc = 'lower left')# 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.legend(LegendItems_SSP_RE,shadow = False, prop={'size':9}, loc = 'upper right')# 'upper right' ,bbox_to_anchor=(1.20, 1))    
    plt.ylabel('Cement production, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('Cement production, by SSP scenario, '+ ScriptConfig['RegionalScope'] + '.', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2017, 2050, 0, 0.30 * ScriptConfig['Plot2Max']])
    plt.show()
    fig_name = 'Cement_Overview'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    #fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    # Recycled steel, RE and no RE
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    for m in range(0,NS):
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), SecondaryProduct[:,0:4,m,0].sum(axis =1), linewidth = linewidth[m], color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.line((0, 0), 1, 1, fc=MyColorCycle[m,:]))
        ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), SecondaryProduct[:,0:4,m,1].sum(axis =1), linewidth = linewidth2[m], linestyle = '--', color = MyColorCycle[ColorOrder[m],:])
        #ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:]))     
    #plt_lgd  = plt.legend(reversed(ProxyHandlesList),reversed(LegendItems_SSP_RE),shadow = False, prop={'size':9}, loc = 'lower left')# 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.legend(LegendItems_SSP_RE,shadow = False, prop={'size':9}, loc = 'upper left')# 'upper right' ,bbox_to_anchor=(1.20, 1))    
    plt.ylabel('Recycled steel and iron, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('Recycled iron and steel, by SSP scenario, '+ ScriptConfig['RegionalScope'] + '.', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2018, 2050, 0, 0.8 * ScriptConfig['Plot3Max']])
    plt.show()
    fig_name = 'SteelRecycling_Overview'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    #fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    # Use phase and indirect emissions, RE and no RE
    fig1, ax1 = plt.subplots()
    ax1.set_prop_cycle('color', MyColorCycle)
    ProxyHandlesList = []
    # Use phase and other ems., SSP1, no RE
    ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), Impacts_UsePhase_7d[0,:,0,0] , linewidth = linewidth[2], color = MyColorCycle[ColorOrder[2],:])
    ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), Impacts_OtherThanUsePhaseDirect[0,:,0,0] ,    linewidth = linewidth[2], color = MyColorCycle[ColorOrder[2],:], linestyle = '--')
    # Use phase and other ems., SSP1, with RE
    ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), Impacts_UsePhase_7d[0,:,0,1] , linewidth = linewidth[2], color = MyColorCycle[ColorOrder[1],:])
    ax1.plot(np.arange(Model_Time_Start,Model_Time_End +1), Impacts_OtherThanUsePhaseDirect[0,:,0,1] ,    linewidth = linewidth[2], color = MyColorCycle[ColorOrder[1],:], linestyle = '--')
    #plt_lgd  = plt.legend(reversed(ProxyHandlesList),reversed(LegendItems_SSP_RE),shadow = False, prop={'size':9}, loc = 'lower left')# 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.legend(LegendItems_SSP_UP,shadow = False, prop={'size':9}, loc = 'upper right')# 'upper right' ,bbox_to_anchor=(1.20, 1))    
    plt.ylabel('GHG emissions, Mt/yr.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    plt.title('GHG emissions by process and scenario, SSP1, '+ ScriptConfig['RegionalScope'] + '.', fontsize = 12) 
    if ScriptConfig['UseGivenPlotBoundaries'] == True:
        plt.axis([2018, 2050, 0, 0.75 * ScriptConfig['Plot1Max']])
    plt.show()
    fig_name = 'GHG_UsePhase_Overview'
    # include figure in logfile:
    fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    Figurecounter += 1
    
    
    # Plot system emissions, by process, stacked.
    # Area plot, stacked, GHG emissions, material production, waste mgt, remelting, etc.
    MyColorCycle = pylab.cm.gist_earth(np.arange(0,1,0.155)) # select 12 colors from the 'Set1' color map.            
    #grey0_9      = np.array([0.9,0.9,0.9,1])
    
    SSPScens   = ['LED','SSP1','SSP2']
    RCPScens   = ['No climate policy','RCP2.6 energy mix']
    Area       = ['use phase','use phase, scope 2 (el)','use phase, other energy, indirect','primary material product.','manufact. & recycling','forest sequestration','total (+ forest sequestr.)']     
    DataAExp   = np.zeros((NS,NR,Nt,6))
    
    
    for mS in range(0,NS): # SSP
        for mR in range(0,NR): # RCP
        
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
            ProxyHandlesList = []   # For legend     
            
            # plot area
            ax1.fill_between(np.arange(2015,2061),np.zeros((Nt)), Impacts_UsePhase_7d[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[1,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[1,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[GWP100_loc,:,mS,mR], Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[2,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[2,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR], Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[3,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR], Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend    
            ax1.fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR] + Impacts_MaterialCycle_5di_9di[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[5,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[5,:])) # create proxy artist for legend    
            ax1.fill_between(np.arange(2016,2061),np.zeros((Nt-1)),Impacts_ForestCO2Uptake[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[6,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[6,:])) # create proxy artist for legend    
            plt.plot(np.arange(2016,2061), Impacts_System_13579di[GWP100_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            plta = Line2D(np.arange(2016,2061), Impacts_System_13579di[GWP100_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            ProxyHandlesList.append(plta) # create proxy artist for legend    
            #plt.text(Data[m,:].min()*0.55, 7.8, 'Baseline: ' + ("%3.0f" % Base[m]) + ' Mt/yr.',fontsize=14,fontweight='bold')
            
            #copy data to export array
            DataAExp[mS,mR,:,0] = Impacts_UsePhase_7d[GWP100_loc,:,mS,mR].copy()
            DataAExp[mS,mR,:,1] = Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR].copy()
            DataAExp[mS,mR,:,2] = Impacts_UsePhase_7i_OtherIndir[GWP100_loc,:,mS,mR].copy()
            DataAExp[mS,mR,:,3] = Impacts_PrimaryMaterial_3di[GWP100_loc,:,mS,mR].copy()
            DataAExp[mS,mR,:,4] = Impacts_MaterialCycle_5di_9di[GWP100_loc,:,mS,mR].copy()
            DataAExp[mS,mR,:,5] = Impacts_System_13579di[GWP100_loc,:,mS,mR].copy()
            
            plt.title('GHG emissions, stacked by process group, \n' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.', fontsize = 18)
            plt.ylabel(r'Mt of CO$_2$-eq.', fontsize = 18)
            plt.xlabel('Year', fontsize = 18)
            plt.xticks(fontsize=18)
            plt.yticks(fontsize=18)
            plt.legend(handles = reversed(ProxyHandlesList),labels = reversed(Area), shadow = False, prop={'size':14},ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
            ax1.set_xlim([2015, 2060])
            #ax1.set_ylim([0, 220])
            
            plt.show()
            fig_name = 'GWP_TimeSeries_AllProcesses_Stacked_' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.png'
            # include figure in logfile:
            fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
            # comment out to save disk space in archive:
            fig.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
            Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
            Figurecounter += 1
            
    # Area plot, for material industries:
    Area2   = ['primary material product.','waste mgt. & recycling','manufacturing']     
    
    
    for mS in range(0,NS): # SSP
        for mR in range(0,NR): # RCP
        
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
            ProxyHandlesList = []   # For legend     
            
            # plot area
            ax1.fill_between(np.arange(2016,2061),np.zeros((Nt-1)), Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061),Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR] + Impacts_WasteMgt_9di_all[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[5,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[5,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061),Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR] + Impacts_WasteMgt_9di_all[GWP100_loc,1::,mS,mR], Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR] + Impacts_WasteMgt_9di_all[GWP100_loc,1::,mS,mR] + Impacts_Manufact_5di_all[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[6,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[6,:])) # create proxy artist for legend
            
            
            plt.title('GHG emissions, stacked by process group, \n' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.', fontsize = 18)
            plt.ylabel('Mt of CO2-eq.', fontsize = 18)
            plt.xlabel('Year', fontsize = 18)
            plt.xticks(fontsize=18)
            plt.yticks(fontsize=18)
            plt.legend(handles = reversed(ProxyHandlesList),labels = reversed(Area2), shadow = False, prop={'size':14},ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
            ax1.set_xlim([2015, 2060])
            
            plt.show()
            fig_name = 'GWP_TimeSeries_Materials_Stacked_' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.png'
            # include figure in logfile:
            fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
            # comment out to save disk space in archive:
            fig.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
            Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
            Figurecounter += 1
            
    # Area plot for three ENVIRONMENTAL INDICATORS
    for mS in range(0,NS): # SSP
        for mR in range(0,NR): # RCP
        
            fig, axs = plt.subplots(nrows=3, ncols=1 , figsize=(7, 7))
            fig.suptitle('Environmental indicators - by process group, \n '+ ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.',fontsize=18)
        
            ProxyHandlesList = []   # For legend     
        
            # plot area
            # GWP
            axs[0].fill_between(np.arange(2015,2061),np.zeros((Nt)), Impacts_UsePhase_7d[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[1,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[1,:])) # create proxy artist for legend
            axs[0].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[GWP100_loc,:,mS,mR], Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[2,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[2,:])) # create proxy artist for legend
            axs[0].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR], Impacts_UsePhase_7d[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,:,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[3,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
            axs[0].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR], Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend    
            axs[0].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR], Impacts_UsePhase_7d[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[GWP100_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[GWP100_loc,1::,mS,mR] + Impacts_MaterialCycle_5di_9di[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[5,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[5,:])) # create proxy artist for legend    
            axs[0].fill_between(np.arange(2016,2061),np.zeros((Nt-1)),Impacts_ForestCO2Uptake[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[6,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[6,:])) # create proxy artist for legend    
            axs[0].plot(np.arange(2016,2061), Impacts_System_13579di[GWP100_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            plta = Line2D(np.arange(2016,2061), Impacts_System_13579di[GWP100_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            ProxyHandlesList.append(plta) # create proxy artist for legend    
            axs[0].set_title('GWP - Mt of CO2-eq')
            # land
            axs[1].fill_between(np.arange(2015,2061),np.zeros((Nt)), Impacts_UsePhase_7d[Land_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[1,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[1,:])) # create proxy artist for legend
            axs[1].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[Land_loc,:,mS,mR], Impacts_UsePhase_7d[Land_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[2,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[2,:])) # create proxy artist for legend
            axs[1].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[Land_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,:,mS,mR], Impacts_UsePhase_7d[Land_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,:,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Land_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[3,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
            axs[1].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Land_loc,1::,mS,mR], Impacts_UsePhase_7d[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Land_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Land_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend    
            axs[1].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Land_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Land_loc,1::,mS,mR], Impacts_UsePhase_7d[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Land_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Land_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Land_loc,1::,mS,mR] + Impacts_MaterialCycle_5di_9di[Land_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[5,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[5,:])) # create proxy artist for legend    
            axs[1].plot(np.arange(2016,2061), Impacts_System_13579di[Land_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            plta = Line2D(np.arange(2016,2061), Impacts_System_13579di[Land_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            ProxyHandlesList.append(plta) # create proxy artist for legend    
            axs[1].set_title('Land occupationn (LOP) - 1000 km2a')
            # water
            axs[2].fill_between(np.arange(2015,2061),np.zeros((Nt)), Impacts_UsePhase_7d[Water_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[1,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[1,:])) # create proxy artist for legend
            axs[2].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[Water_loc,:,mS,mR], Impacts_UsePhase_7d[Water_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[2,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[2,:])) # create proxy artist for legend
            axs[2].fill_between(np.arange(2015,2061),Impacts_UsePhase_7d[Water_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,:,mS,mR], Impacts_UsePhase_7d[Water_loc,:,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,:,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Water_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[3,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
            axs[2].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Water_loc,1::,mS,mR], Impacts_UsePhase_7d[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Water_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Water_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend    
            axs[2].fill_between(np.arange(2016,2061),Impacts_UsePhase_7d[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Water_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Water_loc,1::,mS,mR], Impacts_UsePhase_7d[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_Scope2_El[Water_loc,1::,mS,mR] + Impacts_UsePhase_7i_OtherIndir[Water_loc,1::,mS,mR] + Impacts_PrimaryMaterial_3di[Water_loc,1::,mS,mR] + Impacts_MaterialCycle_5di_9di[Water_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[5,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[5,:])) # create proxy artist for legend    
            axs[2].plot(np.arange(2016,2061), Impacts_System_13579di[Water_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            plta = Line2D(np.arange(2016,2061), Impacts_System_13579di[Water_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            ProxyHandlesList.append(plta) # create proxy artist for legend    
            axs[2].set_title('Water consumption potential (WCP)  - billions m3')
            
            fig.legend(handles = reversed(ProxyHandlesList),labels = reversed(Area), shadow = False, prop={'size':14},ncol=2, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
            plt.tight_layout()
            plt.show()
            fig_name = 'Impacts_TimeSeries_Stacked_' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.png'
            # include figure in logfile:
            fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
            # comment out to save disk space in archive:
            fig.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
            Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
            Figurecounter += 1
     
        
    # Area plot for MATERIAL FOOTPRINT of the entire system.
    Area3   = ['Biomass','Non-metallic minerals','Metal ores','Fossil fuels','All materials']
    
    for mS in range(0,NS): # SSP
        for mR in range(0,NR): # RCP
        
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
            ProxyHandlesList = []   # For legend     
            
            # plot area
            ax1.fill_between(np.arange(2016,2061), np.zeros((Nt-1)), Impacts_System_3579di[Biomass_loc,1::,mS,mR], linestyle = '-', facecolor = 'saddlebrown', linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc='saddlebrown')) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061), Impacts_System_3579di[Biomass_loc,1::,mS,mR], Impacts_System_3579di[Biomass_loc,1::,mS,mR] + Impacts_System_3579di[nMetOres_loc,1::,mS,mR], linestyle = '-', facecolor = 'tan', linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc='tan')) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061), Impacts_System_3579di[Biomass_loc,1::,mS,mR] + Impacts_System_3579di[nMetOres_loc,1::,mS,mR], Impacts_System_3579di[Biomass_loc,1::,mS,mR] + Impacts_System_3579di[nMetOres_loc,1::,mS,mR] + Impacts_System_3579di[MetOres_loc,1::,mS,mR], linestyle = '-', facecolor = 'silver', linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc='silver')) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061), Impacts_System_3579di[Biomass_loc,1::,mS,mR] + Impacts_System_3579di[nMetOres_loc,1::,mS,mR] + Impacts_System_3579di[MetOres_loc,1::,mS,mR],  Impacts_System_3579di[Biomass_loc,1::,mS,mR] + Impacts_System_3579di[nMetOres_loc,1::,mS,mR] + Impacts_System_3579di[MetOres_loc,1::,mS,mR] + Impacts_System_3579di[FosFuel_loc,1::,mS,mR], linestyle = '-', facecolor = 'darkcyan', linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc='darkcyan')) # create proxy artist for legend    
            plt.plot(np.arange(2016,2061), Impacts_System_3579di[AllMat_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            plta = Line2D(np.arange(2016,2061), Impacts_System_3579di[AllMat_loc,1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            ProxyHandlesList.append(plta) # create proxy artist for legend    
            
            plt.title('Raw material input for engineering materials and energy supply, by category, \n' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.', fontsize = 18)
            plt.ylabel('Mt of raw materials', fontsize = 18)
            plt.xlabel('Year', fontsize = 18)
            plt.xticks(fontsize=18)
            plt.yticks(fontsize=18)
            plt.legend(handles = reversed(ProxyHandlesList),labels = reversed(Area3), shadow = False, prop={'size':14},ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
            ax1.set_xlim([2016, 2060])
            
            plt.show()
            fig_name = 'MaterialInput_TimeSeries_Stacked_' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.png'
            # include figure in logfile:
            fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
            # comment out to save disk space in archive:
            fig.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
            Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
            Figurecounter += 1
        
    
    ### 5.2) Export to Excel
    Mylog.info('### 5.2 - Export to Excel')
    # Export list data
    book.save( os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_ModelResults_'      + ScriptConfig['Current_UUID'] + '.xlsx'))
    book2.save(os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_Additional_Results_'+ ScriptConfig['Current_UUID'] + '.xlsx'))
    
    # Export area plot as multi-index Excel file, add to existing pandas export file
    ColIndex       = [str(mmx) for mmx in  range(2015,2061)]
    RowIndex       = pd.MultiIndex.from_product([['use phase','use phase, scope 2 (el)','use phase, other indirect','primary material product.','manufact. & recycling','total (+ forest sequestr.)'],['LED','SSP1','SSP2'],['Base','RCP2_6']], names=('System scope','SSP','RCP'))
    DF_GHGA_global = pd.DataFrame(np.einsum('SRts->sSRt',DataAExp).reshape(36,46), index=RowIndex, columns=ColIndex)
    #DF_GHGA_global.to_excel(os.path.join(ProjectSpecs_Path_Result,'GHG_Area_Data.xlsx'), merge_cells=False)    
    
    # Export total material stock
    ColIndex_m     = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items
    ColIndex_t     = IndexTable.Classification[IndexTable.index.get_loc('Time')].Items
    ColIndex_c     = IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items
    if 'pav' in SectorList:
        RowIndex        = pd.MultiIndex.from_product([IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items,IndexTable.Classification[IndexTable.index.get_loc('Cars')].Items], names=('Region','Stock_Item'))
        DF_matstock2015 = pd.DataFrame(np.einsum('mpr->rpm',TotalMaterialStock_2015_pav).reshape(Nr*Np,Nm), index=RowIndex, columns=ColIndex_m)
        DF_matstock2015.to_excel(pd_xlsx_writer, sheet_name="RECC_matstock_2015_pav_Mt", merge_cells=False) 
    if 'reb' in SectorList:
        RowIndex        = pd.MultiIndex.from_product([IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items,IndexTable.Classification[IndexTable.index.get_loc('ResidentialBuildings')].Items], names=('Region','Stock_Item'))
        DF_matstock2015 = pd.DataFrame(np.einsum('mBr->rBm',TotalMaterialStock_2015_reb).reshape(Nr*NB,Nm), index=RowIndex, columns=ColIndex_m)
        DF_matstock2015.to_excel(pd_xlsx_writer, sheet_name="RECC_matstock_2015_reb_Mt", merge_cells=False)
        # Export pre 2021 age-cohorts building stock, customized export for Jan Streeck building stock comparison
        DF_2020_lockin  = pd.DataFrame(np.einsum('tBr->rBt',Stock_2020_decline_B).reshape(Nr*NB,Nt), index=RowIndex, columns=ColIndex_t)
        DF_2020_lockin.to_excel(pd_xlsx_writer, sheet_name="RECC_2020_lockin_reb_Mm2", merge_cells=False)            
        DF_2020_agestru = pd.DataFrame(np.einsum('cBr->rBc',Stock_2020_agestruct_B).reshape(Nr*NB,Nc), index=RowIndex, columns=ColIndex_c)
        DF_2020_agestru.to_excel(pd_xlsx_writer, sheet_name="RECC_2020_AgeStructure_reb_Mm2", merge_cells=False) 
        RowIndex        = pd.MultiIndex.from_product([IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items,IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items], names=('Region','Material'))
        DF_2020_materia = pd.DataFrame(np.einsum('Btcrm->rmt',RECC_System.StockDict['S_7'].Values[:,0:Ind_2020+1,:,Sector_reb_rge,:,0]).reshape(Nr*Nm,Nt), index=RowIndex, columns=ColIndex_t)
        DF_2020_materia.to_excel(pd_xlsx_writer, sheet_name="RECC_2020_lockin_reb_Mt", merge_cells=False)            
        DF_2020_agestrm = pd.DataFrame(np.einsum('Bcrm->rmc',RECC_System.StockDict['S_7'].Values[Ind_2020-SwitchTime,:,:,Sector_reb_rge,:,0]).reshape(Nr*Nm,Nc), index=RowIndex, columns=ColIndex_c)
        DF_2020_agestrm.to_excel(pd_xlsx_writer, sheet_name="RECC_2020_AgeStructure_reb_Mt", merge_cells=False) 
    if 'nrb' in SectorList:
        RowIndex        = pd.MultiIndex.from_product([IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items,IndexTable.Classification[IndexTable.index.get_loc('NonresidentialBuildings')].Items], names=('Region','Stock_Item'))
        DF_matstock2015 = pd.DataFrame(np.einsum('mNr->rNm',TotalMaterialStock_2015_nrb).reshape(Nr*NN,Nm), index=RowIndex, columns=ColIndex_m)
        DF_matstock2015.to_excel(pd_xlsx_writer, sheet_name="RECC_matstock_2015_nrb_Mt", merge_cells=False)
    
    pd_xlsx_writer.close()
    
    ## 5.3) Export as .mat file
    # Not implemented.
    
    ### 5.4) Model run is finished. Wrap up.
    Mylog.info('### 5.5 - Finishing')
    Mylog.debug("Converting " + os.path.join(ProjectSpecs_Path_Result, '..', log_filename))
    # everything from here on will not be included in the converted log file
    msf.convert_log(os.path.join(ProjectSpecs_Path_Result, log_filename))
    Mylog.info('Script is finished. Terminating logging process and closing all log files.')
    Time_End = time.time()
    Time_Duration = Time_End - Time_Start
    Mylog.info('End of simulation: ' + time.asctime())
    Mylog.info('Duration of simulation: %.1f seconds.' % Time_Duration)
    
    # remove all handlers from logger
    root = log.getLogger()
    root.handlers = []  # required if you don't want to exit the shell
    log.shutdown()
    
    ### 5.5) Create descriptive folder name and rename result folder
    SectList    = eval(ScriptConfig['SectorSelect'])
    DescrString = '__'
    FirstFlag   = True
    for sect in SectList:
        if FirstFlag is True:
            DescrString += sect
            FirstFlag = False
        else:
            DescrString += '_'
            DescrString += sect
    DescrString += '__'        
    
    REStratList = []
    if ScriptConfig['Include_REStrategy_FabYieldImprovement'] == 'True':
        REStratList.append('FYI')
    if ScriptConfig['Include_REStrategy_FabScrapDiversion'] == 'True':
        REStratList.append('FSD')    
    if ScriptConfig['Include_REStrategy_EoL_RR_Improvement'] == 'True':
        REStratList.append('EoL')
    if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'True':
        REStratList.append('MSU')
    if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True':
        REStratList.append('ULD')
    if ScriptConfig['Include_REStrategy_ReUse'] == 'True':
        REStratList.append('RUS')
    if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
        REStratList.append('LTE')
    if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True':
        REStratList.append('MIU')
    if ScriptConfig['Include_REStrategy_CarSharing'] == 'True':
        REStratList.append('CaS')
    if ScriptConfig['Include_REStrategy_RideSharing'] == 'True':
        REStratList.append('RiS')
    if ScriptConfig['IncludeRecycling'] == 'False':
        REStratList.append('NoR')
    if ScriptConfig['No_EE_Improvements'] == 'True':        
        REStratList.append('NoEE')
        
    FirstFlag = True
    if len(REStratList) > 0:
        for REStrat in REStratList:
            if FirstFlag is True:
                DescrString += REStrat
                FirstFlag = False
            else:
                DescrString += '_'
                DescrString += REStrat        
        
    ProjectSpecs_Path_Result_New = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString + DescrString)
    try:
        os.rename(ProjectSpecs_Path_Result,ProjectSpecs_Path_Result_New)
    except:
        Mylog.info('Folder file not renamed. Acces is denied')
            
    
    print('done.')
    
    OutputDict['Name_Scenario'] = Name_Scenario + '__' + TimeString + DescrString # return new scenario folder name to ScenarioControl script
        
    return OutputDict
                    
# code for script to be run as standalone function
#if __name__ == "__main__":
#    main()


# The End.
