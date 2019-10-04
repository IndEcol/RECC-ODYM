# -*- coding: utf-8 -*-
"""
Created on Fri Jun 14 05:18:48 2019

@author: spauliuk
"""

"""

File RECC_ScenarioEvaluate.py

Script that runs the sensitivity and scnenario comparison scripts for different settings.

"""

# Import required libraries:
import os
import xlrd
import openpyxl

import RECC_Paths # Import path file
import RECC_G7IC_Cascade_PassVehicles_V2_2
import RECC_G7IC_Cascade_ResBuildings_V2_2
import RECC_G7IC_Sensitivity_PassVehicles_V2_2
import RECC_G7IC_Sensitivity_ResBuildings_V2_2

#ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
ScenarioSetting = 'Evaluate_RECC_Cascade'
#ScenarioSetting = 'Evaluate_GroupTestRun'
#ScenarioSetting = 'Evaluate_RECC_Sensitivity'

# open scenario sheet
ModelConfigListFile  = xlrd.open_workbook(os.path.join(RECC_Paths.recc_path,'RECC_ModelConfig_List_V2_2.xlsx'))
ModelEvalListSheet = ModelConfigListFile.sheet_by_name(ScenarioSetting)

# open result summary file
mywb  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_Template.xlsx')) # for total emissions
mywb2 = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_Template.xlsx')) # for material-related emissions
mywb3 = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_Template.xlsx')) # for material-related emissions with recycling credit

#Read control lines and execute main model script
Row  = 1

NoofCascadeSteps_pav     = 7 # 7 for vehs. and 6 for buildings
NoofCascadeSteps_reb     = 6 # 7 for vehs. and 6 for buildings
NoofSensitivitySteps_pav = 11 # no of different sensitivity analysis cases for pav
NoofSensitivitySteps_reb = 10 # no of different sensitivity analysis cases for reb

# search for script config list entry
while ModelEvalListSheet.cell_value(Row, 1) != 'ENDOFLIST':
    if ModelEvalListSheet.cell_value(Row, 1) != '':
        PassVehList = []
        ResBldsList = []
        RegionalScope = ModelEvalListSheet.cell_value(Row, 1)
        Setting       = ModelEvalListSheet.cell_value(Row, 2) # cascade or sensitivity
        print(RegionalScope)
        
    if Setting == 'Cascade_pav':
        for m in range(0,int(NoofCascadeSteps_pav)):
            PassVehList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC scenario comparison
        ASummaryV, AvgDecadalEmsV, MatSummaryV, AvgDecadalMatEmsV, MatSummaryVC, AvgDecadalMatEmsVC = RECC_G7IC_Cascade_PassVehicles_V2_2.main(RegionalScope,PassVehList)
        # write results summary to Excel
        Vsheet = mywb[RegionalScope + '_Vehicles']
        for r in range(0,3):
            for c in range(0,7):
                Vsheet.cell(row = r+3,  column = c+5).value  = ASummaryV[r,c]       
                Vsheet.cell(row = r+9,  column = c+5).value  = ASummaryV[r+3,c]       
                Vsheet.cell(row = r+15, column = c+5).value  = ASummaryV[r+6,c]
                Vsheet.cell(row = r+38, column = c+5).value  = ASummaryV[r+9,c]
                for d in range(0,4):
                    Vsheet.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalEmsV[r,c,d]
        Vsheet2 = mywb2[RegionalScope + '_Vehicles']
        for r in range(0,3):
            Vsheet2.cell(row = r+3,  column = 4).value  = 0       
            Vsheet2.cell(row = r+9,  column = 4).value  = 0       
            for c in range(0,7):
                Vsheet2.cell(row = r+3,  column = c+5).value  = MatSummaryV[r,c]       
                Vsheet2.cell(row = r+9,  column = c+5).value  = MatSummaryV[r+3,c]       
                Vsheet2.cell(row = r+15, column = c+5).value  = MatSummaryV[r+6,c]
                Vsheet2.cell(row = r+38, column = c+5).value  = MatSummaryV[r+9,c]
                for d in range(0,4):
                    Vsheet2.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalMatEmsV[r,c,d]                    
        Vsheet3 = mywb3[RegionalScope + '_Vehicles']
        for r in range(0,3):
            Vsheet3.cell(row = r+3,  column = 4).value  = 0       
            Vsheet3.cell(row = r+9,  column = 4).value  = 0       
            for c in range(0,7):
                Vsheet3.cell(row = r+3,  column = c+5).value  = MatSummaryVC[r,c]       
                Vsheet3.cell(row = r+9,  column = c+5).value  = MatSummaryVC[r+3,c]       
                Vsheet3.cell(row = r+15, column = c+5).value  = MatSummaryVC[r+6,c]
                Vsheet3.cell(row = r+38, column = c+5).value  = MatSummaryVC[r+9,c]
                for d in range(0,4):
                    Vsheet3.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalMatEmsVC[r,c,d]          
        
    if Setting == 'Cascade_reb':
        for m in range(0,int(NoofCascadeSteps_reb)):
            ResBldsList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC scenario comparison
        ASummaryB, AvgDecadalEmsB, MatSummaryB, AvgDecadalMatEmsB, MatSummaryBC, AvgDecadalMatEmsBC = RECC_G7IC_Cascade_ResBuildings_V2_2.main(RegionalScope,ResBldsList)
        # write results summary to Excel
        Bsheet = mywb[RegionalScope + '_Buildings']
        for r in range(0,3):
            for c in range(0,6):
                Bsheet.cell(row = r+3, column = c+5).value  = ASummaryB[r,c]       
                Bsheet.cell(row = r+9, column = c+5).value  = ASummaryB[r+3,c]       
                Bsheet.cell(row = r+15, column = c+5).value = ASummaryB[r+6,c]     
                Bsheet.cell(row = r+38, column = c+5).value = ASummaryB[r+9,c]     
                for d in range(0,4):
                    Bsheet.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalEmsB[r,c,d]
        Bsheet2 = mywb2[RegionalScope + '_Buildings']
        for r in range(0,3):
            Bsheet2.cell(row = r+3,  column = 4).value  = 0       
            Bsheet2.cell(row = r+9,  column = 4).value  = 0                   
            for c in range(0,6):
                Bsheet2.cell(row = r+3,  column = c+5).value  = MatSummaryB[r,c]       
                Bsheet2.cell(row = r+9,  column = c+5).value  = MatSummaryB[r+3,c]       
                Bsheet2.cell(row = r+15, column = c+5).value  = MatSummaryB[r+6,c]
                Bsheet2.cell(row = r+38, column = c+5).value  = MatSummaryB[r+9,c]
                for d in range(0,4):
                    Bsheet2.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalMatEmsB[r,c,d]    
        Bsheet3 = mywb3[RegionalScope + '_Buildings']
        for r in range(0,3):
            Bsheet3.cell(row = r+3,  column = 4).value  = 0       
            Bsheet3.cell(row = r+9,  column = 4).value  = 0                   
            for c in range(0,6):
                Bsheet3.cell(row = r+3,  column = c+5).value  = MatSummaryBC[r,c]       
                Bsheet3.cell(row = r+9,  column = c+5).value  = MatSummaryBC[r+3,c]       
                Bsheet3.cell(row = r+15, column = c+5).value  = MatSummaryBC[r+6,c]
                Bsheet3.cell(row = r+38, column = c+5).value  = MatSummaryBC[r+9,c]
                for d in range(0,4):
                    Bsheet3.cell(row = d*3 + r + 21,column = c+5).value  = AvgDecadalMatEmsBC[r,c,d]                      
                    
    if Setting == 'Sensitivity_pav':
        for m in range(0,int(NoofSensitivitySteps_pav)):
            PassVehList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC sensitivity analysis for pav
        CumEmsV_Sens, AnnEmsV2030_Sens, AnnEmsV2050_Sens, AvgDecadalEms, MatCumEmsV_Sens, MatAnnEmsV2030_Sens, MatAnnEmsV2050_Sens, MatAvgDecadalEmsV, MatCumEmsV_SensC, MatAnnEmsV2030_SensC, MatAnnEmsV2050_SensC, MatAvgDecadalEmsC, CumEmsV_Sens2060, MatCumEmsV_Sens2060, MatCumEmsV_SensC2060 = RECC_G7IC_Sensitivity_PassVehicles_V2_2.main(RegionalScope,PassVehList)        
        # write results summary to Excel
        Ssheet = mywb['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,11):
                Ssheet.cell(row = r+4, column = c+6).value  = AnnEmsV2030_Sens[r,c]
                Ssheet.cell(row = r+9, column = c+6).value  = AnnEmsV2050_Sens[r,c]
                Ssheet.cell(row = r+14,column = c+6).value  = CumEmsV_Sens[r,c]
                Ssheet.cell(row = r+68,column = c+6).value  = CumEmsV_Sens2060[r,c]
                for d in range(0,4):
                    Ssheet.cell(row = d*3 + r + 19,column = c+6).value  = AvgDecadalEms[r,c,d]
        Ssheet2 = mywb2['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,11):
                Ssheet2.cell(row = r+4, column = c+6).value  = MatAnnEmsV2030_Sens[r,c]
                Ssheet2.cell(row = r+9, column = c+6).value  = MatAnnEmsV2050_Sens[r,c]
                Ssheet2.cell(row = r+14,column = c+6).value  = MatCumEmsV_Sens[r,c]
                Ssheet2.cell(row = r+68,column = c+6).value  = MatCumEmsV_Sens2060[r,c]
                for d in range(0,4):
                    Ssheet2.cell(row = d*3 + r + 19,column = c+6).value  = MatAvgDecadalEmsV[r,c,d]       
        Ssheet3 = mywb3['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,11):
                Ssheet3.cell(row = r+4, column = c+6).value  = MatAnnEmsV2030_SensC[r,c]
                Ssheet3.cell(row = r+9, column = c+6).value  = MatAnnEmsV2050_SensC[r,c]
                Ssheet3.cell(row = r+14,column = c+6).value  = MatCumEmsV_SensC[r,c]
                Ssheet3.cell(row = r+68,column = c+6).value  = MatCumEmsV_SensC2060[r,c]
                for d in range(0,4):
                    Ssheet3.cell(row = d*3 + r + 19,column = c+6).value  = MatAvgDecadalEmsC[r,c,d]                        
                    
    if Setting == 'Sensitivity_reb':
        for m in range(0,int(NoofSensitivitySteps_reb)):
            ResBldsList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC sensitivity analysis for reb
        CumEmsB_Sens, AnnEmsB2030_Sens, AnnEmsB2050_Sens, AvgDecadalEms, MatCumEmsB_Sens, MatAnnEmsB2030_Sens, MatAnnEmsB2050_Sens, MatAvgDecadalEmsB, MatCumEmsV_SensC, MatAnnEmsV2030_SensC, MatAnnEmsV2050_SensC, MatAvgDecadalEmsC, CumEmsB_Sens2060, MatCumEmsV_Sens2060, MatCumEmsV_SensC2060 = RECC_G7IC_Sensitivity_ResBuildings_V2_2.main(RegionalScope,ResBldsList)        
        # write results summary to Excel
        Ssheet = mywb['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,10):
                Ssheet.cell(row = r+35,column = c+6).value  = AnnEmsB2030_Sens[r,c]
                Ssheet.cell(row = r+40,column = c+6).value  = AnnEmsB2050_Sens[r,c]
                Ssheet.cell(row = r+45,column = c+6).value  = CumEmsB_Sens[r,c]        
                Ssheet.cell(row = r+73,column = c+6).value  = CumEmsB_Sens2060[r,c]        
                for d in range(0,4):
                    Ssheet.cell(row = d*3 + r + 50,column = c+6).value  = AvgDecadalEms[r,c,d]   
        Ssheet2 = mywb2['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,10):
                Ssheet2.cell(row = r+35, column = c+6).value  = MatAnnEmsB2030_Sens[r,c]
                Ssheet2.cell(row = r+40, column = c+6).value  = MatAnnEmsB2050_Sens[r,c]
                Ssheet2.cell(row = r+45,column = c+6).value   = MatCumEmsB_Sens[r,c]
                Ssheet2.cell(row = r+73,column = c+6).value   = MatCumEmsV_Sens2060[r,c]        
                for d in range(0,4):
                    Ssheet2.cell(row = d*3 + r + 50,column = c+6).value  = MatAvgDecadalEmsB[r,c,d]                      
        Ssheet3 = mywb3['Sensitivity_' + RegionalScope]
        print(RegionalScope)
        for r in range(0,3):
            for c in range(0,10):
                Ssheet3.cell(row = r+35, column = c+6).value  = MatAnnEmsV2030_SensC[r,c]
                Ssheet3.cell(row = r+40, column = c+6).value  = MatAnnEmsV2050_SensC[r,c]
                Ssheet3.cell(row = r+45,column = c+6).value   = MatCumEmsV_SensC[r,c]
                Ssheet3.cell(row = r+73,column = c+6).value   = MatCumEmsV_SensC2060[r,c]        
                for d in range(0,4):
                    Ssheet3.cell(row = d*3 + r + 50,column = c+6).value  = MatAvgDecadalEmsC[r,c,d] 
                    
    # forward counter   
    if Setting == 'Cascade_pav':
        Row += NoofCascadeSteps_pav
    if Setting == 'Cascade_reb':
        Row += NoofCascadeSteps_reb        
    if Setting == 'Sensitivity_pav':
        Row += NoofSensitivitySteps_pav
    if Setting == 'Sensitivity_reb':
        Row += NoofSensitivitySteps_reb
        
    
mywb.save(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_SystemGHG_04_October_2019.xlsx'))
mywb2.save(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_MaterialGHG_04_October_2019.xlsx'))    
mywb3.save(os.path.join(RECC_Paths.results_path,'G7_RECC_Results_MaterialGHG_inclRecyclingCredit_04_October_2019.xlsx'))        
    
    
#
#