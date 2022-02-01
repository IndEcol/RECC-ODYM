# -*- coding: utf-8 -*-
"""
Created on Wed Oct 17 10:37:00 2018

@author: spauliuk
"""
def main(RegionalScope,SectorString,ThreeSectoList_Export,SingleSectList,Current_UUID):
    # ThreeSectoList_Export: List of cascade scenarios
    # SingleSectList: List of counterfactual no EE etc. scenarios
    import numpy as np
    import openpyxl
    import matplotlib.pyplot as plt  
    import pylab
    import os
    import pandas as pd
    import RECC_Paths # Import path file   
    
    RECC_Paths.results_path_save = os.path.join(RECC_Paths.results_path_eval,'RECC_Results_' + Current_UUID)
    
    # Scenario order:
    # 1) No energy efficiency, no climate policy scenario
    # 2) No ME, no  climate policy scenario
    # 3) No ME, RCP 2.6 scenario
    # 4) Industrial ME only
    # 5) ALL, including sufficiency-based ME
    
    Region      = RegionalScope
    # first scenario in list below is for center of Eeff-EST-ME cascade, following four scenarios are for right side, last four for left side.
    FolderlistV = [SingleSectList[0],\
                   ThreeSectoList_Export[0],ThreeSectoList_Export[0],SingleSectList[4],ThreeSectoList_Export[-1],\
                   SingleSectList[1],SingleSectList[3],ThreeSectoList_Export[-1],ThreeSectoList_Export[-1]]
    # First line: # no EE, no climpol.  Second line:  # RE Cascade baseline: NoClimPol, RE Cascade baseline: RCP2.6, Industrial ME: FYI_FSD_EoL_MSU_RUS, Full ME.
    
    # Color definition
    MyColorCycle = pylab.cm.tab20(np.arange(0,1,0.05)) # Select 20 colors from the 'tab20' color map.  
    MyColorCycle[4,:]  = np.array([0.960784314, 0.949019608, 0.51372549, 1]) # https://colorbrewer2.org/#type=diverging&scheme=RdYlBu&n=5
    MyColorCycle[0,:]  = np.array([0.17254902,  0.482352941, 0.71372549, 1])
    MyColorCycle[3,:]  = np.array([0.992156863, 0.682352941, 0.380392157,1])
    MyColorCycle[6,:]  = np.array([0.843137255, 0.098039216, 0.109803922,1])
    BaseBrown    = np.array([0.749,0.506,0.176,1])     # Base for GHG before ME reduction
    BaseBlue     = np.array([0.208,0.592,0.561,1])     # Base for GHG after full ME reduction, instread of MyColorCycle[11,:]
    LabelColors  = ['k','k','k','k','k',BaseBlue,BaseBlue]
    
    # read data.
    # get result items:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,ThreeSectoList_Export[0])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,ThreeSectoList_Export[0],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    # Find the index for sysem-wide emissions, the recycling credit and others:
    swe = 1    
    while True:
        if Resultsheet2.cell(swe+1, 1).value == 'GHG emissions, system-wide _3579di':
            break # that gives us the right index to read the recycling credit from the result table.
        swe += 1  
    
    NS = 3  # no of SSP scenarios
    NR = 2  # no of RCP scenarios
    NE = 9  # no of Res. eff. scenarios
    
    CumEmsV           = np.zeros((NS,NR,NE))   # SSP-Scenario x RCP scenario x RES scenario
    CumEmsV2060       = np.zeros((NS,NR,NE))   # SSP-Scenario x RCP scenario x RES scenario
    AnnEmsV2030       = np.zeros((NS,NR,NE))   # SSP-Scenario x RCP scenario x RES scenario
    AnnEmsV2050       = np.zeros((NS,NR,NE))   # SSP-Scenario x RCP scenario x RES scenario
    AvgDecadalEmsV    = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario
    
    for r in range(0,NE): # RE scenario
        # import system-wide GHG and material-related emissions
        ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderlistV[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
        Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderlistV[r],ResFile[0]))
        Resultsheet2 = Resultfile2['Model_Results']    
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    CumEmsV[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    CumEmsV2060[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value                    
                AnnEmsV2030[s,c,r]    = Resultsheet2.cell(swe+ 2*s +c+1,23).value
                AnnEmsV2050[s,c,r]    = Resultsheet2.cell(swe+ 2*s +c+1,43).value
            AvgDecadalEmsV[s,1,r,0]   = sum([Resultsheet2.cell(swe+ 2*s +1+1,t+1).value for i in range(13,23)])/10
            AvgDecadalEmsV[s,1,r,1]   = sum([Resultsheet2.cell(swe+ 2*s +1+1,t+1).value for i in range(23,33)])/10
            AvgDecadalEmsV[s,1,r,2]   = sum([Resultsheet2.cell(swe+ 2*s +1+1,t+1).value for i in range(33,43)])/10
            AvgDecadalEmsV[s,1,r,3]   = sum([Resultsheet2.cell(swe+ 2*s +1+1,t+1).value for i in range(43,53)])/10      
            AvgDecadalEmsV[s,0,r,0]   = sum([Resultsheet2.cell(swe+ 2*s +0+1,t+1).value for i in range(13,23)])/10
            AvgDecadalEmsV[s,0,r,1]   = sum([Resultsheet2.cell(swe+ 2*s +0+1,t+1).value for i in range(23,33)])/10
            AvgDecadalEmsV[s,0,r,2]   = sum([Resultsheet2.cell(swe+ 2*s +0+1,t+1).value for i in range(33,43)])/10
            AvgDecadalEmsV[s,0,r,3]   = sum([Resultsheet2.cell(swe+ 2*s +0+1,t+1).value for i in range(43,53)])/10                  
    
    Sector = ['suff_eff']
    Title  = ['Cum_GHG_2016_2050_Mt','Cum_GHG_2040_2050_Mt','Annual_GHG_2050_Mt']
    Label  = ['Cum. GHG 2016-2050','Cum. GHG 2040-2050','Annual GHG 2050']
    Scens  = ['LED','SSP1','SSP2']
    LWE    = ['No climate policy','Energy efficiency','Low carbon en. supply', 'Industrial ME','Demand-side ME','Residual emissions','Residual emissions']
    DataAL = ['(1): No energy eff., no clim. policy, no ME (reference)','(2): (1) + energy efficiency','(3): (2) + low carbon en. supply', '(4): (3) + industrial ME','(5): (4) + demand-side ME = residual',\
              '(6): (1) + industrial ME','(7): (6) + demand-side ME','(8): (7) + energy efficiency','(9): (8) + low carbon en. supply = residual'] # Labels for export
        
    DataA  = np.zeros((3,9,3)) # for pandas export
    
    for nn in range(0,3):
        Data = np.zeros((3,9))
        if nn == 0:
            Data[:,0] = CumEmsV[:,0,0].copy() # baseline, center
            Data[:,1] = CumEmsV[:,0,1].copy() # right cascade
            Data[:,2] = CumEmsV[:,1,2].copy() # right cascade
            Data[:,3] = CumEmsV[:,1,3].copy() # right cascade
            Data[:,4] = CumEmsV[:,1,4].copy() # right cascade
            Data[:,5] = CumEmsV[:,0,5].copy() # left cascade
            Data[:,6] = CumEmsV[:,0,6].copy() # left cascade
            Data[:,7] = CumEmsV[:,0,7].copy() # left cascade
            Data[:,8] = CumEmsV[:,1,8].copy() # left cascade
            DataA[:,:,0] = Data.copy()

        if nn == 1:
            Data[:,0] = 10*AvgDecadalEmsV[:,0,0,2].copy()
            Data[:,1] = 10*AvgDecadalEmsV[:,0,1,2].copy()
            Data[:,2] = 10*AvgDecadalEmsV[:,1,2,2].copy()
            Data[:,3] = 10*AvgDecadalEmsV[:,1,3,2].copy()
            Data[:,4] = 10*AvgDecadalEmsV[:,1,4,2].copy()            
            Data[:,5] = 10*AvgDecadalEmsV[:,0,5,2].copy() # left cascade  
            Data[:,6] = 10*AvgDecadalEmsV[:,0,6,2].copy() # left cascade
            Data[:,7] = 10*AvgDecadalEmsV[:,0,7,2].copy() # left cascade
            Data[:,8] = 10*AvgDecadalEmsV[:,1,8,2].copy() # left cascade
            DataA[:,:,1] = Data.copy()
            
        if nn == 2:
            Data[:,0] = AnnEmsV2050[:,0,0].copy()
            Data[:,1] = AnnEmsV2050[:,0,1].copy()
            Data[:,2] = AnnEmsV2050[:,1,2].copy()
            Data[:,3] = AnnEmsV2050[:,1,3].copy()
            Data[:,4] = AnnEmsV2050[:,1,4].copy()
            Data[:,5] = AnnEmsV2050[:,0,5].copy() # left cascade
            Data[:,6] = AnnEmsV2050[:,0,6].copy() # left cascade
            Data[:,7] = AnnEmsV2050[:,0,7].copy() # left cascade
            Data[:,8] = AnnEmsV2050[:,1,8].copy() # left cascade
            DataA[:,:,2] = Data.copy()
            
        Data  = Data  / 1000 # Mt -> Gt conversion            
            
        # stacked bar plot with EE-EST-ME cascade on the right side                 
        Xoffs1 = [1,3,5]
        Xoffs2 = [1.7,3.7,5.7]
        bw = 0.5
        # fine-tune Y_max and y axis scaling for annual GHG plot:
        if nn == 2:
            if RegionalScope   == 'Global'       and SectorString == 'pav_reb':
                YMax = 20
            elif RegionalScope == 'Global_North' and SectorString == 'pav_reb':
                YMax = 10
            elif RegionalScope == 'Global_South' and SectorString == 'pav_reb':
                YMax = 10       
            else:
                YMax = Data[2,0] * 1.1        
        else:
            YMax = Data[2,0] * 1.1        
                
        fig  = plt.figure(figsize=(5,8))
        ax1  = plt.axes([0.08,0.08,0.85,0.9])
        
        for ms in range (0,NS):
            # plot bars
            ax1.fill_between([Xoffs1[ms],Xoffs1[ms]+bw], [0,0],[Data[ms,0],Data[ms,0]],linestyle = '--', facecolor =BaseBrown, linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,1],Data[ms,1]],[Data[ms,0],Data[ms,0]],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,2],Data[ms,2]],[Data[ms,1],Data[ms,1]],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,3],Data[ms,3]],[Data[ms,2],Data[ms,2]],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,4],Data[ms,4]],[Data[ms,3],Data[ms,3]],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [0,0],[Data[ms,4],Data[ms,4]],linestyle = '--', facecolor =MyColorCycle[15,:], linewidth = 0.0)
            if ms == 1: 
                ProxyHandlesList = []   # For legend     
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=BaseBrown)) # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[0,:])) # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[6,:])) # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[15,:])) # create proxy artist for legend

        # plot text and labels
        #plt.text(6.85, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
        #plt.text(4.3, 0.94  *Right, Scens[m],fontsize=18,fontweight='bold') 
        #plt.title('Energy, efficiency, and sufficiency, ' + Sector[0] + '.', fontsize = 18)
        plt.ylabel(Label[nn] + r', Gt CO$_2$-eq', fontsize = 18)
        plt.xticks([1.6,3.6,5.6])
        plt.yticks(fontsize =18)
        ax1.set_xticklabels(Scens, rotation =0, fontsize = 21, fontweight = 'normal')
        leg = plt.legend(handles = ProxyHandlesList,labels = LWE,shadow = False, prop={'size':12},ncol=1, loc = 'upper left' ) 
        # Change text color:
        mc = 0
        for text in leg.get_texts():
            plt.setp(text, color = LabelColors[mc])
            mc +=1
        #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
        plt.axis([-0.2, 7, 0, YMax])
    
        plt.show()
        fig_name = Title[nn] + Region + '_' + SectorString + '_' + Sector[0] + '_rightcascade.png'
        fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 500, bbox_inches='tight')    
        fig_name = Title[nn] + Region + '_' + SectorString + '_' + Sector[0] + '_rightcascade.svg'
        fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 500, bbox_inches='tight')   
        
        # stacked bar plot with EE-EST-ME cascades on both sides
        Xoffs0 = [1.0,3.7,6.4]    # left cascade
        Xoffs2 = [2.05,4.75,7.45] # right cascade
        bw = 0.85
        # fine-tune Y_max and y axis scaling for annual GHG plot:
        if nn == 2:
            if RegionalScope == 'Global' and SectorString == 'pav_reb':
                YMax = 20
            elif RegionalScope == 'Global_North' and SectorString == 'pav_reb':
                YMax = 10
            elif RegionalScope == 'Global_South' and SectorString == 'pav_reb':
                YMax = 10      
            else:
                YMax = Data[2,0] * 1.1 
        else:
            YMax = Data[2,0] * 1.1  
            
        fig  = plt.figure(figsize=(5,8))
        ax1  = plt.axes([0.08,0.08,0.85,0.9])
        
        for ms in range (0,NS):
            # plot line
            ax1.plot([Xoffs0[ms],Xoffs2[ms]+bw],[Data[ms,0],Data[ms,0]],color ='k', linewidth = 2.0)
            ax1.plot([Xoffs0[ms],Xoffs2[ms]+bw],[Data[ms,4]-0.023,Data[ms,4]-0.023],color =BaseBlue, linewidth = 2.0)
            # plot bars
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,1],Data[ms,1]],[Data[ms,0],Data[ms,0]],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0) # right
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,2],Data[ms,2]],[Data[ms,1],Data[ms,1]],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,3],Data[ms,3]],[Data[ms,2],Data[ms,2]],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [Data[ms,4],Data[ms,4]],[Data[ms,3],Data[ms,3]],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
            ax1.fill_between([Xoffs2[ms],Xoffs2[ms]+bw], [0,0],[Data[ms,4],Data[ms,4]],linestyle = '--', facecolor =MyColorCycle[15,:], linewidth = 0.0)

            ax1.fill_between([Xoffs0[ms],Xoffs0[ms]+bw], [Data[ms,5],Data[ms,5]],[Data[ms,0],Data[ms,0]],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0) # left
            ax1.fill_between([Xoffs0[ms],Xoffs0[ms]+bw], [Data[ms,6],Data[ms,6]],[Data[ms,5],Data[ms,5]],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
            ax1.fill_between([Xoffs0[ms],Xoffs0[ms]+bw], [Data[ms,7],Data[ms,7]],[Data[ms,6],Data[ms,6]],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0)
            ax1.fill_between([Xoffs0[ms],Xoffs0[ms]+bw], [Data[ms,8],Data[ms,8]],[Data[ms,7],Data[ms,7]],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
            ax1.fill_between([Xoffs0[ms],Xoffs0[ms]+bw], [0,0],[Data[ms,8],Data[ms,8]],linestyle = '--', facecolor =MyColorCycle[15,:], linewidth = 0.0)

            if ms == 1: 
                ProxyHandlesList = []   # For legend     
                PltLegx, = plt.plot([0,0],[1,1], color ='k', linewidth = 2.0) # create proxy artist for legend
                ProxyHandlesList.append(PltLegx)
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:]))  # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[0,:]))  # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:]))  # create proxy artist for legend
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[6,:]))  # create proxy artist for legend
                PltLegx, = plt.plot([0,0],[1,1], color = BaseBlue, linewidth = 2.0) # create proxy artist for legend
                ProxyHandlesList.append(PltLegx)
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[15,:])) # create proxy artist for legend
                
            if ms == 0: 
                # plot ME sequence labels:
                plt.text(1.2,  Data[0,0] * 1.4, 'ME first',rotation = 90, fontsize = 18,fontweight='normal')          
                plt.text(2.25, Data[0,0] * 1.4, 'ME last', rotation = 90, fontsize = 18,fontweight='normal')          
                # plot arrow sequence:
                ax1.plot([Xoffs0[ms]-bw*0.8, Xoffs2[ms]+bw], [Data[ms,0], Data[ms,0]], color ='k', linewidth = 2.0)
                plt.arrow(Xoffs0[ms]-bw*0.6,Data[0,0],0,Data[0,5]-Data[0,0], lw = 1.2, ls = '-', shape = 'full',
                  length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                plt.arrow(Xoffs0[ms]-bw*0.6,Data[0,5],0,Data[0,6]-Data[0,5], lw = 1.2, ls = '-', shape = 'full',
                  length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                plt.arrow(Xoffs0[ms]-bw*0.6,Data[0,6],0,Data[0,7]-Data[0,6], lw = 1.2, ls = '-', shape = 'full',
                  length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                plt.arrow(Xoffs0[ms]-bw*0.6,Data[0,7],0,Data[0,4]-Data[0,7], lw = 1.2, ls = '-', shape = 'full',
                  length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                ax1.plot([Xoffs0[ms]-bw*0.8, Xoffs2[ms]+bw], [Data[ms,4]-0.023, Data[ms,4]-0.023], color = BaseBlue, linewidth = 2.0)

        plt.ylabel(Label[nn] + r', Gt CO$_2$-eq', fontsize = 18)
        plt.xticks([1.95,4.65,7.35])
        plt.yticks(fontsize =18)
        ax1.set_xticklabels(Scens, rotation =0, fontsize = 21, fontweight = 'normal')
        leg = plt.legend(handles = ProxyHandlesList,labels = LWE,shadow = False, prop={'size':12},ncol=1, loc = 'upper left' ) 
        # Change text color:
        mc = 0
        for text in leg.get_texts():
            plt.setp(text, color = LabelColors[mc])
            mc +=1
        #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
        plt.axis([+0.0, 9, 0, YMax])
    
        plt.show()
        fig_name = Title[nn] + Region + '_' + SectorString + '_' + Sector[0] + '_bothcascades.png'
        fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 500, bbox_inches='tight')    
        fig_name = Title[nn] + Region + '_' + SectorString + '_' + Sector[0] + '_bothcascades.svg'
        fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 500, bbox_inches='tight')     
        
    # Save data to xls
    RowIndex       = pd.MultiIndex.from_product([Title,Scens], names=('GHG metric','SSP'))
    DF_Casc_global = pd.DataFrame(np.einsum('SCM->MSC',DataA).reshape(9,9), index=RowIndex, columns=DataAL)
    DF_Casc_global.to_excel(os.path.join(RECC_Paths.results_path_save,'ME_industry_demand_cascade' + Region + '.xls'), merge_cells=False)
    
    return CumEmsV, CumEmsV2060, AnnEmsV2030, AnnEmsV2050, AvgDecadalEmsV

# code for script to be run as standalone function
if __name__ == "__main__":
    main()
