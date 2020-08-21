# -*- coding: utf-8 -*-
"""
Created on Fri Jun 14 05:18:48 2019

@author: spauliuk
"""

"""
File ODYM_RECC_ScenarioEvaluate_V2_4.py

Script that runs the sensitivity and scnenario comparison scripts for different settings.

Section 1: single sector cascade
Section 2: multi-sector cascade
Section 3: Sensitivity plots
Section 4: Bar plot sufficiency
"""

# Import required libraries:
import os
import xlrd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt 
import pylab
import pandas as pd

import RECC_Paths # Import path file

# The following SINGLE REGION scripts are called whenever there is a single cascade (for reb, pav, ...) or sensitivity analysis for a given region.
import ODYM_RECC_Cascade_V2_4
import ODYM_RECC_BarPlot_ME_Industry_Demand_V2_4
import ODYM_RECC_Sensitivity_V2_4
import ODYM_RECC_Table_Extract_V2_4

# The following ALL REGION scripts are called when ALL 20 world regions are present in the result folder list.

# Define list of 20 regions, to be arranged 5 x 5, and corresponding data containers
Pav_axis_5x5 = [5000,5000,2500,2500,2500,800,800,800,50,50,420,420,420,420,420,200,200,200,200,200,100,100,100,100,100]
Pav_RegionList20 = ['Global','Global_North','Global_South','G7','R32USA', \
'EU28','R32CHN','R5.2SSA_Other','Oth_R32EU12-H','R32EU12-M',\
'R32IND','R5.2OECD_Other','R5.2ASIA_Other','R5.2REF_Other','R5.2MNF_Other',\
'R32CAN','R32JPN','R5.2LAM_Other','Oth_R32EU15','Germany',\
'France','Italy','Poland','Spain','UK']
Pav_RegionList20Plot = ['Global','Glob_North','Glob_South','G7','USA', \
'EU28','China','SSA_Other','Oth_EU12-H','EU12-M',\
'India','OECD_Other','ASIA_Other','REF_Other','MNF_Other',\
'Canada','Japan','LAM_Other','Oth_EU15','Germany',\
'France','Italy','Poland','Spain','UK']

Reb_axis_5x5 = [12000,9000,4500,4500,4500,1500,1500,1500,1500,1500,1200,600,600,300,300,200,200,200,200,200,120,120,120,120,120]
Reb_RegionList20 = ['Global','Global_North','Global_South','R32CHN','G7', \
'EU28','R32USA','R5.2ASIA_Other','R5.2REF_Other','R5.2MNF_Other',\
'R32IND','R5.2SSA_Other','R5.2OECD_Other','Germany','R5.2LAM_Other',\
'R32CAN','UK','France','Oth_R32EU15','R32JPN',\
'Italy','Poland','Spain','Oth_R32EU12-H','R32EU12-M']
Reb_RegionList20Plot = ['Global','Glob_North','Glob_South','China','G7', \
'EU28','USA','ASIA_Other','REF_Other','MNF_Other',\
'India','SSA_Other','OECD_Other','Germany','LAM_Other',\
'Canada','UK','France','Oth_EU15','Japan',\
'Italy','Poland','Spain','Oth_EU12-H','EU12-M']

Pav_axis_7x2        = [5000,5000,2500,2500,800,500,500]
Reb_axis_7x2        = [12000,12000,6000,4000,4000,800,800]
All_RegionList7     = ['Global','Global_North','Global_South','G7','R32CHN','R32IND','R5.2SSA_Other']
All_RegionList7Plot = ['Global','Global_North','Global_South','G7','China','India','Sub-Saharan Africa']

PlotOrder_pav       = [] # Will contain positions of countries/regions in 5x5 plot
PlotOrder_reb       = [] # Will contain positions of countries/regions in 5x5 plot
PlotOrder_7_pav     = [] # Will contain positions of countries/regions in 7x2 plot
PlotOrder_7_reb     = [] # Will contain positions of countries/regions in 7x2 plot
TimeSeries_All      = np.zeros((30,45,25,2,3,2)) # NX x Nt x Nr x NV x NS x NR / indicators x time x regions x sectors x SSP x RCP
# 0: system-wide GHG, no RES 1: system-wide GHG, all RES
# 2: material-related GHG, no RES, 3: material-related GHG, all RES,

PlotExpResolution = 300 # dpi 100 for overview or 500 for paper

# Number of scenarios:
NS = 3 # SSP
NR = 2 # RCP
    
###ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
#ScenarioSetting = 'Evaluate_pav_reb_Cascade' # run eval and plot scripts for selected regions and sectors only
ScenarioSetting = 'Evaluate_pav_reb_Cascade_all' # run eval and plot scripts for all regions and sectors
#ScenarioSetting = 'Germany_detail_evaluate' # run eval and plot scripts for Germany case study only
#ScenarioSetting = 'Evaluate_TestRun' # Test run evaluate

# open scenario sheet
ModelConfigListFile  = xlrd.open_workbook(os.path.join(RECC_Paths.recc_path,'RECC_ModelConfig_List_V2_4.xlsx'))
ModelEvalListSheet   = ModelConfigListFile.sheet_by_name(ScenarioSetting)

# open result summary file
mywb  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'RECC_Global_Results_Template_CascSens.xlsx')) # for total emissions
mywb4 = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,'RECC_Global_Results_Template_Overview.xlsx')) # for emissions to be reported in Tables.

#Read control lines and execute main model script
Row   = 1

Table_Annual  = np.zeros((3,8,NS,NR)) # 2050 annual system emissions, cascade steps x SSP scenarios x RCP scenarios.
Table_CumEms  = np.zeros((3,8,NS,NR)) # 2016-2050 (!) cumulative system emissions, cascade steps x SSP scenarios x RCP scenarios.
MatStocksTab1 = np.zeros((9,6)) # Material stocks for table, LED.
MatStocksTab2 = np.zeros((9,6)) # Material stocks for table, SSP1.
MatStocksTab3 = np.zeros((9,6)) # Material stocks for table, SSP2.

CascadeFlag1  = False
CascadeFlag2  = False
SensitiFlag1  = False

SingleSectList           = [] # For model runs not part of sensitivity or cascade, used for efficiency-sufficiency bar plot
SingleSectRegionList     = [] # For regions for eff-suff plot

# search for script config list entry
while ModelEvalListSheet.cell_value(Row, 1)  != 'ENDOFLIST':
    if ModelEvalListSheet.cell_value(Row, 1) != '':
        FolderList       = []
        MultiSectorList  = []
        RegionalScope    = ModelEvalListSheet.cell_value(Row, 1)
        Setting          = ModelEvalListSheet.cell_value(Row, 2) # cascade or sensitivity
        print(RegionalScope)
        
    if Setting == 'Cascade_pav':
        CascadeFlag1     = True
        SectorString     = 'pav'
        Vsheet           = mywb[RegionalScope  + '_Vehicles']
        NE               = 7 # 7 for vehs. and 6 for buildings
        
    if Setting == 'Cascade_reb':        
        CascadeFlag1     = True
        SectorString     = 'reb'
        Vsheet           = mywb[RegionalScope  + '_ResBuildings']
        NE               = 6 # 7 for vehs. and 6 for buildings     

    if Setting == 'Cascade_nrb':        
        CascadeFlag1     = True
        SectorString     = 'nrb'
        Vsheet           = mywb[RegionalScope  + '_NonResBuildings']
        NE               = 6 # 7 for vehs. and 6 for buildings     
        
    if Setting == 'Cascade_pav_reb':
        CascadeFlag2     = True
        SectorString     = 'pav_reb'
        NE               = 8 # 8 for vehs, res and nonres buildings
            
    if Setting == 'Cascade_pav_reb_nrb':
        CascadeFlag2     = True
        SectorString     = 'pav_reb_nrb'
        NE               = 8 # 8 for vehs, res and nonres buildings
        
    CascCols = [5,13]        
        
    if CascadeFlag1 is True: #extract results for this cascade and store
        CascadeFlag1 = False
        Descr = 'Cascade_' + RegionalScope + '_' + SectorString
        print(Descr)
        for m in range(0,NE):
            FolderList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the cascade plot function
        ASummary, AvgDecadalEms, MatSummary, AvgDecadalMatEms, RecCredit, UsePhaseSummary, ManSummary, ForSummary, AvgDecadalUseEms, AvgDecadalManEms, AvgDecadalForEms, AvgDecadalRecEms, CumEms, AnnEms2050, MatStocks, TimeSeries_R, MatEms = ODYM_RECC_Cascade_V2_4.main(RegionalScope,FolderList,SectorString)
        
        # Export cascade results via pandas:
        ColIndex      = [str(mmx) for mmx in  range(2016,2061)]
        MatEma_R_Data = np.einsum('tSRE->ESRt',MatEms).reshape(NE*NS*NR,45)
        if SectorString == 'pav':
            RES_List = ['None','EoL + FSD + FYI','EoL + FSD + FYI + ReU +LTE','EoL + FSD + FYI + ReU +LTE + MSu','EoL + FSD + FYI + ReU +LTE + MSu + LWE','EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS','EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS = ALL']
        else: # for reb and nrb
            RES_List = ['None','EoL + FSD + FYI','EoL + FSD + FYI + ReU +LTE','EoL + FSD + FYI + ReU +LTE + MSu','EoL + FSD + FYI + ReU +LTE + MSu + LWE','EoL + FSD + FYI + ReU +LTE + MSu + LWE + MIU = ALL']
        RowIndex = pd.MultiIndex.from_product([RES_List,['LED','SSP1','SSP2'],['NoNewClimPol','RCP2.6']], names=('res. eff.','SSP','RCP'))
        MatEma_R = pd.DataFrame(MatEma_R_Data, index=RowIndex, columns=ColIndex)
        MatEma_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_GHG_MtCO2.xls'), merge_cells=False)
        # Export material production via pandas
        PP_R_Data = np.einsum('EtSR->ESRt',TimeSeries_R[2,:,:,:,:]).reshape(NE*NS*NR,45)
        PP_R      = pd.DataFrame(PP_R_Data, index=RowIndex, columns=ColIndex)
        PP_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_PrimProd_Mt.xls'), merge_cells=False)
        SP_R_Data = np.einsum('EtSR->ESRt',TimeSeries_R[3,:,:,:,:]).reshape(NE*NS*NR,45)
        SP_R      = pd.DataFrame(SP_R_Data, index=RowIndex, columns=ColIndex)
        SP_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_SecProd_Mt.xls'), merge_cells=False)
        
        # write results summary to Excel
        for R in range(0,NR):
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+3,  column   = c +CascCols[R]).value   = ASummary[r,R,c]       
                    Vsheet.cell(row = r+9,  column   = c +CascCols[R]).value   = ASummary[r+3,R,c]       
                    Vsheet.cell(row = r+15, column   = c +CascCols[R]).value   = ASummary[r+6,R,c]
                    Vsheet.cell(row = r+36, column   = c +CascCols[R]).value   = ASummary[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 21,column = c +CascCols[R]).value  = AvgDecadalEms[r,R,c,d]
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+45,  column  = c +CascCols[R]).value  = UsePhaseSummary[r,R,c]       
                    Vsheet.cell(row = r+51,  column  = c +CascCols[R]).value  = UsePhaseSummary[r+3,R,c]       
                    Vsheet.cell(row = r+57,  column  = c +CascCols[R]).value  = UsePhaseSummary[r+6,R,c]
                    Vsheet.cell(row = r+78,  column  = c +CascCols[R]).value  = UsePhaseSummary[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 63,column = c +CascCols[R]).value  = AvgDecadalUseEms[r,R,c,d]                    
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+87,  column  = c +CascCols[R]).value  = MatSummary[r,R,c]       
                    Vsheet.cell(row = r+93,  column  = c +CascCols[R]).value  = MatSummary[r+3,R,c]       
                    Vsheet.cell(row = r+99,  column  = c +CascCols[R]).value  = MatSummary[r+6,R,c]
                    Vsheet.cell(row = r+120, column  = c +CascCols[R]).value  = MatSummary[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 105,column = c +CascCols[R]).value  = AvgDecadalMatEms[r,R,c,d] 
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+129,  column = c +CascCols[R]).value  = ManSummary[r,R,c]       
                    Vsheet.cell(row = r+135,  column = c +CascCols[R]).value  = ManSummary[r+3,R,c]       
                    Vsheet.cell(row = r+141,  column = c +CascCols[R]).value  = ManSummary[r+6,R,c]
                    Vsheet.cell(row = r+162,  column = c +CascCols[R]).value  = ManSummary[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 147,column = c +CascCols[R]).value  = AvgDecadalManEms[r,R,c,d]                         
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+171, column  = c +CascCols[R]).value  = ForSummary[r,R,c]       
                    Vsheet.cell(row = r+177, column  = c +CascCols[R]).value  = ForSummary[r+3,R,c]       
                    Vsheet.cell(row = r+183, column  = c +CascCols[R]).value  = ForSummary[r+6,R,c]
                    Vsheet.cell(row = r+204, column  = c +CascCols[R]).value  = ForSummary[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 189,column = c +CascCols[R]).value  = AvgDecadalForEms[r,R,c,d]                         
            for r in range(0,3):
                for c in range(0,NE):
                    Vsheet.cell(row = r+213,  column = c +CascCols[R]).value  = RecCredit[r,R,c]       
                    Vsheet.cell(row = r+219,  column = c +CascCols[R]).value  = RecCredit[r+3,R,c]       
                    Vsheet.cell(row = r+225,  column = c +CascCols[R]).value  = RecCredit[r+6,R,c]
                    Vsheet.cell(row = r+246,  column = c +CascCols[R]).value  = RecCredit[r+9,R,c]
                    for d in range(0,4):
                        Vsheet.cell(row = d*3 + r + 231,column = c +CascCols[R]).value  = AvgDecadalRecEms[r,R,c,d]   
        
        # Store results in time series array
        if SectorString == 'pav' or SectorString == 'reb':
            if SectorString == 'pav':
                SectorIndex = 0
                RegPos = Pav_RegionList20.index(RegionalScope)
                PlotOrder_pav.append(RegPos)
            if SectorString == 'reb':
                SectorIndex = 1
                RegPos = Reb_RegionList20.index(RegionalScope)
                PlotOrder_reb.append(RegPos)
            TimeSeries_All[0,:,RegPos,SectorIndex,:,:] = TimeSeries_R[0,0,:,:,:]  # system-wide GHG, no RES
            TimeSeries_All[1,:,RegPos,SectorIndex,:,:] = TimeSeries_R[0,-1,:,:,:] # system-wide GHG, full RES
            TimeSeries_All[2,:,RegPos,SectorIndex,:,:] = TimeSeries_R[1,0,:,:,:]  # matcycle GHG, no RES
            TimeSeries_All[3,:,RegPos,SectorIndex,:,:] = TimeSeries_R[1,-1,:,:,:] # matcycle GHG, full RES
            TimeSeries_All[4,:,RegPos,SectorIndex,:,:] = TimeSeries_R[2,0,:,:,:]  # primary production total, no RES
            TimeSeries_All[5,:,RegPos,SectorIndex,:,:] = TimeSeries_R[2,-1,:,:,:] # primary production total, full RES
            TimeSeries_All[6,:,RegPos,SectorIndex,:,:] = TimeSeries_R[3,0,:,:,:]  # secondary production total, no RES
            TimeSeries_All[7,:,RegPos,SectorIndex,:,:] = TimeSeries_R[3,-1,:,:,:] # secondary production total, full RES
            TimeSeries_All[8,:,RegPos,SectorIndex,:,:] = TimeSeries_R[4,0,:,:,:]  # el + H2 share in use phase, no RES
            TimeSeries_All[9,:,RegPos,SectorIndex,:,:] = TimeSeries_R[4,-1,:,:,:] # el + H2 share in use phase, full RES
            TimeSeries_All[10,:,RegPos,SectorIndex,:,:] =TimeSeries_R[5,0,:,:,:]  # el GHG factor, no RES
            TimeSeries_All[11,:,RegPos,SectorIndex,:,:] =TimeSeries_R[5,-1,:,:,:] # el GHG factor, full RES (same as for no RES)
            TimeSeries_All[12,:,RegPos,SectorIndex,:,:] =TimeSeries_R[6,0,:,:,:]  # all materials stock (sum), no RES
            TimeSeries_All[13,:,RegPos,SectorIndex,:,:] =TimeSeries_R[6,-1,:,:,:] # all materials stock (sum), full RES
            TimeSeries_All[14,:,RegPos,SectorIndex,:,:] =TimeSeries_R[7,0,:,:,:]  # use phase total energy consumption, no RES
            TimeSeries_All[15,:,RegPos,SectorIndex,:,:] =TimeSeries_R[7,-1,:,:,:] # use phase total energy consumption, full RES
            # net GHG impact of wood use: forest uptake + wood-related emissions from waste mgt. Pos sign for flow from system to environment:
            TimeSeries_All[16,:,RegPos,SectorIndex,:,:] =TimeSeries_R[8,0,:,:,:]  # net GHG impact of wood use, no RES
            TimeSeries_All[17,:,RegPos,SectorIndex,:,:] =TimeSeries_R[8,-1,:,:,:] # net GHG impact of wood use, full RES
            TimeSeries_All[18,:,RegPos,SectorIndex,:,:] =TimeSeries_R[9,0,:,:,:]  # passenger-km, no RES
            TimeSeries_All[19,:,RegPos,SectorIndex,:,:] =TimeSeries_R[9,-1,:,:,:] # passenger-km, full RES
            TimeSeries_All[20,:,RegPos,SectorIndex,:,:] =TimeSeries_R[10,0,:,:,:] # heated building space, no RES
            TimeSeries_All[21,:,RegPos,SectorIndex,:,:] =TimeSeries_R[10,-1,:,:,:]# heated building space, full RES
            TimeSeries_All[22,:,RegPos,SectorIndex,:,:] =TimeSeries_R[11,0,:,:,:] # cooled building space, no RES
            TimeSeries_All[23,:,RegPos,SectorIndex,:,:] =TimeSeries_R[11,-1,:,:,:]# cooled building space, full RES
            # calculate service intensities
            TimeSeries_All[24,:,RegPos,SectorIndex,:,:] =TimeSeries_All[18,:,RegPos,SectorIndex,:,:] / TimeSeries_All[0,:,RegPos,SectorIndex,:,:]  # pkm/t GHG
            TimeSeries_All[25,:,RegPos,SectorIndex,:,:] =TimeSeries_All[19,:,RegPos,SectorIndex,:,:] / TimeSeries_All[1,:,RegPos,SectorIndex,:,:]  # pkm/t GHG
            TimeSeries_All[26,:,RegPos,SectorIndex,:,:] =TimeSeries_All[18,:,RegPos,SectorIndex,:,:] / TimeSeries_All[12,:,RegPos,SectorIndex,:,:] # pkm/t Matstocks
            TimeSeries_All[27,:,RegPos,SectorIndex,:,:] =TimeSeries_All[19,:,RegPos,SectorIndex,:,:] / TimeSeries_All[13,:,RegPos,SectorIndex,:,:] # pkm/t Matstocks
            
        RCP_Matstocks = 1 # MatStocks are plotted for RCP2.6 only
        
        if Setting == 'Cascade_pav':                    
            # store other results
            Table_Annual[0,0:-1,1,:]= AnnEms2050[1,:,:].transpose().copy()
            Table_CumEms[0,0:-1,1,:]= CumEms[1,:,:].transpose().copy()
            MatStocksTab1[0,:]      = MatStocks[4,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[1,:]      = MatStocks[34,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[2,:]      = MatStocks[34,:,0,RCP_Matstocks,-1].copy()
            MatStocksTab2[0,:]      = MatStocks[4,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[1,:]      = MatStocks[34,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[2,:]      = MatStocks[34,:,1,RCP_Matstocks,-1].copy()
            MatStocksTab3[0,:]      = MatStocks[4,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[1,:]      = MatStocks[34,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[2,:]      = MatStocks[34,:,2,RCP_Matstocks,-1].copy()
                    
        if Setting == 'Cascade_reb':
            # store other results
            Table_Annual[1,0:5,1,:] = AnnEms2050[1,:,0:-1].transpose().copy()
            Table_Annual[1,7,1,:]   = AnnEms2050[1,:,-1].copy()
            Table_CumEms[1,0:5,1,:] = CumEms[1,:,0:-1].transpose().copy()                    
            Table_CumEms[1,7,1,:]   = CumEms[1,:,-1].copy()        
            MatStocksTab1[3,:]      = MatStocks[4,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[4,:]      = MatStocks[34,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[5,:]      = MatStocks[34,:,0,RCP_Matstocks,-1].copy()            
            MatStocksTab2[3,:]      = MatStocks[4,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[4,:]      = MatStocks[34,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[5,:]      = MatStocks[34,:,1,RCP_Matstocks,-1].copy()            
            MatStocksTab3[3,:]      = MatStocks[4,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[4,:]      = MatStocks[34,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[5,:]      = MatStocks[34,:,2,RCP_Matstocks,-1].copy()            
    
        if Setting == 'Cascade_nrb':                        
            # store other results
            Table_Annual[1,0:5,2,:] = AnnEms2050[1,:,0:-1].transpose().copy()
            Table_Annual[1,7,2,:]   = AnnEms2050[1,:,-1].copy()
            Table_CumEms[1,0:5,2,:] = CumEms[1,:,0:-1].transpose().copy()                    
            Table_CumEms[1,7,2,:]   = CumEms[1,:,-1].copy()      
            MatStocksTab1[6,:]      = MatStocks[4,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[7,:]      = MatStocks[34,:,0,RCP_Matstocks,0].copy()
            MatStocksTab1[8,:]      = MatStocks[34,:,0,RCP_Matstocks,-1].copy()                 
            MatStocksTab2[6,:]      = MatStocks[4,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[7,:]      = MatStocks[34,:,1,RCP_Matstocks,0].copy()
            MatStocksTab2[8,:]      = MatStocks[34,:,1,RCP_Matstocks,-1].copy()                 
            MatStocksTab3[6,:]      = MatStocks[4,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[7,:]      = MatStocks[34,:,2,RCP_Matstocks,0].copy()
            MatStocksTab3[8,:]      = MatStocks[34,:,2,RCP_Matstocks,-1].copy()                 
                    
    if CascadeFlag2 is True: #extract results for this cascade and store    
        CascadeFlag2 = False
        Descr = 'Cascade_' + RegionalScope + '_' + SectorString
        print(Descr)
        for m in range(0,NE):
            MultiSectorList.append(ModelEvalListSheet.cell_value(Row +m, 3))  
        GHG_TableX = ODYM_RECC_Table_Extract_V2_4.main(RegionalScope,MultiSectorList)        
        # write results summary as Table 2 to Excel
        Gsheet = mywb4['GHG_Overview']
        print('GHG_Overview_' + RegionalScope)
        for r in range(0,4):
            for c in range(0,6):
                for R in range(0,2):
                    Gsheet.cell(row = r+4 + 8*R, column = c+4).value  = GHG_TableX[r,c,R]        

        # run the cascade plots for the three sectors
        ASummary, AvgDecadalEms, MatSummary, AvgDecadalMatEms, RecCredit, UsePhaseSummary, ManSummary, ForSummary, AvgDecadalUseEms, AvgDecadalManEms, AvgDecadalForEms, AvgDecadalRecEms, CumEms, AnnEms2050, MatStocks, TimeSeries_R, MatEms = ODYM_RECC_Cascade_V2_4.main(RegionalScope,MultiSectorList,SectorString)              

        # Export cascade results via pandas:
        ColIndex      = [str(mmx) for mmx in  range(2016,2061)]
        MatEma_R_Data = np.einsum('tSRE->ESRt',MatEms).reshape(NE*NS*NR,45)
        RES_List = ['None','EoL + FSD + FYI','EoL + FSD + FYI + ReU +LTE','EoL + FSD + FYI + ReU +LTE + MSu','EoL + FSD + FYI + ReU +LTE + MSu + LWE','EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS','EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS','EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS + MIU = ALL']
        RowIndex = pd.MultiIndex.from_product([RES_List,['LED','SSP1','SSP2'],['NoNewClimPol','RCP2.6']], names=('res. eff.','SSP','RCP'))
        MatEma_R = pd.DataFrame(MatEma_R_Data, index=RowIndex, columns=ColIndex)
        MatEma_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_GHG_MtCO2.xls'), merge_cells=False)
        # Export material production via pandas
        PP_R_Data = np.einsum('EtSR->ESRt',TimeSeries_R[2,:,:,:,:]).reshape(NE*NS*NR,45)
        PP_R      = pd.DataFrame(PP_R_Data, index=RowIndex, columns=ColIndex)
        PP_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_PrimProd_Mt.xls'), merge_cells=False)
        SP_R_Data = np.einsum('EtSR->ESRt',TimeSeries_R[3,:,:,:,:]).reshape(NE*NS*NR,45)
        SP_R      = pd.DataFrame(SP_R_Data, index=RowIndex, columns=ColIndex)
        SP_R.to_excel(os.path.join(RECC_Paths.results_path,Descr + '_Mat_SecProd_Mt.xls'), merge_cells=False)
        
        # plot GHG overview figure global
        if RegionalScope == 'Global':
            ColOrder     = [11,4,0,18,8,16,2,6,15]
            Xoff         = [1,1.7,2.7,3.4,4.4,5.1]
            MyColorCycle = pylab.cm.tab20(np.arange(0,1,0.05)) # select 20 colors from the 'tab20' color map. 
            Data_Cum_Abs = CumEms
            Data_Cum_pc  = Data_Cum_Abs / np.einsum('SR,E->SRE',Data_Cum_Abs[:,:,0],np.ones(8)) # here: pc = percent, not per capita.
            Data_50_Abs  = np.einsum('ESR->SRE',TimeSeries_R[0,:,35,:,:])
            Data_50_pc   = Data_50_Abs / np.einsum('SR,E->SRE',Data_50_Abs[:,:,0],np.ones(8))   # here: pc = percent, not per capita.
            Cum_savings  = Data_Cum_Abs[:,:,0] - Data_Cum_Abs[:,:,-1]
            Ann_savings  = Data_50_Abs[:,:,0]  - Data_50_Abs[:,:,-1]
            LWE          = ['higher yields', 're-use/longer use','material subst.','down-sizing','car-sharing','ride-sharing','More intense bld. use','Bottom line']
            XTicks       = [1.25,1.95,2.95,3.65,4.65,5.35]
            YTicks       = [0,0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9,1.0,1.2,1.3,1.4,1.5,1.6,1.7,1.8,1.9,2.0,2.1,2.2]
            YTickLabels  = ['0','10','20','30','40','50','60','70','80','90','100','0','10','20','30','40','50','60','70','80','90','100']
            # plot results
            bw   = 0.5
            yoff = 1.2
            
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
        
            ProxyHandlesList = []   # For legend     
            # plot bars
            for mS in range(0,3):
                for mR in range(0,2):
                    ax1.fill_between([Xoff[mS*2+mR],Xoff[mS*2+mR]+bw], [yoff,yoff],[yoff+Data_Cum_pc[mS,mR,-1],yoff+Data_Cum_pc[mS,mR,-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[0],:], linewidth = 0.0)
                    for xca in range(1,NE):
                        ax1.fill_between([Xoff[mS*2+mR],Xoff[mS*2+mR]+bw], [yoff+Data_Cum_pc[mS,mR,-xca],yoff+Data_Cum_pc[mS,mR,-xca]],[yoff+Data_Cum_pc[mS,mR,-xca-1],yoff+Data_Cum_pc[mS,mR,-xca-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[xca],:], linewidth = 0.0)
                            
                    ax1.fill_between([Xoff[mS*2+mR],Xoff[mS*2+mR]+bw], [0,0],[Data_50_pc[mS,mR,-1],Data_50_pc[mS,mR,-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[0],:], linewidth = 0.0)
                    for xca in range(1,NE):
                        ax1.fill_between([Xoff[mS*2+mR],Xoff[mS*2+mR]+bw], [Data_50_pc[mS,mR,-xca],Data_50_pc[mS,mR,-xca]],[Data_50_pc[mS,mR,-xca-1],Data_50_pc[mS,mR,-xca-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[xca],:], linewidth = 0.0)
            plt.plot([-1,8],[1.1,1.1],  linestyle = '-',  linewidth = 0.5, color = 'k')
            plt.plot([2.45,2.45],[-1,3],linestyle = '--', linewidth = 0.5, color = 'k')
            plt.plot([4.15,4.15],[-1,3],linestyle = '--', linewidth = 0.5, color = 'k')
            for fca in range(0,NE):
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[ColOrder[fca],:])) # create proxy artist for legend

            # plot emissions reductions
            for mS in range(0,3):
                for mR in range(0,2): 
                    plt.text(Xoff[mS*2+mR]+0.51, 2.19,  ("%3.0f" % (-Cum_savings[mS,mR]/1000))                  + ' Gt',fontsize=14, rotation=90, fontweight='bold')
                    plt.text(Xoff[mS*2+mR]+0.51, 0.99,  ("%3.0f" % (-10*np.round(Ann_savings[mS,mR]/10)))       + ' Mt',fontsize=14, rotation=90, fontweight='bold')
                    plt.text(Xoff[mS*2+mR]+0.2, 1.52,   ("%3.0f" % (10*np.round(Data_Cum_Abs[mS,mR,-1]/10000))) + ' Gt',fontsize=16, rotation=90, fontweight='normal', color =np.array([0.48,0.33,0.22,1]))
                    plt.text(Xoff[mS*2+mR]+0.2, 0.45,   ("%3.0f" % (100*np.round(Data_50_Abs[mS,mR,-1]/100)))   + ' Mt',fontsize=16, rotation=90, fontweight='normal', color =np.array([0.48,0.33,0.22,1]))
                    # top arrow
                    plt.arrow(Xoff[mS*2+mR]+0.42,2.25,0,-0.05, lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                    plt.arrow(Xoff[mS*2+mR]+0.42,1.05,0,-0.05, lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                    # bottom arrow
                    plt.arrow(Xoff[mS*2+mR]+0.42,1.2+Data_Cum_pc[mS,mR,-1]-0.05,0,0.05, lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                    plt.arrow(Xoff[mS*2+mR]+0.42,Data_50_pc[mS,mR,-1]-0.05,0,0.05, lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.06, head_length =0.02, ec = 'k', fc = 'k')
                    #plt.text(Xoff[mS*2+mR]+0.2, 1.9, ("%3.0f" % (Cum_savings[mS,mR]/1000)) + ' Gt',fontsize=16, rotation=90, fontweight='normal')
                    #plt.text(Xoff[mS*2+mR]+0.2, 0.5, ("%3.0f" % Ann_savings[mS,mR]) + ' Mt',fontsize=16, rotation=90, fontweight='normal')
                    
            # plot text and labels
            plt.text(0.85, 0.97, '2050 annual emissions', fontsize=11, rotation=90, fontweight='bold')          
            plt.text(0.85, 2.2,  '2016-50 cum. emissions',fontsize=11, rotation=90, fontweight='bold')          

            #plt.title('RE strats. and GHG emissions, global, pav+reb.', fontsize = 18)
            
            plt.ylabel('GHG reductions, system-wide, %.', fontsize = 18)
            plt.xticks(XTicks)
            plt.yticks(YTicks, fontsize =12)
            ax1.set_xticklabels(['LED+NoPol','LED+RCP2.6','SSP1+NoPol','SSP1+RCP2.6','SSP2+NoPol','SSP2+RCP2.6'], rotation =90, fontsize = 15, fontweight = 'normal')
            ax1.set_yticklabels(YTickLabels, rotation =0, fontsize = 15, fontweight = 'normal')
            plt.legend(handles = reversed(ProxyHandlesList),labels = LWE,shadow = False, prop={'size':12},ncol=1, loc = 'upper right' ,bbox_to_anchor=(1.40, 1)) 
            #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
            plt.axis([0.7, 5.8, -0.1, 2.3])
        
            plt.show()
            fig_name = 'Overview_GHG_Global.png'
            fig.savefig(os.path.join(RECC_Paths.results_path,fig_name), dpi = PlotExpResolution, bbox_inches='tight')   
            fig_name = 'Overview_GHG_Global.svg'
            fig.savefig(os.path.join(RECC_Paths.results_path,fig_name), dpi = PlotExpResolution, bbox_inches='tight')   
                
        if ModelEvalListSheet.cell_value(Row+NE, 2) == 'ME_industry_demandside_Scenario':
            for mmxx in range(0,6):
                SingleSectList.append(ModelEvalListSheet.cell_value(Row+NE+mmxx, 3))
            # run the efficieny_sufficieny plots, with 6 extra single sectors in result list
            CumEmsV, CumEmsV2060, AnnEmsV2030, AnnEmsV2050, AvgDecadalEmsV = ODYM_RECC_BarPlot_ME_Industry_Demand_V2_4.main(RegionalScope,MultiSectorList,SingleSectList)  
            SingleSectList = []
            NE  +=6 # add for extra scenarios for efficiency-sufficiency plot
                
    if Setting == 'Sensitivity_pav':
        SensitiFlag1     = True
        SectorString     = 'pav'
        NE               = 11 # 11 for vehs. and 10 for buildings 
        SensCols         = [6,18] 
        
    if Setting == 'Sensitivity_reb':        
        SensitiFlag1     = True
        SectorString     = 'reb'
        NE               = 10 # 11 for vehs. and 10 for buildings        
        SensCols         = [35,47] 
        
    if Setting == 'Sensitivity_nrb':        
        SensitiFlag1     = True
        SectorString     = 'nrb'
        NE               = 10 # 11 for vehs. and 10 for buildings        
        SensCols         = [63,75] 
    
    SensRows             = [4,9,14,19,24,40,45,50,55,60,76,81,86,91,96,112,117,122,127,132,148,153,158,163,168,184,189,194,199,204]
        
    if SensitiFlag1 is True:
        SensitiFlag1 = False
        for m in range(0,int(NE)):
            FolderList.append(ModelEvalListSheet.cell_value(Row +m, 3))
        # run the ODYM-RECC sensitivity analysis for pav
        CumEms_Sens2050, CumEms_Sens2060, AnnEms2030_Sens, AnnEms2050_Sens, AvgDecadalEms, UseCumEms2050, UseCumEms2060, UseAnnEms2030, UseAnnEms2050, AvgDecadalUseEms, MatCumEms2050, MatCumEms2060, MatAnnEms2030, MatAnnEms2050, AvgDecadalMatEms, ManCumEms2050, ManCumEms2060, ManAnnEms2030, ManAnnEms2050, AvgDecadalManEms, ForCumEms2050, ForCumEms2060, ForAnnEms2030, ForAnnEms2050, AvgDecadalForEms, RecCreditCum2050, RecCreditCum2060, RecCreditAnn2030, RecCreditAnn2050, RecCreditAvgDec = ODYM_RECC_Sensitivity_V2_4.main(RegionalScope,FolderList,SectorString)        

        # write results summary to Excel
        Ssheet  = mywb['Sensitivity_'  + RegionalScope]
        print('Sensitivity_' + RegionalScope + '_' + SectorString)
        for R in range(0,NR):
            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[0], column   = c +SensCols[R]).value   = AnnEms2030_Sens[r,R,c]
                    Ssheet.cell(row = r +SensRows[1], column   = c +SensCols[R]).value   = AnnEms2050_Sens[r,R,c]
                    Ssheet.cell(row = r +SensRows[2], column   = c +SensCols[R]).value   = CumEms_Sens2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[3], column   = c +SensCols[R]).value   = CumEms_Sens2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[4],column  = c +SensCols[R]).value   = AvgDecadalEms[r,R,c,d]
            
            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[5], column  = c +SensCols[R]).value    = UseAnnEms2030[r,R,c]
                    Ssheet.cell(row = r +SensRows[6], column  = c +SensCols[R]).value    = UseAnnEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[7], column  = c +SensCols[R]).value    = UseCumEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[8], column  = c +SensCols[R]).value    = UseCumEms2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[9],column = c +SensCols[R]).value   = AvgDecadalUseEms[r,R,c,d]       
                       
            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[10], column  = c +SensCols[R]).value   = MatAnnEms2030[r,R,c]
                    Ssheet.cell(row = r +SensRows[11], column  = c +SensCols[R]).value   = MatAnnEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[12], column  = c +SensCols[R]).value   = MatCumEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[13], column  = c +SensCols[R]).value   = MatCumEms2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[14],column = c +SensCols[R]).value   = AvgDecadalMatEms[r,R,c,d] 

            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[15], column  = c +SensCols[R]).value   = ManAnnEms2030[r,R,c]
                    Ssheet.cell(row = r +SensRows[16], column  = c +SensCols[R]).value   = ManAnnEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[17], column  = c +SensCols[R]).value   = ManCumEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[18], column  = c +SensCols[R]).value   = ManCumEms2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[19],column = c +SensCols[R]).value   = AvgDecadalManEms[r,R,c,d] 

            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[20], column  = c +SensCols[R]).value   = ForAnnEms2030[r,R,c]
                    Ssheet.cell(row = r +SensRows[21], column  = c +SensCols[R]).value   = ForAnnEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[22], column  = c +SensCols[R]).value   = ForCumEms2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[23], column  = c +SensCols[R]).value   = ForCumEms2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[24],column = c +SensCols[R]).value   = AvgDecadalForEms[r,R,c,d] 

            for r in range(0,3):
                for c in range(0,NE):
                    Ssheet.cell(row = r +SensRows[25], column  = c +SensCols[R]).value   = RecCreditAnn2030[r,R,c]
                    Ssheet.cell(row = r +SensRows[26], column  = c +SensCols[R]).value   = RecCreditAnn2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[27], column  = c +SensCols[R]).value   = RecCreditCum2050[r,R,c]
                    Ssheet.cell(row = r +SensRows[28], column  = c +SensCols[R]).value   = RecCreditCum2060[r,R,c]
                    for d in range(0,4):
                        Ssheet.cell(row = d*3 + r + SensRows[29],column = c +SensCols[R]).value   = RecCreditAvgDec[r,R,c,d]                         
                        
    # forward counter   
    Row += NE
        
# store overview tables
# Done for each region, overwritten each time, data for LAST REGION remain.
WFsheet = mywb4['CascadeBySector']
v = 1 # SSP1
for u in range(0,8):
    for c in range(0,2):
        for z in range(0,3):
            WFsheet.cell(row = u+4,  column = z+3 +4*c).value  = Table_Annual[z,u,v,c]     
            WFsheet.cell(row = u+14, column = z+3 +4*c).value  = Table_CumEms[z,u,v,c]       
            
WFsheet = mywb4['MatStocksBySector']
for u in range(0,9):
    for v in range(0,6):
        WFsheet.cell(row = u+3 , column = v+6).value  = MatStocksTab1[u,v]     
        WFsheet.cell(row = u+14, column = v+6).value  = MatStocksTab2[u,v]     
        WFsheet.cell(row = u+25, column = v+6).value  = MatStocksTab3[u,v]     
    
mywb.save(os.path.join(RECC_Paths.results_path, 'RECC_Global_Results_SystemGHG_V2_4.xlsx'))      
mywb4.save(os.path.join(RECC_Paths.results_path,'RECC_Global_Results_Tables_V2_4.xlsx'))      
   
# plot SSP1 time series in 5x5 plot:
# TimeSeries_All indices: NX x Nt x Nr x NV x NS x NR / indicators x time x regions x sectors x SSP x RCP
# 0: system-wide GHG, no RES 1: system-wide GHG, all RES
# 1: material-related GHG, no RES, 3: material-related GHG, all RES,
# 2: primary materials with and without RES
# 3: secondary material with and without RES

ind_5x5 = [0,2,4,6,8,14,16,18,20,22,24,26]
fin_5x5 = ['GHG_pav_5x5','GHG_reb_5x5','GHGMat_pav_5x5','GHGMat_reb_5x5','PrimMat_5x5_pav','PrimMat_5x5_reb','SecMat_5x5_pav','SecMat_5x5_reb',\
           'ElH2Share_5x5_pav','ElH2Share_5x5_reb','UsePhaseEn_5x5_pav','UsePhaseEn_5x5_reb','WoodCycleGHG_pav','WoodCycleGHG_reb','passenger_km',\
           'no_data_here','no_data_here','heated_m2','no_data_here','cooled_m2','passenger_km_perGHG','no_data_here','passenger_km_perMatStocks','no_data_here']
fit_5x5 = [r'System-wide GHG, pav, Mt CO$_2$-eq/yr,',r'System-wide GHG, reb, Mt CO$_2$-eq/yr,',r'Matcycle GHG, pav, Mt CO$_2$-eq/yr,',r'Matcycle GHG, reb, Mt CO$_2$-eq/yr,','Total primary material, pav, Mt/yr,',\
           'Total primary material, reb, Mt/yr,','Total secondary material, pav, Mt/yr,','Total secondary material, reb, Mt/yr,',r'Share of El and H$_2$ in use phase en. cons, pav, 1,',\
           r'Share of El and H$_2$ in use phase en. cons, reb, 1,','Use phase energy cons, pav, TJ,','Use phase energy cons, reb, TJ,',r'Wood cycle GHG, pav, Mt CO$_2$-eq/yr,',\
           r'Wood cycle GHG, reb, Mt CO$_2$-eq/yr,','passenger-km, Mkm,','no_data','no_data','buildings, heated m², Mm²,','no_data','buildings, cooled m², Mm²,',\
           'passenger-km per GHG km/t,','no_data_here','passenger-km per material stocks, km/t,','no_data_here']

MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.1)) # select 12 colors from the 'Paired' color map. 
BaseBlue     = np.array([0.31,0.51,0.74,1])
R26Green     = np.array([0.03,0.66,0.48,1])
plt.rcParams['axes.labelsize'] = 7
LegendLabels = ['NoNewClimPol, no ME','NoNewClimPol, full ME','RCP2.6, no ME','RCP2.6, full ME']
SEScenLabels = ['LED','SSP1','SSP2']

# System-wide GHG, mat. GHG, and material production, with country names inside plots        
Pav_label_offset = [2017,2023,2018,2050,2045,2040,2017,2026,2020,2033,2017,2017,2017,2017,2017,2033,2037,2015,2026,2028,2040,2045,2040,2040,2050]        
Pav_label_pos    = [0.06,0.83,0.83,0.83,0.83,0.83,0.06,0.06,0.83,0.83,0.83,0.83,0.06,0.06,0.06,0.83,0.83,0.06,0.83,0.83,0.83,0.83,0.83,0.83,0.83]        
Reb_label_offset = [2038,2023,2023,2038,2048,2040,2045,2023,2025,2023,2042,2014,2020,2030,2025,2035,2050,2035,2025,2040,2045,2035,2040,2020,2032]
Reb_label_pos    = [0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.06,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83]        

for mmf in range(0,len(ind_5x5)):
    for Sect in range(0,2):
        for SEScen in range(0,3):
            if Sect == 0:
                RegionList20Plot = Pav_RegionList20Plot
                PlotOrder        = PlotOrder_pav
                AxisMax          = Pav_axis_5x5
            if Sect == 1:
                RegionList20Plot = Reb_RegionList20Plot        
                PlotOrder        = PlotOrder_reb
                AxisMax          = Reb_axis_5x5
            fig, axs = plt.subplots(5, 5, sharex=True, gridspec_kw={'hspace': 0.22, 'wspace': 0.5})
            for plotNo in PlotOrder:
                if mmf == 0: # only for GHG total plot
                    for mmn in range(0,45): # plot grey bar where net emisisons are negative:
                        if TimeSeries_All[1,mmn,plotNo,Sect,SEScen,1] < 0:
                            axs[plotNo//5, plotNo%5].fill_between([2016+mmn,2016+mmn+1], [0,0],[AxisMax[plotNo],AxisMax[plotNo]],linestyle = '--', facecolor =np.array([0.3,0.3,0.3,0.3]), linewidth = 0.0)
                axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,SEScen,0],color=BaseBlue, lw=0.8,  linestyle='-')    # Baseline, no RES
                axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,SEScen,0],color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
                axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,SEScen,1],color=R26Green, lw=0.8,  linestyle='-')    # RCP2.6, no RES
                axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,SEScen,1],color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
                axs[plotNo//5, plotNo%5].set_ylim(bottom=0)
                # Place region labels:
                if mmf == 0 and SEScen == 1:
                    if Sect == 0:
                        axs[plotNo//5, plotNo%5].text(Pav_label_offset[plotNo], Pav_label_pos[plotNo]*AxisMax[plotNo], RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
                    if Sect == 1:
                        axs[plotNo//5, plotNo%5].text(Reb_label_offset[plotNo], Reb_label_pos[plotNo]*AxisMax[plotNo], RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
                else:
                    axs[plotNo//5, plotNo%5].text(2015, 0.12*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,SEScen,:].max(), RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
                axs[plotNo//5, plotNo%5].tick_params(axis='x', labelsize=6)
                axs[plotNo//5, plotNo%5].tick_params(axis='y', labelsize=6)
                for axis in ['top','bottom','left','right']:
                    axs[plotNo//5, plotNo%5].spines[axis].set_linewidth(0.3)
                # Scale axes:
                if mmf == 0 and SEScen == 1:
                    axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, AxisMax[plotNo]])
                else:
                    if mmf == 6: # for wood cycle GHG, which are negative:
                        axs[plotNo//5, plotNo%5].axis([2012, 2063, 1.1*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,SEScen,:].min(), 1.1*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,SEScen,:].max()])
                    else:
                        axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, 1.1*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,SEScen,:].max()])
                axs[plotNo//5, plotNo%5].tick_params(axis='both',width = 0.3)
        
            plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
            plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.99, linestyle='--') # Baseline, full RES
            plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8,  linestyle='-')  # RCP2.6, no RES
            plt.plot([2010,2011],[0,0],color=R26Green, lw=0.99, linestyle='--') # RCP2.6, full RES
            plt.legend(LegendLabels,shadow = False,  prop={'size':7}, loc = 'upper right',bbox_to_anchor=(3.5, 1))    
            
            fig.suptitle(fit_5x5[2*mmf+Sect] +' '+ SEScenLabels[SEScen], fontsize=14)
            for xm in range(0,5):
                plt.sca(axs[4,xm])
                plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 6, fontweight = 'normal')
            plt.show()
            fig_name  = fin_5x5[2*mmf+Sect] +'_'+ SEScenLabels[SEScen]
            fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.png'), dpi = PlotExpResolution, bbox_inches='tight')  
            fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.svg'), dpi = PlotExpResolution, bbox_inches='tight')  
            
# Main paper Fig. 1 (subset of above plots), plot Global, GN, GS, G7, China, India, SSA
Pav_label_offset7 = [2017,2023,2018,2050,2017,2017,2012.5]        
Pav_label_pos7    = [0.06,0.83,0.83,0.83,0.06,0.83,0.06]        
Reb_label_offset7 = [2038,2023,2023,2048,2038,2042,2012.5]
Reb_label_pos7    = [0.83,0.83,0.83,0.83,0.83,0.83,0.83]        
SEScen = 1

for reg in All_RegionList7:
    PlotOrder_7_pav.append(Pav_RegionList20.index(reg))
    PlotOrder_7_reb.append(Reb_RegionList20.index(reg))

#fig, axs = plt.subplots(2, 7, sharex=True, gridspec_kw={'hspace': 0.10, 'wspace': 0.4}, figsize=(15,5))
#for plotNo in np.arange(0,7):
#    # first row: pav
#    Sect = 0
#    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_pav[plotNo],Sect,SEScen,0],color=BaseBlue, lw=0.8,  linestyle='-')   # Baseline, no RES
#    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_pav[plotNo],Sect,SEScen,0],color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
#    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_pav[plotNo],Sect,SEScen,1],color=R26Green, lw=0.8,  linestyle='-')   # RCP2.6, no RES
#    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_pav[plotNo],Sect,SEScen,1],color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
#    axs[0, plotNo].set_ylim(bottom=0)
#    # second row: reb
#    Sect = 1
#    for mmn in range(0,45): # plot grey bar where net emisisons are negative:
#        if TimeSeries_All[1,mmn,PlotOrder_7_reb[plotNo],Sect,SEScen,1] < 0:
#            axs[1, plotNo].fill_between([2016+mmn,2016+mmn+1], [0,0],[Reb_axis_7x2[plotNo],Reb_axis_7x2[plotNo]],linestyle = '--', facecolor =np.array([0.15,0.15,0.15,0.15]), linewidth = 0.0)
#    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_reb[plotNo],Sect,SEScen,0],color=BaseBlue, lw=0.8,  linestyle='-')   # Baseline, no RES
#    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_reb[plotNo],Sect,SEScen,0],color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
#    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_reb[plotNo],Sect,SEScen,1],color=R26Green, lw=0.8,  linestyle='-')   # RCP2.6, no RES
#    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_reb[plotNo],Sect,SEScen,1],color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
#    axs[1, plotNo].set_ylim(bottom=0)
#    
#    axs[0, plotNo].text(Pav_label_offset7[plotNo], Pav_label_pos7[plotNo]*Pav_axis_7x2[plotNo], All_RegionList7Plot[plotNo], fontsize=9, rotation=0, fontweight='normal')
#    axs[1, plotNo].text(Reb_label_offset7[plotNo], Reb_label_pos7[plotNo]*Reb_axis_7x2[plotNo], All_RegionList7Plot[plotNo], fontsize=9, rotation=0, fontweight='normal')
#    
#    axs[0, plotNo].tick_params(axis='x', labelsize=9)
#    axs[0, plotNo].tick_params(axis='y', labelsize=9)
#    axs[1, plotNo].tick_params(axis='x', labelsize=9)
#    axs[1, plotNo].tick_params(axis='y', labelsize=9)
#    
#    for axis in ['top','bottom','left','right']:
#        axs[0, plotNo].spines[axis].set_linewidth(0.5)
#        axs[1, plotNo].spines[axis].set_linewidth(0.5)
#    axs[0, plotNo].axis([2012, 2063, 0, Pav_axis_7x2[plotNo]])
#    axs[0, plotNo].tick_params(axis='both',width = 0.5)
#    axs[1, plotNo].axis([2012, 2063, 0, Reb_axis_7x2[plotNo]])
#    axs[1, plotNo].tick_params(axis='both',width = 0.5)
#
#plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
#plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.99, linestyle='--') # Baseline, full RES
#plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8,  linestyle='-')  # RCP2.6, no RES
#plt.plot([2010,2011],[0,0],color=R26Green, lw=0.99, linestyle='--') # RCP2.6, full RES
#plt.legend(LegendLabels,shadow = False,  prop={'size':9}, loc = 'upper right',bbox_to_anchor=(2.6, 1))    
#
#fig.suptitle(r'System-wide GHG, pav+reb, Mt CO$_2$-eq/yr, SSP1', fontsize=14)
#for xm in range(0,7):
#    plt.sca(axs[1,xm])
#    plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 9, fontweight = 'normal')
#plt.show()
#fig_name  = 'Fig1_select_GHG_pav_reb_SSP1'
#fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.png'), dpi = PlotExpResolution, bbox_inches='tight')  
#fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.svg'), dpi = PlotExpResolution, bbox_inches='tight')  

# Gt version
LegendLabels = ['No new climate policy, no material efficiency strategies','No new climate policy, full material efficiency strategies','RCP2.6 (2°C policy mix), no material efficiency strategies','RCP2.6 (2°C policy mix), full material efficiency strategies']
fig, axs = plt.subplots(2, 7, sharex=True, gridspec_kw={'hspace': 0.10, 'wspace': 0.4}, figsize=(15,5))
for plotNo in np.arange(0,7):
    # first row: pav
    Sect = 0
    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_pav[plotNo],Sect,SEScen,0]/1000,color=BaseBlue, lw=0.8,  linestyle='-')   # Baseline, no RES
    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_pav[plotNo],Sect,SEScen,0]/1000,color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_pav[plotNo],Sect,SEScen,1]/1000,color=R26Green, lw=0.8,  linestyle='-')   # RCP2.6, no RES
    axs[0, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_pav[plotNo],Sect,SEScen,1]/1000,color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
    axs[0, plotNo].set_ylim(bottom=0)
    # second row: reb
    Sect = 1
    for mmn in range(0,45): # plot grey bar where net emisisons are negative:
        if TimeSeries_All[1,mmn,PlotOrder_7_reb[plotNo],Sect,SEScen,1] < 0:
            axs[1, plotNo].fill_between([2016+mmn,2016+mmn+1], [0,0],[Reb_axis_7x2[plotNo]/1000,Reb_axis_7x2[plotNo]/1000],linestyle = '--', facecolor =np.array([0.15,0.15,0.15,0.15]), linewidth = 0.0)
    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_reb[plotNo],Sect,SEScen,0]/1000,color=BaseBlue, lw=0.8,  linestyle='-')   # Baseline, no RES
    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_reb[plotNo],Sect,SEScen,0]/1000,color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[0,:,PlotOrder_7_reb[plotNo],Sect,SEScen,1]/1000,color=R26Green, lw=0.8,  linestyle='-')   # RCP2.6, no RES
    axs[1, plotNo].plot(np.arange(2016,2061), TimeSeries_All[1,:,PlotOrder_7_reb[plotNo],Sect,SEScen,1]/1000,color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
    axs[1, plotNo].set_ylim(bottom=0)
    
    axs[0, plotNo].text(Pav_label_offset7[plotNo], Pav_label_pos7[plotNo]*Pav_axis_7x2[plotNo]/1000, All_RegionList7Plot[plotNo], fontsize=9, rotation=0, fontweight='normal')
    axs[1, plotNo].text(Reb_label_offset7[plotNo], Reb_label_pos7[plotNo]*Reb_axis_7x2[plotNo]/1000, All_RegionList7Plot[plotNo], fontsize=9, rotation=0, fontweight='normal')
    
    axs[0, plotNo].tick_params(axis='x', labelsize=9)
    axs[0, plotNo].tick_params(axis='y', labelsize=9)
    axs[1, plotNo].tick_params(axis='x', labelsize=9)
    axs[1, plotNo].tick_params(axis='y', labelsize=9)
    
    for axis in ['top','bottom','left','right']:
        axs[0, plotNo].spines[axis].set_linewidth(0.5)
        axs[1, plotNo].spines[axis].set_linewidth(0.5)
    axs[0, plotNo].axis([2012, 2063, 0, Pav_axis_7x2[plotNo]/1000])
    axs[0, plotNo].tick_params(axis='both',width = 0.5)
    axs[1, plotNo].axis([2012, 2063, 0, Reb_axis_7x2[plotNo]/1000])
    axs[1, plotNo].tick_params(axis='both',width = 0.5)

plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.99, linestyle='--') # Baseline, full RES
plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8,  linestyle='-')  # RCP2.6, no RES
plt.plot([2010,2011],[0,0],color=R26Green, lw=0.99, linestyle='--') # RCP2.6, full RES
plt.legend(LegendLabels,shadow = False,  prop={'size':9}, loc = 'bottom left',bbox_to_anchor=(-4.8, -0.3))    # x, y 

# fig.suptitle(r'System-wide GHG, pav+reb, Mt CO$_2$-eq/yr, SSP1', fontsize=14)
for xm in range(0,7):
    plt.sca(axs[1,xm])
    plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 9, fontweight = 'normal')
plt.show()
fig_name  = 'Fig1_select_GHG_pav_reb_SSP1_Gt'
fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.png'), dpi = PlotExpResolution, bbox_inches='tight')  
fig.savefig(os.path.join(RECC_Paths.results_path,fig_name+'.svg'), dpi = PlotExpResolution, bbox_inches='tight') 

# System-wide GHG, mat. GHG, and material production, with country names on top of plots
#for mmf in range(0,4):
#    for Sect in range(0,2):
#        if Sect == 0:
#            RegionList20Plot = Pav_RegionList20Plot
#            PlotOrder        = PlotOrder_pav
#            AxisMax          = Pav_axis_5x5
#        if Sect == 1:
#            RegionList20Plot = Reb_RegionList20Plot        
#            PlotOrder        = PlotOrder_reb
#            AxisMax          = Reb_axis_5x5
#        fig, axs = plt.subplots(5, 5, sharex=True, gridspec_kw={'hspace': 0.6, 'wspace': 0.5})
#        for plotNo in PlotOrder:
#            if mmf == 0: # only for GHG total plot
#                for mmn in range(0,45): # plot grey bar where net emisisons are negative:
#                    if TimeSeries_All[1,mmn,plotNo,Sect,1,1] < 0:
#                        axs[plotNo//5, plotNo%5].fill_between([2016+mmn,2016+mmn+1], [0,0],[AxisMax[plotNo],AxisMax[plotNo]],linestyle = '--', facecolor =np.array([0.3,0.3,0.3,0.3]), linewidth = 0.0)
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,1,0],color=BaseBlue, lw=0.8,  linestyle='-')    # Baseline, no RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,1,0],color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,1,1],color=R26Green, lw=0.8,  linestyle='-')    # RCP2.6, no RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,1,1],color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
#            axs[plotNo//5, plotNo%5].set_ylim(ymin=0)
#            axs[plotNo//5, plotNo%5].set_title(RegionList20Plot[plotNo], fontsize=7)
#            #axs[plotNo//5, plotNo%5].set_yticklabels(fontsize = 6)
#            axs[plotNo//5, plotNo%5].tick_params(axis='x', labelsize=6)
#            axs[plotNo//5, plotNo%5].tick_params(axis='y', labelsize=6)
#            for axis in ['top','bottom','left','right']:
#                axs[plotNo//5, plotNo%5].spines[axis].set_linewidth(0.3)
#            if mmf == 0:
#                axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, AxisMax[plotNo]])
#            axs[plotNo//5, plotNo%5].tick_params(axis='both',width = 0.3)
#    
#        plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
#        plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.99, linestyle='--') # Baseline, full RES
#        plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8,  linestyle='-')  # RCP2.6, no RES
#        plt.plot([2010,2011],[0,0],color=R26Green, lw=0.99, linestyle='--') # RCP2.6, full RES
#        plt.legend(LegendLables,shadow = False,  prop={'size':7}, loc = 'upper right',bbox_to_anchor=(3.5, 1))    
#        
#        fig.suptitle(fit_5x5[2*mmf+Sect], fontsize=14)
#        for xm in range(0,5):
#            plt.sca(axs[4,xm])
#            plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 6, fontweight = 'normal')
#        plt.show()
#        fig_name = fin_5x5[2*mmf+Sect]
#        fig.savefig(os.path.join(RECC_Paths.results_path,fig_name), dpi = PlotExpResolution, bbox_inches='tight')  

# OLD version of 5x5 plots for SSP1 only.
## System-wide GHG, mat. GHG, and material production, with country names inside plots        
#Pav_label_offset = [2017,2023,2018,2050,2045,2040,2017,2026,2020,2033,2017,2017,2017,2017,2017,2033,2037,2015,2026,2028,2040,2045,2040,2040,2050]        
#Pav_label_pos    = [0.06,0.83,0.83,0.83,0.83,0.83,0.06,0.06,0.83,0.83,0.83,0.83,0.06,0.06,0.06,0.83,0.83,0.06,0.83,0.83,0.83,0.83,0.83,0.83,0.83]        
#Reb_label_offset = [2038,2023,2023,2038,2048,2040,2045,2023,2025,2023,2042,2014,2020,2030,2025,2035,2050,2035,2025,2040,2045,2035,2040,2020,2032]
#Reb_label_pos    = [0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.06,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83,0.83]        
#
#for mmf in range(0,6):
#    for Sect in range(0,2):
#        if Sect == 0:
#            RegionList20Plot = Pav_RegionList20Plot
#            PlotOrder        = PlotOrder_pav
#            AxisMax          = Pav_axis_5x5
#        if Sect == 1:
#            RegionList20Plot = Reb_RegionList20Plot        
#            PlotOrder        = PlotOrder_reb
#            AxisMax          = Reb_axis_5x5
#        fig, axs = plt.subplots(5, 5, sharex=True, gridspec_kw={'hspace': 0.22, 'wspace': 0.5})
#        for plotNo in PlotOrder:
#            if mmf == 0: # only for GHG total plot
#                for mmn in range(0,45): # plot grey bar where net emisisons are negative:
#                    if TimeSeries_All[1,mmn,plotNo,Sect,1,1] < 0:
#                        axs[plotNo//5, plotNo%5].fill_between([2016+mmn,2016+mmn+1], [0,0],[AxisMax[plotNo],AxisMax[plotNo]],linestyle = '--', facecolor =np.array([0.3,0.3,0.3,0.3]), linewidth = 0.0)
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,1,0],color=BaseBlue, lw=0.8,  linestyle='-')    # Baseline, no RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,1,0],color=BaseBlue, lw=0.99, linestyle='--')  # Baseline, full RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]  ,:,plotNo,Sect,1,1],color=R26Green, lw=0.8,  linestyle='-')    # RCP2.6, no RES
#            axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), TimeSeries_All[ind_5x5[mmf]+1,:,plotNo,Sect,1,1],color=R26Green, lw=0.99, linestyle='--')  # RCP2.6, full RES
#            axs[plotNo//5, plotNo%5].set_ylim(bottom=0)
#            if mmf == 0:
#                if Sect == 0:
#                    axs[plotNo//5, plotNo%5].text(Pav_label_offset[plotNo], Pav_label_pos[plotNo]*AxisMax[plotNo], RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
#                if Sect == 1:
#                    axs[plotNo//5, plotNo%5].text(Reb_label_offset[plotNo], Reb_label_pos[plotNo]*AxisMax[plotNo], RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
#            else:
#                axs[plotNo//5, plotNo%5].text(2015, 0.12*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,1,:].max(), RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
#            axs[plotNo//5, plotNo%5].tick_params(axis='x', labelsize=6)
#            axs[plotNo//5, plotNo%5].tick_params(axis='y', labelsize=6)
#            for axis in ['top','bottom','left','right']:
#                axs[plotNo//5, plotNo%5].spines[axis].set_linewidth(0.3)
#            if mmf == 0:
#                axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, AxisMax[plotNo]])
#            else:
#                axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, 1.1*TimeSeries_All[ind_5x5[mmf]:ind_5x5[mmf]+2,:,plotNo,Sect,1,:].max()])
#            axs[plotNo//5, plotNo%5].tick_params(axis='both',width = 0.3)
#    
#        plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
#        plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.99, linestyle='--') # Baseline, full RES
#        plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8,  linestyle='-')  # RCP2.6, no RES
#        plt.plot([2010,2011],[0,0],color=R26Green, lw=0.99, linestyle='--') # RCP2.6, full RES
#        plt.legend(LegendLables,shadow = False,  prop={'size':7}, loc = 'upper right',bbox_to_anchor=(3.5, 1))    
#        
#        fig.suptitle(fit_5x5[2*mmf+Sect], fontsize=14)
#        for xm in range(0,5):
#            plt.sca(axs[4,xm])
#            plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 6, fontweight = 'normal')
#        plt.show()
#        fig_name  = fin_5x5[2*mmf+Sect]
#        fig.savefig(os.path.join(RECC_Paths.results_path,fig_name),  dpi = PlotExpResolution, bbox_inches='tight')  
#        fig_namev = fiv_5x5[2*mmf+Sect]
#        fig.savefig(os.path.join(RECC_Paths.results_path,fig_namev), dpi = PlotExpResolution, bbox_inches='tight')  
        
#GHG intensity plot
RegionList20Plot = Pav_RegionList20Plot
PlotOrder        = PlotOrder_pav
fig, axs = plt.subplots(5, 5, sharex=True, gridspec_kw={'hspace': 0.22, 'wspace': 0.5})
for plotNo in PlotOrder:
    axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), 3.6 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,0],color=BaseBlue, lw=0.8,  linestyle='-')    # Baseline, no RES
    axs[plotNo//5, plotNo%5].plot(np.arange(2016,2061), 3.6 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,1],color=R26Green, lw=0.8,  linestyle='-')    # RCP2.6, no RES
    axs[plotNo//5, plotNo%5].set_ylim(bottom=0)
    axs[plotNo//5, plotNo%5].text(2015, 0.12 * 3.6 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,:].max(), RegionList20Plot[plotNo], fontsize=6, rotation=0, fontweight='normal')
    axs[plotNo//5, plotNo%5].text(2045, 0.86 * 3.6 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,:].max(), str(np.round(3.6 * 1e6 * TimeSeries_All[10,-1,plotNo,0,1,0])), fontsize=4, rotation=0, fontweight='normal', color = BaseBlue)
    axs[plotNo//5, plotNo%5].text(2045, 0.56 * 3.6 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,:].max(), str(np.round(3.6 * 1e6 * TimeSeries_All[10,-1,plotNo,0,1,1])), fontsize=4, rotation=0, fontweight='normal', color = R26Green)
    axs[plotNo//5, plotNo%5].tick_params(axis='x', labelsize=6)
    axs[plotNo//5, plotNo%5].tick_params(axis='y', labelsize=6)
    for axis in ['top','bottom','left','right']:
        axs[plotNo//5, plotNo%5].spines[axis].set_linewidth(0.3)
    axs[plotNo//5, plotNo%5].axis([2012, 2063, 0, 4.0 * 1e6 * TimeSeries_All[10,:,plotNo,0,1,:].max()])
    axs[plotNo//5, plotNo%5].tick_params(axis='both',width = 0.3)

plt.plot([2010,2011],[0,0],color=BaseBlue, lw=0.8,  linestyle='-')  # Baseline, no RES
plt.plot([2010,2011],[0,0],color=R26Green, lw=0.8, linestyle='-') # RCP2.6, full RES
plt.legend(['NoNewClimPol','RCP2.6'],shadow = False,  prop={'size':7}, loc = 'upper right',bbox_to_anchor=(3.2, 1))    

fig.suptitle(r'GHG intensity of electricity by region, g CO$_2$-eq/kWh', fontsize=14)
for xm in range(0,5):
    plt.sca(axs[4,xm])
    plt.xticks([2020,2030,2040,2050,2060], ['2020','2030','2040','2050','2060'], rotation =90, fontsize = 6, fontweight = 'normal')
plt.show()
fig_name  = 'GHG_intensity.png'
fig.savefig(os.path.join(RECC_Paths.results_path,fig_name),  dpi = PlotExpResolution, bbox_inches='tight')  
        
# Excel export global data via pandas:
# TimeSeries_All indices: NX x Nt x Nr x NV x NS x NR / indicators x time x regions x SSP x RCP 
ColIndex      = [str(mmx) for mmx in  range(2016,2061)]
if len(PlotOrder_pav) == 25:        # only if data for all regions were exported
    # pav:
    DF_Data_pav   = np.einsum('XtrSR->XrSRt',TimeSeries_All[0:2,:,:,0,:,:]).reshape(2*25*3*2,45)
    RowIndex      = pd.MultiIndex.from_product([['no ME','full ME'],Pav_RegionList20Plot,['LED','SSP1','SSP2'],['NoNewClimPol','RCP2.6']], names=('res. eff.','region','SSP','RCP'))
    DF_pav_global = pd.DataFrame(DF_Data_pav, index=RowIndex, columns=ColIndex)
    DF_pav_global.to_excel(os.path.join(RECC_Paths.results_path,'Fig_GHG_pav_5x5.xls'), merge_cells=False)
    #print(ColIndex)
    #print(RowIndex)
if len(PlotOrder_reb) == 25:        # only if data for all regions were exported
    # reb:
    DF_Data_reb   = np.einsum('XtrSR->XrSRt',TimeSeries_All[0:2,:,:,1,:,:]).reshape(2*25*3*2,45)
    RowIndex      = pd.MultiIndex.from_product([['no ME','full ME'],Reb_RegionList20Plot,['LED','SSP1','SSP2'],['NoNewClimPol','RCP2.6']], names=('res. eff.','region','SSP','RCP'))
    DF_reb_global = pd.DataFrame(DF_Data_reb, index=RowIndex, columns=ColIndex)
    DF_reb_global.to_excel(os.path.join(RECC_Paths.results_path,'Fig_GHG_reb_5x5.xls'), merge_cells=False)
    
#
#
#
#
#
#
#
#