# -*- coding: utf-8 -*-
"""
Created on Wed Oct 17 10:37:00 2018

@author: spauliuk
"""
def main(RegionalScope,FolderList,SectorString,Current_UUID):
    
    import openpyxl
    import numpy as np
    import matplotlib.pyplot as plt  
    import pylab
    import pandas as pd
    import os
    import RECC_Paths # Import path file   #
    
    RECC_Paths.results_path_save = os.path.join(RECC_Paths.results_path_eval,'RECC_Results_' + Current_UUID)
    
    PlotExpResolution = 150 # dpi 150 for overview or 500 for paper
    
    # FileOrder needs to be kept:
    # pav:
    # 1) None
    # 2) + EoL + FSD + FYI
    # 3) + EoL + FSD + FYI + ReU +LTE
    # 4) + EoL + FSD + FYI + ReU +LTE + MSu
    # 5) + EoL + FSD + FYI + ReU +LTE + MSu + LWE
    # 6) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS 
    # 7) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS = ALL 
    
    # reb and nrb:
    # 1) None
    # 2) + EoL + FSD + FYI
    # 3) + EoL + FSD + FYI + ReU +LTE
    # 4) + EoL + FSD + FYI + ReU +LTE + MSu
    # 5) + EoL + FSD + FYI + ReU +LTE + MSu + LWE 
    # 6) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + MIU = ALL 
    
    # pav and reb/nrb combined:
    # 1) None
    # 2) + EoL + FSD + FYI
    # 3) + EoL + FSD + FYI + ReU +LTE
    # 4) + EoL + FSD + FYI + ReU +LTE + MSu
    # 5) + EoL + FSD + FYI + ReU +LTE + MSu + LWE
    # 6) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS 
    # 7) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS
    # 8) + EoL + FSD + FYI + ReU +LTE + MSu + LWE + CaS + RiS + MIU = ALL     
    
    # Waterfall plots.
    
    NS = 3  # no of SSP scenarios
    NR = 2  # no of RCP scenarios
    Nt = 45 # no of model years
    Nm = 6  # no of materials for which data are extracted: 0: steel (all kinds), 1: Al (all kinds), 2: copper, 3: cement, 4: plastics, 5: wood
    
    if SectorString == 'pav':
        NE      = 7 # no of Res. eff. scenarios for cascade
        LWE     = ['No ME','higher yields', 're-use/longer use','material subst.','down-sizing','car-sharing','ride-sharing','All ME stratgs.']
        Offset1 = 7.25
        Offset2 = 5.85
        Offset3 = 3.3
        XTicks  = [0.25,1.25,2.25,3.25,4.25,5.25,6.25,7.25]
        Offset4 = 7.7
        LWE_area= ['total, no ME','higher yields', 're-use & LTE','material subst.','down-sizing','car-sharing','ride-sharing','residual, full ME']   
        PlotCtrl= 0 
        ColOrder= [0,1,2,3,4,5,6,7]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.14)) # select 12 colors from the 'Paired' color map.  
        LLeft   = 7.5
        
    if SectorString == 'reb':
        NE      = 6 # no of Res. eff. scenarios for cascade
        LWE     = ['No ME','higher yields', 're-use/longer use','material subst.','light design','more intense use','All ME stratgs.']
        Offset1 = 6.25
        Offset2 = 5.00
        Offset3 = 2.8
        XTicks  = [0.25,1.25,2.25,3.25,4.25,5.25,6.25]
        Offset4 = 6.7
        LWE_area= ['total, no ME','higher yields', 're-use & LTE','material subst.','light design','more intense use','residual, full ME']    
        PlotCtrl= 1
        ColOrder= [0,1,2,3,4,5,6]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.14)) # select 12 colors from the 'Paired' color map.   
        LLeft   = 7.5
        
    if SectorString == 'nrb':
        NE      = 6 # no of Res. eff. scenarios for cascade
        LWE     = ['No ME','higher yields', 're-use/longer use','material subst.','light design','more intense use','All ME stratgs.']
        Offset1 = 6.25
        Offset2 = 5.00
        Offset3 = 2.8
        XTicks  = [0.25,1.25,2.25,3.25,4.25,5.25,6.25]
        Offset4 = 6.7
        LWE_area= ['total, no ME','higher yields', 're-use & LTE','material subst.','light design','more intense use','residual, full ME']    
        PlotCtrl= 1
        ColOrder= [0,1,2,3,4,5,6]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.14)) # select 12 colors from the 'Paired' color map.   
        LLeft   = 7.5
        
    if SectorString == 'pav_reb' or SectorString == 'pav_nrb':
        NE      = 8 # no of Res. eff. scenarios for cascade
        LWE    = ['No ME','higher yields', 're-use/longer use','material subst.','down-sizing','car-sharing','ride-sharing','More intense bld. use','All ME stratgs.']
        Offset1 = 8.25
        Offset2 = 6.85
        Offset3 = 4.3
        XTicks  = [0.25,1.25,2.25,3.25,4.25,5.25,6.25,7.25,8.25]
        Offset4 = 8.7
        LWE_area= ['total, no ME','higher yields', 're-use & LTE','material subst.','down-sizing','car-sharing','ride-sharing','More intense bld. use','residual, full ME']      
        PlotCtrl= 1
        ColOrder= [11,4,0,18,8,16,2,6,15]
        MyColorCycle = pylab.cm.tab20(np.arange(0,1,0.05)) # select 20 colors from the 'tab20' color map. 
        LLeft   = 8.5

    if SectorString == 'pav_reb_nrb':
        NE      = 8 # no of Res. eff. scenarios for cascade
        LWE    = ['No ME','higher yields', 're-use/longer use','material subst.','down-sizing','car-sharing','ride-sharing','More intense bld. use','All ME stratgs.']
        Offset1 = 8.25
        Offset2 = 6.85
        Offset3 = 4.3
        XTicks  = [0.25,1.25,2.25,3.25,4.25,5.25,6.25,7.25,8.25]
        Offset4 = 8.7
        LWE_area= ['total, no ME','higher yields', 're-use & LTE','material subst.','down-sizing','car-sharing','ride-sharing','More intense bld. use','residual, full ME']     
        PlotCtrl= 1
        ColOrder= [11,4,0,18,8,16,2,6,15]
        MyColorCycle = pylab.cm.tab20(np.arange(0,1,0.05)) # select 20 colors from the 'tab20' color map. 
        LLeft   = 8.5
        
    # Population, total over all region (if more than 1):
    Population       = np.zeros((NS,Nt))
    # system-wide emissions:
    CumEms2050       = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario: cum. emissions 2016-2050.
    CumEms2060       = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario: cum. emissions 2016-2060.
    AnnEms2030       = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario: ann. emissions 2030.
    AnnEms2050       = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario: ann. emissions 2050.
    AvgDecadalEms    = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    ASummary         = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.
    # for use phase di emissions:
    UseCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    UseAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalUseEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    UsePhaseSummary  = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.    
    # for material-related emissions:
    MatCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalMatEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    MatSummary       = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.
    # for manufacturing-related emissions:
    ManCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ManAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalManEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    ManSummary       = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.
    # for forestry and wood waste related emissions:
    ForCumEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForCumEms2060    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForAnnEms2030    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    ForAnnEms2050    = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    AvgDecadalForEms = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    ForSummary       = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.    
    # for recycling credit:
    RecCreditCum2050 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditCum2060 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditAnn2030 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    RecCreditAnn2050 = np.zeros((NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario    
    RecCreditAvgDec  = np.zeros((NS,NR,NE,4)) # SSP-Scenario x RCP scenario x RES scenario: avg. emissions per decade 2020-2030 ... 2050-2060
    RecCredit        = np.zeros((12,NR,NE)) # different indices compiled x RCP x RES.
    
    TimeSeries_R     = np.zeros((20,NE,45,3,2)) # NX x NE x Nt x NS x NR / indicators x RES x time x SSP x RCP # starts counting at 2016!
    # 2 system scopes x 3 indicators x SSP-Scenario x RCP scenario x RES scenario 
    # 0: system-wide GHG, 1: material-related GHG, 2: primar production, all materials, 3: secondary production, all materials.
    # 4: share of el + H2 in total use phase energy consumption, 5: electricity suppy GHG use phase, 6: In-use stock, all materials.
    # 7: use phase energy consumption, 8: wood use carbon balance (forest and waste mgt.),
    # 9: passenger-km, 10: heated building space, 11: cooled building space.
    CascDataExp      = np.zeros((2,3,NS,NR,NE)) 

    # get result items:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[0])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[0],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    # Find the index for sysem-wide emissions, the recycling credit and others:
    swe = 1    
    while True:
        if Resultsheet2.cell(swe+1, 1).value == 'GHG emissions, system-wide _3579di':
            break # that gives us the right index to read the recycling credit from the result table.
        swe += 1 

    for r in range(0,NE): # RE scenario                
        # import system-wide GHG and material-related emissions
        ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
        Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[r],ResFile[0]))
        Resultsheet2 = Resultfile2['Model_Results']
        # system-wide emissions results
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time
                    CumEms2050[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value
                for t in range(0,45): # time
                    CumEms2060[s,c,r] += Resultsheet2.cell(swe+ 2*s +c+1,t+9).value    
                AnnEms2030[s,c,r]   = Resultsheet2.cell(swe+ 2*s +c+1,23).value
                AnnEms2050[s,c,r]   = Resultsheet2.cell(swe+ 2*s +c+1,43).value
                AvgDecadalEms[s,c,r,0]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(13,23)])/10
                AvgDecadalEms[s,c,r,1]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(23,33)])/10
                AvgDecadalEms[s,c,r,2]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(33,43)])/10
                AvgDecadalEms[s,c,r,3]   = sum([Resultsheet2.cell(swe+ 2*s +c+1,t+1).value for i in range(43,53)])/10        

    ASummary[0:3,:] = AnnEms2030.copy()
    ASummary[3:6,:] = AnnEms2050.copy()
    ASummary[6:9,:] = CumEms2050.copy()
    ASummary[9::,:] = CumEms2060.copy()                        
    
    # Waterfall plot            
    Title  = ['CumGHG_16_50','CumGHG_40_50','AnnGHG_50']
    Scens  = ['LED','SSP1','SSP2']
    Rcens  = ['Base','RCP2_6']
    
    for nn in range(0,3):
        for m in range(0,NS): # SSP
            for rcp in range(0,NR): # RCP
                if nn == 0:
                    Data = np.einsum('SE->ES',CumEms2050[:,rcp,:])
                    CascDataExp[0,0,:,:,:] = CumEms2050.copy()
                if nn == 1:
                    Data = np.einsum('SE->ES',10*AvgDecadalEms[:,rcp,:,2])
                    CascDataExp[0,1,:,:,:] = 10*AvgDecadalEms[:,:,:,2].copy()
                if nn == 2:
                    Data = np.einsum('SE->ES',AnnEms2050[:,rcp,:])
                    CascDataExp[0,2,:,:,:] = AnnEms2050.copy()
                    
                inc = -100 * (Data[0,m] - Data[-1,m])/Data[0,m]
            
                Left  = Data[0,m]
                Right = Data[-1,m]
                # plot results
                bw = 0.5
            
                fig  = plt.figure(figsize=(5,8))
                ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
                ProxyHandlesList = []   # For legend     
                # plot bars
                ax1.fill_between([0,0+bw], [0,0],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[0],:], linewidth = 0.0)
                ax1.fill_between([1,1+bw], [Data[1,m],Data[1,m]],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[1],:], linewidth = 0.0)
                for xca in range(2,NE):
                    ax1.fill_between([xca,xca+bw], [Data[xca,m],Data[xca,m]],[Data[xca-1,m],Data[xca-1,m]],linestyle = '--', facecolor =MyColorCycle[ColOrder[xca],:], linewidth = 0.0)
                ax1.fill_between([NE,NE+bw], [0,0],[Data[NE-1,m],Data[NE-1,m]],linestyle = '--', facecolor =MyColorCycle[ColOrder[NE],:], linewidth = 0.0)                
                    
                for fca in range(0,NE+1):
                    ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[ColOrder[fca],:])) # create proxy artist for legend
                
                # plot lines:
                plt.plot([0,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
                for yca in range(1,NE):
                    plt.plot([yca,yca +1.5],[Data[yca,m],Data[yca,m]],linestyle = '-', linewidth = 0.5, color = 'k')
                    
                plt.arrow(Offset1, Data[NE-1,m],0, Data[0,m]-Data[NE-1,m], lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
                plt.arrow(Offset1,Data[0,m],0,Data[NE-1,m]-Data[0,m], lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
                    
                # plot text and labels
                plt.text(Offset2, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
                plt.text(Offset3, 0.94  *Right, Scens[m],fontsize=18,fontweight='bold') 
                plt.title('RE strats. and GHG emissions, ' + SectorString + '.', fontsize = 18)
                plt.ylabel(Title[nn] + r', Mt CO$_2$-eq.', fontsize = 18)
                plt.xticks(XTicks)
                plt.yticks(fontsize =18)
                ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
                plt.legend(handles = ProxyHandlesList,labels = LWE,shadow = False, prop={'size':12},ncol=1, loc = 'upper right' ,bbox_to_anchor=(1.91, 1)) 
                #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
                plt.axis([-0.2, Offset4, 0, 1.02*Left])
            
                plt.show()
                fig_name = RegionalScope + '_' + SectorString + '_' + Title[nn] + '_' + Scens[m] + '_' + Rcens[rcp] + '.png'
                fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = PlotExpResolution, bbox_inches='tight')             
                
    ### Area plot RE
    AnnEms             = np.zeros((Nt,NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatEms             = np.zeros((Nt,NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
    MatStocks          = np.zeros((Nt,Nm,NS,NR,NE))
    MatProduction_Prim = np.zeros((Nt,Nm,NS,NR,NE))
    MatProduction_Sec  = np.zeros((Nt,Nm,NS,NR,NE))
    
    # First, get the position indices for the different result variables:
    ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[0])) if filename.startswith('ODYM_RECC_ModelResults_')]
    Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[0],ResFile[0]))
    Resultsheet2 = Resultfile2['Model_Results']
    
    rci = 1
    while True:
        if Resultsheet2.cell(rci+1,1).value == 'GHG emissions, recycling credits':
            break # that gives us the right index to read the recycling credit from the result table.
        rci += 1
    mci = 1
    while True:
        if Resultsheet2.cell(mci+1,1).value == 'GHG emissions, material cycle industries and their energy supply _3di_9di':
            break # that gives us the right index to read the recycling credit from the result table.
        mci += 1
    mp1 = 1
    while True:
        if Resultsheet2.cell(mp1+1,1).value == 'Primary materials, total':
            break # that gives us the right index to read the recycling credit from the result table.
        mp1 += 1    
    ms1 = 1
    while True:
        if Resultsheet2.cell(ms1+1,1).value == 'In-use stock, construction grade steel':
            break # that gives us the right index from the result table.
        ms1 += 1            
    ms2 = 1
    while True:
        if Resultsheet2.cell(ms2+1,1).value == 'In-use stock, automotive steel':
            break # that gives us the right index from the result table.
        ms2 += 1 
    ms3 = 1
    while True:
        if Resultsheet2.cell(ms3+1,1).value == 'In-use stock, stainless steel':
            break # that gives us the right index from the result table.
        ms3 += 1 
    ms4 = 1
    while True:
        if Resultsheet2.cell(ms4+1,1).value == 'In-use stock, cast iron':
            break # that gives us the right index from the result table.
        ms4 += 1 
    ms5 = 1
    while True:
        if Resultsheet2.cell(ms5+1,1).value == 'In-use stock, wrought Al':
            break # that gives us the right index from the result table.
        ms5 += 1 
    ms6 = 1
    while True:
        if Resultsheet2.cell(ms6+1,1).value == 'In-use stock, cast Al':
            break # that gives us the right index from the result table.
        ms6 += 1 
    ms7 = 1
    while True:
        if Resultsheet2.cell(ms7+1,1).value == 'In-use stock, copper electric grade':
            break # that gives us the right index from the result table.
        ms7 += 1 
    ms8 = 1
    while True:
        if Resultsheet2.cell(ms8+1,1).value == 'In-use stock, plastics':
            break # that gives us the right index from the result table.
        ms8 += 1 
    ms9 = 1
    while True:
        if Resultsheet2.cell(ms9+1,1).value == 'In-use stock, cement':
            break # that gives us the right index from the result table.
        ms9 += 1 
    ms10 = 1
    while True:
        if Resultsheet2.cell(ms10+1,1).value == 'In-use stock, wood and wood products':
            break # that gives us the right index from the result table.
        ms10 += 1 

    mc1 = 1
    while True:
        if Resultsheet2.cell(mc1+1,1).value == 'Primary steel production':
            break # that gives us the right index from the result table.
        mc1 += 1
    mc2 = 1
    while True:
        if Resultsheet2.cell(mc2+1,1).value == 'Primary Al production':
            break # that gives us the right index from the result table.
        mc2 += 1
    mc3 = 1
    while True:
        if Resultsheet2.cell(mc3+1,1).value == 'Primary Cu production':
            break # that gives us the right index from the result table.
        mc3 += 1
    mc4 = 1
    while True:
        if Resultsheet2.cell(mc4+1,1).value == 'Cement production':
            break # that gives us the right index from the result table.
        mc4 += 1
    mc5 = 1
    while True:
        if Resultsheet2.cell(mc5+1,1).value == 'Primary plastics production':
            break # that gives us the right index from the result table.
        mc5 += 1
    mc6 = 1
    while True:
        if Resultsheet2.cell(mc6+1,1).value == 'Wood, from forests':
            break # that gives us the right index from the result table.
        mc6 += 1
    mc7 = 1
    while True:
        if Resultsheet2.cell(mc7+1,1).value == 'Secondary steel':
            break # that gives us the right index from the result table.
        mc7 += 1
    mc8 = 1
    while True:
        if Resultsheet2.cell(mc8+1,1).value == 'Secondary Al':
            break # that gives us the right index from the result table.
        mc8 += 1            
    mc9 = 1
    while True:
        if Resultsheet2.cell(mc9+1,1).value == 'Secondary copper':
            break # that gives us the right index from the result table.
        mc9 += 1   
    mc10 = 1
    while True:
        if Resultsheet2.cell(mc10+1,1).value == 'Secondary plastics':
            break # that gives us the right index from the result table.
        mc10 += 1   
    mc11 = 1
    while True:
        if Resultsheet2.cell(mc11+1,1).value == 'Recycled wood':
            break # that gives us the right index from the result table.
        mc11 += 1               
        
    ru1 = 1
    while True:
        if Resultsheet2.cell(ru1+1,1).value == 'ReUse of materials in products, construction grade steel':
            break # that gives us the right index from the result table.
        ru1 += 1 
    ru2 = 1
    while True:
        if Resultsheet2.cell(ru2+1,1).value == 'ReUse of materials in products, automotive steel':
            break # that gives us the right index from the result table.
        ru2 += 1             
    ru3 = 1
    while True:
        if Resultsheet2.cell(ru3+1,1).value == 'ReUse of materials in products, stainless steel':
            break # that gives us the right index from the result table.
        ru3 += 1 
    ru4 = 1
    while True:
        if Resultsheet2.cell(ru4+1,1).value == 'ReUse of materials in products, cast iron':
            break # that gives us the right index from the result table.
        ru4 += 1 
    ru5 = 1
    while True:
        if Resultsheet2.cell(ru5+1,1).value == 'ReUse of materials in products, wrought Al':
            break # that gives us the right index from the result table.
        ru5 += 1 
    ru6 = 1
    while True:
        if Resultsheet2.cell(ru6+1,1).value == 'ReUse of materials in products, cast Al':
            break # that gives us the right index from the result table.
        ru6 += 1
    ru7 = 1
    while True:
        if Resultsheet2.cell(ru7+1,1).value == 'ReUse of materials in products, copper electric grade':
            break # that gives us the right index from the result table.
        ru7 += 1
    ru8 = 1
    while True:
        if Resultsheet2.cell(ru8+1,1).value == 'ReUse of materials in products, plastics':
            break # that gives us the right index from the result table.
        ru8 += 1
    ru9 = 1
    while True:
        if Resultsheet2.cell(ru9+1,1).value == 'ReUse of materials in products, cement':
            break # that gives us the right index from the result table.
        ru9 += 1 
    ru10 = 1
    while True:
        if Resultsheet2.cell(ru10+1,1).value == 'ReUse of materials in products, wood and wood products':
            break # that gives us the right index from the result table.
        ru10 += 1    
        
    mp2 = 1
    while True:
        if Resultsheet2.cell(mp2+1,1).value == 'Secondary materials, total':
            break # that gives us the right index from the result table.
        mp2 += 1  
    
    up1i = 1
    while True:
        if Resultsheet2.cell(up1i+1,1).value == 'GHG emissions, use phase _7d':
            break # that gives us the right index from the result table.
        up1i += 1  
    up2i = 1
    while True:
        if Resultsheet2.cell(up2i+1,1).value == 'GHG emissions, use phase scope 2 (electricity) _7i':
            break # that gives us the right index from the result table.
        up2i += 1  
    up3i = 1
    while True:
        if Resultsheet2.cell(up3i+1,1).value == 'GHG emissions, use phase other indirect (non-el.) _7i':
            break # that gives us the right index from the result table.
        up3i += 1  

    mfi = 1
    while True:
        if Resultsheet2.cell(mfi+1,1).value == 'GHG emissions, manufacturing _5i, all':
            break # that gives us the right index from the result table.
        mfi += 1 
    fci = 1
    while True:
        if Resultsheet2.cell(fci+1,1).value == 'GHG emissions, energy recovery from waste wood (biogenic C plus energy substitution within System)':
            break # that gives us the right index from the result table.
        fci += 1               
    wci = 1
    while True:
        if Resultsheet2.cell(wci+1,1).value == 'GHG sequestration by forests (w. neg. sign)':
            break # that gives us the right index from the result table.
        wci += 1         
    en1 = 1
    while True:
        if Resultsheet2.cell(en1+1,1).value == 'energy consumption, use phase: electricity':
            break # that gives us the right index from the result table.
        en1 += 1         
    en2 = 1
    while True:
        if Resultsheet2.cell(en2+1,1).value == 'energy consumption, use phase: hydrogen':
            break # that gives us the right index from the result table.
        en2 += 2         
    en3 = 1
    while True:
        if Resultsheet2.cell(en3+1,1).value == 'energy consumption, use phase: all':
            break # that gives us the right index from the result table.
        en3 += 1
    en4 = 1
    while True:
        if Resultsheet2.cell(en4+1,1).value == 'GHG emissions, use phase scope 2 (electricity) _7i':
            break # that gives us the right index from the result table.
        en4 += 1    
    am1 = 1
    while True:
        if Resultsheet2.cell(am1+1,1).value == 'In-use stock, all materials':
            break # that gives us the right index from the result table.
        am1 += 1 
    popc = 1
    while True:
        if Resultsheet2.cell(popc+1,1).value == 'Population':
            break # that gives us the right index from the result table.
        popc += 1                
    if SectorString.find('pav') >= 0:
        pkm = 1
        while True:
            if Resultsheet2.cell(pkm+1,1).value == 'passenger-km supplied by pass. vehicles':
                break # that gives us the right index from the result table.
            pkm += 1            
    if SectorString.find('reb') >= 0:            
        bs1 = 1
        while True:
            if Resultsheet2.cell(bs1+1,1).value == 'Total heated floor space, res. buildings':
                break # that gives us the right index from the result table.
            bs1 += 1 
        bs2 = 1
        while True:
            if Resultsheet2.cell(bs2+1,1).value == 'Total cooled floor space, res. buildings':
                break # that gives us the right index from the result table.
            bs2 += 1 
        
    for r in range(0,NE): # RE scenario
        ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
        Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[r],ResFile[0]))
        Resultsheet2 = Resultfile2['Model_Results']
            
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,45): # time
                    AnnEms[t,s,c,r]       = Resultsheet2.cell(swe+ 2*s +c+1,t+9).value
                    MatEms[t,s,c,r]       = Resultsheet2.cell(mci+ 2*s +c+1,t+9).value
        # Use phase results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    UseCumEms2050[s,c,r] += Resultsheet2.cell(up1i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up2i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up3i+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    UseCumEms2060[s,c,r] += Resultsheet2.cell(up1i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up2i+ 2*s +c+1,t+9).value + Resultsheet2.cell(up3i+ 2*s +c+1,t+9).value                    
                UseAnnEms2030[s,c,r]      = Resultsheet2.cell(up1i+ 2*s +c+1,23).value  + Resultsheet2.cell(up2i+ 2*s +c+1,23).value  + Resultsheet2.cell(up3i+ 2*s +c+1,23).value  
                UseAnnEms2050[s,c,r]      = Resultsheet2.cell(up1i+ 2*s +c+1,43).value  + Resultsheet2.cell(up2i+ 2*s +c+1,43).value  + Resultsheet2.cell(up3i+ 2*s +c+1,43).value  
                AvgDecadalUseEms[s,c,r,0] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalUseEms[s,c,r,1] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalUseEms[s,c,r,2] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalUseEms[s,c,r,3] = sum([Resultsheet2.cell(up1i+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(up2i+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(up3i+ 2*s +c+1,t+1).value for t in range(43,53)])/10          
        # Material results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    MatCumEms2050[s,c,r] += Resultsheet2.cell(mci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    MatCumEms2060[s,c,r] += Resultsheet2.cell(mci+ 2*s +c+1,t+9).value                    
                    TimeSeries_R[1,r,t,s,c] = Resultsheet2.cell(mci+ 2*s +c+1,t+9).value 
                    TimeSeries_R[2,r,t,s,c] = Resultsheet2.cell(mp1+ 2*s +c+1,t+9).value 
                    TimeSeries_R[3,r,t,s,c] = Resultsheet2.cell(mp2+ 2*s +c+1,t+9).value 
                    try: # works only if total energy is not 0.
                        ElH2share = (Resultsheet2.cell(en1+ 2*s +c+1,t+9).value + Resultsheet2.cell(en2+ 2*s +c+1,t+9).value)/Resultsheet2.cell(en3+ 2*s +c+1,t+9).value
                        TimeSeries_R[4,r,t,s,c] = ElH2share
                    except:
                        None
                    try: 
                        TimeSeries_R[5,r,t,s,c] = Resultsheet2.cell(en4+ 2*s +c+1,t+9).value / Resultsheet2.cell(en1+ 2*s +c+1,t+9).value # ton/MJ
                    except: 
                        None
                    TimeSeries_R[6,r,t,s,c] = Resultsheet2.cell(am1+ 2*s +c+1,t+9).value   
                    TimeSeries_R[7,r,t,s,c] = Resultsheet2.cell(en3+ 2*s +c+1,t+9).value     
                    TimeSeries_R[8,r,t,s,c] = Resultsheet2.cell(fci+ 2*s +c+1,t+9).value + Resultsheet2.cell(wci+ 2*s +c+1,t+9).value
                    if SectorString.find('pav') >= 0:
                        TimeSeries_R[9,r,t,s,c] = Resultsheet2.cell(pkm+ 2*s +c+1,t+9).value 
                    if SectorString.find('reb') >= 0:
                        TimeSeries_R[10,r,t,s,c]= Resultsheet2.cell(bs1+ 2*s +c+1,t+9).value 
                        TimeSeries_R[11,r,t,s,c]= Resultsheet2.cell(bs2+ 2*s +c+1,t+9).value 
                MatAnnEms2030[s,c,r]      = Resultsheet2.cell(mci+ 2*s +c+1,23).value
                MatAnnEms2050[s,c,r]      = Resultsheet2.cell(mci+ 2*s +c+1,43).value
                AvgDecadalMatEms[s,c,r,0] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalMatEms[s,c,r,1] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalMatEms[s,c,r,2] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalMatEms[s,c,r,3] = sum([Resultsheet2.cell(mci+ 2*s +c+1,t+1).value for t in range(43,53)])/10    
        # Manufacturing results export 
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    ManCumEms2050[s,c,r] += Resultsheet2.cell(mfi+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    ManCumEms2060[s,c,r] += Resultsheet2.cell(mfi+ 2*s +c+1,t+9).value                    
                ManAnnEms2030[s,c,r]      = Resultsheet2.cell(mfi+ 2*s +c+1,23).value
                ManAnnEms2050[s,c,r]      = Resultsheet2.cell(mfi+ 2*s +c+1,43).value
                AvgDecadalManEms[s,c,r,0] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalManEms[s,c,r,1] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalManEms[s,c,r,2] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalManEms[s,c,r,3] = sum([Resultsheet2.cell(mfi+ 2*s +c+1,t+1).value for t in range(43,53)])/10                 
        # Forestry results export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR): # RCP scenario
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    ForCumEms2050[s,c,r] += Resultsheet2.cell(fci+ 2*s +c+1,t+9).value + Resultsheet2.cell(wci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    ForCumEms2060[s,c,r] += Resultsheet2.cell(fci+ 2*s +c+1,t+9).value + Resultsheet2.cell(wci+ 2*s +c+1,t+9).value                    
                ForAnnEms2030[s,c,r]      = Resultsheet2.cell(fci+ 2*s +c+1,23).value  + Resultsheet2.cell(wci+ 2*s +c+1,23).value
                ForAnnEms2050[s,c,r]      = Resultsheet2.cell(fci+ 2*s +c+1,43).value  + Resultsheet2.cell(wci+ 2*s +c+1,43).value
                AvgDecadalForEms[s,c,r,0] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(13,23)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(13,23)])/10
                AvgDecadalForEms[s,c,r,1] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(23,33)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(23,33)])/10
                AvgDecadalForEms[s,c,r,2] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(33,43)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(33,43)])/10
                AvgDecadalForEms[s,c,r,3] = sum([Resultsheet2.cell(fci+ 2*s +c+1,t+1).value for t in range(43,53)])/10 + sum([Resultsheet2.cell(wci+ 2*s +c+1,t+1).value for t in range(43,53)])/10              
        # recycling credit
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,35): # time until 2050 only!!! Cum. emissions until 2050.
                    RecCreditCum2050[s,c,r]+= Resultsheet2.cell(rci+ 2*s +c+1,t+9).value
                for t in range(0,45): # time until 2060.
                    RecCreditCum2060[s,c,r]+= Resultsheet2.cell(rci+ 2*s +c+1,t+9).value
                RecCreditAnn2030[s,c,r]     = Resultsheet2.cell(rci+ 2*s +c+1,23).value
                RecCreditAnn2050[s,c,r]     = Resultsheet2.cell(rci+ 2*s +c+1,43).value
                RecCreditAvgDec[s,c,r,0]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(13,23)])/10
                RecCreditAvgDec[s,c,r,1]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(23,33)])/10
                RecCreditAvgDec[s,c,r,2]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(33,43)])/10
                RecCreditAvgDec[s,c,r,3]= sum([Resultsheet2.cell(rci+ 2*s +2,t+1).value for t in range(43,53)])/10                       

        # Material stocks export
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,45): # time until 2060
                    MatStocks[t,0,s,c,r]  = Resultsheet2.cell(ms1+ 2*s +c+1,t+9).value + Resultsheet2.cell(ms2+ 2*s +c+1,t+9).value + Resultsheet2.cell(ms3+ 2*s +c+1,t+9).value + Resultsheet2.cell(ms4+ 2*s +c+1,t+9).value
                    MatStocks[t,1,s,c,r]  = Resultsheet2.cell(ms5+ 2*s +c+1,t+9).value + Resultsheet2.cell(ms6+ 2*s +c+1,t+9).value
                    MatStocks[t,2,s,c,r]  = Resultsheet2.cell(ms7+ 2*s +c+1,t+9).value
                    MatStocks[t,3,s,c,r]  = Resultsheet2.cell(ms9+ 2*s +c+1,t+9).value
                    MatStocks[t,4,s,c,r]  = Resultsheet2.cell(ms8+ 2*s +c+1,t+9).value
                    MatStocks[t,5,s,c,r]  = Resultsheet2.cell(ms10+ 2*s +c+1,t+9).value
                    
        # Material results export, prim. and secondary prod.
        for s in range(0,NS): # SSP scenario
            for c in range(0,NR):
                for t in range(0,45): # time until 2060
                    MatProduction_Prim[t,0,s,c,r] = Resultsheet2.cell(mc1+ 2*s +c+1,t+9).value
                    MatProduction_Prim[t,1,s,c,r] = Resultsheet2.cell(mc2+ 2*s +c+1,t+9).value
                    MatProduction_Prim[t,2,s,c,r] = Resultsheet2.cell(mc3+ 2*s +c+1,t+9).value
                    MatProduction_Prim[t,3,s,c,r] = Resultsheet2.cell(mc4+ 2*s +c+1,t+9).value
                    MatProduction_Prim[t,4,s,c,r] = Resultsheet2.cell(mc5+ 2*s +c+1,t+9).value
                    MatProduction_Prim[t,5,s,c,r] = Resultsheet2.cell(mc6+ 2*s +c+1,t+9).value
                    
                    MatProduction_Sec[t,0,s,c,r]  = Resultsheet2.cell(mc7+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru1+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru2+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru3+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru4+ 2*s +c+1,t+9).value
                    MatProduction_Sec[t,1,s,c,r]  = Resultsheet2.cell(mc8+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru5+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru6+ 2*s +c+1,t+9).value
                    MatProduction_Sec[t,2,s,c,r]  = Resultsheet2.cell(mc9+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru7+ 2*s +c+1,t+9).value
                    MatProduction_Sec[t,3,s,c,r]  = Resultsheet2.cell(ru9+ 2*s +c+1,t+9).value
                    MatProduction_Sec[t,4,s,c,r]  = Resultsheet2.cell(mc10+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru8+ 2*s +c+1,t+9).value
                    MatProduction_Sec[t,5,s,c,r]  = Resultsheet2.cell(mc11+ 2*s +c+1,t+9).value + Resultsheet2.cell(ru10+ 2*s +c+1,t+9).value    
        
        # Population
        # Here, the regional total is not exported, which is why it has to be summed up from the individual countries/regions in the larger regions.
        # Unlike for the other indicators, where there are always six entries, for population, it is 6 * N, where N is the No. of indiv. countries/regions in the larger region.
        # parser below scans through all 'population entries' until it breaks, and does a double modulo division: first, by six, to single out each indiv. region,
        # and then by 2, to single out the SSP scenario (Pop. is the same for all RCP scenarios)
        popcr = 0
        while True:
            if Resultsheet2.cell(popc+1, 1).value != 'Population':
                break # that includes all population values
            else:
                if (popcr % 6) % 2 == 0:
                    for t in range(0,45): # time until 2060
                        Population[(popcr % 6) // 2,t] += Resultsheet2.cell(popc+1,t+9).value
            popc  += 1  
            popcr += 1
        
    UsePhaseSummary[0:3,:,:] = UseAnnEms2030.copy()
    UsePhaseSummary[3:6,:,:] = UseAnnEms2050.copy()
    UsePhaseSummary[6:9,:,:] = UseCumEms2050.copy()
    UsePhaseSummary[9::,:,:] = UseCumEms2060.copy()
    
    MatSummary[0:3,:,:] = MatAnnEms2030.copy()
    MatSummary[3:6,:,:] = MatAnnEms2050.copy()
    MatSummary[6:9,:,:] = MatCumEms2050.copy()
    MatSummary[9::,:,:] = MatCumEms2060.copy()
    
    ManSummary[0:3,:,:] = ManAnnEms2030.copy()
    ManSummary[3:6,:,:] = ManAnnEms2050.copy()
    ManSummary[6:9,:,:] = ManCumEms2050.copy()
    ManSummary[9::,:,:] = ManCumEms2060.copy()
    
    ForSummary[0:3,:,:] = ForAnnEms2030.copy()
    ForSummary[3:6,:,:] = ForAnnEms2050.copy()
    ForSummary[6:9,:,:] = ForCumEms2050.copy()
    ForSummary[9::,:,:] = ForCumEms2060.copy()    
    
    RecCredit[0:3,:,:]= RecCreditAnn2030.copy()
    RecCredit[3:6,:,:]= RecCreditAnn2050.copy()
    RecCredit[6:9,:,:]= RecCreditCum2050.copy()
    RecCredit[9::,:,:]= RecCreditCum2060.copy()
    
    # Waterfall plot for material-related GHG          
    Title  = ['MatCumGHG_16_50','MatCumGHG_40_50','MatAnnGHG_50']
    Scens  = ['LED','SSP1','SSP2']
    Rcens  = ['Base','RCP2_6']
    
    for nn in range(0,3):
        for m in range(0,NS): # SSP
            for rcp in range(0,NR): # RCP
                if nn == 0:
                    Data = np.einsum('SE->ES',MatCumEms2050[:,rcp,:])
                    CascDataExp[1,0,:,:,:] = MatCumEms2050.copy()
                if nn == 1:
                    Data = np.einsum('SE->ES',10*AvgDecadalMatEms[:,rcp,:,2])
                    CascDataExp[1,1,:,:,:] = 10*AvgDecadalMatEms[:,:,:,2].copy()
                if nn == 2:
                    Data = np.einsum('SE->ES',MatAnnEms2050[:,rcp,:])
                    CascDataExp[1,2,:,:,:] = MatAnnEms2050.copy()
                    
                inc = -100 * (Data[0,m] - Data[-1,m])/Data[0,m]
            
                Left  = Data[0,m]
                Right = Data[-1,m]
                # plot results
                bw = 0.5
            
                fig  = plt.figure(figsize=(5,8))
                ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
                ProxyHandlesList = []   # For legend     
                # plot bars
                ax1.fill_between([0,0+bw], [0,0],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[0],:], linewidth = 0.0)
                ax1.fill_between([1,1+bw], [Data[1,m],Data[1,m]],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[1],:], linewidth = 0.0)
                for xca in range(2,NE):
                    ax1.fill_between([xca,xca+bw], [Data[xca,m],Data[xca,m]],[Data[xca-1,m],Data[xca-1,m]],linestyle = '--', facecolor =MyColorCycle[ColOrder[xca],:], linewidth = 0.0)
                ax1.fill_between([NE,NE+bw], [0,0],[Data[NE-1,m],Data[NE-1,m]],linestyle = '--', facecolor =MyColorCycle[ColOrder[NE],:], linewidth = 0.0)                
                    
                for fca in range(0,NE+1):
                    ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[ColOrder[fca],:])) # create proxy artist for legend
                
                # plot lines:
                plt.plot([0,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
                for yca in range(1,NE):
                    plt.plot([yca,yca +1.5],[Data[yca,m],Data[yca,m]],linestyle = '-', linewidth = 0.5, color = 'k')
                    
                plt.arrow(Offset1, Data[NE-1,m],0, Data[0,m]-Data[NE-1,m], lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
                plt.arrow(Offset1,Data[0,m],0,Data[NE-1,m]-Data[0,m], lw = 0.8, ls = '-', shape = 'full',
                      length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
                    
                # plot text and labels
                plt.text(Offset2, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
                plt.text(Offset3, 0.94  *Right, Scens[m],fontsize=18,fontweight='bold') 
                plt.title('RE strats. and mat GHG emissions, ' + SectorString + '.', fontsize = 18)
                plt.ylabel(Title[nn] + r', Mt CO$_2$-eq.', fontsize = 18)
                plt.xticks(XTicks)
                plt.yticks(fontsize =18)
                ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
                plt.legend(handles = ProxyHandlesList,labels = LWE,shadow = False, prop={'size':12},ncol=1, loc = 'upper right' ,bbox_to_anchor=(1.91, 1)) 
                #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
                plt.axis([-0.2, Offset4, 0, 1.02*Left])
            
                plt.show()
                fig_name = RegionalScope + '_' + SectorString + '_' + Title[nn] + '_' + Scens[m] + '_' + Rcens[rcp] + '.png'
                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = PlotExpResolution, bbox_inches='tight') 

    # Area plot, stacked, GHG emissions, system and material production
    MyColorCycle = pylab.cm.Set1(np.arange(0,1,0.1)) # select colors from the 'Paired' color map.            
    grey0_9      = np.array([0.8,0.8,0.8,1])
    
    Title      = ['GHG, system-wide','GHG, material cycles']
    FName      = ['GHG_System','GHG_matcycles']
    Scens      = ['LED','SSP1','SSP2']
    Rcens      = ['Base','RCP2_6']   
    DataArea   = np.zeros((2,Nt,NS,NR,NE)) # 2 system scopes x Nt x SSP x RCP x RE scenarios
    
    for nn in range(0,len(Title)):
        #mS = 1
        #mR = 1
        for mRCP in range(0,NR):
            for mS in range(0,NS): # SSP               
                if nn == 0:
                    Data                 = AnnEms[:,mS,mRCP,:]
                    DataArea[nn,:,:,:,:] = AnnEms.copy()
                
                if nn == 1:
                    Data                 = MatEms[:,mS,mRCP,:]                
                    DataArea[nn,:,:,:,:] = MatEms.copy()
                    
                fig  = plt.figure(figsize=(8,5))
                ax1  = plt.axes([0.08,0.08,0.85,0.9])
                
                ProxyHandlesList = []   # For legend     
                
                # plot area
                ax1.fill_between(np.arange(2016,2061),np.zeros((Nt)), Data[:,-1], linestyle = '-', facecolor = grey0_9, linewidth = 1.0, alpha=0.5)
                ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=grey0_9)) # create proxy artist for legend
                for m in range(NE-1,0,-1):
                    ax1.fill_between(np.arange(2016,2061),Data[:,m], Data[:,m-1], linestyle = '-', facecolor = MyColorCycle[m,:], linewidth = 1.0, alpha=0.5)
                    ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:], alpha=0.75)) # create proxy artist for legend
                    ax1.plot(np.arange(2016,2061),Data[:,m],linestyle = '--', color = MyColorCycle[m,:], linewidth = 1.1,)                
                ax1.plot(np.arange(2016,2061),Data[:,0],linestyle = '--', color = 'k', linewidth = 1.1,)               
                PltLegx, = plt.plot([0,1],[0,1],linestyle = '--', color = 'k', linewidth = 1.1)
                ProxyHandlesList.append(PltLegx)
                #plt.text(Data[m,:].min()*0.55, 7.8, 'Baseline: ' + ("%3.0f" % Base[m]) + ' Mt/yr.',fontsize=14,fontweight='bold')
                plt.text(2027,Data[m,:].max()*1.02, 'Colors may deviate from legend colors due to overlap of RES wedges.',fontsize=8.5,fontweight='bold')
                
                plt.title(Title[nn] + ' \n' + RegionalScope + ', ' + SectorString + ', ' + Scens[mS] + ', ' + Rcens[mRCP] + '.', fontsize = 18)
                plt.ylabel(r'Mt of CO$_2$-eq.', fontsize = 18)
                plt.xlabel('Year', fontsize = 18)
                plt.xticks(fontsize=18)
                plt.yticks(fontsize=18)
                if PlotCtrl == 0: # vehicles, legend lower left
                    plt.legend(handles = reversed(ProxyHandlesList),labels = LWE_area, shadow = False, prop={'size':12},ncol=1, loc = 'lower left')# ,bbox_to_anchor=(1.91, 1)) 
                if PlotCtrl == 1: # buildings, upper right
                    plt.legend(handles = reversed(ProxyHandlesList),labels = LWE_area, shadow = False, prop={'size':12},ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
                ax1.set_xlim([2015, 2061])
#                if nn == 0:
#                    ax1.set_ylim([0, 220])
#                if nn == 1:
#                    ax1.set_ylim([0, 10.5])
                plt.show()
                fig_name = RegionalScope + '_' + SectorString + '_' + FName[nn] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '.png'
                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = PlotExpResolution, bbox_inches='tight')             
               
    ################################################################          
    #####      Overview plots metal production and stocks      #####
    ################################################################
    
    MyColorCycle = pylab.cm.tab20(np.arange(0,1,0.05)).copy() # select colors from the 'tab20' color map.  
    # Manually adjust colors:          
    MyColorCycle[0,:] = np.array([0.094117647,0.360784314,0.541176471,1]) # steel prim
    MyColorCycle[1,:] = np.array([0.329411765,0.662745098,0.88627451,1])  # steel sec
    MyColorCycle[2,:] = np.array([0.635294118,0.301960784,0,1])           # Al prim
    MyColorCycle[3,:] = np.array([1,0.498039216,0.054901961,1])           # Al sec
    MyColorCycle[4,:] = np.array([0.125490196,0.462745098,0.125490196,1]) # Cu prim
    MyColorCycle[5,:] = np.array([0.423529412,0.839215686,0.423529412,1]) # Cu sec
    MyColorCycle[6,:] = np.array([0.250980392,0.250980392,0.250980392,1]) # Cement prim
    MyColorCycle[7,:] = np.array([0.721568627,0.721568627,0.721568627,1]) # Cement sec
    MyColorCycle[8,:] = np.array([0.545098039,0.098039216,0.098039216,1]) # Plastics prim
    MyColorCycle[9,:] = np.array([0.901960784,0.462745098,0.462745098,1]) # Plastics sec
    MyColorCycle[10,:]= np.array([0.341176471,0.278431373,0.184313725,1])  # Wood prim
    MyColorCycle[11,:]= np.array([0.68627451,0.576470588,0.411764706,1])  # Wood sec
    grey0_9      = np.array([0.9,0.9,0.9,1])
    
    Title      = ['Materials']
    Sector     = ['pav_reb_nrb'] # also works for pav_reb 
    Scens      = ['LED','SSP1','SSP2']
    Rcens      = ['Base','RCP2_6']      
    
    # (1) Bar plot of metal production, primary and secondary, decadal average.
    bw = 0.7
    #mS = 0
    #mR = 0
#    for mRCP in range(0,NR): # RCP
#        for mS in range(0,NS): # SSP
#            for mR in range(0,1): # pav-reb-nrb
#                              
#                fig, ((ax1, ax2, ax3), (ax4, ax5, ax6)) = plt.subplots(2, 3, sharex=True, gridspec_kw={'hspace': 0.3, 'wspace': 0.35})
#                
#                ax1.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,0,mS,mRCP,0],MatProduction_Prim[4,0,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
#                ax1.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,0,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,0,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
#                ax1.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,0,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,0,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[0,:], linewidth = 0.0)
#                ax1.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,0,mS,mRCP,0],MatProduction_Sec[4,0,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[1,:], linewidth = 0.0)
#                ax1.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,0,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,0,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[1,:], linewidth = 0.0)
#                ax1.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,0,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,0,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[1,:], linewidth = 0.0)
#                ax1.set_title('Steel')
#                ax2.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,1,mS,mRCP,0],MatProduction_Prim[4,1,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[2,:], linewidth = 0.0)
#                ax2.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,1,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,1,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[2,:], linewidth = 0.0)
#                ax2.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,1,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,1,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[2,:], linewidth = 0.0)
#                ax2.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,1,mS,mRCP,0],MatProduction_Sec[4,1,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0)
#                ax2.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,1,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,1,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0)
#                ax2.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,1,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,1,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[3,:], linewidth = 0.0)
#                ax2.set_title('Aluminium')
#                ax3.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,2,mS,mRCP,0],MatProduction_Prim[4,2,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0)
#                ax3.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,2,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,2,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0)
#                ax3.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,2,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,2,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[4,:], linewidth = 0.0)
#                ax3.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,2,mS,mRCP,0],MatProduction_Sec[4,2,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[5,:], linewidth = 0.0)
#                ax3.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,2,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,2,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[5,:], linewidth = 0.0)
#                ax3.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,2,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,2,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[5,:], linewidth = 0.0)
#                ax3.set_title('Copper')
#                ax4.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,3,mS,mRCP,0],MatProduction_Prim[4,3,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
#                ax4.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,3,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,3,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
#                ax4.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,3,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,3,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[6,:], linewidth = 0.0)
#                ax4.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,3,mS,mRCP,0],MatProduction_Sec[4,3,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[7,:], linewidth = 0.0)
#                ax4.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,3,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,3,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[7,:], linewidth = 0.0)
#                ax4.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,3,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,3,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[7,:], linewidth = 0.0)
#                ax4.set_title('Cement')
#                ax5.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,4,mS,mRCP,0],MatProduction_Prim[4,4,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[8,:], linewidth = 0.0)
#                ax5.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,4,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,4,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[8,:], linewidth = 0.0)
#                ax5.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,4,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,4,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[8,:], linewidth = 0.0)
#                ax5.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,4,mS,mRCP,0],MatProduction_Sec[4,4,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[9,:], linewidth = 0.0)
#                ax5.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,4,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,4,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[9,:], linewidth = 0.0)
#                ax5.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,4,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,4,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[9,:], linewidth = 0.0)
#                ax5.set_title('Plastics')
#                ax6.fill_between([1,1+bw], [0,0],[MatProduction_Prim[4,5,mS,mRCP,0],MatProduction_Prim[4,5,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[10,:], linewidth = 0.0)
#                ax6.fill_between([1.8,1.8+bw], [0,0],[MatProduction_Prim[24:34,5,mS,mRCP,0].sum()/10,MatProduction_Prim[24:34,5,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[10,:], linewidth = 0.0)
#                ax6.fill_between([2.6,2.6+bw], [0,0],[MatProduction_Prim[24:34,5,mS,mRCP,-1].sum()/10,MatProduction_Prim[24:34,5,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[10,:], linewidth = 0.0)
#                ax6.fill_between([4,4+bw], [0,0],[MatProduction_Sec[4,5,mS,mRCP,0],MatProduction_Sec[4,5,mS,mRCP,0]],linestyle = '--', facecolor =MyColorCycle[11,:], linewidth = 0.0)
#                ax6.fill_between([4.8,4.8+bw], [0,0],[MatProduction_Sec[24:34,5,mS,mRCP,0].sum()/10,MatProduction_Sec[24:34,5,mS,mRCP,0].sum()/10],linestyle = '--', facecolor =MyColorCycle[11,:], linewidth = 0.0)
#                ax6.fill_between([5.6,5.6+bw], [0,0],[MatProduction_Sec[24:34,5,mS,mRCP,-1].sum()/10,MatProduction_Sec[24:34,5,mS,mRCP,-1].sum()/10],linestyle = '--', facecolor =MyColorCycle[11,:], linewidth = 0.0)
#                ax6.set_title('Wood')
#    
#                plt.sca(ax4)
#                plt.xticks([1.4,2.2,3.0,4.4,5.2,6.0], ['2020','2040-50, no ME','2040-50, ME','2020','2040-50, no ME','2040-50, ME'], rotation =90, fontsize = 10, fontweight = 'normal')
#                plt.sca(ax5)
#                plt.xticks([1.4,2.2,3.0,4.4,5.2,6.0], ['2020','2040-50, no ME','2040-50, ME','2020','2040-50, no ME','2040-50, ME'], rotation =90, fontsize = 10, fontweight = 'normal')
#                plt.sca(ax6)
#                plt.xticks([1.4,2.2,3.0,4.4,5.2,6.0], ['2020','2040-50, no ME','2040-50, ME','2020','2040-50, no ME','2040-50, ME'], rotation =90, fontsize = 10, fontweight = 'normal')
#    
#                plt.sca(ax1)
#                plt.ylabel('Mt/yr', fontsize = 12)
#                plt.sca(ax4)
#                plt.ylabel('Mt/yr', fontsize = 12)
#                
#                plt.show()
#                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '_bar.png'
#                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')  
                
    # (2) Line plot of metal production, primary and secondary, decadal average, all socec scenarios
#    LegendLabels = ['Primary material production, no ME','Primary material production, full ME','Secondary material production, no ME','Secondary material production, full ME']
#    if RegionalScope == 'Global':
#        LWI = [0.8,1.4,0.8]
#        for mRCP in range(0,NR):  # RCP
#            for mR in range(0,1): # pav-reb-nrb
#                              
#                fig, ((ax1, ax2, ax3), (ax4, ax5, ax6)) = plt.subplots(2, 3, sharex=True, gridspec_kw={'hspace': 0.3, 'wspace': 0.35})
#    
#                for mS in range(0,NS): # SSP
#                    ax1.plot(np.arange(2016,2061,1),MatProduction_Prim[:,0,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[0,:], linewidth = LWI[mS])
#                    ax1.plot(np.arange(2016,2061,1),MatProduction_Prim[:,0,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[0,:], linewidth = LWI[mS])
#                    ax1.plot(np.arange(2016,2061,1),MatProduction_Sec[:,0,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[1,:], linewidth = LWI[mS])
#                    ax1.plot(np.arange(2016,2061,1),MatProduction_Sec[:,0,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[1,:], linewidth = LWI[mS])
#                ax1.set_title('Steel')
#                for mS in range(0,NS): # SSP
#                    ax2.plot(np.arange(2016,2061,1),MatProduction_Prim[:,1,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[2,:], linewidth = LWI[mS])
#                    ax2.plot(np.arange(2016,2061,1),MatProduction_Prim[:,1,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[2,:], linewidth = LWI[mS])
#                    ax2.plot(np.arange(2016,2061,1),MatProduction_Sec[:,1,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[3,:], linewidth = LWI[mS])
#                    ax2.plot(np.arange(2016,2061,1),MatProduction_Sec[:,1,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[3,:], linewidth = LWI[mS])
#                ax2.set_title('Aluminium')
#                for mS in range(0,NS): # SSP
#                    ax3.plot(np.arange(2016,2061,1),MatProduction_Prim[:,2,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[4,:], linewidth = LWI[mS])
#                    ax3.plot(np.arange(2016,2061,1),MatProduction_Prim[:,2,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[4,:], linewidth = LWI[mS])
#                    ax3.plot(np.arange(2016,2061,1),MatProduction_Sec[:,2,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[5,:], linewidth = LWI[mS])
#                    ax3.plot(np.arange(2016,2061,1),MatProduction_Sec[:,2,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[5,:], linewidth = LWI[mS])
#                ax3.set_title('Copper')
#                for mS in range(0,NS): # SSP
#                    ax4.plot(np.arange(2016,2061,1),MatProduction_Prim[:,3,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[6,:], linewidth = LWI[mS])
#                    ax4.plot(np.arange(2016,2061,1),MatProduction_Prim[:,3,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[6,:], linewidth = LWI[mS])
#                    ax4.plot(np.arange(2016,2061,1),MatProduction_Sec[:,3,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[7,:], linewidth = LWI[mS])
#                    ax4.plot(np.arange(2016,2061,1),MatProduction_Sec[:,3,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[7,:], linewidth = LWI[mS])
#                ax4.set_title('Cement')
#                for mS in range(0,NS): # SSP
#                    ax5.plot(np.arange(2016,2061,1),MatProduction_Prim[:,4,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[8,:], linewidth = LWI[mS])
#                    ax5.plot(np.arange(2016,2061,1),MatProduction_Prim[:,4,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[8,:], linewidth = LWI[mS])
#                    ax5.plot(np.arange(2016,2061,1),MatProduction_Sec[:,4,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[9,:], linewidth = LWI[mS])
#                    ax5.plot(np.arange(2016,2061,1),MatProduction_Sec[:,4,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[9,:], linewidth = LWI[mS])
#                ax5.set_title('Plastics')
#                for mS in range(0,NS): # SSP
#                    ax6.plot(np.arange(2016,2061,1),MatProduction_Prim[:,5,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[10,:], linewidth = LWI[mS])
#                    ax6.plot(np.arange(2016,2061,1),MatProduction_Prim[:,5,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[10,:], linewidth = LWI[mS])
#                    ax6.plot(np.arange(2016,2061,1),MatProduction_Sec[:,5,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[11,:], linewidth = LWI[mS])
#                    ax6.plot(np.arange(2016,2061,1),MatProduction_Sec[:,5,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[11,:], linewidth = LWI[mS])
#                ax6.set_title('Wood')
#    
#                plt.sca(ax1)
#                plt.ylabel('Mt/yr', fontsize = 12)
#                plt.sca(ax4)
#                plt.ylabel('Mt/yr', fontsize = 12)
#
#                plt.plot(2016,0,color=np.array([0,0,0,1]),       lw=LWI[1],  linestyle='-')
#                plt.plot(2016,0,color=np.array([0,0,0,1]),       lw=LWI[1],  linestyle='--')
#                plt.plot(2016,0,color=np.array([0.3,0.3,0.3,1]), lw=LWI[1],  linestyle='-')
#                plt.plot(2016,0,color=np.array([0.3,0.3,0.3,1]), lw=LWI[1],  linestyle='--') 
#                plt.legend(LegendLabels,shadow = False, prop={'size':7}, loc = 'upper right',bbox_to_anchor=(2.5, 1))  
#    
#                plt.show()
#                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Rcens[mRCP] + '_line.png'
#                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')      
                    
            
    # (3a) 3x2 Line plot of metal production, primary and secondary. Same data, but with line plot for each SSP
#    LWI = [0.8,1.4,0.8]
#    for mRCP in range(0,NR): # RCP
#        for mS in range(0,NS): # SSP
#            for mR in range(0,1): # pav-reb-nrb
#                          
#                fig, ((ax1, ax2, ax3), (ax4, ax5, ax6)) = plt.subplots(2, 3, sharex=True, gridspec_kw={'hspace': 0.3, 'wspace': 0.35})
#
#                ax1.plot(np.arange(2016,2061,1),MatProduction_Prim[:,0,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[0,:], linewidth = LWI[mS])
#                ax1.plot(np.arange(2016,2061,1),MatProduction_Prim[:,0,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[0,:], linewidth = LWI[mS])
#                ax1.plot(np.arange(2016,2061,1),MatProduction_Sec[:,0,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[1,:], linewidth = LWI[mS])
#                ax1.plot(np.arange(2016,2061,1),MatProduction_Sec[:,0,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[1,:], linewidth = LWI[mS])
#                ax1.set_title('Steel')
#                ax2.plot(np.arange(2016,2061,1),MatProduction_Prim[:,1,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[2,:], linewidth = LWI[mS])
#                ax2.plot(np.arange(2016,2061,1),MatProduction_Prim[:,1,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[2,:], linewidth = LWI[mS])
#                ax2.plot(np.arange(2016,2061,1),MatProduction_Sec[:,1,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[3,:], linewidth = LWI[mS])
#                ax2.plot(np.arange(2016,2061,1),MatProduction_Sec[:,1,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[3,:], linewidth = LWI[mS])
#                ax2.set_title('Aluminium')
#                ax3.plot(np.arange(2016,2061,1),MatProduction_Prim[:,2,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[4,:], linewidth = LWI[mS])
#                ax3.plot(np.arange(2016,2061,1),MatProduction_Prim[:,2,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[4,:], linewidth = LWI[mS])
#                ax3.plot(np.arange(2016,2061,1),MatProduction_Sec[:,2,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[5,:], linewidth = LWI[mS])
#                ax3.plot(np.arange(2016,2061,1),MatProduction_Sec[:,2,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[5,:], linewidth = LWI[mS])
#                ax3.set_title('Copper')
#                ax4.plot(np.arange(2016,2061,1),MatProduction_Prim[:,3,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[6,:], linewidth = LWI[mS])
#                ax4.plot(np.arange(2016,2061,1),MatProduction_Prim[:,3,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[6,:], linewidth = LWI[mS])
#                ax4.plot(np.arange(2016,2061,1),MatProduction_Sec[:,3,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[7,:], linewidth = LWI[mS])
#                ax4.plot(np.arange(2016,2061,1),MatProduction_Sec[:,3,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[7,:], linewidth = LWI[mS])
#                ax4.set_title('Cement')
#                ax5.plot(np.arange(2016,2061,1),MatProduction_Prim[:,4,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[8,:], linewidth = LWI[mS])
#                ax5.plot(np.arange(2016,2061,1),MatProduction_Prim[:,4,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[8,:], linewidth = LWI[mS])
#                ax5.plot(np.arange(2016,2061,1),MatProduction_Sec[:,4,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[9,:], linewidth = LWI[mS])
#                ax5.plot(np.arange(2016,2061,1),MatProduction_Sec[:,4,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[9,:], linewidth = LWI[mS])
#                ax5.set_title('Plastics')
#                ax6.plot(np.arange(2016,2061,1),MatProduction_Prim[:,5,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[10,:], linewidth = LWI[mS])
#                ax6.plot(np.arange(2016,2061,1),MatProduction_Prim[:,5,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[10,:], linewidth = LWI[mS])
#                ax6.plot(np.arange(2016,2061,1),MatProduction_Sec[:,5,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[11,:], linewidth = LWI[mS])
#                ax6.plot(np.arange(2016,2061,1),MatProduction_Sec[:,5,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[11,:], linewidth = LWI[mS])
#                ax6.set_title('Wood')
#
#                plt.sca(ax1)
#                plt.ylabel('Mt/yr', fontsize = 12)
#                plt.sca(ax4)
#                plt.ylabel('Mt/yr', fontsize = 12)
#    
#                plt.show()
#                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '_line.png'
#                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')               
#                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '_line.svg'
#                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')      

    # (3b) 6x1 Line plot of metal production, primary and secondary. Same data, but with line plot for each SSP
    LWI = [0.8,1.4,0.8]
    for mRCP in range(0,NR): # RCP
        for mS in range(0,NS): # SSP
            for mR in range(0,1): # pav-reb-nrb
                          
                fig, ((ax1, ax2, ax3, ax4, ax5, ax6)) = plt.subplots(1, 6, sharex=True, gridspec_kw={'hspace': 0.3, 'wspace': 0.35},figsize=(15,5))

                ax1.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,0,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[0,:], linewidth = LWI[mS])
                ax1.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,0,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[0,:], linewidth = LWI[mS])
                ax1.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,0,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[1,:], linewidth = LWI[mS])
                ax1.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,0,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[1,:], linewidth = LWI[mS])
                ax1.set_title('Steel', fontsize = 14)
                ax2.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,1,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[2,:], linewidth = LWI[mS])
                ax2.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,1,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[2,:], linewidth = LWI[mS])
                ax2.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,1,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[3,:], linewidth = LWI[mS])
                ax2.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,1,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[3,:], linewidth = LWI[mS])
                ax2.set_title('Aluminium', fontsize = 14)
                ax3.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,2,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[4,:], linewidth = LWI[mS])
                ax3.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,2,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[4,:], linewidth = LWI[mS])
                ax3.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,2,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[5,:], linewidth = LWI[mS])
                ax3.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,2,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[5,:], linewidth = LWI[mS])
                ax3.set_title('Copper', fontsize = 14)
                ax4.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,3,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[6,:], linewidth = LWI[mS])
                ax4.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,3,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[6,:], linewidth = LWI[mS])
                ax4.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,3,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[7,:], linewidth = LWI[mS])
                ax4.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,3,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[7,:], linewidth = LWI[mS])
                ax4.set_title('Cement', fontsize = 14)
                ax5.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,4,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[8,:], linewidth = LWI[mS])
                ax5.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,4,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[8,:], linewidth = LWI[mS])
                ax5.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,4,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[9,:], linewidth = LWI[mS])
                ax5.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,4,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[9,:], linewidth = LWI[mS])
                ax5.set_title('Plastics', fontsize = 14)
                ax6.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,5,mS,mRCP,0], linestyle = '--',  color =MyColorCycle[10,:], linewidth = LWI[mS])
                ax6.plot(np.arange(2016,2053,1),MatProduction_Prim[0:37,5,mS,mRCP,-1],linestyle = '-', color =MyColorCycle[10,:], linewidth = LWI[mS])
                ax6.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,5,mS,mRCP,0],  linestyle = '--',  color =MyColorCycle[11,:], linewidth = LWI[mS])
                ax6.plot(np.arange(2016,2053,1),MatProduction_Sec[0:37,5,mS,mRCP,-1], linestyle = '-', color =MyColorCycle[11,:], linewidth = LWI[mS])
                ax6.set_title('Wood', fontsize = 14)
                
                ax1.set_xlim([2015, 2053])
                ax2.set_xlim([2015, 2053])
                ax3.set_xlim([2015, 2053])
                ax4.set_xlim([2015, 2053])
                ax5.set_xlim([2015, 2053])
                ax6.set_xlim([2015, 2053])
                
                ax1.set_xticks([2020,2030,2040,2050])
                ax2.set_xticks([2020,2030,2040,2050])
                ax3.set_xticks([2020,2030,2040,2050])
                ax4.set_xticks([2020,2030,2040,2050])
                ax5.set_xticks([2020,2030,2040,2050])
                ax6.set_xticks([2020,2030,2040,2050])
                ax1.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                ax2.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                ax3.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                ax4.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                ax5.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                ax6.set_xticklabels(['2020','2030','2040','2050'], rotation =90, fontsize = 9, fontweight = 'normal')
                    
                
                # For global paper only:
                if RegionalScope == 'Global' and mS == 1 and mRCP == 1: # format axes
                    ax1.set_ylim([0, 500])    
                    ax2.set_ylim([0, 25])    
                    ax3.set_ylim([0, 25])    
                    ax4.set_ylim([0, 1000])    
                    ax5.set_ylim([0, 250])    
                    ax6.set_ylim([0, 500])    
                    
                plt.sca(ax1)
                plt.ylabel('Mt/yr', fontsize = 14)
    
                plt.show()
                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '_line_v2.png'
                fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')               
                fig_name = RegionalScope + '_' + Sector[mR] + '_' + Title[0] + '_' + Scens[mS] + '_' + Rcens[mRCP] + '_line_v2.svg'
                # fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = 400, bbox_inches='tight')      
    
    ### (4)                  
    # None
    
    ### (5) Line plot overview of primary steel and steel recycling
    if RegionalScope == 'Global':
        MyColorCycle = pylab.cm.Paired(np.arange(0,1,0.2))
        #linewidth = [1.2,2.4,1.2,1.2,1.2]
        linewidth  = [1.2,2,1.2]
        linewidth2 = [1.2,2,1.2]
        
        ColorOrder = [1,0,3]
                
        # Primary steel
        AnnEmsV_PrimarySteel   = np.zeros((Nt,NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
        AnnEmsV_SecondarySteel = np.zeros((Nt,NS,NR,NE)) # SSP-Scenario x RCP scenario x RES scenario
        
        for r in range(0,NE): # RE scenario
            ResFile = [filename for filename in os.listdir(os.path.join(RECC_Paths.results_path,FolderList[r])) if filename.startswith('ODYM_RECC_ModelResults_')]
            Resultfile2  = openpyxl.load_workbook(os.path.join(RECC_Paths.results_path,FolderList[r],ResFile[0]))
            Resultsheet2 = Resultfile2['Model_Results']
            # Find the index for materials
            pps = 1
            while True:
                if Resultsheet2.cell(pps+1, 1).value == 'Primary steel production':
                    break # that gives us the right index to read the recycling credit from the result table.
                pps += 1
            sps = 1
            while True:
                if Resultsheet2.cell(sps+1, 1).value == 'Secondary steel':
                    break # that gives us the right index to read the recycling credit from the result table.
                sps += 1
     
            for s in range(0,NS): # SSP scenario
                for c in range(0,NR):
                    for t in range(0,45): # timeAnnEmsV_SecondarySteel[t,s,c,r] = Resultsheet2.cell(152+ 2*s +c,t+9).value
                        AnnEmsV_PrimarySteel[t,s,c,r]   = Resultsheet2.cell(pps+ 2*s +c+1,t+9).value
                        AnnEmsV_SecondarySteel[t,s,c,r] = Resultsheet2.cell(sps+ 2*s +c+1,t+9).value
                        
        Title      = ['primary_steel','secondary_steel']            
        ScensL     = ['SSP2, no ME','SSP2, full ME spectrum','SSP1, no ME','SSP1, full ME spectrum','LED, no ME','LED, full ME spectrum']
        
        #mS = 1
        #mR = 1
        for nn in range(0,2):
            mRCP = 1 # select RCP2.6, which has full implementation of RE strategies by 2050.
            
            if nn == 0:
                Data = AnnEmsV_PrimarySteel[:,:,mRCP,:]
            if nn == 1:
                Data = AnnEmsV_SecondarySteel[:,:,mRCP,:]
        
        
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            
            ProxyHandlesList = []   # For legend     
            
            for mS in range(NS-1,-1,-1):
                ax1.plot(np.arange(2016,2061), Data[:,mS,0],  linewidth = linewidth[mS],  linestyle = '-',  color = MyColorCycle[ColorOrder[mS],:])
                #ProxyHandlesList.append(plt.line((0, 0), 1, 1, fc=MyColorCycle[m,:]))
                ax1.plot(np.arange(2016,2061), Data[:,mS,-1], linewidth = linewidth2[mS], linestyle = '--', color = MyColorCycle[ColorOrder[mS],:])
                #ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[m,:]))     
            plt.legend(ScensL,shadow = False, prop={'size':12}, loc = 'upper left',bbox_to_anchor=(1.05, 1))    
            plt.ylabel('Mt/yr.', fontsize = 18) 
            plt.xlabel('year', fontsize = 18)         
            plt.title(Title[nn] + ', by socio-economic scenario, \n' + RegionalScope + ', ' + SectorString + '.', fontsize = 18)
            plt.xticks(fontsize=18)
            plt.yticks(fontsize=18)
            ax1.set_xlim([2015, 2061])
            plt.gca().set_ylim(bottom=0)
            
            plt.show()
            fig_name = RegionalScope + '_' + SectorString + '_' + Title[nn] + '.png'
            fig.savefig(os.path.join(RECC_Paths.results_path_save,fig_name), dpi = PlotExpResolution, bbox_inches='tight')             
            
    # Save data to xls
    # GHG data, area plots
    ColIndex           = [str(mmx) for mmx in  range(2016,2061)]
    RowIndex           = pd.MultiIndex.from_product([['System-wide GHG','Material cycle GHG'],['LED','SSP1','SSP2'],['Base','RCP2_6'],LWE_area[0:-1]], names=('System scope','SSP','RCP','ME strategy'))
    DF_GHGA_global     = pd.DataFrame(np.einsum('ItSRE->ISREt',DataArea).reshape(12*NE,45), index=RowIndex, columns=ColIndex)
    DF_GHGA_global.to_excel(os.path.join(RECC_Paths.results_path_save,'GHG_Area_Data_' + SectorString + '_' + RegionalScope + '.xls'), merge_cells=False)
    
    # Primary and secondary material production, by material and RE strategy cascade
    ColIndex           = [str(mmx) for mmx in  range(2016,2061)]
    if SectorString == 'pav_reb' or SectorString == 'pav_nrb' or SectorString == 'pav_reb_nrb':
        RowIndex       = pd.MultiIndex.from_product([['Steel','Aluminium','Copper','Cement','Plastics','Timber'],['LED','SSP1','SSP2'],['Base','RCP2_6'],['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ down-sizing','+ car-sharing','+ ride-sharing','+ more intense bld. use = All ME stratgs.']], names=('Material','SSP','RCP','ME cascade steps'))
        DF_PriM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Prim).reshape(288,45), index=RowIndex, columns=ColIndex)
        DF_SecM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Sec).reshape(288,45),  index=RowIndex, columns=ColIndex)
    if SectorString == 'pav':
        RowIndex       = pd.MultiIndex.from_product([['Steel','Aluminium','Copper','Cement','Plastics','Timber'],['LED','SSP1','SSP2'],['Base','RCP2_6'],['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ down-sizing','+ car-sharing','+ ride-sharing = All ME stratgs.']], names=('Material','SSP','RCP','ME cascade steps'))
        DF_PriM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Prim).reshape(252,45), index=RowIndex, columns=ColIndex)
        DF_SecM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Sec).reshape(252,45),  index=RowIndex, columns=ColIndex)
    if SectorString == 'reb':
        RowIndex       = pd.MultiIndex.from_product([['Steel','Aluminium','Copper','Cement','Plastics','Timber'],['LED','SSP1','SSP2'],['Base','RCP2_6'],['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ light design','+ more intense bld. use = All ME stratgs.']], names=('Material','SSP','RCP','ME cascade steps'))
        DF_PriM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Prim).reshape(216,45), index=RowIndex, columns=ColIndex)
        DF_SecM_global = pd.DataFrame(np.einsum('tmSRE->mSREt',MatProduction_Sec).reshape(216,45),  index=RowIndex, columns=ColIndex)
    DF_PriM_global.to_excel(os.path.join(RECC_Paths.results_path_save,'PrimaryMaterial_'   + SectorString + '_' + RegionalScope + '.xls'), merge_cells=False)
    DF_SecM_global.to_excel(os.path.join(RECC_Paths.results_path_save,'SecondaryMaterial_' + SectorString + '_' + RegionalScope + '.xls'), merge_cells=False)        
        
    RowIndex           = pd.MultiIndex.from_product([['System-wide GHG','Material cycle GHG'],['Cumulative GHG, 2016-2050','Avg. annual GHG, 2040-2050','Annual GHG, 2050'],['LED','SSP1','SSP2'],['Base','RCP2_6']], names=('System scope','Indicator','SSP','RCP'))
    if SectorString == 'pav_reb' or SectorString == 'pav_nrb' or SectorString == 'pav_reb_nrb':
        ColIndex       = ['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ down-sizing','+ car-sharing','+ ride-sharing','+ more intense bld. use = All ME stratgs.']
        DF_GHGC_global = pd.DataFrame(CascDataExp.reshape(36,8), index=RowIndex, columns=ColIndex)
    if SectorString == 'pav':
        ColIndex       = ['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ down-sizing','+ car-sharing','+ ride-sharing = All ME stratgs.']
        DF_GHGC_global = pd.DataFrame(CascDataExp.reshape(36,7), index=RowIndex, columns=ColIndex)        
    if SectorString == 'reb':
        ColIndex       = ['No ME','+ higher yields', '+ re-use/longer use','+ material subst.','+ light design','+ more intense bld. use = All ME stratgs.']
        DF_GHGC_global = pd.DataFrame(CascDataExp.reshape(36,6), index=RowIndex, columns=ColIndex)        
    DF_GHGC_global.to_excel(os.path.join(RECC_Paths.results_path_save,'GHG_Cascade_Data_'  + SectorString + '_' + RegionalScope + '.xls'), merge_cells=False)        
        
    return ASummary, AvgDecadalEms, MatSummary, AvgDecadalMatEms, RecCredit, UsePhaseSummary, ManSummary, ForSummary, AvgDecadalUseEms, AvgDecadalManEms, AvgDecadalForEms, RecCreditAvgDec, CumEms2050, CumEms2060, AnnEms2050, MatStocks, TimeSeries_R, MatEms, Population

# code for script to be run as standalone function
if __name__ == "__main__":
    main()

#
#
#
    
    
    