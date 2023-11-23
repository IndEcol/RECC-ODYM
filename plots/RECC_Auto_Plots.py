# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 13:54:19 2023

@author: spauliuk

This script loads previously compiled results and then compiles selected results 
into different visualisations, like line plots and bar charts.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select.xlxs

Need to run ODYM_RECC_Export_xlxs_Combine_Select.py first!

Documentation and how to in RECCv2.5_EXPORT_Combine_Select.xlxs

"""
import os
import plotnine
import openpyxl
from plotnine import *
import pandas as pd
import pylab
import matplotlib.pyplot as plt
import numpy as np
import RECC_Paths # Import path file


CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')   
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value

r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define Plots':
        break    
    r += 1
r += 1

ptitles = []
ptypes  = []
pinds   = []
pregs   = []
pscens  = []
prange  = []
pflags  = []
indlab  = [] # indicator short labels for plots
scelab  = [] # scenario  short labels for plots
colors  = [] # List with color strings
while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ptitles.append(CF[CS].cell(r,2).value)
    ptypes.append(CF[CS].cell(r,3).value)
    pinds.append(CF[CS].cell(r,4).value)
    pregs.append(CF[CS].cell(r,5).value)
    pscens.append(CF[CS].cell(r,6).value)
    prange.append(CF[CS].cell(r,7).value)
    pflags.append(CF[CS].cell(r,8).value)
    indlab.append(CF[CS].cell(r,9).value)
    scelab.append(CF[CS].cell(r,10).value)
    colors.append(CF[CS].cell(r,11).value)
    r += 1

# open data file with results
fn = os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '_sep.xlsx')
ps = pd.read_excel(fn, sheet_name='Results', index_col=0) # plot sheet
pc = pd.read_excel(fn, sheet_name='Results_Cumulative', index_col=0) # plot sheet cumulative


for m in range(0,len(ptitles)):
    if ptypes[m] == 'line_fixedIndicator_fixedRegion_varScenario':
    # Plot single indicator for one region and all scenarios
        selectI = [pinds[m]]
        selectR = [pregs[m]]
        if pscens[m] == 'All':
            pst       = ps[ps['Indicator'].isin(selectI) & ps['Region'].isin(selectR)].T # Select the specified data and transpose them for plotting
            title_add = '_all_scenarios'
        else:
            selectS = pscens[m].split(';')
            pst     = ps[ps['Indicator'].isin(selectI) & ps['Region'].isin(selectR) & ps['Scenario'].isin(selectS)].T # Select the specified data and transpose them for plotting
            title_add = '_select_scenarios_' + str(len(selectS))
        pst.columns = pst.iloc[2] # Set scenario column (with unique labels) as column names
        unit    = pst.iloc[4][1] 
        pst.drop(['Region','Indicator','Scenario','Sectors','Unit'], inplace=True) # Delete labels that are not needed
        pst.plot(kind = 'line', figsize=(10,5), ) # plot data, configure plot, and save results
        plt.xlabel('Year')
        plt.ylabel(unit)
        title = ptitles[m] + '_' + selectR[0] + title_add
        plt.title(title)
        plt.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '.png'), dpi=150, bbox_inches='tight')

    if ptypes[m] == 'line_fixedIndicator_varRegion_fixedScenario':
    # Plot single indicator for one scenario and all regions
        selectI = [pinds[m]]
        selectS = [pscens[m]]
        if pregs[m] == 'All':
            pst    = ps[ps['Indicator'].isin(selectI) & ps['Scenario'].isin(selectS)].T # Select the specified data and transpose them for plotting
            title_add = '_all_regions'
        else:
            selectR = pregs[m].split(';')
            pst    = ps[ps['Indicator'].isin(selectI) & ps['Region'].isin(selectR) & ps['Scenario'].isin(selectS)].T # Select the specified data and transpose them for plotting
            title_add = '_select_regions_' + str(len(selectS))
        pst.columns = pst.iloc[0] # Set scenario column (with unique labels) as column names
        unit    = pst.iloc[4][1] 
        pst.drop(['Region','Indicator','Scenario','Sectors','Unit'], inplace=True) # Delete labels that are not needed
        pst.plot(kind = 'line', figsize=(10,5), ) # plot data, configure plot, and save results
        plt.xlabel('Year')
        plt.ylabel(unit)
        title = ptitles[m] + '_' + selectS[0] + title_add
        plt.title(title)
        plt.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title +'.png'), dpi=150, bbox_inches='tight')

    if ptypes[m] == 'hbar_cum_fixedIndicator_fixedRegion_varScenario':
        # Plot bar graph with cumulative indicator by scenario
        selectI = [pinds[m]]
        selectR = [pregs[m]]
        if pscens[m] == 'All':
            pst    = pc[pc['Indicator'].isin(selectI) & pc['Region'].isin(selectR)] # Select the specified data and transpose them for plotting
            title_add = '_all_scenarios'
        else:
            selectS = pscens[m].split(';')
            pst    = pc[pc['Indicator'].isin(selectI) & pc['Region'].isin(selectR) & pc['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
            title_add = '_select_scenarios_' + str(len(selectS))
        pst.set_index('Scenario', inplace=True)
        pst[prange[m]]
        unit = pst.iloc[0][3]
        pst.plot.barh(y=prange[m])
        plt.xlabel(unit)
        title = ptitles[m] + '_' + selectR[0] + title_add
        plt.title(title)
        plt.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title +'.png'), dpi=150, bbox_inches='tight')
        
    if ptypes[m] == 'hbar_cum_fixedIndicator_varRegion_fixedScenario':
        # Plot bar graph with cumulative indicator by scenario
        selectI = [pinds[m]]
        selectS = [pscens[m]]
        if pregs[m] == 'All':
            pst    = pc[pc['Indicator'].isin(selectI) & pc['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
            title_add = '_all_regions'
        else:
            selectR = pregs[m].split(';')
            pst    = pc[pc['Indicator'].isin(selectI) & pc['Region'].isin(selectR) & pc['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
            title_add = '_select_regions_' + str(len(selectR))
        pst.set_index('Region', inplace=True)
        pst[prange[m]]
        unit = pst.iloc[0][3]
        pst.plot.barh(y=prange[m])
        plt.xlabel(unit)
        title = ptitles[m] + '_' + selectR[0] + title_add
        plt.title(title)
        plt.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title +'.png'), dpi=150, bbox_inches='tight')
        
    if ptypes[m] == 'CE_strategy_sensitivity':
        # Plot bar graph with sensitivity analysis of a given indicator by scenario
        selectI = pinds[m].split(';')  # For different indicators, separated by ;
        selectS = pscens[m].split(';') # For different scenarios, separated by ;
        selectR = [pregs[m]]           # For one region
        NooI    = len(selectI)
        NooS    = len(selectS)
        title_add = '_' + selectR[0]
        # Select data sheet acc. to flag set:
        if pflags[m] == 'annual':
            ddf = ps
        if pflags[m] == 'cumulative':
            ddf = pc
        # Create figure
        fig, axs = plt.subplots(nrows=NooI, ncols=NooS , figsize=(3*NooS, 3*NooI))        
        fig.suptitle(ptitles[m] + title_add + ', ' + prange[m] + ', ' + selectR[0],fontsize=18)
        for iI in range(0,NooI):
            for iS in range(0,NooS):
                pst    = ddf[ddf['Indicator'].isin([selectI[iI]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[iS]])] # Select the specified data
                pst.set_index('Region', inplace=True)
                pst[prange[m]]
                unit = pst.iloc[0][3]
                pstref    = ddf[ddf['Indicator'].isin([selectI[iI]]) & ddf['Region'].isin(selectR) & ddf['Scenario'].isin([selectS[0]])] # Select the specified data
                pstref.set_index('Region', inplace=True)
                pstref[prange[m]]
                valueref = pstref.iloc[0][prange[m]]
                pstplot = pstref.drop(['Indicator', 'Scenario', 'Sectors', 'Unit'], axis=1)-pst.drop(['Indicator', 'Scenario', 'Sectors', 'Unit'], axis=1)
                if iS > 0:
                    pstplot.plot.barh(ax=axs[iI,iS], y=prange[m], legend=False, color = colors[m].split(';')[iI])
                else: #left plot
                    axs[iI,iS].plot()
                    axs[iI,iS].text(0,0,int(np.rint(valueref)), fontsize = 15, style = 'oblique')
                if iI == 0:
                    axs[iI,iS].set_title( scelab[m].split(';')[iS], fontsize = 11)
                if iS == 0:
                    axs[iI,iS].set_ylabel(indlab[m].split(';')[iI], fontsize = 11)
                else:
                    axs[iI,iS].set_ylabel('')
                axs[iI,iS].set_xlabel(unit)
                axs[iI,iS].set_yticks([])
        # adjust x axis limits
        for iI in range(0,NooI):
            xaleft  = [0]
            xaright = [0]
            for iS in range(1,NooS):
                xaleft.append( axs[iI,iS].get_xlim()[0])
                xaright.append(axs[iI,iS].get_xlim()[1])
            xmin = min(xaleft)
            xmax = max(xaright)
            for iS in range(1,NooS):
                if xaleft[iS] == 0:
                    axs[iI,iS].set_xlim([0, xmax])
                if xaright[iS] == 0:
                    axs[iI,iS].set_xlim([xmin, 0])
        ftitle = ptitles[m] + title_add
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ftitle +'.png'), dpi=150, bbox_inches='tight')

    if ptypes[m] == 'Cascade_fixedRegion_varScenario':
        # Plot cascade with indicator by scenario
        selectI = [pinds[m]]
        selectR = [pregs[m]]
        selectS = pscens[m].split(';')
        pst     = ps[ps['Indicator'].isin(selectI) & ps['Region'].isin(selectR) & ps['Scenario'].isin(selectS)] # Select the specified data and transpose them for plotting
        title_add = '_select_scenarios_' + str(len(selectS))
        pst.set_index('Scenario', inplace=True)
        unit = pst.iloc[0]['Unit']
        CData=pst[prange[m]]
        CLabels = [CData.axes[0].values[i] for i in range(0,len(CData.axes[0].values))]
        Data    = CData.values
        nD      = len(CLabels)
        # Prepare plot
        ColOrder= [i for i in range(0,nD+1)]
        MyColorCycle = pylab.cm.Set1(np.arange(0,1,1/(nD+1))) # select colors from the 'Paired' color map.  
        Left  = Data[0]
        Right = Data[-1]
        inc = -100 * (Data[0] - Data[-1])/Data[0]
        # plot results
        bw = 0.5
        
        LLeft   = nD+bw
        XTicks  = [0.25 + i for i in range(0,nD+1)]
        
        fig  = plt.figure(figsize=(5,8))
        ax1  = plt.axes([0.08,0.08,0.85,0.9])
    
        ProxyHandlesList = []   # For legend     
        # plot bars
        ax1.fill_between([0,0+bw], [0,0],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[0],:], linewidth = 0.0)
        ax1.fill_between([1,1+bw], [Data[1],Data[1]],[Left,Left],linestyle = '--', facecolor =MyColorCycle[ColOrder[1],:], linewidth = 0.0)
        for xca in range(2,nD):
            ax1.fill_between([xca,xca+bw], [Data[xca],Data[xca]],[Data[xca-1],Data[xca-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[xca],:], linewidth = 0.0)
        ax1.fill_between([nD,nD+bw], [0,0],[Data[nD-1],Data[nD-1]],linestyle = '--', facecolor =MyColorCycle[ColOrder[nD],:], linewidth = 0.0)                
            
        for fca in range(0,nD):
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[ColOrder[fca],:])) # create proxy artist for legend
        
        # plot lines:
        plt.plot([0,LLeft],[Left,Left],linestyle = '-', linewidth = 0.5, color = 'k')
        for yca in range(1,nD):
            plt.plot([yca,yca +1.5],[Data[yca],Data[yca]],linestyle = '-', linewidth = 0.5, color = 'k')
            
        plt.arrow(XTicks[-1], Data[nD-1],0, Data[0]-Data[nD-1], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
        plt.arrow(XTicks[-1],Data[0],0,Data[nD-1]-Data[0], lw = 0.5, ls = '-', shape = 'full',
              length_includes_head = True, head_width =0.1, head_length =0.01*Left, ec = 'k', fc = 'k')
            
        # plot text and labels
        plt.text(nD-1, 0.94 *Left, ("%3.0f" % inc) + ' %',fontsize=18,fontweight='bold')          
        title = ptitles[m] + '_' + selectR[0] + title_add
        plt.title(title)
        plt.ylabel(unit, fontsize = 18)
        plt.xticks(XTicks)
        plt.yticks(fontsize =18)
        ax1.set_xticklabels([], rotation =90, fontsize = 21, fontweight = 'normal')
        plt.legend(handles = ProxyHandlesList,labels = CLabels,shadow = False, prop={'size':12},ncol=1, loc = 'upper right' ,bbox_to_anchor=(2.18, 1)) 
        #plt.axis([-0.2, 7.7, 0.9*Right, 1.02*Left])
        plt.axis([-0.2, LLeft+bw/2, 0, 1.02*Left])
    
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), 'Cascade' + title +'.png'), dpi=150, bbox_inches='tight')

#
#
#
#
# The end.
#
#    