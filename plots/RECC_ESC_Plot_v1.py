# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 13:54:19 2023

@author: spauliuk

This script loads previously compiled results and then compiles selected results 
into different visualisations of the energy service cascade.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select.xlxs

Documentation and how to in RECCv2.5_EXPORT_Combine_Select.xlxs
"""

import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
import numpy as np
import RECC_Paths # Import path file

def get_esc_data_from_pandas(ps,selectI,selectR,cscenss):
    # return numpy array with ESC data for given single indicator, region, and scenario list
    pst     = ps[ps['Indicator'].isin([selectI]) & ps['Region'].isin([selectR]) & ps['Scenario'].isin([cscenss])] # Select the specified data and transpose them for plotting
    unit    = pst.iloc[0]['Unit']
    Data    = pst.drop(['Scenario','Indicator', 'Region', 'Sectors', 'Unit'], axis=1).values
    return Data, unit

plt.style.use('default') # set all plotting parameters to their default values

# Definitions/Specifications
CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')   
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value


scen = [] # list of target scenarios

# Read scenarios
r = 4
while True:
    if CF[CS].cell(r,1).value is None:
        break
    if CF[CS].cell(r,1).value == 1: # Only if the SELECT flag in col. A is set to 1.
        scen.append(CF[CS].cell(r,2).value)
    r += 1

r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define ESC plot':
        break    
    r += 1
r += 1

ctitles = []
ctypes  = []
cregs   = []
cscens  = []
colors  = [] # List with color strings

while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ctitles.append(CF[CS].cell(r,2).value)
    ctypes.append(CF[CS].cell(r,3).value)
    cregs.append(CF[CS].cell(r,4).value)
    cscens.append(CF[CS].cell(r,5).value)
    colors.append(CF[CS].cell(r,11).value)
    r += 1


# open data file with results
fn = os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '_sep.xlsx')
ps = pd.read_excel(fn, sheet_name='Results', index_col=0) # plot sheet

# determine ESC indicators and plot ESC cascades
# find ESC parameters in Res array: outer index: Ar, middle index: ti, inner indicator: scen

for c in range(0,len(ctitles)):
    if ctypes[c] == 'version_1':
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = cregs[c]
        
        # Define data container
        esc_data = np.zeros((6,46,nocs)) # 6 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Decoupling 1: Lower stock levels
        # Decoupling 2: Operational energy per stock, etc.
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Energy cons., use phase, res+non-res buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_rebx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, res. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_nrbx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, nonres. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Decoupling 3: Build-up material per stock
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Decoupling 4: circular material use rate
        Data_rec1 = np.concatenate([get_esc_data_from_pandas(ps,'ReUse of materials in products, construction grade steel',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_rec2 = np.concatenate([get_esc_data_from_pandas(ps,'ReUse of materials in products, concrete',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_rec3 = np.concatenate([get_esc_data_from_pandas(ps,'ReUse of materials in products, wood and wood products',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_rec4 = np.concatenate([get_esc_data_from_pandas(ps,'Secondary construction steel',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Decoupling 6: GHG per energy use
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg2 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, non-res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res+non-res buildings, energy supply',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Decoupling 5: material footprint per final consumption
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = (Data_rebx + Data_nrbx).transpose()
        esc_data[1,:,:] = (Data_edx  / (Data_rebx + Data_nrbx)).transpose()
        esc_data[2,:,:] = (Data_matm / (Data_rebx + Data_nrbx)).transpose() 
        esc_data[3,:,:] = ((Data_rec1 + Data_rec2 + Data_rec3 + Data_rec4)/Data_matm).transpose()
        esc_data[4,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[5,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_edx).transpose()
        
        # normalize esc data so that all time series are in relation to SSP2 (maximum):
        for mm in range(0,46):
            esc_data[0,mm,:] = esc_data[0,mm,:] / esc_data[0,mm,:].max()

        # normalize the esc_data
        esc_data[:,0,:] = 0 # start year is 2016, 2015 data are not considered.
        esc_data_Divisor = np.einsum('cs,t->cts',esc_data[:,1,:],np.ones(46))
        esc_data_n = np.divide(esc_data, esc_data_Divisor, out=np.zeros_like(esc_data_Divisor), where=esc_data_Divisor!=0)
        
        # Define colors
        cc = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=6 , figsize=(21, 3))        
        fig.suptitle('Energy service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", cc)
        
        axs[0].plot(np.arange(2016,2061), esc_data_n[0,1::,:],   linewidth = 2.3)
        plta = Line2D(np.arange(2016,2061), esc_data_n[0,1::,:], linewidth = 2.3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[0].set_title('(1) Stock per service')
        
        axs[1].plot(np.arange(2016,2061), esc_data_n[1,1::,:], linewidth = 2.3)  
        axs[1].set_title('(2) Operational energy per stock')
        
        axs[2].plot(np.arange(2016,2061), esc_data_n[2,1::,:], linewidth = 2.3) 
        axs[2].set_title('(3) Build-up material per stock')
        
        axs[3].plot(np.arange(2016,2061), esc_data_n[3,1::,:], linewidth = 2.3)  
        axs[3].set_title('(4) Circlar material use rate')
        
        axs[4].plot(np.arange(2016,2061), esc_data_n[4,1::,:], linewidth = 2.3)   
        axs[4].set_title('(5) material footprint per final consumption')        
        
        axs[5].plot(np.arange(2016,2061), esc_data_n[5,1::,:], linewidth = 2.3) 
        axs[5].set_title('(6) GHG emissions per energy use')
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=1, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '.png'), dpi=150, bbox_inches='tight')


    if ctypes[c] == 'version_2_blds': # Energy service cascade_GHG for buildings
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = cregs[c]
        
        # Define data container
        esc_data = np.zeros((10,46,nocs)) # 10 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg2 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, non-res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res+non-res buildings, energy supply',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Energy cons., use phase, res+non-res buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_rebx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, res. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_nrbx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, nonres. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_rebf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all res. building types together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_nrbf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all nonres. building types together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)])        
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx / (Data_rebx + Data_nrbx)).transpose() 
        esc_data[3,:,:] = ((Data_rebx + Data_nrbx)/Data_pop).transpose()
        esc_data[4,:,:] = ((Data_matm) / (Data_rebf + Data_nrbf)).transpose()
        esc_data[5,:,:] = (Data_ghg4 / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4 / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = ((Data_rebf + Data_nrbf)/(Data_rebx + Data_nrbx)).transpose()
        
        # Define maximal GHG/cap
        maxGHG = np.max(esc_data[[0,6],1::,:])
        
        # Define colors
        ccol = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple        
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=8 , figsize=(24, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", ccol)
        
        axs[0].plot(np.arange(2016,2061), esc_data[0,1::,:],   linewidth = 3)
        plta = Line2D(np.arange(2016,2061), esc_data[0,1::,:], linewidth = 3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[0].set_title('Energy-GHG per capita     = \n (Scope 1 + 2 emissions)      ', weight='bold')
        axs[0].set_ylabel('t CO2-eq/yr', fontsize = 12)
        #axs[0].set_facecolor((221/255, 235/255, 247/255))
        axs[0].set_facecolor((197/255, 221/255, 241/255))
        axs[0].set_ylim(bottom=0)
        axs[0].set_ylim(top=1.05 * maxGHG)
        
        axs[1].plot(np.arange(2016,2061), esc_data[1,1::,:] * 1e6, linewidth = 2.0)  
        axs[1].set_title('GHG per final energy     *')
        axs[1].set_ylabel('g CO2-eq/MJ', fontsize = 12)
        axs[1].set_facecolor((238/255, 245/255, 252/255))
        axs[1].set_ylim(bottom=0)
        
        axs[2].plot(np.arange(2016,2061), esc_data[2,1::,:], linewidth = 2.0) 
        axs[2].set_title('Final energy per stock     *')
        axs[2].set_ylabel('MJ/(m²·yr)', fontsize = 12)
        axs[2].set_facecolor((238/255, 245/255, 252/255))
        axs[2].set_ylim(bottom=0)
        
        axs[3].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[3].set_title(' <-- | --> \n Stock per capita', weight='bold')
        axs[3].set_ylabel('m²', fontsize = 12)
        axs[3].set_facecolor((237/255, 226/255, 246/255))  
        axs[3].set_ylim(bottom=0)
        
        axs[4].plot(np.arange(2016,2061), esc_data[9,1::,:], linewidth = 2.0)   
        axs[4].set_title('*     inflow per stock')        
        axs[4].set_ylabel('1/yr', fontsize = 12)
        axs[4].set_facecolor((253/255, 239/255, 231/255))            
        axs[4].set_ylim(bottom=0)
        
        axs[5].plot(np.arange(2016,2061), esc_data[4,1::,:] * 1e3, linewidth = 2.0)   
        axs[5].set_title('*     material intensity of inflow')        
        axs[5].set_ylabel('kg/m²', fontsize = 12)
        axs[5].set_facecolor((253/255, 239/255, 231/255))            
        axs[5].set_ylim(bottom=0)
        
        axs[6].plot(np.arange(2016,2061), esc_data[5,1::,:], linewidth = 2.0) 
        axs[6].set_title('*     material-GHG per material')
        axs[6].set_ylabel('t CO2-eq/t', fontsize = 12)
        axs[6].set_facecolor((253/255, 239/255, 231/255))            
        axs[6].set_ylim(bottom=0)
        
        axs[7].plot(np.arange(2016,2061), esc_data[6,1::,:], linewidth = 3.0) 
        axs[7].set_title('=     material-GHG per capita \n(Scope 3 emissions)', weight='bold')
        axs[7].set_ylabel('t CO2-eq/yr', fontsize = 12)
        #axs[7].set_facecolor((252/255, 228/255, 214/255))         
        axs[7].set_facecolor((249/255, 203/255, 177/255))   
        axs[7].set_ylim(bottom=0)
        axs[7].set_ylim(top=1.05 * maxGHG)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '_' + selectR + '.png'), dpi=150, bbox_inches='tight')


    if ctypes[c] == 'version_3_blds': # Energy service cascade_RMI for buildings
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = cregs[c]
        
        # Define data container
        esc_data = np.zeros((10,46,nocs)) # 10 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg2 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, non-res. buildings, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, res+non-res buildings, energy supply',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Energy cons., use phase, res+non-res buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_rebx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, res. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_nrbx = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, nonres. buildings',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_rebf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all res. building types together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_nrbf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all nonres. building types together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)])        
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx / (Data_rebx + Data_nrbx)).transpose() 
        esc_data[3,:,:] = ((Data_rebx + Data_nrbx)/Data_pop).transpose()
        esc_data[4,:,:] = ((Data_matm) / (Data_rebf + Data_nrbf)).transpose()
        esc_data[5,:,:] = (Data_ghg4 / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4 / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = ((Data_rebf + Data_nrbf)/(Data_rebx + Data_nrbx)).transpose()        
                
        # Define colors
        cc = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=5 , figsize=(16, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", cc)
                
        axs[0].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[0].set_title('Stock per capita', weight='bold')
        axs[0].set_ylabel('m²', fontsize = 12)
        axs[0].set_facecolor((237/255, 226/255, 246/255))  
        axs[0].set_ylim(bottom=0)
        
        axs[1].plot(np.arange(2016,2061), esc_data[9,1::,:], linewidth = 2.0)   
        axs[1].set_title('*     inflow per stock')        
        axs[1].set_ylabel('1/yr', fontsize = 12)
        axs[1].set_facecolor((253/255, 239/255, 231/255))            
        axs[1].set_ylim(bottom=0)

        axs[2].plot(np.arange(2016,2061), esc_data[4,1::,:] * 1e3, linewidth = 2.0)   
        axs[2].set_title('*     material intensity of inflow')        
        axs[2].set_ylabel('kg/m²', fontsize = 12)
        axs[2].set_facecolor((253/255, 239/255, 231/255))            
        axs[2].set_ylim(bottom=0)        
        
        axs[3].plot(np.arange(2016,2061), esc_data[7,1::,:], linewidth = 2.0) 
        axs[3].set_title('*     RMI per material')
        axs[3].set_ylabel('t/t', fontsize = 12)
        axs[3].set_facecolor((253/255, 239/255, 231/255))            
        axs[3].set_ylim(bottom=0)
        
        axs[4].plot(np.arange(2016,2061), esc_data[8,1::,:], linewidth = 3.0) 
        axs[4].set_title('=     RMI per capita', weight='bold')
        axs[4].set_ylabel('t/yr', fontsize = 12)
        #axs[4].set_facecolor((252/255, 228/255, 214/255))         
        axs[4].set_facecolor((249/255, 203/255, 177/255))         
        axs[4].set_ylim(bottom=0)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '_' + selectR + '.png'), dpi=150, bbox_inches='tight')
       
        
    if ctypes[c] == 'version_2_pav': # Energy service cascade_GHG for vehicles
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = cregs[c]
        
        # Define data container
        esc_data = np.zeros((12,46,nocs)) # 12 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Service
        Data_pkm = np.concatenate([get_esc_data_from_pandas(ps,'passenger-km supplied by pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_vkm = np.concatenate([get_esc_data_from_pandas(ps,'vehicle-km driven by pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, vehicles, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, vehicles, energy supply',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Energy cons., use phase, vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_pavs = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_pavc = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all drive technologies together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx   / Data_pavs).transpose() 
        esc_data[3,:,:] = (Data_pavs  / Data_vkm).transpose()
        esc_data[4,:,:] = (Data_matm  / Data_pavc).transpose()
        esc_data[5,:,:] = (Data_ghg4  / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4  / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = (Data_pavc  / Data_pavs).transpose()
        esc_data[10,:,:] = (Data_vkm  / Data_pkm).transpose()
        esc_data[11,:,:] = (Data_pkm  / Data_pop).transpose()
        
        # Define maximal GHG/cap
        maxGHG = np.max(esc_data[[0,6],1::,:])
        
        # Define colors
        ccol = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple        
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=6 , figsize=(18, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", ccol)
        
        axs[0].plot(np.arange(2016,2061), esc_data[0,1::,:],   linewidth = 3)
        plta = Line2D(np.arange(2016,2061), esc_data[0,1::,:], linewidth = 3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[0].set_title('Energy-GHG per capita     = \n (Scope 1 + 2 emissions)      ', weight='bold')
        axs[0].set_ylabel('t CO2-eq/yr', fontsize = 12)
        #axs[0].set_facecolor((221/255, 235/255, 247/255))
        axs[0].set_facecolor((197/255, 221/255, 241/255))
        axs[0].set_ylim(bottom=0)
        axs[0].set_ylim(top=1.05 * maxGHG)
        
        axs[1].plot(np.arange(2016,2061), esc_data[1,1::,:] * 1e6, linewidth = 2.0)  
        axs[1].set_title('GHG per final energy     *')
        axs[1].set_ylabel('g CO2-eq/MJ', fontsize = 12)
        axs[1].set_facecolor((238/255, 245/255, 252/255))
        axs[1].set_ylim(bottom=0)
        
        axs[2].plot(np.arange(2016,2061), esc_data[2,1::,:]/1000, linewidth = 2.0) 
        axs[2].set_title('Final energy per stock     *')
        axs[2].set_ylabel('GJ/(vehicle·yr)', fontsize = 12)
        axs[2].set_facecolor((238/255, 245/255, 252/255))
        axs[2].set_ylim(bottom=0)
        
        axs[3].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[3].set_title('vehicles per driven km     *')
        axs[3].set_ylabel('vehicle/vkm', fontsize = 12)
        axs[3].set_facecolor((238/255, 245/255, 252/255))  
        axs[3].set_ylim(bottom=0)
        
        axs[4].plot(np.arange(2016,2061), esc_data[10,1::,:], linewidth = 3.0)  
        axs[4].set_title('vehicle-km per passenger-km     *')
        axs[4].set_ylabel('vkm/pkm', fontsize = 12)
        axs[4].set_facecolor((238/255, 245/255, 252/255))  
        axs[4].set_ylim(bottom=0)

        axs[5].plot(np.arange(2016,2061), esc_data[11,1::,:], linewidth = 3.0)  
        axs[5].set_title('passenger-km per capita', weight='bold')
        axs[5].set_ylabel('pkm/cap', fontsize = 12)
        axs[5].set_facecolor((237/255, 226/255, 246/255))  
        axs[5].set_ylim(bottom=0)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '_' + selectR + '.png'), dpi=150, bbox_inches='tight')
        
    if ctypes[c] == 'version_3_pav': # Energy service cascade_RMI for vehicles
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = cregs[c]
        
        # Define data container
        esc_data = np.zeros((12,46,nocs)) # 10 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Service
        Data_pkm = np.concatenate([get_esc_data_from_pandas(ps,'passenger-km supplied by pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_vkm = np.concatenate([get_esc_data_from_pandas(ps,'vehicle-km driven by pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, vehicles, use phase',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, vehicles, energy supply',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Energy cons., use phase, vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_pavs = np.concatenate([get_esc_data_from_pandas(ps,'In-use stock, pass. vehicles',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_pavc = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all drive technologies together',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,cscenss[mscen])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx   / Data_pavs).transpose() 
        esc_data[3,:,:] = (Data_pavs  / Data_vkm).transpose()
        esc_data[4,:,:] = (Data_matm  / Data_pavc).transpose()
        esc_data[5,:,:] = (Data_ghg4  / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4  / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = (Data_pavc  / Data_pavs).transpose()
        esc_data[10,:,:] = (Data_vkm  / Data_pkm).transpose()
        esc_data[11,:,:] = (Data_pkm  / Data_pop).transpose()     
                
        # Define colors
        cc = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=7 , figsize=(21, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", cc)
                
        axs[0].plot(np.arange(2016,2061), esc_data[11,1::,:], linewidth = 3.0)  
        axs[0].set_title('passenger-km per capita', weight='bold')
        axs[0].set_ylabel('pkm/cap', fontsize = 12)
        axs[0].set_facecolor((237/255, 226/255, 246/255))  
        axs[0].set_ylim(bottom=0)
        
        axs[1].plot(np.arange(2016,2061), esc_data[10,1::,:], linewidth = 3.0)  
        axs[1].set_title('*     vehicle-km per passenger-km')
        axs[1].set_ylabel('vkm/pkm', fontsize = 12)
        axs[1].set_facecolor((253/255, 239/255, 231/255))  
        axs[1].set_ylim(bottom=0)

        axs[2].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[2].set_title('* vehicles per driven km')
        axs[2].set_ylabel('vehicle/vkm', fontsize = 12)
        axs[2].set_facecolor((253/255, 239/255, 231/255))  
        axs[2].set_ylim(bottom=0)
        
        axs[3].plot(np.arange(2016,2061), esc_data[9,1::,:], linewidth = 2.0)   
        axs[3].set_title('*     inflow per vehicle')        
        axs[3].set_ylabel('1/yr', fontsize = 12)
        axs[3].set_facecolor((253/255, 239/255, 231/255))            
        axs[3].set_ylim(bottom=0)

        axs[4].plot(np.arange(2016,2061), esc_data[4,1::,:], linewidth = 2.0)   
        axs[4].set_title('*     material intensity of inflow')        
        axs[4].set_ylabel('t/vehicle', fontsize = 12)
        axs[4].set_facecolor((253/255, 239/255, 231/255))            
        axs[4].set_ylim(bottom=0)        
        
        axs[5].plot(np.arange(2016,2061), esc_data[7,1::,:], linewidth = 2.0) 
        axs[5].set_title('*     RMI per material')
        axs[5].set_ylabel('t/t', fontsize = 12)
        axs[5].set_facecolor((253/255, 239/255, 231/255))            
        axs[5].set_ylim(bottom=0)
        
        axs[6].plot(np.arange(2016,2061), esc_data[8,1::,:], linewidth = 3.0) 
        axs[6].set_title('=     RMI per capita', weight='bold')
        axs[6].set_ylabel('t/yr', fontsize = 12)
        axs[6].set_facecolor((249/255, 203/255, 177/255))         
        axs[6].set_ylim(bottom=0)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), title + '_' + selectR + '.png'), dpi=150, bbox_inches='tight')
              
#
#
#
# The end.
#
#
#