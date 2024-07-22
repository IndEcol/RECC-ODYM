# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 13:54:19 2023

@author: spauliuk

This script loads previously compiled results and then compiles selected results 
into different visualisations, like line plots and bar charts.

Works together with control workbook
RECCv2.5_EXPORT_Combine_Select.xlxs

Need to run ODYM_RECC_Export_xlxs_Combine_Select.py first!

Documentation and how to in RECCv2.5_EXPORT_Combine_Select.xlxs

"""
import os
import plotnine
import openpyxl
from plotnine import *
import pandas as pd
import pylab
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import matplotlib.lines as mlines
from matplotlib.lines import Line2D
import numpy as np
import RECC_Paths # Import path file

plt.style.use('default') # set all plotting parameters to their default values

CP            = os.path.join(RECC_Paths.results_path,'RECCv2.5_EXPORT_Combine_Select.xlsx')   
CF            = openpyxl.load_workbook(CP)
CS            = CF['Cover'].cell(4,4).value
outpath       = CF[CS].cell(1,2).value
fn_add        = CF[CS].cell(1,4).value

r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define Plots':
        break    
    r += 1
r += 1

ptitles = []
ptypes  = []
pinds   = []
pregs   = []
pscens  = []
prange  = []
pflags  = []
indlab  = [] # indicator short labels for plots
scelab  = [] # scenario  short labels for plots
colors  = [] # List with color strings
while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ptitles.append(CF[CS].cell(r,2).value)
    ptypes.append(CF[CS].cell(r,3).value)
    pinds.append(CF[CS].cell(r,4).value)
    pregs.append(CF[CS].cell(r,5).value)
    pscens.append(CF[CS].cell(r,6).value)
    prange.append(CF[CS].cell(r,7).value)
    pflags.append(CF[CS].cell(r,8).value)
    indlab.append(CF[CS].cell(r,9).value)
    scelab.append(CF[CS].cell(r,10).value)
    colors.append(CF[CS].cell(r,11).value)
    r += 1

# open data file with results
fn = os.path.join(RECC_Paths.export_path,outpath,'Results_Extracted_RECCv2.5_' + fn_add + '_sep.xlsx')
ps = pd.read_excel(fn, sheet_name='Results', index_col=0) # plot sheet
pc = pd.read_excel(fn, sheet_name='Results_Cumulative', index_col=0) # plot sheet cumulative

regions = ['R5.2SSA','R5.2LAM','EU_UK','China','India','R5.2ASIA_Other','R5.2MNF','R5.2REF','R5.2OECD_Other','R32USACAN','Global']

# prepare Excel export:
book = openpyxl.Workbook() 

for m in range(0,len(ptitles)):
    if ptypes[m] == 'RECC_CRAFT_Sensitivity_1': # cumulative indicators
        senscaseno   = 5 # number of sensitivity plots
        corner_scens = ['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood']
        sens_scens   = [[['SSP2_Base_LowPop','SSP2_Wood_LowPop','LEMD_FullCE_LowPop','LEMD_FullCE_Wood_LowPop'],
                         ['SSP2_Base_HighPop','SSP2_Wood_HighPop','LEMD_FullCE_HighPop','LEMD_FullCE_Wood_HighPop']],
                        [['SSP2_Base_LowYield','SSP2_Wood_LowYield','LEMD_FullCE_LowYield','LEMD_FullCE_Wood_LowYield'],
                         ['SSP2_Base_HighYield','SSP2_Wood_HighYield','LEMD_FullCE_HighYield','LEMD_FullCE_Wood_HighYield']],
                        [['SSP2_Base_NoCascade','SSP2_Wood_NoCascade','LEMD_FullCE_NoCascade','LEMD_FullCE_Wood_NoCascade'],
                         ['SSP2_Base_MoreCascade','SSP2_Wood_MoreCascade','LEMD_FullCE_MoreCascade','LEMD_FullCE_Wood_MoreCascade']],
                        [['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood'],
                         ['SSP2_Base_LateCE','SSP2_Wood_LateCE','LEMD_FullCE_LateCE','LEMD_FullCE_Wood_LateCE']],
                        [['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood'],
                         ['SSP2_Base_FullCE','SSP2_Wood_FullCE','LEMD_Base','LEMD_Wood']]]
        cornsceno    = len(corner_scens) # number of corner scenarios
        # indicator: pinds[m]
        # region: pregs[m]
        cornerdata   = np.zeros((cornsceno)) # corner scenario data
        sensdata     = np.zeros((senscaseno,cornsceno,2)) # sensitivity cases x corner scenrios x high-low
        scelabs      = scelab[m].split(';')
        # fetch data
        for cs in range(0,cornsceno):
            pst    = pc[pc['Indicator'].isin([pinds[m]]) & pc['Region'].isin([pregs[m]]) & pc['Scenario'].isin([corner_scens[cs]])] # Select the specified data and compile them for plotting        
            cornerdata[cs] = pst.iloc[0][[prange[m]]]/1000 # Gt of C    
        for sc in range(0,senscaseno):
            for cs in range(0,cornsceno):
                for hl in range(0,2):
                    pst    = pc[pc['Indicator'].isin([pinds[m]]) & pc['Region'].isin([pregs[m]]) & pc['Scenario'].isin([sens_scens[sc][hl][cs]])] # Select the specified data and compile them for plotting        
                    sensdata[sc,cs,hl] = pst.iloc[0][[prange[m]]]/1000 # Gt of C    
                    
                    
        fig, axs = plt.subplots(nrows=5, ncols=1 , figsize=(4, 10), sharex=True)        
        plt.subplots_adjust(wspace=0, hspace=0)
        #fig.tight_layout(rect=[0, 0.03, 1, 0.85])
        fig.suptitle(indlab[m] + ', ' + pregs[m], fontsize=13)
        # plot corner scenarios
        for pi in range(0,senscaseno):
            for pj in range(0,cornsceno):
                axs[pi].barh(pj,cornerdata[pj], color = 'skyblue')
            axs[pi].set_ylim([-0.5,5])
            axs[pi].set_yticks([0,1,2,3])
            axs[pi].set_yticklabels(corner_scens, rotation =0, fontsize = 8, fontweight = 'normal')
            axs[pi].text(5, 4, scelabs[pi], fontsize=13, fontweight='normal', rotation = 0) 
        # plot sensitivity ranges
        for pi in range(0,senscaseno):
            for pj in range(0,cornsceno):
                # see details here: https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.errorbar.html
                xerr = np.zeros((2,1))
                x_min = min(cornerdata[pj], sensdata[pi,pj,0], sensdata[pi,pj,1])
                x_max = max(cornerdata[pj], sensdata[pi,pj,0], sensdata[pi,pj,1])
                xerr[0,0] = cornerdata[pj] - x_min
                xerr[1,0] = x_max - cornerdata[pj]
                axs[pi].errorbar(cornerdata[pj],pj, xerr=xerr, capsize=7, ecolor='black', ls='', lw=3, capthick=3, fmt='none')
        axs[4].set_xlabel(pflags[m], fontsize = 12)                    
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ptitles[m] +'.png'), dpi=150, bbox_inches='tight')
        #xlsx export
        ws = book.create_sheet(ptitles[m][0:20])
        ws.cell(row=1, column=1).value = indlab[m] + ', ' + pregs[m] # Title
        ws.cell(row=1, column=2).value = pflags[m]
        ws.cell(row=3, column=1).value = 'Scenario'
        for msce in range(0,4):
            ws.cell(row=msce+4, column=1).value = corner_scens[msce]     
        ws.cell(row=3, column=2).value = 'Base value of scenario'
        for msce in range(0,4):
            ws.cell(row=msce+4, column=2).value = cornerdata[msce]
        for msca in range(0,5):
            ws.cell(row=3, column=4+3*msca).value = scelabs[msca]
            for mscb in range(0,4):
                for mscc in range(0,2):
                    ws.cell(row=mscb+4, column=4+3*msca+mscc).value = sensdata[msca,mscb,mscc]
             
    if ptypes[m] == 'RECC_CRAFT_Sensitivity_2': # annual indicator
        senscaseno   = 5 # number of sensitivity plots
        corner_scens = ['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood']
        sens_scens   = [[['SSP2_Base_LowPop','SSP2_Wood_LowPop','LEMD_FullCE_LowPop','LEMD_FullCE_Wood_LowPop'],
                         ['SSP2_Base_HighPop','SSP2_Wood_HighPop','LEMD_FullCE_HighPop','LEMD_FullCE_Wood_HighPop']],
                        [['SSP2_Base_LowYield','SSP2_Wood_LowYield','LEMD_FullCE_LowYield','LEMD_FullCE_Wood_LowYield'],
                         ['SSP2_Base_HighYield','SSP2_Wood_HighYield','LEMD_FullCE_HighYield','LEMD_FullCE_Wood_HighYield']],
                        [['SSP2_Base_NoCascade','SSP2_Wood_NoCascade','LEMD_FullCE_NoCascade','LEMD_FullCE_Wood_NoCascade'],
                         ['SSP2_Base_MoreCascade','SSP2_Wood_MoreCascade','LEMD_FullCE_MoreCascade','LEMD_FullCE_Wood_MoreCascade']],
                        [['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood'],
                         ['SSP2_Base_LateCE','SSP2_Wood_LateCE','LEMD_FullCE_LateCE','LEMD_FullCE_Wood_LateCE']],
                        [['SSP2_Base','SSP2_Wood','LEMD_FullCE','LEMD_FullCE_Wood'],
                         ['SSP2_Base_FullCE','SSP2_Wood_FullCE','LEMD_Base','LEMD_Wood']]]
        cornsceno    = len(corner_scens) # number of corner scenarios
        # indicator: pinds[m]
        # region: pregs[m]
        cornerdata   = np.zeros((cornsceno)) # corner scenario data
        sensdata     = np.zeros((senscaseno,cornsceno,2)) # sensitivity cases x corner scenrios x high-low
        scelabs      = scelab[m].split(';')
        # fetch data
        for cs in range(0,cornsceno):
            pst    = ps[ps['Indicator'].isin([pinds[m]]) & ps['Region'].isin([pregs[m]]) & ps['Scenario'].isin([corner_scens[cs]])] # Select the specified data and compile them for plotting        
            cornerdata[cs] = pst.iloc[0][[prange[m]]]/1000 # Gt of C    
        for sc in range(0,senscaseno):
            for cs in range(0,cornsceno):
                for hl in range(0,2):
                    pst    = ps[ps['Indicator'].isin([pinds[m]]) & ps['Region'].isin([pregs[m]]) & ps['Scenario'].isin([sens_scens[sc][hl][cs]])] # Select the specified data and compile them for plotting        
                    sensdata[sc,cs,hl] = pst.iloc[0][[prange[m]]]/1000 # Gt of C    
                    
                    
        fig, axs = plt.subplots(nrows=5, ncols=1 , figsize=(4, 10), sharex=True)        
        plt.subplots_adjust(wspace=0, hspace=0)
        #fig.tight_layout(rect=[0, 0.03, 1, 0.85])
        fig.suptitle(indlab[m] + ', ' + pregs[m], fontsize=13)
        # plot corner scenarios
        for pi in range(0,senscaseno):
            for pj in range(0,cornsceno):
                axs[pi].barh(pj,cornerdata[pj], color = 'skyblue')
            axs[pi].set_ylim([-0.5,5])
            axs[pi].set_yticks([0,1,2,3])
            axs[pi].set_yticklabels(corner_scens, rotation =0, fontsize = 8, fontweight = 'normal')
            axs[pi].text(5, 4, scelabs[pi], fontsize=13, fontweight='normal', rotation = 0) 
        # plot sensitivity ranges
        for pi in range(0,senscaseno):
            for pj in range(0,cornsceno):
                # see details here: https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.errorbar.html
                xerr = np.zeros((2,1))
                x_min = min(cornerdata[pj], sensdata[pi,pj,0], sensdata[pi,pj,1])
                x_max = max(cornerdata[pj], sensdata[pi,pj,0], sensdata[pi,pj,1])
                xerr[0,0] = cornerdata[pj] - x_min
                xerr[1,0] = x_max - cornerdata[pj]
                axs[pi].errorbar(cornerdata[pj],pj, xerr=xerr, capsize=7, ecolor='black', ls='', lw=3, capthick=3, fmt='none')
        axs[4].set_xlabel(pflags[m], fontsize = 12)                    
        plt.show()
        fig.savefig(os.path.join(os.path.join(RECC_Paths.export_path,outpath), ptitles[m] +'.png'), dpi=150, bbox_inches='tight')
        #xlsx export
        ws = book.create_sheet(ptitles[m][0:20])
        ws.cell(row=1, column=1).value = indlab[m] + ', ' + pregs[m] # Title
        ws.cell(row=1, column=2).value = pflags[m]
        ws.cell(row=3, column=1).value = 'Scenario'
        for msce in range(0,4):
            ws.cell(row=msce+4, column=1).value = corner_scens[msce]     
        ws.cell(row=3, column=2).value = 'Base value of scenario'
        for msce in range(0,4):
            ws.cell(row=msce+4, column=2).value = cornerdata[msce]
        for msca in range(0,5):
            ws.cell(row=3, column=4+3*msca).value = scelabs[msca]
            for mscb in range(0,4):
                for mscc in range(0,2):
                    ws.cell(row=mscb+4, column=4+3*msca+mscc).value = sensdata[msca,mscb,mscc]
                                      
# Save plot data to xlsx:
book.save('RECC_CRAFT_Sensitivity.xlsx')                
                         
#
#
#
#
# The end.
#
#    